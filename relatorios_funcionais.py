"""
Sistema de Relatórios Funcionais - SIGE v3.1
Implementação completa dos relatórios com exportação PDF/Excel/CSV
"""

from flask import Blueprint, request, jsonify, make_response
from flask_login import login_required
from app import db
from models import *
from datetime import datetime, date
from sqlalchemy import func
from io import BytesIO
import csv
import io

# Criar Blueprint para relatórios
relatorios_bp = Blueprint('relatorios', __name__)

@relatorios_bp.route('/gerar/<tipo>', methods=['POST'])
@login_required
def gerar_relatorio(tipo):
    """Gerar relatórios dinâmicos com filtros"""
    filtros = request.get_json()
    
    # Processar filtros
    data_inicio = datetime.strptime(filtros.get('dataInicio', ''), '%Y-%m-%d').date() if filtros.get('dataInicio') else None
    data_fim = datetime.strptime(filtros.get('dataFim', ''), '%Y-%m-%d').date() if filtros.get('dataFim') else None
    obra_id = int(filtros.get('obra')) if filtros.get('obra') else None
    departamento_id = int(filtros.get('departamento')) if filtros.get('departamento') else None
    
    if tipo == 'funcionarios':
        return _relatorio_funcionarios(departamento_id)
    elif tipo == 'ponto':
        return _relatorio_ponto(data_inicio, data_fim, obra_id, departamento_id)
    elif tipo == 'horas-extras':
        return _relatorio_horas_extras(data_inicio, data_fim, obra_id, departamento_id)
    elif tipo == 'alimentacao':
        return _relatorio_alimentacao(data_inicio, data_fim, obra_id, departamento_id)
    elif tipo == 'obras':
        return _relatorio_obras(obra_id)
    elif tipo == 'custos-obra':
        return _relatorio_custos_obra(data_inicio, data_fim, obra_id)
    elif tipo == 'veiculos':
        return _relatorio_veiculos()
    elif tipo == 'dashboard-executivo':
        return _relatorio_dashboard_executivo()
    elif tipo == 'progresso-obras':
        return _relatorio_progresso_obras()
    elif tipo == 'rentabilidade':
        return _relatorio_rentabilidade()
    else:
        return jsonify({
            'titulo': 'Relatório não implementado',
            'html': '<div class="alert alert-info">Este tipo de relatório ainda não foi implementado.</div>'
        })

def _relatorio_funcionarios(departamento_id=None):
    """Relatório de funcionários"""
    query = Funcionario.query
    if departamento_id:
        query = query.filter_by(departamento_id=departamento_id)
    funcionarios = query.all()
    
    html = '<div class="table-responsive"><table class="table table-striped">'
    html += '<thead><tr><th>Código</th><th>Nome</th><th>CPF</th><th>Departamento</th><th>Função</th><th>Data Admissão</th><th>Salário</th><th>Status</th></tr></thead><tbody>'
    
    for f in funcionarios:
        status_badge = '<span class="badge bg-success">Ativo</span>' if f.ativo else '<span class="badge bg-danger">Inativo</span>'
        html += f'<tr><td>{f.codigo or "-"}</td><td>{f.nome}</td><td>{f.cpf}</td><td>{f.departamento_ref.nome if f.departamento_ref else "-"}</td>'
        html += f'<td>{f.funcao_ref.nome if f.funcao_ref else "-"}</td><td>{f.data_admissao.strftime("%d/%m/%Y") if f.data_admissao else "-"}</td>'
        html += f'<td>R$ {f.salario:,.2f}</td><td>{status_badge}</td></tr>'
    
    html += '</tbody></table></div>'
    
    return jsonify({
        'titulo': f'Lista de Funcionários ({len(funcionarios)} registros)',
        'html': html
    })

def _relatorio_ponto(data_inicio, data_fim, obra_id=None, departamento_id=None):
    """Relatório de registros de ponto"""
    query = RegistroPonto.query
    if data_inicio:
        query = query.filter(RegistroPonto.data >= data_inicio)
    if data_fim:
        query = query.filter(RegistroPonto.data <= data_fim)
    if obra_id:
        query = query.filter_by(obra_id=obra_id)
    if departamento_id:
        query = query.join(Funcionario).filter(Funcionario.departamento_id == departamento_id)
    
    registros = query.join(Funcionario).order_by(RegistroPonto.data.desc()).all()
    
    html = '<div class="table-responsive"><table class="table table-striped">'
    html += '<thead><tr><th>Data</th><th>Funcionário</th><th>Obra</th><th>Entrada</th><th>Saída</th><th>Horas Trabalhadas</th><th>Horas Extras</th><th>Atrasos</th></tr></thead><tbody>'
    
    for r in registros:
        html += f'<tr><td>{r.data.strftime("%d/%m/%Y")}</td>'
        html += f'<td>{r.funcionario.nome}</td>'
        html += f'<td>{r.obra.nome if r.obra else "-"}</td>'
        html += f'<td>{r.hora_entrada.strftime("%H:%M") if r.hora_entrada else "-"}</td>'
        html += f'<td>{r.hora_saida.strftime("%H:%M") if r.hora_saida else "-"}</td>'
        html += f'<td>{r.horas_trabalhadas:.2f}h</td>'
        html += f'<td>{r.horas_extras:.2f}h</td>'
        html += f'<td>{r.total_atraso_minutos or 0} min</td></tr>'
    
    html += '</tbody></table></div>'
    
    return jsonify({
        'titulo': f'Relatório de Ponto ({len(registros)} registros)',
        'html': html
    })

def _relatorio_horas_extras(data_inicio, data_fim, obra_id=None, departamento_id=None):
    """Relatório de horas extras"""
    query = RegistroPonto.query.filter(RegistroPonto.horas_extras > 0)
    if data_inicio:
        query = query.filter(RegistroPonto.data >= data_inicio)
    if data_fim:
        query = query.filter(RegistroPonto.data <= data_fim)
    if obra_id:
        query = query.filter_by(obra_id=obra_id)
    if departamento_id:
        query = query.join(Funcionario).filter(Funcionario.departamento_id == departamento_id)
    
    registros = query.join(Funcionario).order_by(RegistroPonto.data.desc()).all()
    
    html = '<div class="table-responsive"><table class="table table-striped">'
    html += '<thead><tr><th>Data</th><th>Funcionário</th><th>Obra</th><th>Horas Extras</th><th>Valor Estimado</th></tr></thead><tbody>'
    
    total_horas = 0
    total_valor = 0
    
    for r in registros:
        valor_hora = (r.funcionario.salario / 220) * 1.5 if r.funcionario.salario else 0
        valor_extras = r.horas_extras * valor_hora
        total_horas += r.horas_extras
        total_valor += valor_extras
        
        html += f'<tr><td>{r.data.strftime("%d/%m/%Y")}</td>'
        html += f'<td>{r.funcionario.nome}</td>'
        html += f'<td>{r.obra.nome if r.obra else "-"}</td>'
        html += f'<td>{r.horas_extras:.2f}h</td>'
        html += f'<td>R$ {valor_extras:.2f}</td></tr>'
    
    html += f'<tr class="table-info"><td colspan="3"><strong>TOTAL</strong></td><td><strong>{total_horas:.2f}h</strong></td><td><strong>R$ {total_valor:.2f}</strong></td></tr>'
    html += '</tbody></table></div>'
    
    return jsonify({
        'titulo': f'Relatório de Horas Extras ({len(registros)} registros)',
        'html': html
    })

def _relatorio_alimentacao(data_inicio, data_fim, obra_id=None, departamento_id=None):
    """Relatório de alimentação"""
    query = RegistroAlimentacao.query
    if data_inicio:
        query = query.filter(RegistroAlimentacao.data >= data_inicio)
    if data_fim:
        query = query.filter(RegistroAlimentacao.data <= data_fim)
    if obra_id:
        query = query.filter_by(obra_id=obra_id)
    if departamento_id:
        query = query.join(Funcionario).filter(Funcionario.departamento_id == departamento_id)
    
    registros = query.join(Funcionario).order_by(RegistroAlimentacao.data.desc()).all()
    
    html = '<div class="table-responsive"><table class="table table-striped">'
    html += '<thead><tr><th>Data</th><th>Funcionário</th><th>Tipo</th><th>Restaurante</th><th>Obra</th><th>Valor</th></tr></thead><tbody>'
    
    total_valor = 0
    
    for r in registros:
        total_valor += r.valor
        html += f'<tr><td>{r.data.strftime("%d/%m/%Y")}</td>'
        html += f'<td>{r.funcionario.nome}</td>'
        html += f'<td>{r.tipo.title()}</td>'
        html += f'<td>{r.restaurante.nome if r.restaurante else "-"}</td>'
        html += f'<td>{r.obra.nome if r.obra else "-"}</td>'
        html += f'<td>R$ {r.valor:.2f}</td></tr>'
    
    html += f'<tr class="table-info"><td colspan="5"><strong>TOTAL</strong></td><td><strong>R$ {total_valor:.2f}</strong></td></tr>'
    html += '</tbody></table></div>'
    
    return jsonify({
        'titulo': f'Relatório de Alimentação ({len(registros)} registros)',
        'html': html
    })

def _relatorio_obras(obra_id=None):
    """Relatório de obras"""
    query = Obra.query
    if obra_id:
        query = query.filter_by(id=obra_id)
    
    obras = query.all()
    
    html = '<div class="table-responsive"><table class="table table-striped">'
    html += '<thead><tr><th>Nome</th><th>Responsável</th><th>Data Início</th><th>Previsão Fim</th><th>Orçamento</th><th>Status</th><th>Funcionários</th></tr></thead><tbody>'
    
    for obra in obras:
        funcionarios_obra = db.session.query(func.count(FuncionarioObra.id)).filter_by(obra_id=obra.id).scalar() or 0
        
        html += f'<tr><td>{obra.nome}</td>'
        html += f'<td>{obra.responsavel.nome if obra.responsavel else "-"}</td>'
        html += f'<td>{obra.data_inicio.strftime("%d/%m/%Y") if obra.data_inicio else "-"}</td>'
        html += f'<td>{obra.data_previsao_fim.strftime("%d/%m/%Y") if obra.data_previsao_fim else "-"}</td>'
        html += f'<td>R$ {obra.orcamento:,.2f}</td>'
        html += f'<td><span class="badge bg-primary">{obra.status}</span></td>'
        html += f'<td>{funcionarios_obra}</td></tr>'
    
    html += '</tbody></table></div>'
    
    return jsonify({
        'titulo': f'Lista de Obras ({len(obras)} registros)',
        'html': html
    })

def _relatorio_custos_obra(data_inicio, data_fim, obra_id=None):
    """Relatório de custos por obra"""
    query = CustoObra.query
    if data_inicio:
        query = query.filter(CustoObra.data >= data_inicio)
    if data_fim:
        query = query.filter(CustoObra.data <= data_fim)
    if obra_id:
        query = query.filter_by(obra_id=obra_id)
    
    custos = query.join(Obra).order_by(CustoObra.data.desc()).all()
    
    html = '<div class="table-responsive"><table class="table table-striped">'
    html += '<thead><tr><th>Data</th><th>Obra</th><th>Tipo</th><th>Descrição</th><th>Valor</th></tr></thead><tbody>'
    
    total_custos = 0
    
    for custo in custos:
        total_custos += custo.valor
        html += f'<tr><td>{custo.data.strftime("%d/%m/%Y")}</td>'
        html += f'<td>{custo.obra.nome}</td>'
        html += f'<td>{custo.tipo.title()}</td>'
        html += f'<td>{custo.descricao or "-"}</td>'
        html += f'<td>R$ {custo.valor:.2f}</td></tr>'
    
    html += f'<tr class="table-info"><td colspan="4"><strong>TOTAL</strong></td><td><strong>R$ {total_custos:.2f}</strong></td></tr>'
    html += '</tbody></table></div>'
    
    return jsonify({
        'titulo': f'Custos por Obra ({len(custos)} registros)',
        'html': html
    })

def _relatorio_veiculos():
    """Relatório de veículos"""
    veiculos = Veiculo.query.all()
    
    html = '<div class="table-responsive"><table class="table table-striped">'
    html += '<thead><tr><th>Placa</th><th>Marca/Modelo</th><th>Ano</th><th>Tipo</th><th>KM Atual</th><th>Status</th><th>Próxima Manutenção</th></tr></thead><tbody>'
    
    for veiculo in veiculos:
        html += f'<tr><td>{veiculo.placa}</td>'
        html += f'<td>{veiculo.marca} {veiculo.modelo}</td>'
        html += f'<td>{veiculo.ano or "-"}</td>'
        html += f'<td>{veiculo.tipo}</td>'
        html += f'<td>{veiculo.km_atual or 0:,} km</td>'
        html += f'<td><span class="badge bg-info">{veiculo.status}</span></td>'
        html += f'<td>{veiculo.data_proxima_manutencao.strftime("%d/%m/%Y") if veiculo.data_proxima_manutencao else "-"}</td></tr>'
    
    html += '</tbody></table></div>'
    
    return jsonify({
        'titulo': f'Relatório de Veículos ({len(veiculos)} registros)',
        'html': html
    })

def _relatorio_dashboard_executivo():
    """Dashboard executivo"""
    total_funcionarios = Funcionario.query.filter_by(ativo=True).count()
    total_obras = Obra.query.filter_by(status='Em andamento').count()
    total_veiculos = Veiculo.query.count()
    
    # Custos do mês atual
    hoje = date.today()
    primeiro_dia_mes = hoje.replace(day=1)
    
    custos_mes = db.session.query(func.sum(CustoObra.valor)).filter(
        CustoObra.data >= primeiro_dia_mes,
        CustoObra.data <= hoje
    ).scalar() or 0
    
    # Horas trabalhadas do mês
    horas_mes = db.session.query(func.sum(RegistroPonto.horas_trabalhadas)).filter(
        RegistroPonto.data >= primeiro_dia_mes,
        RegistroPonto.data <= hoje
    ).scalar() or 0
    
    # Alimentação do mês
    alimentacao_mes = db.session.query(func.sum(RegistroAlimentacao.valor)).filter(
        RegistroAlimentacao.data >= primeiro_dia_mes,
        RegistroAlimentacao.data <= hoje
    ).scalar() or 0
    
    html = '<div class="row">'
    html += f'<div class="col-md-3"><div class="card text-center"><div class="card-body"><h2 class="text-primary">{total_funcionarios}</h2><p>Funcionários Ativos</p></div></div></div>'
    html += f'<div class="col-md-3"><div class="card text-center"><div class="card-body"><h2 class="text-success">{total_obras}</h2><p>Obras em Andamento</p></div></div></div>'
    html += f'<div class="col-md-3"><div class="card text-center"><div class="card-body"><h2 class="text-warning">{total_veiculos}</h2><p>Veículos</p></div></div></div>'
    html += f'<div class="col-md-3"><div class="card text-center"><div class="card-body"><h2 class="text-info">R$ {custos_mes:,.0f}</h2><p>Custos do Mês</p></div></div></div>'
    html += '</div>'
    
    html += '<div class="row mt-4">'
    html += f'<div class="col-md-6"><div class="card"><div class="card-body"><h5>Resumo Mensal</h5><p><strong>Horas Trabalhadas:</strong> {horas_mes:.0f}h</p><p><strong>Custo por Hora:</strong> R$ {(custos_mes/horas_mes if horas_mes > 0 else 0):.2f}</p></div></div></div>'
    html += f'<div class="col-md-6"><div class="card"><div class="card-body"><h5>Alimentação</h5><p><strong>Gasto Mensal:</strong> R$ {alimentacao_mes:,.2f}</p><p><strong>Média por Funcionário:</strong> R$ {(alimentacao_mes/total_funcionarios if total_funcionarios > 0 else 0):.2f}</p></div></div></div>'
    html += '</div>'
    
    return jsonify({
        'titulo': 'Dashboard Executivo',
        'html': html
    })

def _relatorio_progresso_obras():
    """Relatório de progresso das obras"""
    obras = Obra.query.all()
    
    html = '<div class="table-responsive"><table class="table table-striped">'
    html += '<thead><tr><th>Obra</th><th>Progresso</th><th>Orçamento</th><th>Gasto Atual</th><th>Saldo</th><th>% Utilizado</th></tr></thead><tbody>'
    
    for obra in obras:
        gasto_atual = db.session.query(func.sum(CustoObra.valor)).filter_by(obra_id=obra.id).scalar() or 0
        saldo = obra.orcamento - gasto_atual
        percentual = (gasto_atual / obra.orcamento * 100) if obra.orcamento > 0 else 0
        
        cor_saldo = 'text-success' if saldo > 0 else 'text-danger'
        
        html += f'<tr><td>{obra.nome}</td>'
        html += f'<td><span class="badge bg-info">{obra.status}</span></td>'
        html += f'<td>R$ {obra.orcamento:,.2f}</td>'
        html += f'<td>R$ {gasto_atual:,.2f}</td>'
        html += f'<td class="{cor_saldo}">R$ {saldo:,.2f}</td>'
        html += f'<td>{percentual:.1f}%</td></tr>'
    
    html += '</tbody></table></div>'
    
    return jsonify({
        'titulo': f'Progresso das Obras ({len(obras)} registros)',
        'html': html
    })

def _relatorio_rentabilidade():
    """Análise de rentabilidade por obra"""
    obras = Obra.query.all()
    
    html = '<div class="table-responsive"><table class="table table-striped">'
    html += '<thead><tr><th>Obra</th><th>Receita Prevista</th><th>Custos</th><th>Margem</th><th>% Margem</th><th>Status</th></tr></thead><tbody>'
    
    for obra in obras:
        custos_obra = db.session.query(func.sum(CustoObra.valor)).filter_by(obra_id=obra.id).scalar() or 0
        receita_prevista = obra.orcamento * 1.3  # Assumindo 30% de margem padrão
        margem = receita_prevista - custos_obra
        percentual_margem = (margem / receita_prevista * 100) if receita_prevista > 0 else 0
        
        cor_margem = 'text-success' if margem > 0 else 'text-danger'
        
        html += f'<tr><td>{obra.nome}</td>'
        html += f'<td>R$ {receita_prevista:,.2f}</td>'
        html += f'<td>R$ {custos_obra:,.2f}</td>'
        html += f'<td class="{cor_margem}">R$ {margem:,.2f}</td>'
        html += f'<td class="{cor_margem}">{percentual_margem:.1f}%</td>'
        html += f'<td><span class="badge bg-primary">{obra.status}</span></td></tr>'
    
    html += '</tbody></table></div>'
    
    return jsonify({
        'titulo': f'Análise de Rentabilidade ({len(obras)} obras)',
        'html': html
    })

@relatorios_bp.route('/exportar/<formato>', methods=['POST'])
@login_required
def exportar_relatorio(formato):
    """Exportar relatórios em diferentes formatos"""
    # Processar filtros
    data_inicio = request.form.get('dataInicio')
    data_fim = request.form.get('dataFim')
    obra_id = request.form.get('obra')
    departamento_id = request.form.get('departamento')
    
    # Converter datas se fornecidas
    if data_inicio:
        data_inicio = datetime.strptime(data_inicio, '%Y-%m-%d').date()
    if data_fim:
        data_fim = datetime.strptime(data_fim, '%Y-%m-%d').date()
    
    if formato == 'csv':
        return _exportar_csv(data_inicio, data_fim, obra_id, departamento_id)
    elif formato == 'excel':
        return _exportar_excel(data_inicio, data_fim, obra_id, departamento_id)
    elif formato == 'pdf':
        return _exportar_pdf(data_inicio, data_fim, obra_id, departamento_id)
    else:
        return jsonify({'error': 'Formato não suportado'}), 400

def _exportar_csv(data_inicio, data_fim, obra_id, departamento_id):
    """Exportar relatório consolidado em CSV"""
    output = BytesIO()
    output.write('\ufeff'.encode('utf-8'))  # BOM for UTF-8
    
    csv_content = io.StringIO()
    writer = csv.writer(csv_content)
    
    # Cabeçalho principal
    writer.writerow(['RELATÓRIO CONSOLIDADO SIGE - SISTEMA INTEGRADO DE GESTÃO EMPRESARIAL'])
    writer.writerow(['Data de Geração:', datetime.now().strftime('%d/%m/%Y %H:%M')])
    writer.writerow(['Período:', f'{data_inicio.strftime("%d/%m/%Y") if data_inicio else "N/A"} até {data_fim.strftime("%d/%m/%Y") if data_fim else "N/A"}'])
    writer.writerow([])
    
    # Relatório de Funcionários
    writer.writerow(['=== FUNCIONÁRIOS ==='])
    writer.writerow(['Código', 'Nome', 'CPF', 'Departamento', 'Função', 'Data Admissão', 'Salário', 'Status'])
    
    query = Funcionario.query
    if departamento_id:
        query = query.filter_by(departamento_id=departamento_id)
    funcionarios = query.all()
    
    for f in funcionarios:
        writer.writerow([
            f.codigo or '',
            f.nome,
            f.cpf,
            f.departamento_ref.nome if f.departamento_ref else '',
            f.funcao_ref.nome if f.funcao_ref else '',
            f.data_admissao.strftime('%d/%m/%Y') if f.data_admissao else '',
            f'R$ {f.salario:,.2f}',
            'Ativo' if f.ativo else 'Inativo'
        ])
    
    # Finalizar CSV
    output.write(csv_content.getvalue().encode('utf-8'))
    output.seek(0)
    
    response = make_response(output.getvalue())
    response.headers["Content-Disposition"] = f"attachment; filename=relatorio_sige_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    response.headers["Content-type"] = "text/csv; charset=utf-8"
    
    return response

def _exportar_excel(data_inicio, data_fim, obra_id, departamento_id):
    """Exportar relatório em Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório SIGE"
    
    # Cabeçalho
    headers = ['Código', 'Nome', 'CPF', 'Departamento', 'Função', 'Data Admissão', 'Salário', 'Status']
    ws.append(headers)
    
    # Estilizar cabeçalho
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Dados dos funcionários
    query = Funcionario.query
    if departamento_id:
        query = query.filter_by(departamento_id=departamento_id)
    funcionarios = query.all()
    
    for f in funcionarios:
        ws.append([
            f.codigo or '',
            f.nome,
            f.cpf,
            f.departamento_ref.nome if f.departamento_ref else '',
            f.funcao_ref.nome if f.funcao_ref else '',
            f.data_admissao.strftime('%d/%m/%Y') if f.data_admissao else '',
            f.salario,
            'Ativo' if f.ativo else 'Inativo'
        ])
    
    # Salvar Excel
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = make_response(output.getvalue())
    response.headers["Content-Disposition"] = f"attachment; filename=relatorio_sige_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response.headers["Content-type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    
    return response

def _exportar_pdf(data_inicio, data_fim, obra_id, departamento_id):
    """Exportar relatório em PDF"""
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.navy,
        spaceAfter=30,
        alignment=1  # Center
    )
    
    # Elementos do PDF
    elements = []
    
    # Título
    title = Paragraph("RELATÓRIO CONSOLIDADO SIGE", title_style)
    subtitle = Paragraph("Sistema Integrado de Gestão Empresarial - Estruturas do Vale", styles['Normal'])
    
    elements.append(title)
    elements.append(subtitle)
    elements.append(Spacer(1, 20))
    
    # Dados dos funcionários (primeiros 20)
    query = Funcionario.query
    if departamento_id:
        query = query.filter_by(departamento_id=departamento_id)
    funcionarios = query.all()
    
    func_data = [['Código', 'Nome', 'Departamento', 'Status']]
    for f in funcionarios[:20]:  # Limitar para evitar PDF muito grande
        func_data.append([
            f.codigo or '-',
            f.nome[:25] + '...' if len(f.nome) > 25 else f.nome,
            (f.departamento_ref.nome[:15] + '...') if f.departamento_ref and len(f.departamento_ref.nome) > 15 else (f.departamento_ref.nome if f.departamento_ref else '-'),
            'Ativo' if f.ativo else 'Inativo'
        ])
    
    func_table = Table(func_data, colWidths=[1*inch, 3*inch, 2*inch, 1*inch])
    func_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
    ]))
    
    elements.append(func_table)
    
    # Construir PDF
    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()
    
    response = make_response(pdf)
    response.headers["Content-Disposition"] = f"attachment; filename=relatorio_sige_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    response.headers["Content-Type"] = "application/pdf"
    return response

def gerar_relatorio_funcional(tipo, formato, filtros=None):
    """
    Função principal para gerar relatórios funcionais com exportação
    Suporta formatos: csv, excel, pdf
    """
    if filtros is None:
        filtros = {}
    
    # Processar filtros
    data_inicio = datetime.strptime(filtros.get('dataInicio', ''), '%Y-%m-%d').date() if filtros.get('dataInicio') else None
    data_fim = datetime.strptime(filtros.get('dataFim', ''), '%Y-%m-%d').date() if filtros.get('dataFim') else None
    obra_id = int(filtros.get('obra')) if filtros.get('obra') else None
    departamento_id = int(filtros.get('departamento')) if filtros.get('departamento') else None
    
    # Gerar dados do relatório
    dados = []
    nome_arquivo = f"relatorio_{tipo}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    
    if tipo == 'funcionarios':
        query = Funcionario.query
        if departamento_id:
            query = query.filter_by(departamento_id=departamento_id)
        funcionarios = query.all()
        
        for func in funcionarios:
            dados.append({
                'Código': func.codigo or '',
                'Nome': func.nome,
                'CPF': func.cpf,
                'Departamento': func.departamento_ref.nome if func.departamento_ref else '',
                'Função': func.funcao_ref.nome if func.funcao_ref else '',
                'Data Admissão': func.data_admissao.strftime('%d/%m/%Y') if func.data_admissao else '',
                'Salário': f'R$ {func.salario:,.2f}' if func.salario else '',
                'Status': 'Ativo' if func.ativo else 'Inativo'
            })
    
    elif tipo == 'ponto':
        query = RegistroPonto.query
        if data_inicio:
            query = query.filter(RegistroPonto.data >= data_inicio)
        if data_fim:
            query = query.filter(RegistroPonto.data <= data_fim)
        if departamento_id:
            query = query.join(Funcionario).filter(Funcionario.departamento_id == departamento_id)
        
        registros = query.all()
        for reg in registros:
            dados.append({
                'Data': reg.data.strftime('%d/%m/%Y'),
                'Funcionário': reg.funcionario_ref.nome,
                'Entrada': reg.hora_entrada.strftime('%H:%M') if reg.hora_entrada else '',
                'Saída': reg.hora_saida.strftime('%H:%M') if reg.hora_saida else '',
                'Horas Trabalhadas': f"{reg.horas_trabalhadas:.2f}" if reg.horas_trabalhadas else '0.00',
                'Horas Extras': f"{reg.horas_extras:.2f}" if reg.horas_extras else '0.00',
                'Atraso (min)': str(reg.total_atraso_minutos or 0),
                'Observações': reg.observacoes or ''
            })
    
    # Gerar arquivo baseado no formato
    if formato == 'csv':
        return _gerar_csv_export(dados, nome_arquivo)
    elif formato == 'excel':
        return _gerar_excel_export(dados, nome_arquivo)
    elif formato == 'pdf':
        return _gerar_pdf_export(dados, nome_arquivo, tipo)
    else:
        return jsonify({'erro': 'Formato não suportado'}), 400

def _gerar_csv_export(dados, nome_arquivo):
    """Gera arquivo CSV"""
    if not dados:
        return jsonify({'erro': 'Nenhum dado encontrado'}), 400
    
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=dados[0].keys())
    writer.writeheader()
    writer.writerows(dados)
    
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'text/csv'
    response.headers['Content-Disposition'] = f'attachment; filename="{nome_arquivo}.csv"'
    return response

def _gerar_excel_export(dados, nome_arquivo):
    """Gera arquivo Excel usando openpyxl"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        
        if not dados:
            return jsonify({'erro': 'Nenhum dado encontrado'}), 400
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Relatório"
        
        # Cabeçalhos
        headers = list(dados[0].keys())
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Dados
        for row, item in enumerate(dados, 2):
            for col, header in enumerate(headers, 1):
                ws.cell(row=row, column=col, value=item[header])
        
        # Salvar em BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = make_response(output.read())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename="{nome_arquivo}.xlsx"'
        return response
        
    except ImportError:
        return jsonify({'erro': 'Biblioteca openpyxl não disponível'}), 500

def _gerar_pdf_export(dados, nome_arquivo, tipo):
    """Gera arquivo PDF usando reportlab"""
    try:
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        
        if not dados:
            return jsonify({'erro': 'Nenhum dado encontrado'}), 400
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        
        # Título
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=1  # Center
        )
        title = Paragraph(f"Relatório - {tipo.replace('-', ' ').title()}", title_style)
        
        # Preparar dados da tabela
        table_data = []
        headers = list(dados[0].keys())
        table_data.append(headers)
        
        for item in dados:
            row = [str(item[header]) for header in headers]
            table_data.append(row)
        
        # Criar tabela
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
        ]))
        
        # Construir documento
        elements = [title, Spacer(1, 12), table]
        doc.build(elements)
        
        buffer.seek(0)
        response = make_response(buffer.read())
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename="{nome_arquivo}.pdf"'
        return response
        
    except ImportError:
        return jsonify({'erro': 'Biblioteca reportlab não disponível'}), 500
    
    return response