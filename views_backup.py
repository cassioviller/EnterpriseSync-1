from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify
from flask_login import login_required, current_user
from app import db
from models import *
from forms import *
from utils import calcular_horas_trabalhadas, calcular_custo_real_obra, calcular_custos_mes, calcular_kpis_funcionarios_geral, calcular_kpis_funcionario_periodo, calcular_kpis_funcionario_completo, calcular_ocorrencias_funcionario, processar_meio_periodo_exemplo
from datetime import datetime, date
from sqlalchemy import func
from kpis_engine_simple import kpis_engine
import os
from werkzeug.utils import secure_filename

main_bp = Blueprint('main', __name__)

# Rotas adicionais de veículos serão adicionadas diretamente aqui

@main_bp.route('/')
@login_required
def dashboard():
    # Filtros de data dos parâmetros
    data_inicio = request.args.get('data_inicio')
    data_fim = request.args.get('data_fim')
    
    # Definir período padrão (mês atual)
    if not data_inicio:
        data_inicio = date.today().replace(day=1)
    else:
        data_inicio = datetime.strptime(data_inicio, '%Y-%m-%d').date()
    
    if not data_fim:
        data_fim = date.today()
    else:
        data_fim = datetime.strptime(data_fim, '%Y-%m-%d').date()
    
    # KPIs para o dashboard
    total_funcionarios = Funcionario.query.filter_by(ativo=True).count()
    total_obras = Obra.query.filter(Obra.status.in_(['Em andamento', 'Pausada'])).count()
    total_veiculos = Veiculo.query.count()
    
    # Custos do período usando função corrigida
    custos_detalhados = calcular_custos_mes(data_inicio, data_fim)
    custos_mes = custos_detalhados['total']
    
    # Obras em andamento
    obras_andamento = Obra.query.filter_by(status='Em andamento').limit(5).all()
    
    # Funcionários por departamento
    funcionarios_dept = db.session.query(
        Departamento.nome,
        func.count(Funcionario.id).label('total')
    ).join(Funcionario).filter(Funcionario.ativo == True).group_by(Departamento.nome).all()
    
    # Custos por obra no período
    custos_recentes = []
    for obra in Obra.query.limit(10).all():
        custo_obra = calcular_custo_real_obra(obra.id, data_inicio, data_fim)
        if custo_obra['custo_total'] > 0:
            custos_recentes.append({
                'nome': obra.nome,
                'total_custo': custo_obra['custo_total']
            })
    
    return render_template('dashboard.html',
                         total_funcionarios=total_funcionarios,
                         total_obras=total_obras,
                         total_veiculos=total_veiculos,
                         custos_mes=custos_mes,
                         custos_detalhados=custos_detalhados,
                         data_inicio=data_inicio,
                         data_fim=data_fim,
                         obras_andamento=obras_andamento,
                         funcionarios_dept=funcionarios_dept,
                         custos_recentes=custos_recentes)

# Funcionários
@main_bp.route('/funcionarios')
@login_required
def funcionarios():
    # Filtros de data dos parâmetros
    data_inicio = request.args.get('data_inicio')
    data_fim = request.args.get('data_fim')
    
    # Definir período padrão (mês atual)
    if not data_inicio:
        data_inicio = date.today().replace(day=1)
    else:
        data_inicio = datetime.strptime(data_inicio, '%Y-%m-%d').date()
    
    if not data_fim:
        data_fim = date.today()
    else:
        data_fim = datetime.strptime(data_fim, '%Y-%m-%d').date()
    
    # Calcular KPIs gerais dos funcionários para o período
    kpis_geral = calcular_kpis_funcionarios_geral(data_inicio, data_fim)
    
    return render_template('funcionarios.html', 
                         funcionarios_kpis=kpis_geral['funcionarios_kpis'],
                         kpis_geral=kpis_geral,
                         data_inicio=data_inicio,
                         data_fim=data_fim)

@main_bp.route('/funcionarios/novo', methods=['GET', 'POST'])
@login_required
def novo_funcionario():
    form = FuncionarioForm()
    form.departamento_id.choices = [(0, 'Selecione...')] + [(d.id, d.nome) for d in Departamento.query.all()]
    form.funcao_id.choices = [(0, 'Selecione...')] + [(f.id, f.nome) for f in Funcao.query.all()]
    form.horario_trabalho_id.choices = [(0, 'Selecione...')] + [(h.id, h.nome) for h in HorarioTrabalho.query.all()]
    
    if form.validate_on_submit():
        funcionario = Funcionario(
            nome=form.nome.data,
            cpf=form.cpf.data,
            rg=form.rg.data,
            data_nascimento=form.data_nascimento.data,
            endereco=form.endereco.data,
            telefone=form.telefone.data,
            email=form.email.data,
            data_admissao=form.data_admissao.data,
            salario=form.salario.data or 0.0,
            departamento_id=form.departamento_id.data if form.departamento_id.data > 0 else None,
            funcao_id=form.funcao_id.data if form.funcao_id.data > 0 else None,
            horario_trabalho_id=form.horario_trabalho_id.data if form.horario_trabalho_id.data > 0 else None,
            ativo=form.ativo.data
        )
        db.session.add(funcionario)
        db.session.flush()  # Para obter o ID antes do commit
        
        # Processar upload de foto
        if form.foto.data:
            foto = form.foto.data
            if foto.filename:
                # Criar diretório se não existir
                import os
                upload_dir = 'static/uploads/funcionarios'
                if not os.path.exists(upload_dir):
                    os.makedirs(upload_dir)
                
                # Salvar arquivo
                filename = f"funcionario_{funcionario.id}_{foto.filename}"
                foto_path = os.path.join(upload_dir, filename)
                foto.save(foto_path)
                funcionario.foto = f"uploads/funcionarios/{filename}"
        
        db.session.commit()
        flash('Funcionário cadastrado com sucesso!', 'success')
        return redirect(url_for('main.funcionarios'))
    
    return render_template('funcionarios.html', form=form, funcionarios=Funcionario.query.all())

@main_bp.route('/funcionarios/<int:id>/editar', methods=['GET', 'POST'])
@login_required
def editar_funcionario(id):
    funcionario = Funcionario.query.get_or_404(id)
    
    if request.method == 'POST':
        try:
            # Atualizar dados do funcionário
            funcionario.nome = request.form.get('nome')
            funcionario.cpf = request.form.get('cpf')
            funcionario.rg = request.form.get('rg')
            funcionario.endereco = request.form.get('endereco')
            funcionario.telefone = request.form.get('telefone')
            funcionario.email = request.form.get('email')
            funcionario.salario = float(request.form.get('salario', 0) or 0)
            
            # Data de nascimento
            data_nascimento = request.form.get('data_nascimento')
            if data_nascimento:
                funcionario.data_nascimento = datetime.strptime(data_nascimento, '%Y-%m-%d').date()
            
            # Data de admissão
            data_admissao = request.form.get('data_admissao')
            if data_admissao:
                funcionario.data_admissao = datetime.strptime(data_admissao, '%Y-%m-%d').date()
            
            # IDs opcionais
            departamento_id = request.form.get('departamento_id')
            funcionario.departamento_id = int(departamento_id) if departamento_id and departamento_id != '0' else None
            
            funcao_id = request.form.get('funcao_id')
            funcionario.funcao_id = int(funcao_id) if funcao_id and funcao_id != '0' else None
            
            horario_id = request.form.get('horario_trabalho_id')
            funcionario.horario_trabalho_id = int(horario_id) if horario_id and horario_id != '0' else None
            
            funcionario.ativo = bool(request.form.get('ativo'))
            
            # Processar upload de foto
            if 'foto' in request.files:
                foto = request.files['foto']
                if foto and foto.filename:
                    # Criar diretório se não existir
                    import os
                    upload_dir = 'static/uploads/funcionarios'
                    if not os.path.exists(upload_dir):
                        os.makedirs(upload_dir)
                    
                    # Salvar arquivo
                    filename = f"funcionario_{funcionario.id}_{foto.filename}"
                    foto_path = os.path.join(upload_dir, filename)
                    foto.save(foto_path)
                    funcionario.foto = f"uploads/funcionarios/{filename}"
            
            db.session.commit()
            flash('Funcionário atualizado com sucesso!', 'success')
            return redirect(url_for('main.funcionario_perfil', id=funcionario.id))
            
        except Exception as e:
            flash(f'Erro ao atualizar funcionário: {str(e)}', 'error')
            return redirect(url_for('main.funcionario_perfil', id=id))
    
    # Para GET, redirecionar para a página de perfil com modo de edição
    return redirect(url_for('main.funcionario_perfil', id=id, edit=1))

def calcular_kpis_funcionario(funcionario_id):
    """Calcula KPIs individuais do funcionário para o mês atual"""
    hoje = date.today()
    primeiro_dia_mes = hoje.replace(day=1)
    
    # Buscar registros de ponto do mês atual
    registros_ponto = RegistroPonto.query.filter(
        RegistroPonto.funcionario_id == funcionario_id,
        RegistroPonto.data >= primeiro_dia_mes,
        RegistroPonto.data <= hoje
    ).all()
    
    # Calcular KPIs
    horas_trabalhadas = sum(r.horas_trabalhadas or 0 for r in registros_ponto)
    horas_extras = sum(r.horas_extras or 0 for r in registros_ponto)
    
    # Contar dias úteis no mês (aproximação: 22 dias úteis)
    dias_uteis_mes = 22
    dias_com_registro = len([r for r in registros_ponto if r.hora_entrada])
    
    # Calcular faltas e atrasos
    faltas = max(0, dias_uteis_mes - dias_com_registro)
    atrasos = len([r for r in registros_ponto if r.hora_entrada and r.hora_entrada.hour > 8])
    
    # Calcular absenteísmo
    absenteismo = (faltas / dias_uteis_mes) * 100 if dias_uteis_mes > 0 else 0
    
    # Calcular média de horas diárias
    media_horas_diarias = horas_trabalhadas / dias_com_registro if dias_com_registro > 0 else 0
    
    return {
        'horas_trabalhadas': horas_trabalhadas,
        'horas_extras': horas_extras,
        'faltas': faltas,
        'atrasos': atrasos,
        'absenteismo': absenteismo,
        'media_horas_diarias': media_horas_diarias
    }

def obter_dados_graficos_funcionario(funcionario_id):
    """Obtém dados para gráficos de desempenho do funcionário"""
    # Últimos 6 meses
    meses = []
    horas_trabalhadas = []
    absenteismo = []
    
    hoje = date.today()
    
    for i in range(6):
        # Calcular mês
        mes = hoje.month - i
        ano = hoje.year
        if mes <= 0:
            mes += 12
            ano -= 1
        
        primeiro_dia = date(ano, mes, 1)
        if mes == 12:
            ultimo_dia = date(ano + 1, 1, 1)
        else:
            ultimo_dia = date(ano, mes + 1, 1)
        
        # Buscar registros do mês
        registros = RegistroPonto.query.filter(
            RegistroPonto.funcionario_id == funcionario_id,
            RegistroPonto.data >= primeiro_dia,
            RegistroPonto.data < ultimo_dia
        ).all()
        
        # Calcular totais
        horas_mes = sum(r.horas_trabalhadas or 0 for r in registros)
        dias_com_registro = len([r for r in registros if r.hora_entrada])
        
        # Nome do mês
        nomes_meses = [
            'Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
            'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro'
        ]
        nome_mes = nomes_meses[mes - 1]
        
        meses.insert(0, nome_mes)
        horas_trabalhadas.insert(0, horas_mes)
        
        # Calcular absenteísmo (aproximação: 22 dias úteis)
        absenteismo_mes = ((22 - dias_com_registro) / 22) * 100 if dias_com_registro < 22 else 0
        absenteismo.insert(0, absenteismo_mes)
    
    return {
        'meses': meses,
        'horas_trabalhadas': horas_trabalhadas,
        'absenteismo': absenteismo
    }

@main_bp.route('/funcionarios/<int:id>/perfil')
@login_required
def funcionario_perfil(id):
    funcionario = Funcionario.query.get_or_404(id)
    
    # Filtros de data dos parâmetros
    data_inicio = request.args.get('data_inicio')
    data_fim = request.args.get('data_fim')
    obra_filtro = request.args.get('obra')
    
    # Definir período padrão (mês atual)
    if not data_inicio:
        data_inicio = date.today().replace(day=1)
    else:
        data_inicio = datetime.strptime(data_inicio, '%Y-%m-%d').date()
    
    if not data_fim:
        data_fim = date.today()
    else:
        data_fim = datetime.strptime(data_fim, '%Y-%m-%d').date()
    
    # Calcular KPIs individuais para o período (usando engine v3.0)
    from kpis_engine_v3 import calcular_kpis_funcionario_v3
    kpis = calcular_kpis_funcionario_v3(id, data_inicio, data_fim)
    
    # Buscar registros de ponto com filtros e identificação de faltas
    from kpis_engine_v3 import processar_registros_ponto_com_faltas
    
    if obra_filtro:
        # Se há filtro de obra, usar query tradicional
        query_ponto = RegistroPonto.query.filter_by(funcionario_id=id).filter(
            RegistroPonto.data >= data_inicio,
            RegistroPonto.data <= data_fim,
            RegistroPonto.obra_id == obra_filtro
        )
        registros_ponto = query_ponto.order_by(RegistroPonto.data.desc()).all()
        
        # Adicionar informação de falta manualmente para registros filtrados por obra
        from kpis_engine_v3 import identificar_faltas_periodo
        faltas = identificar_faltas_periodo(id, data_inicio, data_fim)
        
        # Lista de feriados 2025
        feriados_2025 = {
            date(2025, 1, 1),   # Ano Novo
            date(2025, 2, 17),  # Carnaval (Segunda-feira)
            date(2025, 2, 18),  # Carnaval (Terça-feira)
            date(2025, 4, 18),  # Paixão de Cristo (Sexta-feira Santa)
            date(2025, 4, 21),  # Tiradentes
            date(2025, 5, 1),   # Dia do Trabalhador
            date(2025, 6, 19),  # Corpus Christi
            date(2025, 9, 7),   # Independência
            date(2025, 10, 12), # Nossa Senhora Aparecida
            date(2025, 11, 2),  # Finados
            date(2025, 11, 15), # Proclamação da República
            date(2025, 12, 25)  # Natal
        }
        
        for registro in registros_ponto:
            registro.is_falta = (registro.data in faltas)
            registro.is_feriado = (registro.data in feriados_2025)
    else:
        # Usar função que já identifica faltas
        registros_ponto, faltas = processar_registros_ponto_com_faltas(id, data_inicio, data_fim)
        
        # Lista de feriados 2025 para quando não há filtro de obra
        feriados_2025 = {
            date(2025, 1, 1),   # Ano Novo
            date(2025, 2, 17),  # Carnaval (Segunda-feira)
            date(2025, 2, 18),  # Carnaval (Terça-feira)
            date(2025, 4, 18),  # Paixão de Cristo (Sexta-feira Santa)
            date(2025, 4, 21),  # Tiradentes
            date(2025, 5, 1),   # Dia do Trabalhador
            date(2025, 6, 19),  # Corpus Christi
            date(2025, 9, 7),   # Independência
            date(2025, 10, 12), # Nossa Senhora Aparecida
            date(2025, 11, 2),  # Finados
            date(2025, 11, 15), # Proclamação da República
            date(2025, 12, 25)  # Natal
        }
        
        # Adicionar informação de feriado para todos os registros
        for registro in registros_ponto:
            registro.is_feriado = (registro.data in feriados_2025)
    
    # Buscar ocorrências (sem filtro de data por enquanto)
    ocorrencias = Ocorrencia.query.filter_by(funcionario_id=id).order_by(Ocorrencia.data_inicio.desc()).all()
    
    # Buscar registros de alimentação com filtros
    query_alimentacao = RegistroAlimentacao.query.filter_by(funcionario_id=id).filter(
        RegistroAlimentacao.data >= data_inicio,
        RegistroAlimentacao.data <= data_fim
    )
    
    if obra_filtro:
        query_alimentacao = query_alimentacao.filter_by(obra_id=obra_filtro)
    
    registros_alimentacao = query_alimentacao.order_by(RegistroAlimentacao.data.desc()).all()
    
    # Buscar obras para os dropdowns
    obras = Obra.query.filter_by(status='Em andamento').all()
    
    # Obter dados para gráficos
    graficos = obter_dados_graficos_funcionario(id)
    
    # Criar objeto para KPIs (simular uma estrutura)
    class KPIData:
        def __init__(self, data):
            for key, value in data.items():
                setattr(self, key, value)
    
    kpis_obj = KPIData(kpis) if kpis else KPIData({
        'horas_trabalhadas': 0,
        'horas_extras': 0,
        'faltas': 0,
        'atrasos': 0,
        'absenteismo': 0,
        'horas_extras_valor': 0,
        'media_horas_diarias': 0,
        'total_atrasos': 0,
        'pontualidade': 100,
        'custo_total': 0,
        'custo_mao_obra': 0,
        'custo_alimentacao': 0,
        'custo_transporte': 0,
        'dias_trabalhados': 0,
        'dias_uteis': 0
    })
    
    # Buscar dados adicionais para o modal de edição
    departamentos = Departamento.query.all()
    funcoes = Funcao.query.all()
    horarios = HorarioTrabalho.query.all()
    
    return render_template('funcionario_perfil.html',
                         funcionario=funcionario,
                         kpis=kpis_obj,
                         registros_ponto=registros_ponto,
                         ocorrencias=ocorrencias,
                         registros_alimentacao=registros_alimentacao,
                         obras=obras,
                         graficos=graficos,
                         data_inicio=data_inicio,
                         data_fim=data_fim,
                         obra_filtro=obra_filtro,
                         departamentos=departamentos,
                         funcoes=funcoes,
                         horarios=horarios)

@main_bp.route('/funcionarios/<int:id>/excluir', methods=['POST'])
@login_required
def excluir_funcionario(id):
    funcionario = Funcionario.query.get_or_404(id)
    db.session.delete(funcionario)
    db.session.commit()
    flash('Funcionário excluído com sucesso!', 'success')
    return redirect(url_for('main.funcionarios'))

# Departamentos
@main_bp.route('/departamentos')
@login_required
def departamentos():
    departamentos = Departamento.query.all()
    return render_template('departamentos.html', departamentos=departamentos)

@main_bp.route('/departamentos/novo', methods=['GET', 'POST'])
@login_required
def novo_departamento():
    form = DepartamentoForm()
    if form.validate_on_submit():
        departamento = Departamento(
            nome=form.nome.data,
            descricao=form.descricao.data
        )
        db.session.add(departamento)
        db.session.commit()
        flash('Departamento cadastrado com sucesso!', 'success')
        return redirect(url_for('main.departamentos'))
    
    return render_template('departamentos.html', form=form, departamentos=Departamento.query.all())

@main_bp.route('/departamentos/<int:id>/editar', methods=['GET', 'POST'])
@login_required
def editar_departamento(id):
    departamento = Departamento.query.get_or_404(id)
    form = DepartamentoForm(obj=departamento)
    
    if form.validate_on_submit():
        departamento.nome = form.nome.data
        departamento.descricao = form.descricao.data
        db.session.commit()
        flash('Departamento atualizado com sucesso!', 'success')
        return redirect(url_for('main.departamentos'))
    
    return render_template('departamentos.html', form=form, departamento=departamento, departamentos=Departamento.query.all())

@main_bp.route('/departamentos/<int:id>/excluir', methods=['POST'])
@login_required
def excluir_departamento(id):
    departamento = Departamento.query.get_or_404(id)
    db.session.delete(departamento)
    db.session.commit()
    flash('Departamento excluído com sucesso!', 'success')
    return redirect(url_for('main.departamentos'))

# Funções
@main_bp.route('/funcoes')
@login_required
def funcoes():
    funcoes = Funcao.query.all()
    return render_template('funcoes.html', funcoes=funcoes)

@main_bp.route('/funcoes/novo', methods=['GET', 'POST'])
@login_required
def nova_funcao():
    form = FuncaoForm()
    if form.validate_on_submit():
        funcao = Funcao(
            nome=form.nome.data,
            descricao=form.descricao.data,
            salario_base=form.salario_base.data or 0.0
        )
        db.session.add(funcao)
        db.session.commit()
        flash('Função cadastrada com sucesso!', 'success')
        return redirect(url_for('main.funcoes'))
    
    return render_template('funcoes.html', form=form, funcoes=Funcao.query.all())

@main_bp.route('/funcoes/<int:id>/editar', methods=['GET', 'POST'])
@login_required
def editar_funcao(id):
    funcao = Funcao.query.get_or_404(id)
    form = FuncaoForm(obj=funcao)
    
    if form.validate_on_submit():
        funcao.nome = form.nome.data
        funcao.descricao = form.descricao.data
        funcao.salario_base = form.salario_base.data or 0.0
        db.session.commit()
        flash('Função atualizada com sucesso!', 'success')
        return redirect(url_for('main.funcoes'))
    
    return render_template('funcoes.html', form=form, funcao=funcao, funcoes=Funcao.query.all())

@main_bp.route('/funcoes/<int:id>/excluir', methods=['POST'])
@login_required
def excluir_funcao(id):
    funcao = Funcao.query.get_or_404(id)
    db.session.delete(funcao)
    db.session.commit()
    flash('Função excluída com sucesso!', 'success')
    return redirect(url_for('main.funcoes'))

# Obras
@main_bp.route('/obras')
@login_required
def obras():
    from datetime import date, timedelta
    from sqlalchemy import func
    
    obras = Obra.query.all()
    
    # Adicionar KPIs básicos para cada obra (últimos 30 dias)
    data_fim = date.today()
    data_inicio = data_fim - timedelta(days=30)
    
    for obra in obras:
        # Calcular RDOs
        total_rdos = RDO.query.filter_by(obra_id=obra.id).count()
        
        # Calcular dias trabalhados (registros de ponto únicos)
        dias_trabalhados = db.session.query(RegistroPonto.data).filter(
            RegistroPonto.obra_id == obra.id,
            RegistroPonto.data.between(data_inicio, data_fim),
            RegistroPonto.hora_entrada.isnot(None)
        ).distinct().count()
        
        # Calcular custo total básico (últimos 30 dias)
        custo_alimentacao = db.session.query(func.sum(RegistroAlimentacao.valor)).filter(
            RegistroAlimentacao.obra_id == obra.id,
            RegistroAlimentacao.data.between(data_inicio, data_fim)
        ).scalar() or 0.0
        
        # Criar objeto KPI simples
        obra.kpis = type('KPIs', (), {
            'total_rdos': total_rdos,
            'dias_trabalhados': dias_trabalhados,
            'custo_total': custo_alimentacao  # Simplificado para exibição nos cards
        })()
    
    return render_template('obras.html', obras=obras)

@main_bp.route('/obras/novo', methods=['GET', 'POST'])
@login_required
def nova_obra():
    form = ObraForm()
    form.responsavel_id.choices = [(0, 'Selecione...')] + [(f.id, f.nome) for f in Funcionario.query.filter_by(ativo=True).all()]
    
    if form.validate_on_submit():
        obra = Obra(
            nome=form.nome.data,
            endereco=form.endereco.data,
            data_inicio=form.data_inicio.data,
            data_previsao_fim=form.data_previsao_fim.data,
            orcamento=form.orcamento.data or 0.0,
            status=form.status.data,
            responsavel_id=form.responsavel_id.data if form.responsavel_id.data > 0 else None
        )
        db.session.add(obra)
        db.session.commit()
        flash('Obra cadastrada com sucesso!', 'success')
        return redirect(url_for('main.obras'))
    
    return render_template('obras.html', form=form, obras=Obra.query.all())

@main_bp.route('/obras/<int:id>/editar', methods=['GET', 'POST'])
@login_required
def editar_obra(id):
    obra = Obra.query.get_or_404(id)
    form = ObraForm(obj=obra)
    form.responsavel_id.choices = [(0, 'Selecione...')] + [(f.id, f.nome) for f in Funcionario.query.filter_by(ativo=True).all()]
    
    if form.validate_on_submit():
        obra.nome = form.nome.data
        obra.endereco = form.endereco.data
        obra.data_inicio = form.data_inicio.data
        obra.data_previsao_fim = form.data_previsao_fim.data
        obra.orcamento = form.orcamento.data or 0.0
        obra.status = form.status.data
        obra.responsavel_id = form.responsavel_id.data if form.responsavel_id.data > 0 else None
        
        db.session.commit()
        flash('Obra atualizada com sucesso!', 'success')
        return redirect(url_for('main.obras'))
    
    return render_template('obras.html', form=form, obra=obra, obras=Obra.query.all())

@main_bp.route('/obra/<int:id>')
@login_required
def detalhes_obra(id):
    from datetime import datetime, date, timedelta
    from sqlalchemy import func
    
    obra = Obra.query.get_or_404(id)
    
    # Obter filtros de data da query string
    data_inicio_filtro = request.args.get('data_inicio')
    data_fim_filtro = request.args.get('data_fim')
    
    # Definir período padrão (últimos 30 dias)
    if not data_inicio_filtro or not data_fim_filtro:
        data_fim = date.today()
        data_inicio = data_fim - timedelta(days=30)
    else:
        data_inicio = datetime.strptime(data_inicio_filtro, '%Y-%m-%d').date()
        data_fim = datetime.strptime(data_fim_filtro, '%Y-%m-%d').date()
    
    # ===== CÁLCULO DOS KPIS =====
    
    # 1. Custos de Transporte (Veículos)
    # Por enquanto, vamos usar apenas custos de veículos sem vinculação específica à obra
    custo_transporte = db.session.query(func.sum(CustoVeiculo.valor)).filter(
        CustoVeiculo.data_custo.between(data_inicio, data_fim)
    ).scalar() or 0.0
    
    # 2. Custos de Alimentação
    custo_alimentacao = db.session.query(func.sum(RegistroAlimentacao.valor)).filter(
        RegistroAlimentacao.obra_id == id,
        RegistroAlimentacao.data.between(data_inicio, data_fim)
    ).scalar() or 0.0
    
    # 3. Custos de Mão de Obra
    # Buscar registros de ponto da obra no período
    registros_ponto = db.session.query(RegistroPonto).join(Funcionario).filter(
        RegistroPonto.obra_id == id,
        RegistroPonto.data.between(data_inicio, data_fim),
        RegistroPonto.hora_entrada.isnot(None)  # Só dias trabalhados
    ).all()
    
    custo_mao_obra = 0.0
    total_horas = 0.0
    dias_trabalhados = len(set(rp.data for rp in registros_ponto))
    
    for registro in registros_ponto:
        if registro.hora_entrada and registro.hora_saida:
            # Calcular horas trabalhadas
            entrada = datetime.combine(registro.data, registro.hora_entrada)
            saida = datetime.combine(registro.data, registro.hora_saida)
            
            # Subtrair tempo de almoço (1 hora padrão)
            horas_dia = (saida - entrada).total_seconds() / 3600 - 1
            horas_dia = max(0, horas_dia)  # Não pode ser negativo
            total_horas += horas_dia
            
            # Calcular custo baseado no salário do funcionário
            if registro.funcionario_ref.salario:
                valor_hora = registro.funcionario_ref.salario / 220  # 220 horas/mês aprox
                custo_mao_obra += horas_dia * valor_hora
    
    # 4. Custo Total da Obra
    custo_total = custo_transporte + custo_alimentacao + custo_mao_obra
    
    # ===== RDOs =====
    rdos_periodo = RDO.query.filter(
        RDO.obra_id == id,
        RDO.data_relatorio.between(data_inicio, data_fim)
    ).order_by(RDO.data_relatorio.desc()).all()
    
    rdos_recentes = RDO.query.filter_by(obra_id=id).order_by(RDO.data_relatorio.desc()).limit(5).all()
    total_rdos = RDO.query.filter_by(obra_id=id).count()
    rdos_finalizados = RDO.query.filter_by(obra_id=id, status='Finalizado').count()
    
    # ===== MÉTRICAS ADICIONAIS =====
    
    # Progresso da obra
    progresso_obra = 0
    if total_rdos > 0:
        progresso_obra = min(100, (rdos_finalizados / total_rdos) * 100)
    
    # Calcular dias da obra
    hoje = date.today()
    dias_decorridos = (hoje - obra.data_inicio).days if obra.data_inicio else 0
    dias_restantes = 0
    if obra.data_previsao_fim:
        dias_restantes = max(0, (obra.data_previsao_fim - hoje).days)
    
    # Funcionários únicos que trabalharam na obra no período
    funcionarios_periodo = db.session.query(Funcionario).join(RegistroPonto).filter(
        RegistroPonto.obra_id == id,
        RegistroPonto.data.between(data_inicio, data_fim),
        RegistroPonto.hora_entrada.isnot(None)
    ).distinct().all()
    
    # Custos da obra no período
    custos_obra = CustoObra.query.filter(
        CustoObra.obra_id == id,
        CustoObra.data.between(data_inicio, data_fim)
    ).order_by(CustoObra.data.desc()).all()
    
    # ===== CUSTOS DE TRANSPORTE DETALHADOS =====
    custos_transporte = db.session.query(CustoVeiculo).filter(
        CustoVeiculo.data_custo.between(data_inicio, data_fim)
    ).order_by(CustoVeiculo.data_custo.desc()).all()
    
    custos_transporte_total = sum(c.valor for c in custos_transporte)
    
    # ===== CUSTOS DE MÃO DE OBRA DETALHADOS =====
    custos_mao_obra = []
    for registro in registros_ponto:
        if registro.hora_entrada and registro.hora_saida:
            # Calcular horas trabalhadas
            entrada = datetime.combine(registro.data, registro.hora_entrada)
            saida = datetime.combine(registro.data, registro.hora_saida)
            
            # Subtrair tempo de almoço (1 hora padrão)
            horas_dia = (saida - entrada).total_seconds() / 3600 - 1
            horas_dia = max(0, horas_dia)  # Não pode ser negativo
            
            # Calcular custo baseado no salário do funcionário
            if registro.funcionario_ref.salario:
                valor_hora = registro.funcionario_ref.salario / 220  # 220 horas/mês aprox
                total_dia = horas_dia * valor_hora
                
                custos_mao_obra.append({
                    'data': registro.data,
                    'funcionario_nome': registro.funcionario_ref.nome,
                    'horas_trabalhadas': horas_dia,
                    'salario_hora': valor_hora,
                    'total_dia': total_dia
                })
    
    # KPIs organizados
    kpis = {
        'custo_transporte': custo_transporte,
        'custo_alimentacao': custo_alimentacao,
        'custo_mao_obra': custo_mao_obra,
        'custo_total': custo_total,
        'dias_trabalhados': dias_trabalhados,
        'total_horas': round(total_horas, 1),
        'funcionarios_periodo': len(funcionarios_periodo),
        'rdos_periodo': len(rdos_periodo)
    }
    
    return render_template('obras/detalhes_obra.html', 
                         obra=obra,
                         kpis=kpis,
                         custos_obra=custos_obra,
                         custos_transporte=custos_transporte,
                         custos_transporte_total=custos_transporte_total,
                         custos_mao_obra=custos_mao_obra,
                         rdos_periodo=rdos_periodo,
                         rdos_recentes=rdos_recentes,
                         total_rdos=total_rdos,
                         rdos_finalizados=rdos_finalizados,
                         progresso_obra=int(progresso_obra),
                         dias_decorridos=dias_decorridos,
                         dias_restantes=dias_restantes,
                         funcionarios_periodo=funcionarios_periodo,
                         data_inicio=data_inicio,
                         data_fim=data_fim)

@main_bp.route('/obras/<int:id>/excluir', methods=['POST'])
@login_required
def excluir_obra(id):
    obra = Obra.query.get_or_404(id)
    db.session.delete(obra)
    db.session.commit()
    flash('Obra excluída com sucesso!', 'success')
    return redirect(url_for('main.obras'))

# Veículos
@main_bp.route('/veiculos')
@login_required
def veiculos():
    veiculos = Veiculo.query.all()
    funcionarios = Funcionario.query.filter_by(ativo=True).all()
    obras = Obra.query.filter_by(status='Em andamento').all()
    return render_template('veiculos.html', 
                         veiculos=veiculos, 
                         funcionarios=funcionarios, 
                         obras=obras)

@main_bp.route('/veiculos/novo', methods=['GET', 'POST'])
@login_required
def novo_veiculo():
    form = VeiculoForm()
    if form.validate_on_submit():
        veiculo = Veiculo(
            placa=form.placa.data,
            marca=form.marca.data,
            modelo=form.modelo.data,
            ano=form.ano.data,
            tipo=form.tipo.data,
            status=form.status.data,
            km_atual=form.km_atual.data or 0,
            data_ultima_manutencao=form.data_ultima_manutencao.data,
            data_proxima_manutencao=form.data_proxima_manutencao.data
        )
        db.session.add(veiculo)
        db.session.commit()
        flash('Veículo cadastrado com sucesso!', 'success')
        return redirect(url_for('main.veiculos'))
    
    return render_template('veiculos.html', form=form, veiculos=Veiculo.query.all())

@main_bp.route('/veiculos/<int:id>/editar', methods=['GET', 'POST'])
@login_required
def editar_veiculo(id):
    veiculo = Veiculo.query.get_or_404(id)
    form = VeiculoForm(obj=veiculo)
    
    if form.validate_on_submit():
        veiculo.placa = form.placa.data
        veiculo.marca = form.marca.data
        veiculo.modelo = form.modelo.data
        veiculo.ano = form.ano.data
        veiculo.tipo = form.tipo.data
        veiculo.status = form.status.data
        veiculo.km_atual = form.km_atual.data or 0
        veiculo.data_ultima_manutencao = form.data_ultima_manutencao.data
        veiculo.data_proxima_manutencao = form.data_proxima_manutencao.data
        
        db.session.commit()
        flash('Veículo atualizado com sucesso!', 'success')
        return redirect(url_for('main.veiculos'))
    
    return render_template('veiculos.html', form=form, veiculo=veiculo, veiculos=Veiculo.query.all())

@main_bp.route('/veiculos/<int:id>/excluir', methods=['POST'])
@login_required
def excluir_veiculo(id):
    veiculo = Veiculo.query.get_or_404(id)
    db.session.delete(veiculo)
    db.session.commit()
    flash('Veículo excluído com sucesso!', 'success')
    return redirect(url_for('main.veiculos'))

@main_bp.route('/veiculos/<int:id>/detalhes')
@login_required
def detalhes_veiculo(id):
    veiculo = Veiculo.query.get_or_404(id)
    
    # Buscar registros de uso
    usos = UsoVeiculo.query.filter_by(veiculo_id=id).order_by(UsoVeiculo.data_uso.desc()).all()
    
    # Buscar registros de custo
    custos = CustoVeiculo.query.filter_by(veiculo_id=id).order_by(CustoVeiculo.data_custo.desc()).all()
    
    # Dados para os formulários
    funcionarios = Funcionario.query.filter_by(ativo=True).order_by(Funcionario.nome).all()
    obras = Obra.query.filter(Obra.status.in_(['Em andamento', 'Pausada'])).order_by(Obra.nome).all()
    
    # KPIs do veículo
    custo_total = sum(custo.valor for custo in custos)
    kpis = {
        'custo_total': custo_total,
        'total_usos': len(usos),
        'total_custos': len(custos)
    }
    
    # Dados para gráfico de uso por obra
    from sqlalchemy import func
    uso_por_obra = db.session.query(
        Obra.nome,
        func.count(UsoVeiculo.id).label('total_usos')
    ).join(UsoVeiculo, Obra.id == UsoVeiculo.obra_id).filter(
        UsoVeiculo.veiculo_id == id
    ).group_by(Obra.id, Obra.nome).all()
    
    # Preparar dados para o gráfico de pizza
    grafico_uso_obras = {
        'labels': [uso[0] for uso in uso_por_obra],
        'data': [uso[1] for uso in uso_por_obra],
        'total': sum([uso[1] for uso in uso_por_obra])
    }
    
    return render_template('detalhes_veiculo.html', 
                         veiculo=veiculo, 
                         usos=usos, 
                         custos=custos,
                         kpis=kpis,
                         funcionarios=funcionarios,
                         obras=obras,
                         grafico_uso_obras=grafico_uso_obras)

@main_bp.route('/veiculos/novo-uso-direto', methods=['POST'])
@login_required
def novo_uso_veiculo_direto():
    """Registra novo uso de veículo na página de detalhes"""
    try:
        veiculo_id = request.form.get('veiculo_id')
        funcionario_id = request.form.get('funcionario_id')
        obra_id = request.form.get('obra_id')
        data_uso = datetime.strptime(request.form.get('data_uso'), '%Y-%m-%d').date()
        km_inicial = int(request.form.get('km_inicial')) if request.form.get('km_inicial') else None
        km_final = int(request.form.get('km_final')) if request.form.get('km_final') else None
        horario_saida = datetime.strptime(request.form.get('horario_saida'), '%H:%M').time() if request.form.get('horario_saida') else None
        horario_chegada = datetime.strptime(request.form.get('horario_chegada'), '%H:%M').time() if request.form.get('horario_chegada') else None
        finalidade = request.form.get('finalidade')
        observacoes = request.form.get('observacoes')
        
        if not obra_id:
            flash('Obra é obrigatória para registrar uso do veículo!', 'error')
            return redirect(url_for('main.detalhes_veiculo', id=veiculo_id))
        
        novo_uso = UsoVeiculo(
            veiculo_id=veiculo_id,
            funcionario_id=funcionario_id,
            obra_id=obra_id,
            data_uso=data_uso,
            km_inicial=km_inicial,
            km_final=km_final,
            horario_saida=horario_saida,
            horario_chegada=horario_chegada,
            finalidade=finalidade,
            observacoes=observacoes
        )
        
        db.session.add(novo_uso)
        
        # Atualizar KM do veículo se fornecido
        if km_final:
            veiculo = Veiculo.query.get(veiculo_id)
            if veiculo:
                veiculo.km_atual = km_final
        
        db.session.commit()
        
        flash('Uso do veículo registrado com sucesso!', 'success')
    except Exception as e:
        flash(f'Erro ao registrar uso: {str(e)}', 'error')
    
    return redirect(url_for('main.detalhes_veiculo', id=veiculo_id))

@main_bp.route('/veiculos/novo-custo', methods=['POST'])
@login_required
def novo_custo_veiculo():
    """Registra novo custo de veículo"""
    try:
        veiculo_id = request.form.get('veiculo_id')
        data_custo = datetime.strptime(request.form.get('data_custo'), '%Y-%m-%d').date()
        valor = float(request.form.get('valor'))
        tipo_custo = request.form.get('tipo_custo')
        obra_id = request.form.get('obra_id')
        fornecedor = request.form.get('fornecedor')
        descricao = request.form.get('descricao')
        
        if not obra_id:
            flash('Obra é obrigatória para registrar custo do veículo!', 'error')
            return redirect(url_for('main.detalhes_veiculo', id=veiculo_id))
        
        novo_custo = CustoVeiculo(
            veiculo_id=veiculo_id,
            data_custo=data_custo,
            valor=valor,
            tipo_custo=tipo_custo,
            obra_id=obra_id,
            fornecedor=fornecedor,
            descricao=descricao
        )
        
        db.session.add(novo_custo)
        db.session.commit()
        
        flash('Custo do veículo registrado com sucesso!', 'success')
    except Exception as e:
        flash(f'Erro ao registrar custo: {str(e)}', 'error')
    
    return redirect(url_for('main.detalhes_veiculo', id=veiculo_id))

@main_bp.route('/veiculos/<int:id>/dados')
@login_required
def dados_veiculo(id):
    """Retorna dados do veículo em JSON para edição via AJAX"""
    veiculo = Veiculo.query.get_or_404(id)
    return jsonify({
        'id': veiculo.id,
        'placa': veiculo.placa,
        'marca': veiculo.marca,
        'modelo': veiculo.modelo,
        'ano': veiculo.ano,
        'tipo': veiculo.tipo,
        'km_atual': veiculo.km_atual,
        'status': veiculo.status,
        'data_ultima_manutencao': veiculo.data_ultima_manutencao.strftime('%Y-%m-%d') if veiculo.data_ultima_manutencao else '',
        'data_proxima_manutencao': veiculo.data_proxima_manutencao.strftime('%Y-%m-%d') if veiculo.data_proxima_manutencao else ''
    })

@main_bp.route('/veiculos/uso', methods=['POST'])
@login_required
def novo_uso_veiculo():
    """Registra novo uso de veículo"""
    form = UsoVeiculoForm()
    
    # Preencher choices dinamicamente
    form.veiculo_id.choices = [(0, 'Selecione...')] + [(v.id, f"{v.placa} - {v.marca} {v.modelo}") for v in Veiculo.query.filter_by(status='Disponível').all()]
    form.funcionario_id.choices = [(0, 'Selecione...')] + [(f.id, f.nome) for f in Funcionario.query.filter_by(ativo=True).all()]
    form.obra_id.choices = [(0, 'Selecione...')] + [(o.id, o.nome) for o in Obra.query.all()]
    
    if form.validate_on_submit():
        uso = UsoVeiculo(
            veiculo_id=form.veiculo_id.data,
            funcionario_id=form.funcionario_id.data,
            obra_id=form.obra_id.data if form.obra_id.data else None,
            data_uso=form.data_uso.data,
            km_inicial=form.km_inicial.data,
            km_final=form.km_final.data,
            finalidade=form.finalidade.data,
            observacoes=form.observacoes.data
        )
        db.session.add(uso)
        
        # Atualizar status do veículo para "Em uso"
        veiculo = Veiculo.query.get(form.veiculo_id.data)
        if veiculo:
            veiculo.status = 'Em uso'
            # Atualizar KM se fornecido
            if form.km_final.data:
                veiculo.km_atual = form.km_final.data
        
        db.session.commit()
        flash('Uso de veículo registrado com sucesso!', 'success')
    else:
        flash('Erro ao registrar uso do veículo. Verifique os dados.', 'error')
    
    return redirect(url_for('main.veiculos'))

@main_bp.route('/veiculos/custo', methods=['POST'])
@login_required
def novo_custo_veiculo_lista():
    """Registra novo custo de veículo a partir da lista principal"""
    form = CustoVeiculoForm()
    
    # Preencher choices dinamicamente
    form.veiculo_id.choices = [(0, 'Selecione...')] + [(v.id, f"{v.placa} - {v.marca} {v.modelo}") for v in Veiculo.query.all()]
    
    if form.validate_on_submit():
        custo = CustoVeiculo(
            veiculo_id=form.veiculo_id.data,
            data_custo=form.data_custo.data,
            valor=form.valor.data,
            tipo_custo=form.tipo_custo.data,
            descricao=form.descricao.data,
            km_atual=form.km_atual.data,
            fornecedor=form.fornecedor.data
        )
        db.session.add(custo)
        
        # Atualizar KM do veículo se fornecido
        if form.km_atual.data:
            veiculo = Veiculo.query.get(form.veiculo_id.data)
            if veiculo:
                veiculo.km_atual = form.km_atual.data
        
        db.session.commit()
        flash('Custo de veículo registrado com sucesso!', 'success')
    else:
        flash('Erro ao registrar custo do veículo. Verifique os dados.', 'error')
    
    return redirect(url_for('main.veiculos'))

# Serviços
@main_bp.route('/servicos')
@login_required
def servicos():
    servicos = Servico.query.all()
    return render_template('servicos.html', servicos=servicos)

@main_bp.route('/servicos/novo', methods=['GET', 'POST'])
@login_required
def novo_servico():
    form = ServicoForm()
    if form.validate_on_submit():
        servico = Servico(
            nome=form.nome.data,
            descricao=form.descricao.data,
            preco_unitario=form.preco_unitario.data or 0.0
        )
        db.session.add(servico)
        db.session.commit()
        flash('Serviço cadastrado com sucesso!', 'success')
        return redirect(url_for('main.servicos'))
    
    return render_template('servicos.html', form=form, servicos=Servico.query.all())

@main_bp.route('/servicos/<int:id>/editar', methods=['GET', 'POST'])
@login_required
def editar_servico(id):
    servico = Servico.query.get_or_404(id)
    form = ServicoForm(obj=servico)
    
    if form.validate_on_submit():
        servico.nome = form.nome.data
        servico.descricao = form.descricao.data
        servico.preco_unitario = form.preco_unitario.data or 0.0
        
        db.session.commit()
        flash('Serviço atualizado com sucesso!', 'success')
        return redirect(url_for('main.servicos'))
    
    return render_template('servicos.html', form=form, servico=servico, servicos=Servico.query.all())

@main_bp.route('/servicos/<int:id>/excluir', methods=['POST'])
@login_required
def excluir_servico(id):
    servico = Servico.query.get_or_404(id)
    db.session.delete(servico)
    db.session.commit()
    flash('Serviço excluído com sucesso!', 'success')
    return redirect(url_for('main.servicos'))

# Ponto
@main_bp.route('/ponto')
@login_required
def ponto():
    registros = RegistroPonto.query.order_by(RegistroPonto.data.desc()).limit(50).all()
    return render_template('ponto.html', registros=registros)

@main_bp.route('/ponto/novo', methods=['GET', 'POST'])
@login_required
def novo_ponto():
    form = RegistroPontoForm()
    form.funcionario_id.choices = [(f.id, f.nome) for f in Funcionario.query.filter_by(ativo=True).all()]
    form.obra_id.choices = [(0, 'Selecione...')] + [(o.id, o.nome) for o in Obra.query.filter_by(status='Em andamento').all()]
    
    if form.validate_on_submit():
        registro = RegistroPonto(
            funcionario_id=form.funcionario_id.data,
            obra_id=form.obra_id.data if form.obra_id.data > 0 else None,
            data=form.data.data,
            hora_entrada=form.hora_entrada.data,
            hora_saida=form.hora_saida.data,
            hora_almoco_saida=form.hora_almoco_saida.data,
            hora_almoco_retorno=form.hora_almoco_retorno.data,
            observacoes=form.observacoes.data
        )
        
        # Calcular horas trabalhadas
        if registro.hora_entrada and registro.hora_saida:
            horas_trabalhadas = calcular_horas_trabalhadas(
                registro.hora_entrada, registro.hora_saida,
                registro.hora_almoco_saida, registro.hora_almoco_retorno
            )
            registro.horas_trabalhadas = horas_trabalhadas['total']
            registro.horas_extras = horas_trabalhadas['extras']
        
        db.session.add(registro)
        db.session.commit()
        
        # Atualizar cálculos automáticos do registro
        from kpis_engine_v3 import atualizar_calculos_ponto
        atualizar_calculos_ponto(registro.id)
        
        flash('Registro de ponto adicionado com sucesso!', 'success')
        return redirect(url_for('main.ponto'))
    
    return render_template('ponto.html', form=form, registros=RegistroPonto.query.order_by(RegistroPonto.data.desc()).limit(50).all())

# Alimentação
@main_bp.route('/alimentacao')
@login_required
def alimentacao():
    from datetime import date
    registros = RegistroAlimentacao.query.order_by(RegistroAlimentacao.data.desc()).limit(50).all()
    return render_template('alimentacao.html', registros=registros, date=date)

@main_bp.route('/alimentacao/novo', methods=['GET', 'POST'])
@login_required
def nova_alimentacao():
    form = AlimentacaoForm()
    form.funcionario_id.choices = [(f.id, f.nome) for f in Funcionario.query.filter_by(ativo=True).all()]
    form.obra_id.choices = [(0, 'Selecione...')] + [(o.id, o.nome) for o in Obra.query.filter_by(status='Em andamento').all()]
    form.restaurante_id.choices = [(0, 'Selecione...')] + [(r.id, r.nome) for r in Restaurante.query.filter_by(ativo=True).all()]
    
    if form.validate_on_submit():
        registro = RegistroAlimentacao(
            funcionario_id=form.funcionario_id.data,
            obra_id=form.obra_id.data if form.obra_id.data > 0 else None,
            restaurante_id=form.restaurante_id.data if form.restaurante_id.data > 0 else None,
            data=form.data.data,
            tipo=form.tipo.data,
            valor=form.valor.data,
            observacoes=form.observacoes.data
        )
        
        db.session.add(registro)
        
        # Adicionar custo à obra se especificada
        if registro.obra_id:
            custo = CustoObra(
                obra_id=registro.obra_id,
                tipo='alimentacao',
                descricao=f'Alimentação - {registro.tipo} - {registro.funcionario_ref.nome}',
                valor=registro.valor,
                data=registro.data
            )
            db.session.add(custo)
        
        db.session.commit()
        flash('Registro de alimentação adicionado com sucesso!', 'success')
        return redirect(url_for('main.alimentacao'))
    
    return render_template('alimentacao.html', form=form, registros=RegistroAlimentacao.query.order_by(RegistroAlimentacao.data.desc()).limit(50).all())

# Relatórios
@main_bp.route('/relatorios')
@login_required
def relatorios():
    # Buscar dados para filtros
    obras = Obra.query.all()
    departamentos = Departamento.query.all()
    
    return render_template('relatorios.html', obras=obras, departamentos=departamentos)

@main_bp.route('/relatorios/dados-graficos', methods=['POST'])
@login_required
def dados_graficos():
    from datetime import datetime, timedelta
    import json
    from sqlalchemy import func, extract
    
    filtros = request.get_json()
    
    # Processar filtros de data
    data_inicio = datetime.strptime(filtros.get('dataInicio', ''), '%Y-%m-%d').date() if filtros.get('dataInicio') else None
    data_fim = datetime.strptime(filtros.get('dataFim', ''), '%Y-%m-%d').date() if filtros.get('dataFim') else None
    obra_id = filtros.get('obra') if filtros.get('obra') else None
    departamento_id = filtros.get('departamento') if filtros.get('departamento') else None
    
    # Data padrão (últimos 6 meses)
    if not data_inicio or not data_fim:
        hoje = date.today()
        data_fim = hoje
        data_inicio = date(hoje.year, hoje.month - 5 if hoje.month > 5 else hoje.month + 7, 1)
        if hoje.month <= 5:
            data_inicio = data_inicio.replace(year=hoje.year - 1)
    
    # 1. Evolução de Custos
    custos_query = db.session.query(
        extract('month', CustoObra.data).label('mes'),
        extract('year', CustoObra.data).label('ano'),
        CustoObra.tipo,
        func.sum(CustoObra.valor).label('total')
    ).filter(
        CustoObra.data >= data_inicio,
        CustoObra.data <= data_fim
    )
    
    if obra_id:
        custos_query = custos_query.filter(CustoObra.obra_id == obra_id)
    
    custos_dados = custos_query.group_by(
        extract('year', CustoObra.data),
        extract('month', CustoObra.data),
        CustoObra.tipo
    ).all()
    
    # Processar dados de custos
    meses_labels = []
    mao_obra_dados = []
    alimentacao_dados = []
    veiculos_dados = []
    
    # Gerar lista de meses
    current_date = data_inicio.replace(day=1)
    while current_date <= data_fim:
        mes_nome = ['Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun',
                   'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez'][current_date.month - 1]
        meses_labels.append(f"{mes_nome}/{current_date.year}")
        
        # Buscar dados do mês
        mao_obra_mes = sum(d.total for d in custos_dados if d.mes == current_date.month and d.ano == current_date.year and d.tipo == 'mao_obra')
        alimentacao_mes = sum(d.total for d in custos_dados if d.mes == current_date.month and d.ano == current_date.year and d.tipo == 'alimentacao')
        veiculos_mes = sum(d.total for d in custos_dados if d.mes == current_date.month and d.ano == current_date.year and d.tipo == 'veiculo')
        
        mao_obra_dados.append(float(mao_obra_mes))
        alimentacao_dados.append(float(alimentacao_mes))
        veiculos_dados.append(float(veiculos_mes))
        
        # Próximo mês
        if current_date.month == 12:
            current_date = current_date.replace(year=current_date.year + 1, month=1)
        else:
            current_date = current_date.replace(month=current_date.month + 1)
    
    # 2. Produtividade por Departamento
    produtividade_query = db.session.query(
        Departamento.nome,
        func.avg(RegistroPonto.horas_trabalhadas).label('media_horas')
    ).join(
        Funcionario, Departamento.id == Funcionario.departamento_id
    ).join(
        RegistroPonto, Funcionario.id == RegistroPonto.funcionario_id
    ).filter(
        RegistroPonto.data >= data_inicio,
        RegistroPonto.data <= data_fim
    )
    
    if departamento_id:
        produtividade_query = produtividade_query.filter(Departamento.id == departamento_id)
    
    produtividade_dados = produtividade_query.group_by(Departamento.nome).all()
    
    # 3. Distribuição de Custos
    distribuicao_dados = db.session.query(
        CustoObra.tipo,
        func.sum(CustoObra.valor).label('total')
    ).filter(
        CustoObra.data >= data_inicio,
        CustoObra.data <= data_fim
    )
    
    if obra_id:
        distribuicao_dados = distribuicao_dados.filter(CustoObra.obra_id == obra_id)
    
    distribuicao_dados = distribuicao_dados.group_by(CustoObra.tipo).all()
    
    # 4. Horas Trabalhadas vs Extras
    horas_query = db.session.query(
        extract('month', RegistroPonto.data).label('mes'),
        extract('year', RegistroPonto.data).label('ano'),
        func.sum(RegistroPonto.horas_trabalhadas).label('horas_normais'),
        func.sum(RegistroPonto.horas_extras).label('horas_extras')
    ).filter(
        RegistroPonto.data >= data_inicio,
        RegistroPonto.data <= data_fim
    )
    
    if departamento_id:
        horas_query = horas_query.join(Funcionario).filter(Funcionario.departamento_id == departamento_id)
    
    horas_dados = horas_query.group_by(
        extract('year', RegistroPonto.data),
        extract('month', RegistroPonto.data)
    ).all()
    
    # Processar dados de horas
    horas_normais_dados = []
    horas_extras_dados = []
    
    current_date = data_inicio.replace(day=1)
    while current_date <= data_fim:
        horas_mes = next((h for h in horas_dados if h.mes == current_date.month and h.ano == current_date.year), None)
        
        horas_normais_dados.append(float(horas_mes.horas_normais or 0) if horas_mes else 0)
        horas_extras_dados.append(float(horas_mes.horas_extras or 0) if horas_mes else 0)
        
        # Próximo mês
        if current_date.month == 12:
            current_date = current_date.replace(year=current_date.year + 1, month=1)
        else:
            current_date = current_date.replace(month=current_date.month + 1)
    
    return jsonify({
        'evolucao': {
            'labels': meses_labels,
            'mao_obra': mao_obra_dados,
            'alimentacao': alimentacao_dados,
            'veiculos': veiculos_dados
        },
        'produtividade': {
            'labels': [p.nome for p in produtividade_dados],
            'valores': [float(p.media_horas or 0) for p in produtividade_dados]
        },
        'distribuicao': {
            'valores': [float(d.total) for d in distribuicao_dados]
        },
        'horas': {
            'labels': meses_labels,
            'normais': horas_normais_dados,
            'extras': horas_extras_dados
        }
    })

@main_bp.route('/relatorios/gerar/<tipo>', methods=['POST'])
@login_required
def gerar_relatorio(tipo):
    filtros = request.get_json()
    
    # Processar filtros
    data_inicio = datetime.strptime(filtros.get('dataInicio', ''), '%Y-%m-%d').date() if filtros.get('dataInicio') else None
    data_fim = datetime.strptime(filtros.get('dataFim', ''), '%Y-%m-%d').date() if filtros.get('dataFim') else None
    obra_id = int(filtros.get('obra')) if filtros.get('obra') else None
    departamento_id = int(filtros.get('departamento')) if filtros.get('departamento') else None
    
    if tipo == 'funcionarios':
        query = Funcionario.query
        if departamento_id:
            query = query.filter_by(departamento_id=departamento_id)
        funcionarios = query.all()
        
        html = '<div class="table-responsive"><table class="table table-striped">'
        html += '<thead><tr><th>Código</th><th>Nome</th><th>CPF</th><th>Departamento</th><th>Função</th><th>Data Admissão</th><th>Salário</th><th>Status</th></tr></thead><tbody>'
        
        for f in funcionarios:
            status_badge = '<span class="badge bg-success">Ativo</span>' if f.ativo else '<span class="badge bg-danger">Inativo</span>'
            html += f'<tr><td>{f.codigo or "-"}</td><td>{f.nome}</td><td>{f.cpf}</td><td>{f.departamento_ref.nome if f.departamento_ref else "-"}</td>'
            html += f'<td>{f.funcao_ref.nome if f.funcao_ref else "-"}</td><td>{f.data_admissao.strftime("%d/%m/%Y") if f.data_admissao else "-"}</td>'
            html += f'<td>R$ {f.salario:,.2f}</td><td>{status_badge}</td></tr>'
        
        html += '</tbody></table></div>'
        
        return jsonify({
            'titulo': f'Lista de Funcionários ({len(funcionarios)} registros)',
            'html': html
        })
    
    elif tipo == 'ponto':
        query = RegistroPonto.query
        if data_inicio:
            query = query.filter(RegistroPonto.data >= data_inicio)
        if data_fim:
            query = query.filter(RegistroPonto.data <= data_fim)
        if obra_id:
            query = query.filter_by(obra_id=obra_id)
        if departamento_id:
            query = query.join(Funcionario).filter(Funcionario.departamento_id == departamento_id)
        
        registros = query.join(Funcionario).order_by(RegistroPonto.data.desc()).all()
        
        html = '<div class="table-responsive"><table class="table table-striped">'
        html += '<thead><tr><th>Data</th><th>Funcionário</th><th>Obra</th><th>Entrada</th><th>Saída</th><th>Horas Trabalhadas</th><th>Horas Extras</th><th>Atrasos</th></tr></thead><tbody>'
        
        for r in registros:
            html += f'<tr><td>{r.data.strftime("%d/%m/%Y")}</td>'
            html += f'<td>{r.funcionario.nome}</td>'
            html += f'<td>{r.obra.nome if r.obra else "-"}</td>'
            html += f'<td>{r.hora_entrada.strftime("%H:%M") if r.hora_entrada else "-"}</td>'
            html += f'<td>{r.hora_saida.strftime("%H:%M") if r.hora_saida else "-"}</td>'
            html += f'<td>{r.horas_trabalhadas:.2f}h</td>'
            html += f'<td>{r.horas_extras:.2f}h</td>'
            html += f'<td>{r.total_atraso_minutos or 0} min</td></tr>'
        
        html += '</tbody></table></div>'
        
        return jsonify({
            'titulo': f'Relatório de Ponto ({len(registros)} registros)',
            'html': html
        })
    
    elif tipo == 'horas-extras':
        query = RegistroPonto.query.filter(RegistroPonto.horas_extras > 0)
        if data_inicio:
            query = query.filter(RegistroPonto.data >= data_inicio)
        if data_fim:
            query = query.filter(RegistroPonto.data <= data_fim)
        if obra_id:
            query = query.filter_by(obra_id=obra_id)
        if departamento_id:
            query = query.join(Funcionario).filter(Funcionario.departamento_id == departamento_id)
        
        registros = query.join(Funcionario).order_by(RegistroPonto.data.desc()).all()
        
        html = '<div class="table-responsive"><table class="table table-striped">'
        html += '<thead><tr><th>Data</th><th>Funcionário</th><th>Obra</th><th>Horas Extras</th><th>Valor Estimado</th></tr></thead><tbody>'
        
        total_horas = 0
        total_valor = 0
        
        for r in registros:
            valor_hora = (r.funcionario.salario / 220) * 1.5 if r.funcionario.salario else 0
            valor_extras = r.horas_extras * valor_hora
            total_horas += r.horas_extras
            total_valor += valor_extras
            
            html += f'<tr><td>{r.data.strftime("%d/%m/%Y")}</td>'
            html += f'<td>{r.funcionario.nome}</td>'
            html += f'<td>{r.obra.nome if r.obra else "-"}</td>'
            html += f'<td>{r.horas_extras:.2f}h</td>'
            html += f'<td>R$ {valor_extras:.2f}</td></tr>'
        
        html += f'<tr class="table-info"><td colspan="3"><strong>TOTAL</strong></td><td><strong>{total_horas:.2f}h</strong></td><td><strong>R$ {total_valor:.2f}</strong></td></tr>'
        html += '</tbody></table></div>'
        
        return jsonify({
            'titulo': f'Relatório de Horas Extras ({len(registros)} registros)',
            'html': html
        })
    
    elif tipo == 'alimentacao':
        query = RegistroAlimentacao.query
        if data_inicio:
            query = query.filter(RegistroAlimentacao.data >= data_inicio)
        if data_fim:
            query = query.filter(RegistroAlimentacao.data <= data_fim)
        if obra_id:
            query = query.filter_by(obra_id=obra_id)
        if departamento_id:
            query = query.join(Funcionario).filter(Funcionario.departamento_id == departamento_id)
        
        registros = query.join(Funcionario).order_by(RegistroAlimentacao.data.desc()).all()
        
        html = '<div class="table-responsive"><table class="table table-striped">'
        html += '<thead><tr><th>Data</th><th>Funcionário</th><th>Tipo</th><th>Restaurante</th><th>Obra</th><th>Valor</th></tr></thead><tbody>'
        
        total_valor = 0
        
        for r in registros:
            total_valor += r.valor
            html += f'<tr><td>{r.data.strftime("%d/%m/%Y")}</td>'
            html += f'<td>{r.funcionario.nome}</td>'
            html += f'<td>{r.tipo.title()}</td>'
            html += f'<td>{r.restaurante.nome if r.restaurante else "-"}</td>'
            html += f'<td>{r.obra.nome if r.obra else "-"}</td>'
            html += f'<td>R$ {r.valor:.2f}</td></tr>'
        
        html += f'<tr class="table-info"><td colspan="5"><strong>TOTAL</strong></td><td><strong>R$ {total_valor:.2f}</strong></td></tr>'
        html += '</tbody></table></div>'
        
        return jsonify({
            'titulo': f'Relatório de Alimentação ({len(registros)} registros)',
            'html': html
        })
    
    elif tipo == 'obras':
        query = Obra.query
        if obra_id:
            query = query.filter_by(id=obra_id)
        
        obras = query.all()
        
        html = '<div class="table-responsive"><table class="table table-striped">'
        html += '<thead><tr><th>Nome</th><th>Responsável</th><th>Data Início</th><th>Previsão Fim</th><th>Orçamento</th><th>Status</th><th>Funcionários</th></tr></thead><tbody>'
        
        for obra in obras:
            funcionarios_obra = db.session.query(func.count(FuncionarioObra.id)).filter_by(obra_id=obra.id).scalar()
            
            html += f'<tr><td>{obra.nome}</td>'
            html += f'<td>{obra.responsavel.nome if obra.responsavel else "-"}</td>'
            html += f'<td>{obra.data_inicio.strftime("%d/%m/%Y") if obra.data_inicio else "-"}</td>'
            html += f'<td>{obra.data_previsao_fim.strftime("%d/%m/%Y") if obra.data_previsao_fim else "-"}</td>'
            html += f'<td>R$ {obra.orcamento:,.2f}</td>'
            html += f'<td><span class="badge bg-primary">{obra.status}</span></td>'
            html += f'<td>{funcionarios_obra}</td></tr>'
        
        html += '</tbody></table></div>'
        
        return jsonify({
            'titulo': f'Lista de Obras ({len(obras)} registros)',
            'html': html
        })
    
    elif tipo == 'custos-obra':
        query = CustoObra.query
        if data_inicio:
            query = query.filter(CustoObra.data >= data_inicio)
        if data_fim:
            query = query.filter(CustoObra.data <= data_fim)
        if obra_id:
            query = query.filter_by(obra_id=obra_id)
        
        custos = query.join(Obra).order_by(CustoObra.data.desc()).all()
        
        html = '<div class="table-responsive"><table class="table table-striped">'
        html += '<thead><tr><th>Data</th><th>Obra</th><th>Tipo</th><th>Descrição</th><th>Valor</th></tr></thead><tbody>'
        
        total_custos = 0
        
        for custo in custos:
            total_custos += custo.valor
            html += f'<tr><td>{custo.data.strftime("%d/%m/%Y")}</td>'
            html += f'<td>{custo.obra.nome}</td>'
            html += f'<td>{custo.tipo.title()}</td>'
            html += f'<td>{custo.descricao or "-"}</td>'
            html += f'<td>R$ {custo.valor:.2f}</td></tr>'
        
        html += f'<tr class="table-info"><td colspan="4"><strong>TOTAL</strong></td><td><strong>R$ {total_custos:.2f}</strong></td></tr>'
        html += '</tbody></table></div>'
        
        return jsonify({
            'titulo': f'Custos por Obra ({len(custos)} registros)',
            'html': html
        })
    
    elif tipo == 'veiculos':
        query = Veiculo.query
        veiculos = query.all()
        
        html = '<div class="table-responsive"><table class="table table-striped">'
        html += '<thead><tr><th>Placa</th><th>Marca/Modelo</th><th>Ano</th><th>Tipo</th><th>KM Atual</th><th>Status</th><th>Próxima Manutenção</th></tr></thead><tbody>'
        
        for veiculo in veiculos:
            html += f'<tr><td>{veiculo.placa}</td>'
            html += f'<td>{veiculo.marca} {veiculo.modelo}</td>'
            html += f'<td>{veiculo.ano or "-"}</td>'
            html += f'<td>{veiculo.tipo}</td>'
            html += f'<td>{veiculo.km_atual or 0:,} km</td>'
            html += f'<td><span class="badge bg-info">{veiculo.status}</span></td>'
            html += f'<td>{veiculo.data_proxima_manutencao.strftime("%d/%m/%Y") if veiculo.data_proxima_manutencao else "-"}</td></tr>'
        
        html += '</tbody></table></div>'
        
        return jsonify({
            'titulo': f'Relatório de Veículos ({len(veiculos)} registros)',
            'html': html
        })
    
    elif tipo == 'dashboard-executivo':
        # Dados consolidados para dashboard executivo
        total_funcionarios = Funcionario.query.filter_by(ativo=True).count()
        total_obras = Obra.query.filter_by(status='Em andamento').count()
        total_veiculos = Veiculo.query.count()
        
        # Custos do mês atual
        hoje = date.today()
        primeiro_dia_mes = hoje.replace(day=1)
        
        custos_mes = db.session.query(func.sum(CustoObra.valor)).filter(
            CustoObra.data >= primeiro_dia_mes,
            CustoObra.data <= hoje
        ).scalar() or 0
        
        # Horas trabalhadas do mês
        horas_mes = db.session.query(func.sum(RegistroPonto.horas_trabalhadas)).filter(
            RegistroPonto.data >= primeiro_dia_mes,
            RegistroPonto.data <= hoje
        ).scalar() or 0
        
        # Alimentação do mês
        alimentacao_mes = db.session.query(func.sum(RegistroAlimentacao.valor)).filter(
            RegistroAlimentacao.data >= primeiro_dia_mes,
            RegistroAlimentacao.data <= hoje
        ).scalar() or 0
        
        html = '<div class="row">'
        html += f'<div class="col-md-3"><div class="card text-center"><div class="card-body"><h2 class="text-primary">{total_funcionarios}</h2><p>Funcionários Ativos</p></div></div></div>'
        html += f'<div class="col-md-3"><div class="card text-center"><div class="card-body"><h2 class="text-success">{total_obras}</h2><p>Obras em Andamento</p></div></div></div>'
        html += f'<div class="col-md-3"><div class="card text-center"><div class="card-body"><h2 class="text-warning">{total_veiculos}</h2><p>Veículos</p></div></div></div>'
        html += f'<div class="col-md-3"><div class="card text-center"><div class="card-body"><h2 class="text-info">R$ {custos_mes:,.0f}</h2><p>Custos do Mês</p></div></div></div>'
        html += '</div>'
        
        html += '<div class="row mt-4">'
        html += f'<div class="col-md-6"><div class="card"><div class="card-body"><h5>Resumo Mensal</h5><p><strong>Horas Trabalhadas:</strong> {horas_mes:.0f}h</p><p><strong>Custo por Hora:</strong> R$ {(custos_mes/horas_mes if horas_mes > 0 else 0):.2f}</p></div></div></div>'
        html += f'<div class="col-md-6"><div class="card"><div class="card-body"><h5>Alimentação</h5><p><strong>Gasto Mensal:</strong> R$ {alimentacao_mes:,.2f}</p><p><strong>Média por Funcionário:</strong> R$ {(alimentacao_mes/total_funcionarios if total_funcionarios > 0 else 0):.2f}</p></div></div></div>'
        html += '</div>'
        
        return jsonify({
            'titulo': 'Dashboard Executivo',
            'html': html
        })
    
    elif tipo == 'progresso-obras':
        obras = Obra.query.all()
        
        html = '<div class="table-responsive"><table class="table table-striped">'
        html += '<thead><tr><th>Obra</th><th>Progresso</th><th>Orçamento</th><th>Gasto Atual</th><th>Saldo</th><th>% Utilizado</th></tr></thead><tbody>'
        
        for obra in obras:
            gasto_atual = db.session.query(func.sum(CustoObra.valor)).filter_by(obra_id=obra.id).scalar() or 0
            saldo = obra.orcamento - gasto_atual
            percentual = (gasto_atual / obra.orcamento * 100) if obra.orcamento > 0 else 0
            
            cor_saldo = 'text-success' if saldo > 0 else 'text-danger'
            
            html += f'<tr><td>{obra.nome}</td>'
            html += f'<td><span class="badge bg-info">{obra.status}</span></td>'
            html += f'<td>R$ {obra.orcamento:,.2f}</td>'
            html += f'<td>R$ {gasto_atual:,.2f}</td>'
            html += f'<td class="{cor_saldo}">R$ {saldo:,.2f}</td>'
            html += f'<td>{percentual:.1f}%</td></tr>'
        
        html += '</tbody></table></div>'
        
        return jsonify({
            'titulo': f'Progresso das Obras ({len(obras)} registros)',
            'html': html
        })
    
    elif tipo == 'rentabilidade':
        # Análise de rentabilidade por obra
        obras = Obra.query.all()
        
        html = '<div class="table-responsive"><table class="table table-striped">'
        html += '<thead><tr><th>Obra</th><th>Receita Prevista</th><th>Custos</th><th>Margem</th><th>% Margem</th><th>Status</th></tr></thead><tbody>'
        
        for obra in obras:
            custos_obra = db.session.query(func.sum(CustoObra.valor)).filter_by(obra_id=obra.id).scalar() or 0
            receita_prevista = obra.orcamento * 1.3  # Assumindo 30% de margem padrão
            margem = receita_prevista - custos_obra
            percentual_margem = (margem / receita_prevista * 100) if receita_prevista > 0 else 0
            
            cor_margem = 'text-success' if margem > 0 else 'text-danger'
            
            html += f'<tr><td>{obra.nome}</td>'
            html += f'<td>R$ {receita_prevista:,.2f}</td>'
            html += f'<td>R$ {custos_obra:,.2f}</td>'
            html += f'<td class="{cor_margem}">R$ {margem:,.2f}</td>'
            html += f'<td class="{cor_margem}">{percentual_margem:.1f}%</td>'
            html += f'<td><span class="badge bg-primary">{obra.status}</span></td></tr>'
        
        html += '</tbody></table></div>'
        
        return jsonify({
            'titulo': f'Análise de Rentabilidade ({len(obras)} obras)',
            'html': html
        })
    
    else:
        return jsonify({
            'titulo': 'Relatório não implementado',
            'html': '<div class="alert alert-info">Este tipo de relatório ainda não foi implementado.</div>'
        })
        query = RegistroPonto.query
        if data_inicio:
            query = query.filter(RegistroPonto.data >= data_inicio)
        if data_fim:
            query = query.filter(RegistroPonto.data <= data_fim)
        if obra_id:
            query = query.filter_by(obra_id=obra_id)
        
        registros = query.order_by(RegistroPonto.data.desc()).limit(100).all()
        
        html = '<div class="table-responsive"><table class="table table-striped">'
        html += '<thead><tr><th>Funcionário</th><th>Data</th><th>Entrada</th><th>Saída Almoço</th><th>Retorno Almoço</th><th>Saída</th><th>Horas Trabalhadas</th><th>Horas Extras</th><th>Obra</th></tr></thead><tbody>'
        
        for r in registros:
            html += f'<tr><td>{r.funcionario_ref.nome}</td><td>{r.data.strftime("%d/%m/%Y")}</td>'
            html += f'<td>{r.hora_entrada.strftime("%H:%M") if r.hora_entrada else "-"}</td>'
            html += f'<td>{r.hora_almoco_saida.strftime("%H:%M") if r.hora_almoco_saida else "-"}</td>'
            html += f'<td>{r.hora_almoco_retorno.strftime("%H:%M") if r.hora_almoco_retorno else "-"}</td>'
            html += f'<td>{r.hora_saida.strftime("%H:%M") if r.hora_saida else "-"}</td>'
            horas_trab = f"{r.horas_trabalhadas:.2f}h" if r.horas_trabalhadas else "0.00h"
            horas_ext = f"{r.horas_extras:.2f}h" if r.horas_extras else "0.00h"
            html += f'<td>{horas_trab}</td><td>{horas_ext}</td>'
            html += f'<td>{r.obra_ref.nome if r.obra_ref else "-"}</td></tr>'
        
        html += '</tbody></table></div>'
        
        return jsonify({
            'titulo': 'Relatório de Ponto',
            'html': html
        })
    
    elif tipo == 'horas-extras':
        query = RegistroPonto.query.filter(RegistroPonto.horas_extras > 0)
        if data_inicio:
            query = query.filter(RegistroPonto.data >= data_inicio)
        if data_fim:
            query = query.filter(RegistroPonto.data <= data_fim)
        if departamento_id:
            query = query.join(Funcionario).filter(Funcionario.departamento_id == departamento_id)
        if obra_id:
            query = query.filter_by(obra_id=obra_id)
        
        registros = query.order_by(RegistroPonto.data.desc()).limit(100).all()
        
        html = '<div class="table-responsive"><table class="table table-striped">'
        html += '<thead><tr><th>Funcionário</th><th>Data</th><th>Horas Extras</th><th>Valor Aproximado</th><th>Obra</th><th>Departamento</th></tr></thead><tbody>'
        
        total_horas = 0
        total_valor = 0
        
        for r in registros:
            valor_hora = (r.funcionario_ref.salario / 220) * 1.5 if r.funcionario_ref.salario else 0  # 50% adicional
            valor_extra = (r.horas_extras or 0) * valor_hora
            total_horas += r.horas_extras or 0
            total_valor += valor_extra
            
            html += f'<tr><td>{r.funcionario_ref.nome}</td><td>{r.data.strftime("%d/%m/%Y")}</td>'
            html += f'<td>{r.horas_extras:.2f}h</td><td>R$ {valor_extra:.2f}</td>'
            html += f'<td>{r.obra_ref.nome if r.obra_ref else "-"}</td>'
            html += f'<td>{r.funcionario_ref.departamento_ref.nome if r.funcionario_ref.departamento_ref else "-"}</td></tr>'
        
        html += '</tbody></table></div>'
        html += f'<div class="mt-3"><strong>Total: {total_horas:.2f} horas extras • R$ {total_valor:,.2f}</strong></div>'
        
        return jsonify({
            'titulo': 'Relatório de Horas Extras',
            'html': html
        })
    
    elif tipo == 'alimentacao':
        query = RegistroAlimentacao.query
        if data_inicio:
            query = query.filter(RegistroAlimentacao.data >= data_inicio)
        if data_fim:
            query = query.filter(RegistroAlimentacao.data <= data_fim)
        if obra_id:
            query = query.filter_by(obra_id=obra_id)
        if departamento_id:
            query = query.join(Funcionario).filter(Funcionario.departamento_id == departamento_id)
        
        registros = query.order_by(RegistroAlimentacao.data.desc()).limit(200).all()
        
        html = '<div class="table-responsive"><table class="table table-striped">'
        html += '<thead><tr><th>Funcionário</th><th>Data</th><th>Tipo</th><th>Valor</th><th>Restaurante</th><th>Obra</th></tr></thead><tbody>'
        
        total_valor = 0
        
        for r in registros:
            total_valor += r.valor or 0
            tipo_labels = {
                'cafe': 'Café da Manhã',
                'almoco': 'Almoço', 
                'jantar': 'Jantar',
                'lanche': 'Lanche',
                'marmita': 'Marmita',
                'refeicao_local': 'Refeição no Local',
                'outros': 'Outros'
            }
            
            html += f'<tr><td>{r.funcionario_ref.nome}</td><td>{r.data.strftime("%d/%m/%Y")}</td>'
            html += f'<td>{tipo_labels.get(r.tipo, r.tipo)}</td><td>R$ {r.valor:.2f}</td>'
            html += f'<td>{r.restaurante_ref.nome if r.restaurante_ref else "-"}</td>'
            html += f'<td>{r.obra_ref.nome if r.obra_ref else "-"}</td></tr>'
        
        html += '</tbody></table></div>'
        html += f'<div class="mt-3"><strong>Total Gasto: R$ {total_valor:,.2f}</strong></div>'
        
        return jsonify({
            'titulo': 'Relatório de Alimentação',
            'html': html
        })
    
    elif tipo == 'obras':
        query = Obra.query
        if obra_id:
            query = query.filter_by(id=obra_id)
        obras = query.all()
        
        html = '<div class="table-responsive"><table class="table table-striped">'
        html += '<thead><tr><th>Nome</th><th>Endereço</th><th>Data Início</th><th>Previsão Fim</th><th>Orçamento</th><th>Status</th><th>Responsável</th></tr></thead><tbody>'
        
        for o in obras:
            status_badges = {
                'Em andamento': 'bg-primary',
                'Concluída': 'bg-success',
                'Pausada': 'bg-warning',
                'Cancelada': 'bg-danger'
            }
            status_badge = f'<span class="badge {status_badges.get(o.status, "bg-secondary")}">{o.status}</span>'
            
            html += f'<tr><td>{o.nome}</td><td>{o.endereco or "-"}</td>'
            html += f'<td>{o.data_inicio.strftime("%d/%m/%Y") if o.data_inicio else "-"}</td>'
            html += f'<td>{o.data_previsao_fim.strftime("%d/%m/%Y") if o.data_previsao_fim else "-"}</td>'
            html += f'<td>R$ {o.orcamento:,.2f}</td><td>{status_badge}</td>'
            
            responsavel = Funcionario.query.get(o.responsavel_id) if o.responsavel_id else None
            html += f'<td>{responsavel.nome if responsavel else "-"}</td></tr>'
        
        html += '</tbody></table></div>'
        
        return jsonify({
            'titulo': 'Lista de Obras',
            'html': html
        })
    
    elif tipo == 'custos-obra':
        query = CustoObra.query
        if data_inicio:
            query = query.filter(CustoObra.data >= data_inicio)
        if data_fim:
            query = query.filter(CustoObra.data <= data_fim)
        if obra_id:
            query = query.filter_by(obra_id=obra_id)
        
        custos = query.order_by(CustoObra.data.desc()).limit(100).all()
        
        html = '<div class="table-responsive"><table class="table table-striped">'
        html += '<thead><tr><th>Data</th><th>Obra</th><th>Tipo</th><th>Descrição</th><th>Valor</th><th>Observações</th></tr></thead><tbody>'
        
        total_custos = 0
        tipos_labels = {
            'mao_obra': 'Mão de Obra',
            'material': 'Material',
            'servico': 'Serviço',
            'veiculo': 'Veículo',
            'alimentacao': 'Alimentação'
        }
        
        for c in custos:
            total_custos += c.valor or 0
            html += f'<tr><td>{c.data.strftime("%d/%m/%Y")}</td>'
            html += f'<td>{c.obra_ref.nome if c.obra_ref else "-"}</td>'
            html += f'<td>{tipos_labels.get(c.tipo, c.tipo)}</td>'
            html += f'<td>{c.descricao or "-"}</td><td>R$ {c.valor:,.2f}</td>'
            html += f'<td>{c.observacoes or "-"}</td></tr>'
        
        html += '</tbody></table></div>'
        html += f'<div class="mt-3"><strong>Total de Custos: R$ {total_custos:,.2f}</strong></div>'
        
        return jsonify({
            'titulo': 'Custos por Obra',
            'html': html
        })
    
    elif tipo == 'progresso-obras':
        obras = Obra.query.all()
        from datetime import date
        
        html = '<div class="table-responsive"><table class="table table-striped">'
        html += '<thead><tr><th>Obra</th><th>Status</th><th>Progresso</th><th>Orçamento</th><th>Gastos</th><th>Restante</th><th>Prazo</th></tr></thead><tbody>'
        
        for obra in obras:
            # Calcular gastos totais
            gastos_total = db.session.query(func.sum(CustoObra.valor)).filter_by(obra_id=obra.id).scalar() or 0
            restante = (obra.orcamento or 0) - gastos_total
            
            # Calcular progresso baseado no orçamento
            progresso = (gastos_total / obra.orcamento * 100) if obra.orcamento > 0 else 0
            progresso = min(progresso, 100)
            
            # Status do prazo
            if obra.data_previsao_fim:
                dias_restantes = (obra.data_previsao_fim - date.today()).days
                prazo_status = "No prazo" if dias_restantes > 0 else f"Atrasado ({abs(dias_restantes)} dias)"
                prazo_class = "text-success" if dias_restantes > 0 else "text-danger"
            else:
                prazo_status = "Sem prazo definido"
                prazo_class = "text-muted"
            
            status_badges = {
                'Em andamento': 'bg-primary',
                'Concluída': 'bg-success',
                'Pausada': 'bg-warning',
                'Cancelada': 'bg-danger'
            }
            status_badge = f'<span class="badge {status_badges.get(obra.status, "bg-secondary")}">{obra.status}</span>'
            
            html += f'<tr><td>{obra.nome}</td><td>{status_badge}</td>'
            html += f'<td><div class="progress"><div class="progress-bar" style="width: {progresso:.1f}%">{progresso:.1f}%</div></div></td>'
            html += f'<td>R$ {obra.orcamento:,.2f}</td><td>R$ {gastos_total:,.2f}</td><td>R$ {restante:,.2f}</td>'
            html += f'<td><span class="{prazo_class}">{prazo_status}</span></td></tr>'
        
        html += '</tbody></table></div>'
        
        return jsonify({
            'titulo': 'Progresso das Obras',
            'html': html
        })
    
    elif tipo == 'veiculos':
        veiculos = Veiculo.query.all()
        
        html = '<div class="table-responsive"><table class="table table-striped">'
        html += '<thead><tr><th>Placa</th><th>Marca/Modelo</th><th>Ano</th><th>Tipo</th><th>Status</th><th>KM Atual</th><th>Última Manutenção</th><th>Próxima Manutenção</th></tr></thead><tbody>'
        
        for v in veiculos:
            status_badges = {
                'Disponível': 'bg-success',
                'Em uso': 'bg-primary',
                'Manutenção': 'bg-warning',
                'Indisponível': 'bg-danger'
            }
            status_badge = f'<span class="badge {status_badges.get(v.status, "bg-secondary")}">{v.status}</span>'
            
            html += f'<tr><td>{v.placa}</td><td>{v.marca} {v.modelo}</td><td>{v.ano or "-"}</td>'
            html += f'<td>{v.tipo}</td><td>{status_badge}</td><td>{v.km_atual or "-"}</td>'
            html += f'<td>{v.data_ultima_manutencao.strftime("%d/%m/%Y") if v.data_ultima_manutencao else "-"}</td>'
            html += f'<td>{v.data_proxima_manutencao.strftime("%d/%m/%Y") if v.data_proxima_manutencao else "-"}</td></tr>'
        
        html += '</tbody></table></div>'
        
        return jsonify({
            'titulo': 'Relatório de Veículos',
            'html': html
        })
    
    elif tipo == 'rentabilidade':
        obras = Obra.query.all()
        
        html = '<div class="table-responsive"><table class="table table-striped">'
        html += '<thead><tr><th>Obra</th><th>Orçamento</th><th>Custos</th><th>Margem</th><th>Rentabilidade</th><th>Status</th></tr></thead><tbody>'
        
        total_orcamento = 0
        total_custos = 0
        
        for obra in obras:
            custos_obra = db.session.query(func.sum(CustoObra.valor)).filter_by(obra_id=obra.id).scalar() or 0
            margem = (obra.orcamento or 0) - custos_obra
            rentabilidade = (margem / obra.orcamento * 100) if obra.orcamento > 0 else 0
            
            total_orcamento += obra.orcamento or 0
            total_custos += custos_obra
            
            rent_class = "text-success" if rentabilidade > 0 else "text-danger"
            
            html += f'<tr><td>{obra.nome}</td><td>R$ {obra.orcamento:,.2f}</td>'
            html += f'<td>R$ {custos_obra:,.2f}</td><td>R$ {margem:,.2f}</td>'
            html += f'<td><span class="{rent_class}">{rentabilidade:.1f}%</span></td>'
            html += f'<td>{obra.status}</td></tr>'
        
        margem_total = total_orcamento - total_custos
        rentabilidade_total = (margem_total / total_orcamento * 100) if total_orcamento > 0 else 0
        
        html += '</tbody></table></div>'
        html += f'<div class="mt-3"><strong>Totais: Orçamento R$ {total_orcamento:,.2f} • Custos R$ {total_custos:,.2f} • Margem R$ {margem_total:,.2f} • Rentabilidade {rentabilidade_total:.1f}%</strong></div>'
        
        return jsonify({
            'titulo': 'Rentabilidade por Obra',
            'html': html
        })
    
    elif tipo == 'dashboard-executivo':
        # Dashboard executivo com KPIs gerais
        total_funcionarios = Funcionario.query.filter_by(ativo=True).count()
        total_obras = Obra.query.filter_by(status='Em andamento').count()
        total_veiculos = Veiculo.query.count()
        
        # Custos do mês atual
        hoje = date.today()
        primeiro_dia_mes = date(hoje.year, hoje.month, 1)
        custos_mes = db.session.query(func.sum(CustoObra.valor)).filter(
            CustoObra.data >= primeiro_dia_mes,
            CustoObra.data <= hoje
        ).scalar() or 0
        
        # Horas trabalhadas no mês
        horas_mes = db.session.query(func.sum(RegistroPonto.horas_trabalhadas)).filter(
            RegistroPonto.data >= primeiro_dia_mes,
            RegistroPonto.data <= hoje
        ).scalar() or 0
        
        html = '<div class="row">'
        html += f'<div class="col-md-3"><div class="card text-center"><div class="card-body"><h2 class="text-primary">{total_funcionarios}</h2><p>Funcionários Ativos</p></div></div></div>'
        html += f'<div class="col-md-3"><div class="card text-center"><div class="card-body"><h2 class="text-success">{total_obras}</h2><p>Obras em Andamento</p></div></div></div>'
        html += f'<div class="col-md-3"><div class="card text-center"><div class="card-body"><h2 class="text-warning">{total_veiculos}</h2><p>Veículos</p></div></div></div>'
        html += f'<div class="col-md-3"><div class="card text-center"><div class="card-body"><h2 class="text-info">R$ {custos_mes:,.0f}</h2><p>Custos do Mês</p></div></div></div>'
        html += '</div>'
        html += f'<div class="mt-4"><div class="card"><div class="card-body"><h5>Resumo do Mês</h5><p><strong>Horas Trabalhadas:</strong> {horas_mes:.0f}h</p><p><strong>Custo por Hora:</strong> R$ {(custos_mes/horas_mes if horas_mes > 0 else 0):.2f}</p></div></div></div>'
        
        return jsonify({
            'titulo': 'Dashboard Executivo',
            'html': html
        })
    
    else:
        return jsonify({
            'titulo': 'Relatório não implementado',
            'html': '<p>Este relatório ainda não foi implementado.</p>'
        })

@main_bp.route('/relatorios/exportar/<formato>', methods=['POST'])
@login_required
def exportar_relatorio(formato):
    from io import BytesIO
    from flask import make_response
    import csv
    from datetime import datetime
    
    # Processar filtros
    data_inicio = request.form.get('dataInicio')
    data_fim = request.form.get('dataFim')
    obra_id = request.form.get('obra')
    departamento_id = request.form.get('departamento')
    tipo_relatorio = request.form.get('tipo', 'geral')  # Adicionar tipo de relatório
    
    # Converter datas se fornecidas
    if data_inicio:
        data_inicio = datetime.strptime(data_inicio, '%Y-%m-%d').date()
    if data_fim:
        data_fim = datetime.strptime(data_fim, '%Y-%m-%d').date()
    
    if formato == 'csv':
        # Exportação CSV consolidada
        output = BytesIO()
        output.write('\ufeff'.encode('utf-8'))  # BOM for UTF-8
        
        import io
        csv_content = io.StringIO()
        writer = csv.writer(csv_content)
        
        # Cabeçalho principal
        writer.writerow(['RELATÓRIO CONSOLIDADO SIGE - SISTEMA INTEGRADO DE GESTÃO EMPRESARIAL'])
        writer.writerow(['Data de Geração:', datetime.now().strftime('%d/%m/%Y %H:%M')])
        writer.writerow(['Período:', f'{data_inicio.strftime("%d/%m/%Y") if data_inicio else "N/A"} até {data_fim.strftime("%d/%m/%Y") if data_fim else "N/A"}'])
        writer.writerow([])
        
        # Relatório de Funcionários
        writer.writerow(['=== FUNCIONÁRIOS ==='])
        writer.writerow(['Código', 'Nome', 'CPF', 'Departamento', 'Função', 'Data Admissão', 'Salário', 'Status'])
        
        query = Funcionario.query
        if departamento_id:
            query = query.filter_by(departamento_id=departamento_id)
        funcionarios = query.all()
        
        for f in funcionarios:
            writer.writerow([
                f.codigo or '',
                f.nome,
                f.cpf,
                f.departamento_ref.nome if f.departamento_ref else '',
                f.funcao_ref.nome if f.funcao_ref else '',
                f.data_admissao.strftime('%d/%m/%Y') if f.data_admissao else '',
                f'R$ {f.salario:,.2f}',
                'Ativo' if f.ativo else 'Inativo'
            ])
        
        writer.writerow([])
        
        # Relatório de Ponto (se houver filtro de data)
        if data_inicio and data_fim:
            writer.writerow(['=== REGISTROS DE PONTO ==='])
            writer.writerow(['Data', 'Funcionário', 'Obra', 'Entrada', 'Saída', 'Horas Trabalhadas', 'Horas Extras', 'Atrasos (min)'])
            
            query = RegistroPonto.query.filter(
                RegistroPonto.data >= data_inicio,
                RegistroPonto.data <= data_fim
            )
            if obra_id:
                query = query.filter_by(obra_id=obra_id)
            if departamento_id:
                query = query.join(Funcionario).filter(Funcionario.departamento_id == departamento_id)
            
            registros = query.join(Funcionario).order_by(RegistroPonto.data.desc()).all()
            
            for r in registros:
                writer.writerow([
                    r.data.strftime('%d/%m/%Y'),
                    r.funcionario.nome,
                    r.obra.nome if r.obra else '',
                    r.hora_entrada.strftime('%H:%M') if r.hora_entrada else '',
                    r.hora_saida.strftime('%H:%M') if r.hora_saida else '',
                    f'{r.horas_trabalhadas:.2f}h',
                    f'{r.horas_extras:.2f}h',
                    r.total_atraso_minutos or 0
                ])
            
            writer.writerow([])
        
        # Relatório de Alimentação (se houver filtro de data)
        if data_inicio and data_fim:
            writer.writerow(['=== ALIMENTAÇÃO ==='])
            writer.writerow(['Data', 'Funcionário', 'Tipo', 'Restaurante', 'Obra', 'Valor'])
            
            query = RegistroAlimentacao.query.filter(
                RegistroAlimentacao.data >= data_inicio,
                RegistroAlimentacao.data <= data_fim
            )
            if obra_id:
                query = query.filter_by(obra_id=obra_id)
            if departamento_id:
                query = query.join(Funcionario).filter(Funcionario.departamento_id == departamento_id)
            
            registros = query.join(Funcionario).order_by(RegistroAlimentacao.data.desc()).all()
            
            total_alimentacao = 0
            for r in registros:
                total_alimentacao += r.valor
                writer.writerow([
                    r.data.strftime('%d/%m/%Y'),
                    r.funcionario.nome,
                    r.tipo.title(),
                    r.restaurante.nome if r.restaurante else '',
                    r.obra.nome if r.obra else '',
                    f'R$ {r.valor:.2f}'
                ])
            
            writer.writerow(['', '', '', '', 'TOTAL', f'R$ {total_alimentacao:.2f}'])
        
        # Finalizar CSV
        output.write(csv_content.getvalue().encode('utf-8'))
        output.seek(0)
        
        response = make_response(output.getvalue())
        response.headers["Content-Disposition"] = f"attachment; filename=relatorio_sige_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        response.headers["Content-type"] = "text/csv; charset=utf-8"
        
        return response
    
    elif formato == 'excel':
        # Exportação Excel usando openpyxl
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        wb = Workbook()
        
        # Planilha de Funcionários
        ws_funcionarios = wb.active
        ws_funcionarios.title = "Funcionários"
        
        # Cabeçalho
        headers = ['Código', 'Nome', 'CPF', 'Departamento', 'Função', 'Data Admissão', 'Salário', 'Status']
        ws_funcionarios.append(headers)
        
        # Estilizar cabeçalho
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws_funcionarios.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Dados dos funcionários
        query = Funcionario.query
        if departamento_id:
            query = query.filter_by(departamento_id=departamento_id)
        funcionarios = query.all()
        
        for f in funcionarios:
            ws_funcionarios.append([
                f.codigo or '',
                f.nome,
                f.cpf,
                f.departamento_ref.nome if f.departamento_ref else '',
                f.funcao_ref.nome if f.funcao_ref else '',
                f.data_admissao.strftime('%d/%m/%Y') if f.data_admissao else '',
                f.salario,
                'Ativo' if f.ativo else 'Inativo'
            ])
        
        # Ajustar largura das colunas
        for column in ws_funcionarios.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_funcionarios.column_dimensions[column_letter].width = adjusted_width
        
        # Planilha de Ponto (se houver filtro de data)
        if data_inicio and data_fim:
            ws_ponto = wb.create_sheet(title="Registros de Ponto")
            
            headers_ponto = ['Data', 'Funcionário', 'Obra', 'Entrada', 'Saída', 'Horas Trabalhadas', 'Horas Extras', 'Atrasos (min)']
            ws_ponto.append(headers_ponto)
            
            # Estilizar cabeçalho
            for col_num, header in enumerate(headers_ponto, 1):
                cell = ws_ponto.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # Dados de ponto
            query = RegistroPonto.query.filter(
                RegistroPonto.data >= data_inicio,
                RegistroPonto.data <= data_fim
            )
            if obra_id:
                query = query.filter_by(obra_id=obra_id)
            if departamento_id:
                query = query.join(Funcionario).filter(Funcionario.departamento_id == departamento_id)
            
            registros = query.join(Funcionario).order_by(RegistroPonto.data.desc()).all()
            
            for r in registros:
                ws_ponto.append([
                    r.data.strftime('%d/%m/%Y'),
                    r.funcionario.nome,
                    r.obra.nome if r.obra else '',
                    r.hora_entrada.strftime('%H:%M') if r.hora_entrada else '',
                    r.hora_saida.strftime('%H:%M') if r.hora_saida else '',
                    f'{r.horas_trabalhadas:.2f}h',
                    f'{r.horas_extras:.2f}h',
                    r.total_atraso_minutos or 0
                ])
            
            # Ajustar largura das colunas
            for column in ws_ponto.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                ws_ponto.column_dimensions[column_letter].width = adjusted_width
        
        # Salvar Excel
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = make_response(output.getvalue())
        response.headers["Content-Disposition"] = f"attachment; filename=relatorio_sige_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response.headers["Content-type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        
        return response
    
    elif formato == 'pdf':
        # Exportação PDF usando reportlab
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
        
        # Estilos
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.navy,
            spaceAfter=30,
            alignment=1  # Center
        )
        
        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Heading2'],
            fontSize=12,
            textColor=colors.darkblue,
            spaceAfter=12
        )
        
        # Elementos do PDF
        elements = []
        
        # Título
        title = Paragraph("RELATÓRIO CONSOLIDADO SIGE", title_style)
        subtitle = Paragraph("Sistema Integrado de Gestão Empresarial - Estruturas do Vale", styles['Normal'])
        
        elements.append(title)
        elements.append(subtitle)
        elements.append(Spacer(1, 20))
        
        # Informações do relatório
        info_data = [
            ['Data de Geração:', datetime.now().strftime('%d/%m/%Y %H:%M')],
            ['Período:', f'{data_inicio.strftime("%d/%m/%Y") if data_inicio else "N/A"} até {data_fim.strftime("%d/%m/%Y") if data_fim else "N/A"}'],
            ['Filtros:', f'Obra: {obra_id or "Todas"} | Departamento: {departamento_id or "Todos"}']
        ]
        
        info_table = Table(info_data, colWidths=[2*inch, 4*inch])
        info_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(info_table)
        elements.append(Spacer(1, 20))
        
        # Relatório de Funcionários
        elements.append(Paragraph("Funcionários", subtitle_style))
        
        # Dados dos funcionários
        query = Funcionario.query
        if departamento_id:
            query = query.filter_by(departamento_id=departamento_id)
        funcionarios = query.all()
        
        func_data = [['Código', 'Nome', 'Departamento', 'Função', 'Status']]
        for f in funcionarios[:20]:  # Limitar para evitar PDF muito grande
            func_data.append([
                f.codigo or '-',
                f.nome[:25] + '...' if len(f.nome) > 25 else f.nome,
                (f.departamento_ref.nome[:15] + '...') if f.departamento_ref and len(f.departamento_ref.nome) > 15 else (f.departamento_ref.nome if f.departamento_ref else '-'),
                (f.funcao_ref.nome[:15] + '...') if f.funcao_ref and len(f.funcao_ref.nome) > 15 else (f.funcao_ref.nome if f.funcao_ref else '-'),
                'Ativo' if f.ativo else 'Inativo'
            ])
        
        func_table = Table(func_data, colWidths=[0.8*inch, 2.2*inch, 1.5*inch, 1.5*inch, 0.8*inch])
        func_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
        ]))
        
        elements.append(func_table)
        
        if len(funcionarios) > 20:
            elements.append(Spacer(1, 10))
            elements.append(Paragraph(f"... e mais {len(funcionarios) - 20} funcionários (consulte relatório completo em Excel/CSV)", styles['Normal']))
        
        elements.append(Spacer(1, 20))
        
        # Resumo estatístico
        if data_inicio and data_fim:
            elements.append(Paragraph("Resumo do Período", subtitle_style))
            
            # Calcular estatísticas
            total_horas = db.session.query(func.sum(RegistroPonto.horas_trabalhadas)).filter(
                RegistroPonto.data >= data_inicio,
                RegistroPonto.data <= data_fim
            ).scalar() or 0
            
            total_alimentacao = db.session.query(func.sum(RegistroAlimentacao.valor)).filter(
                RegistroAlimentacao.data >= data_inicio,
                RegistroAlimentacao.data <= data_fim
            ).scalar() or 0
            
            resumo_data = [
                ['Funcionários Ativos:', str(len([f for f in funcionarios if f.ativo]))],
                ['Total Horas Trabalhadas:', f'{total_horas:.0f}h'],
                ['Total Gasto Alimentação:', f'R$ {total_alimentacao:,.2f}'],
                ['Média Horas/Funcionário:', f'{(total_horas/len([f for f in funcionarios if f.ativo]) if len([f for f in funcionarios if f.ativo]) > 0 else 0):.1f}h']
            ]
            
            resumo_table = Table(resumo_data, colWidths=[3*inch, 2*inch])
            resumo_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.lightblue),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            elements.append(resumo_table)
        
        # Construir PDF
        doc.build(elements)
        pdf = buffer.getvalue()
        buffer.close()
        
        response = make_response(pdf)
        response.headers["Content-Disposition"] = f"attachment; filename=relatorio_sige_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        response.headers["Content-Type"] = "application/pdf"
        
        return response
    
    else:
        return jsonify({'error': 'Formato não suportado'}), 400ta_fim.strftime("%d/%m/%Y") if data_fim else "N/A"}'])
        writer.writerow([])
        
        # 1. Dados de Funcionários
        writer.writerow(['=== FUNCIONÁRIOS ==='])
        writer.writerow(['Nome', 'CPF', 'Departamento', 'Função', 'Data Admissão', 'Salário', 'Status'])
        
        query = Funcionario.query
        if departamento_id:
            query = query.filter_by(departamento_id=departamento_id)
        
        for f in query.all():
            writer.writerow([
                f.nome,
                f.cpf,
                f.departamento_ref.nome if f.departamento_ref else '',
                f.funcao_ref.nome if f.funcao_ref else '',
                f.data_admissao.strftime('%d/%m/%Y') if f.data_admissao else '',
                f'R$ {f.salario:,.2f}' if f.salario else '',
                'Ativo' if f.ativo else 'Inativo'
            ])
        
        writer.writerow([])
        
        # 2. Dados de Ponto (se filtros de data fornecidos)
        if data_inicio and data_fim:
            writer.writerow(['=== REGISTROS DE PONTO ==='])
            writer.writerow(['Funcionário', 'Data', 'Entrada', 'Saída Almoço', 'Retorno Almoço', 'Saída', 'Horas Trabalhadas', 'Horas Extras', 'Obra'])
            
            query = RegistroPonto.query.filter(
                RegistroPonto.data >= data_inicio,
                RegistroPonto.data <= data_fim
            )
            if obra_id:
                query = query.filter_by(obra_id=obra_id)
            
            for r in query.order_by(RegistroPonto.data.desc()).limit(500).all():
                writer.writerow([
                    r.funcionario_ref.nome,
                    r.data.strftime('%d/%m/%Y'),
                    r.hora_entrada.strftime('%H:%M') if r.hora_entrada else '',
                    r.hora_almoco_saida.strftime('%H:%M') if r.hora_almoco_saida else '',
                    r.hora_almoco_retorno.strftime('%H:%M') if r.hora_almoco_retorno else '',
                    r.hora_saida.strftime('%H:%M') if r.hora_saida else '',
                    f'{r.horas_trabalhadas:.2f}h' if r.horas_trabalhadas else '',
                    f'{r.horas_extras:.2f}h' if r.horas_extras else '',
                    r.obra_ref.nome if r.obra_ref else ''
                ])
            
            writer.writerow([])
        
        # 3. Obras
        writer.writerow(['=== OBRAS ==='])
        writer.writerow(['Nome', 'Endereço', 'Data Início', 'Previsão Fim', 'Orçamento', 'Status', 'Responsável'])
        
        query = Obra.query
        if obra_id:
            query = query.filter_by(id=obra_id)
        
        for o in query.all():
            responsavel = Funcionario.query.get(o.responsavel_id) if o.responsavel_id else None
            writer.writerow([
                o.nome,
                o.endereco or '',
                o.data_inicio.strftime('%d/%m/%Y') if o.data_inicio else '',
                o.data_previsao_fim.strftime('%d/%m/%Y') if o.data_previsao_fim else '',
                f'R$ {o.orcamento:,.2f}' if o.orcamento else '',
                o.status,
                responsavel.nome if responsavel else ''
            ])
        
        # Salvar CSV
        csv_string = csv_content.getvalue()
        output.write(csv_string.encode('utf-8'))
        output.seek(0)
        
        response = make_response(output.read())
        response.headers['Content-Type'] = 'text/csv; charset=utf-8'
        response.headers['Content-Disposition'] = f'attachment; filename=relatorio_sige_{datetime.now().strftime("%Y%m%d_%H%M")}.csv'
        
        return response
    
    elif formato == 'excel':
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = Workbook()
        
        # Aba 1: Funcionários
        ws1 = wb.active
        ws1.title = "Funcionários"
        
        # Cabeçalho
        ws1['A1'] = 'RELATÓRIO DE FUNCIONÁRIOS - SIGE'
        ws1['A1'].font = Font(bold=True, size=14)
        ws1['A2'] = f'Gerado em: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
        
        # Headers
        headers = ['Nome', 'CPF', 'Departamento', 'Função', 'Data Admissão', 'Salário', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws1.cell(row=4, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        
        # Dados
        query = Funcionario.query
        if departamento_id:
            query = query.filter_by(departamento_id=departamento_id)
        
        for row, f in enumerate(query.all(), 5):
            ws1.cell(row=row, column=1, value=f.nome)
            ws1.cell(row=row, column=2, value=f.cpf)
            ws1.cell(row=row, column=3, value=f.departamento_ref.nome if f.departamento_ref else '')
            ws1.cell(row=row, column=4, value=f.funcao_ref.nome if f.funcao_ref else '')
            ws1.cell(row=row, column=5, value=f.data_admissao.strftime('%d/%m/%Y') if f.data_admissao else '')
            ws1.cell(row=row, column=6, value=f.salario if f.salario else 0)
            ws1.cell(row=row, column=7, value='Ativo' if f.ativo else 'Inativo')
        
        # Aba 2: Obras
        ws2 = wb.create_sheet("Obras")
        ws2['A1'] = 'RELATÓRIO DE OBRAS - SIGE'
        ws2['A1'].font = Font(bold=True, size=14)
        ws2['A2'] = f'Gerado em: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
        
        headers = ['Nome', 'Endereço', 'Data Início', 'Previsão Fim', 'Orçamento', 'Status', 'Responsável']
        for col, header in enumerate(headers, 1):
            cell = ws2.cell(row=4, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        
        query = Obra.query
        if obra_id:
            query = query.filter_by(id=obra_id)
        
        for row, o in enumerate(query.all(), 5):
            responsavel = Funcionario.query.get(o.responsavel_id) if o.responsavel_id else None
            ws2.cell(row=row, column=1, value=o.nome)
            ws2.cell(row=row, column=2, value=o.endereco or '')
            ws2.cell(row=row, column=3, value=o.data_inicio.strftime('%d/%m/%Y') if o.data_inicio else '')
            ws2.cell(row=row, column=4, value=o.data_previsao_fim.strftime('%d/%m/%Y') if o.data_previsao_fim else '')
            ws2.cell(row=row, column=5, value=o.orcamento if o.orcamento else 0)
            ws2.cell(row=row, column=6, value=o.status)
            ws2.cell(row=row, column=7, value=responsavel.nome if responsavel else '')
        
        # Aba 3: Ponto (se filtros de data fornecidos)
        if data_inicio and data_fim:
            ws3 = wb.create_sheet("Ponto")
            ws3['A1'] = 'RELATÓRIO DE PONTO - SIGE'
            ws3['A1'].font = Font(bold=True, size=14)
            ws3['A2'] = f'Período: {data_inicio.strftime("%d/%m/%Y")} até {data_fim.strftime("%d/%m/%Y")}'
            
            headers = ['Funcionário', 'Data', 'Entrada', 'Saída Almoço', 'Retorno Almoço', 'Saída', 'Horas Trabalhadas', 'Horas Extras', 'Obra']
            for col, header in enumerate(headers, 1):
                cell = ws3.cell(row=4, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
            
            query = RegistroPonto.query.filter(
                RegistroPonto.data >= data_inicio,
                RegistroPonto.data <= data_fim
            )
            if obra_id:
                query = query.filter_by(obra_id=obra_id)
            
            for row, r in enumerate(query.order_by(RegistroPonto.data.desc()).limit(500).all(), 5):
                ws3.cell(row=row, column=1, value=r.funcionario_ref.nome)
                ws3.cell(row=row, column=2, value=r.data.strftime('%d/%m/%Y'))
                ws3.cell(row=row, column=3, value=r.hora_entrada.strftime('%H:%M') if r.hora_entrada else '')
                ws3.cell(row=row, column=4, value=r.hora_almoco_saida.strftime('%H:%M') if r.hora_almoco_saida else '')
                ws3.cell(row=row, column=5, value=r.hora_almoco_retorno.strftime('%H:%M') if r.hora_almoco_retorno else '')
                ws3.cell(row=row, column=6, value=r.hora_saida.strftime('%H:%M') if r.hora_saida else '')
                ws3.cell(row=row, column=7, value=r.horas_trabalhadas if r.horas_trabalhadas else 0)
                ws3.cell(row=row, column=8, value=r.horas_extras if r.horas_extras else 0)
                ws3.cell(row=row, column=9, value=r.obra_ref.nome if r.obra_ref else '')
        
        # Salvar Excel
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = make_response(output.read())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=relatorio_sige_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
        
        return response
    
    elif formato == 'pdf':
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        
        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=landscape(A4))
        elements = []
        styles = getSampleStyleSheet()
        
        # Título
        title = Paragraph(f'<b>RELATÓRIO SIGE - Sistema Integrado de Gestão Empresarial</b>', styles['Title'])
        elements.append(title)
        elements.append(Spacer(1, 12))
        
        subtitle = Paragraph(f'Gerado em: {datetime.now().strftime("%d/%m/%Y %H:%M")}<br/>Período: {data_inicio.strftime("%d/%m/%Y") if data_inicio else "N/A"} até {data_fim.strftime("%d/%m/%Y") if data_fim else "N/A"}', styles['Normal'])
        elements.append(subtitle)
        elements.append(Spacer(1, 20))
        
        # 1. Tabela de Funcionários
        func_title = Paragraph('<b>FUNCIONÁRIOS</b>', styles['Heading2'])
        elements.append(func_title)
        elements.append(Spacer(1, 12))
        
        query = Funcionario.query
        if departamento_id:
            query = query.filter_by(departamento_id=departamento_id)
        
        func_data = [['Nome', 'CPF', 'Departamento', 'Função', 'Salário', 'Status']]
        for f in query.limit(20).all():  # Limitar para caber na página
            func_data.append([
                f.nome[:20] + '...' if len(f.nome) > 20 else f.nome,
                f.cpf,
                (f.departamento_ref.nome[:15] + '...' if len(f.departamento_ref.nome) > 15 else f.departamento_ref.nome) if f.departamento_ref else '',
                (f.funcao_ref.nome[:15] + '...' if len(f.funcao_ref.nome) > 15 else f.funcao_ref.nome) if f.funcao_ref else '',
                f'R$ {f.salario:,.0f}' if f.salario else '',
                'Ativo' if f.ativo else 'Inativo'
            ])
        
        func_table = Table(func_data)
        func_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(func_table)
        elements.append(Spacer(1, 20))
        
        # 2. Tabela de Obras
        obras_title = Paragraph('<b>OBRAS</b>', styles['Heading2'])
        elements.append(obras_title)
        elements.append(Spacer(1, 12))
        
        query = Obra.query
        if obra_id:
            query = query.filter_by(id=obra_id)
        
        obras_data = [['Nome', 'Data Início', 'Previsão Fim', 'Orçamento', 'Status']]
        for o in query.limit(15).all():
            obras_data.append([
                o.nome[:25] + '...' if len(o.nome) > 25 else o.nome,
                o.data_inicio.strftime('%d/%m/%Y') if o.data_inicio else '',
                o.data_previsao_fim.strftime('%d/%m/%Y') if o.data_previsao_fim else '',
                f'R$ {o.orcamento:,.0f}' if o.orcamento else '',
                o.status
            ])
        
        obras_table = Table(obras_data)
        obras_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(obras_table)
        
        # Build PDF
        doc.build(elements)
        output.seek(0)
        
        response = make_response(output.read())
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename=relatorio_sige_{datetime.now().strftime("%Y%m%d_%H%M")}.pdf'
        
        return response
    
    else:
        return jsonify({'error': 'Formato não suportado'}), 400

@main_bp.route('/alimentacao/restaurantes/<int:restaurante_id>/lancamento', methods=['POST'])
@login_required
def criar_lancamento_restaurante(restaurante_id):
    """Criar lançamento de alimentação para múltiplos funcionários em um restaurante"""
    from datetime import datetime
    
    # Verificar se o restaurante existe e está ativo
    restaurante = Restaurante.query.get_or_404(restaurante_id)
    if not restaurante.ativo:
        flash('Restaurante não está ativo para novos lançamentos.', 'error')
        return redirect(url_for('main.detalhes_restaurante', restaurante_id=restaurante_id))
    
    try:
        # Obter dados do formulário
        data_lancamento = datetime.strptime(request.form.get('data_lancamento'), '%Y-%m-%d').date()
        tipo_refeicao = request.form.get('tipo_refeicao')
        valor_refeicao = float(request.form.get('valor_refeicao'))
        obra_id = request.form.get('obra_lancamento') if request.form.get('obra_lancamento') else None
        observacoes = request.form.get('observacoes_lancamento')
        funcionarios_ids = request.form.getlist('funcionarios[]')
        
        # Validações
        if not funcionarios_ids:
            flash('Selecione pelo menos um funcionário para o lançamento.', 'error')
            return redirect(url_for('main.detalhes_restaurante', restaurante_id=restaurante_id))
        
        if valor_refeicao <= 0:
            flash('O valor da refeição deve ser maior que zero.', 'error')
            return redirect(url_for('main.detalhes_restaurante', restaurante_id=restaurante_id))
        
        # Converter obra_id para inteiro se fornecido
        if obra_id:
            obra_id = int(obra_id)
        
        # Criar registros de alimentação para cada funcionário
        registros_criados = 0
        registros_duplicados = 0
        
        for funcionario_id in funcionarios_ids:
            funcionario_id = int(funcionario_id)
            
            # Verificar se já existe registro para este funcionário, data, tipo e restaurante
            registro_existente = RegistroAlimentacao.query.filter_by(
                funcionario_id=funcionario_id,
                data=data_lancamento,
                tipo=tipo_refeicao,
                restaurante_id=restaurante_id
            ).first()
            
            if registro_existente:
                registros_duplicados += 1
                continue
            
            # Criar novo registro
            novo_registro = RegistroAlimentacao(
                funcionario_id=funcionario_id,
                restaurante_id=restaurante_id,
                obra_id=obra_id,
                data=data_lancamento,
                tipo=tipo_refeicao,
                valor=valor_refeicao,
                observacoes=observacoes
            )
            
            db.session.add(novo_registro)
            registros_criados += 1
        
        # Salvar no banco de dados
        db.session.commit()
        
        # Mensagem de sucesso
        if registros_criados > 0:
            valor_total = registros_criados * valor_refeicao
            msg = f'Lançamento realizado com sucesso! {registros_criados} registro(s) criado(s) no valor total de R$ {valor_total:.2f}.'
            if registros_duplicados > 0:
                msg += f' {registros_duplicados} registro(s) já existiam e foram ignorados.'
            flash(msg, 'success')
        else:
            flash('Nenhum registro foi criado. Todos os funcionários já possuem lançamento para esta data, tipo e restaurante.', 'warning')
        
    except ValueError as e:
        flash('Erro nos dados fornecidos. Verifique os valores e tente novamente.', 'error')
    except Exception as e:
        db.session.rollback()
        flash('Erro interno ao processar o lançamento. Tente novamente.', 'error')
    
    return redirect(url_for('main.detalhes_restaurante', restaurante_id=restaurante_id))

@main_bp.route('/api/dashboard-data')
@login_required
def dashboard_data():
    # Dados para gráficos do dashboard
    funcionarios_dept = db.session.query(
        Departamento.nome,
        func.count(Funcionario.id).label('total')
    ).join(Funcionario).filter(Funcionario.ativo == True).group_by(Departamento.nome).all()
    
    custos_obra = db.session.query(
        Obra.nome,
        func.sum(CustoObra.valor).label('total_custo')
    ).join(CustoObra).group_by(Obra.nome).limit(10).all()
    
    return jsonify({
        'funcionarios_departamento': {
            'labels': [item[0] for item in funcionarios_dept],
            'data': [item[1] for item in funcionarios_dept]
        },
        'custos_obra': {
            'labels': [item[0] for item in custos_obra],
            'data': [float(item[1]) for item in custos_obra]
        }
    })

# Routes para Ocorrências
@main_bp.route('/funcionarios/<int:funcionario_id>/ocorrencia/nova', methods=['POST'])
@login_required
def nova_ocorrencia(funcionario_id):
    from models import Ocorrencia
    
    funcionario = Funcionario.query.get_or_404(funcionario_id)
    
    # Criar nova ocorrência
    ocorrencia = Ocorrencia()
    ocorrencia.funcionario_id = funcionario_id
    ocorrencia.tipo = request.form.get('tipo')
    ocorrencia.data_inicio = datetime.strptime(request.form.get('data_inicio'), '%Y-%m-%d').date()
    if request.form.get('data_fim'):
        ocorrencia.data_fim = datetime.strptime(request.form.get('data_fim'), '%Y-%m-%d').date()
    ocorrencia.descricao = request.form.get('descricao')
    ocorrencia.status = request.form.get('status', 'Pendente')
    
    db.session.add(ocorrencia)
    db.session.commit()
    
    flash('Ocorrência registrada com sucesso!', 'success')
    return redirect(url_for('main.funcionario_perfil', id=funcionario_id))

# Horários de Trabalho
@main_bp.route('/horarios')
@login_required
def horarios():
    horarios = HorarioTrabalho.query.all()
    return render_template('horarios.html', horarios=horarios)

@main_bp.route('/horarios/novo', methods=['GET', 'POST'])
@login_required
def novo_horario():
    form = HorarioTrabalhoForm()
    
    if form.validate_on_submit():
        horario = HorarioTrabalho(
            nome=form.nome.data,
            entrada=form.entrada.data,
            saida_almoco=form.saida_almoco.data,
            retorno_almoco=form.retorno_almoco.data,
            saida=form.saida.data,
            dias_semana=form.dias_semana.data
        )
        db.session.add(horario)
        db.session.commit()
        flash('Horário de trabalho cadastrado com sucesso!', 'success')
        return redirect(url_for('main.horarios'))
    
    return render_template('horarios.html', form=form, horarios=HorarioTrabalho.query.all())

@main_bp.route('/horarios/editar/<int:id>', methods=['GET', 'POST'])
@login_required
def editar_horario(id):
    horario = HorarioTrabalho.query.get_or_404(id)
    form = HorarioTrabalhoForm(obj=horario)
    
    if form.validate_on_submit():
        horario.nome = form.nome.data
        horario.entrada = form.entrada.data
        horario.saida_almoco = form.saida_almoco.data
        horario.retorno_almoco = form.retorno_almoco.data
        horario.saida = form.saida.data
        horario.dias_semana = form.dias_semana.data
        
        db.session.commit()
        flash('Horário de trabalho atualizado com sucesso!', 'success')
        return redirect(url_for('main.horarios'))
    
    return render_template('horarios.html', form=form, horarios=HorarioTrabalho.query.all(), edit_id=id)

@main_bp.route('/horarios/excluir/<int:id>', methods=['POST'])
@login_required
def excluir_horario(id):
    horario = HorarioTrabalho.query.get_or_404(id)
    
    # Verificar se há funcionários usando este horário
    funcionarios_usando = Funcionario.query.filter_by(horario_trabalho_id=id).count()
    if funcionarios_usando > 0:
        flash(f'Não é possível excluir este horário. {funcionarios_usando} funcionário(s) estão usando este horário.', 'error')
        return redirect(url_for('main.horarios'))
    
    db.session.delete(horario)
    db.session.commit()
    flash('Horário de trabalho excluído com sucesso!', 'success')
    return redirect(url_for('main.horarios'))


# ============ ROTAS RDO ============

@main_bp.route('/rdo')
@login_required
def lista_rdos():
    """Lista todos os RDOs com filtros"""
    from forms import RDOFiltroForm
    from models import RDO, Obra
    
    form = RDOFiltroForm()
    
    # Populando choices das obras
    form.obra_id.choices = [(0, 'Todas as Obras')] + [(obra.id, obra.nome) for obra in Obra.query.all()]
    
    # Query base
    query = RDO.query.join(Obra).join(Funcionario, RDO.criado_por_id == Funcionario.id)
    
    # Aplicar filtros se fornecidos
    if request.method == 'POST' and form.validate():
        if form.data_inicio.data:
            query = query.filter(RDO.data_relatorio >= form.data_inicio.data)
        if form.data_fim.data:
            query = query.filter(RDO.data_relatorio <= form.data_fim.data)
        if form.obra_id.data and form.obra_id.data != 0:
            query = query.filter(RDO.obra_id == form.obra_id.data)
        if form.status.data:
            query = query.filter(RDO.status == form.status.data)
    
    # Ordenar por data mais recente
    rdos = query.order_by(RDO.data_relatorio.desc()).all()
    
    return render_template('rdo/lista_rdos.html', rdos=rdos, form=form)


@main_bp.route('/rdo/novo')
@main_bp.route('/rdo/novo/<int:obra_id>')
@login_required
def novo_rdo(obra_id=None):
    """Criar novo RDO"""
    from forms import RDOForm
    from models import Obra, Funcionario
    import json
    
    form = RDOForm()
    
    # Populando choices das obras
    form.obra_id.choices = [(obra.id, obra.nome) for obra in Obra.query.all()]
    
    # Se obra_id foi fornecida, pré-selecionar
    if obra_id:
        form.obra_id.data = obra_id
    
    # Preparar dados para JavaScript
    funcionarios = [
        {'id': f.id, 'nome': f.nome, 'funcao': f.funcao_ref.nome if f.funcao_ref else 'Sem função'}
        for f in Funcionario.query.filter_by(ativo=True).all()
    ]
    
    obras = [
        {'id': o.id, 'nome': o.nome}
        for o in Obra.query.all()
    ]
    
    funcionarios_json = json.dumps(funcionarios)
    obras_json = json.dumps(obras)
    
    return render_template('rdo/formulario_rdo.html', 
                         form=form, 
                         modo='criar',
                         funcionarios_json=funcionarios_json,
                         obras_json=obras_json)


@main_bp.route('/rdo/criar', methods=['POST'])
@login_required
def criar_rdo():
    """Processar criação de RDO"""
    from forms import RDOForm
    from models import RDO, Obra
    import uuid
    from datetime import datetime
    
    form = RDOForm()
    form.obra_id.choices = [(obra.id, obra.nome) for obra in Obra.query.all()]
    
    if form.validate_on_submit():
        # Gerar número único do RDO
        numero_rdo = f"RDO-{datetime.now().year}-{str(uuid.uuid4())[:8].upper()}"
        
        rdo = RDO(
            numero_rdo=numero_rdo,
            data_relatorio=form.data_relatorio.data,
            obra_id=form.obra_id.data,
            criado_por_id=current_user.id,  # Assumindo que current_user é um funcionário
            tempo_manha=form.tempo_manha.data,
            tempo_tarde=form.tempo_tarde.data,
            tempo_noite=form.tempo_noite.data,
            observacoes_meteorologicas=form.observacoes_meteorologicas.data,
            comentario_geral=form.comentario_geral.data,
            status=form.status.data
        )
        
        db.session.add(rdo)
        db.session.commit()
        
        flash('RDO criado com sucesso!', 'success')
        return redirect(url_for('main.visualizar_rdo', id=rdo.id))
    
    return render_template('rdo/formulario_rdo.html', form=form, modo='criar')


@main_bp.route('/rdo/<int:id>')
@login_required
def visualizar_rdo(id):
    """Visualizar detalhes de um RDO"""
    from models import RDO
    
    rdo = RDO.query.get_or_404(id)
    return render_template('rdo/visualizar_rdo.html', rdo=rdo)


@main_bp.route('/rdo/<int:id>/editar')
@login_required
def editar_rdo(id):
    """Editar RDO existente"""
    from forms import RDOForm
    from models import RDO, Obra
    
    rdo = RDO.query.get_or_404(id)
    form = RDOForm(obj=rdo)
    
    # Populando choices das obras
    form.obra_id.choices = [(obra.id, obra.nome) for obra in Obra.query.all()]
    
    return render_template('rdo/formulario_rdo.html', form=form, rdo=rdo, modo='editar')


@main_bp.route('/rdo/<int:id>/atualizar', methods=['POST'])
@login_required
def atualizar_rdo(id):
    """Processar atualização de RDO"""
    from forms import RDOForm
    from models import RDO, Obra
    
    rdo = RDO.query.get_or_404(id)
    form = RDOForm()
    form.obra_id.choices = [(obra.id, obra.nome) for obra in Obra.query.all()]
    
    if form.validate_on_submit():
        rdo.data_relatorio = form.data_relatorio.data
        rdo.obra_id = form.obra_id.data
        rdo.tempo_manha = form.tempo_manha.data
        rdo.tempo_tarde = form.tempo_tarde.data
        rdo.tempo_noite = form.tempo_noite.data
        rdo.observacoes_meteorologicas = form.observacoes_meteorologicas.data
        rdo.comentario_geral = form.comentario_geral.data
        rdo.status = form.status.data
        
        db.session.commit()
        
        flash('RDO atualizado com sucesso!', 'success')
        return redirect(url_for('main.visualizar_rdo', id=rdo.id))
    
    return render_template('rdo/formulario_rdo.html', form=form, rdo=rdo, modo='editar')


@main_bp.route('/rdo/<int:id>/excluir', methods=['POST'])
@login_required
def excluir_rdo(id):
    """Excluir RDO"""
    from models import RDO
    
    rdo = RDO.query.get_or_404(id)
    
    db.session.delete(rdo)
    db.session.commit()
    
    flash('RDO excluído com sucesso!', 'success')
    return redirect(url_for('main.lista_rdos'))


# Rotas para seções dinâmicas do RDO
@main_bp.route('/rdo/<int:rdo_id>/mao-obra')
@login_required
def gerenciar_mao_obra(rdo_id):
    """Gerenciar mão de obra do RDO"""
    from models import RDO, RDOMaoObra, Funcionario
    
    rdo = RDO.query.get_or_404(rdo_id)
    funcionarios = Funcionario.query.filter_by(ativo=True).all()
    
    return render_template('rdo/secoes/mao_obra.html', 
                         rdo=rdo, 
                         funcionarios=funcionarios)


@main_bp.route('/obras/<int:id>/rdos')
@login_required
def rdos_por_obra(id):
    """Lista RDOs de uma obra específica"""
    from models import Obra, RDO
    
    obra = Obra.query.get_or_404(id)
    rdos = RDO.query.filter_by(obra_id=id).order_by(RDO.data_relatorio.desc()).all()
    
    return render_template('rdo/rdos_obra.html', obra=obra, rdos=rdos)


# ========== NOVAS FUNCIONALIDADES: ALIMENTAÇÃO APRIMORADA ==========

@main_bp.route('/alimentacao/restaurantes')
@main_bp.route('/alimentacao/restaurantes', methods=['POST'])
@login_required
def alimentacao_restaurantes():
    """Página principal de alimentação com cards de restaurantes e KPIs"""
    from forms import FiltroDataForm, AlimentacaoMultiplaForm, RestauranteForm
    from models import Restaurante, RegistroAlimentacao, Funcionario, Obra
    from datetime import date, datetime, timedelta
    from sqlalchemy import func
    import json
    
    # Formulários
    filtro_form = FiltroDataForm()
    alimentacao_form = AlimentacaoMultiplaForm()
    restaurante_form = RestauranteForm()
    
    # Definir período padrão (mês atual)
    data_inicio = date.today().replace(day=1)
    data_fim = date.today()
    
    # Aplicar filtros se fornecidos
    if request.method == 'POST' and filtro_form.validate_on_submit():
        if filtro_form.data_inicio.data:
            data_inicio = filtro_form.data_inicio.data
        if filtro_form.data_fim.data:
            data_fim = filtro_form.data_fim.data
    
    # Popular formulários
    alimentacao_form.obra_id.choices = [(0, 'Selecione...')] + [(obra.id, obra.nome) for obra in Obra.query.all()]
    alimentacao_form.restaurante_id.choices = [(0, 'Selecione...')] + [(r.id, r.nome) for r in Restaurante.query.filter_by(ativo=True).all()]
    
    # Calcular KPIs gerais
    query_kpis = RegistroAlimentacao.query.filter(
        RegistroAlimentacao.data >= data_inicio,
        RegistroAlimentacao.data <= data_fim
    )
    
    custo_total = db.session.query(func.sum(RegistroAlimentacao.valor)).filter(
        RegistroAlimentacao.data >= data_inicio,
        RegistroAlimentacao.data <= data_fim
    ).scalar() or 0
    
    registros_hoje = RegistroAlimentacao.query.filter_by(data=date.today()).count()
    
    funcionarios_alimentados = db.session.query(func.count(func.distinct(RegistroAlimentacao.funcionario_id))).filter(
        RegistroAlimentacao.data >= data_inicio,
        RegistroAlimentacao.data <= data_fim
    ).scalar() or 0
    
    dias_periodo = (data_fim - data_inicio).days + 1
    media_diaria = custo_total / dias_periodo if dias_periodo > 0 else 0
    
    kpis = {
        'custo_total': custo_total,
        'media_diaria': media_diaria,
        'registros_hoje': registros_hoje,
        'funcionarios_alimentados': funcionarios_alimentados
    }
    
    # Buscar restaurantes com KPIs
    restaurantes = []
    for restaurante in Restaurante.query.all():
        custo_restaurante = db.session.query(func.sum(RegistroAlimentacao.valor)).filter(
            RegistroAlimentacao.restaurante_id == restaurante.id,
            RegistroAlimentacao.data >= data_inicio,
            RegistroAlimentacao.data <= data_fim
        ).scalar() or 0
        
        total_registros = RegistroAlimentacao.query.filter(
            RegistroAlimentacao.restaurante_id == restaurante.id,
            RegistroAlimentacao.data >= data_inicio,
            RegistroAlimentacao.data <= data_fim
        ).count()
        
        restaurante.kpis = {
            'custo_total': custo_restaurante,
            'total_registros': total_registros
        }
        restaurantes.append(restaurante)
    
    # Dados para JavaScript
    funcionarios = [
        {'id': f.id, 'nome': f.nome, 'funcao': f.funcao_ref.nome if f.funcao_ref else 'Sem função'}
        for f in Funcionario.query.filter_by(ativo=True).all()
    ]
    funcionarios_json = json.dumps(funcionarios)
    
    return render_template('alimentacao/restaurantes.html',
                         restaurantes=restaurantes,
                         kpis=kpis,
                         filtro_form=filtro_form,
                         alimentacao_form=alimentacao_form,
                         restaurante_form=restaurante_form,
                         funcionarios_json=funcionarios_json)


@main_bp.route('/alimentacao/restaurante/criar', methods=['POST'])
@login_required
def criar_restaurante():
    """Criar novo restaurante"""
    from forms import RestauranteForm
    from models import Restaurante
    
    form = RestauranteForm()
    
    if form.validate_on_submit():
        restaurante = Restaurante(
            nome=form.nome.data,
            endereco=form.endereco.data,
            telefone=form.telefone.data,
            email=form.email.data,
            contato_responsavel=form.contato_responsavel.data,
            ativo=form.ativo.data
        )
        
        db.session.add(restaurante)
        db.session.commit()
        
        flash('Restaurante cadastrado com sucesso!', 'success')
    else:
        flash('Erro ao cadastrar restaurante. Verifique os dados.', 'error')
    
    return redirect(url_for('main.alimentacao_restaurantes'))


@main_bp.route('/alimentacao/multipla/criar', methods=['POST'])
@login_required
def criar_alimentacao_multipla():
    """Criar lançamento de alimentação para múltiplos funcionários"""
    from forms import AlimentacaoMultiplaForm
    from models import RegistroAlimentacao
    import json
    
    form = AlimentacaoMultiplaForm()
    
    # Popular choices para validação
    from models import Obra, Restaurante
    form.obra_id.choices = [(0, 'Selecione...')] + [(obra.id, obra.nome) for obra in Obra.query.all()]
    form.restaurante_id.choices = [(0, 'Selecione...')] + [(r.id, r.nome) for r in Restaurante.query.filter_by(ativo=True).all()]
    
    if form.validate_on_submit():
        try:
            funcionarios_ids = json.loads(form.funcionarios_selecionados.data)
            
            if not funcionarios_ids:
                flash('Selecione pelo menos um funcionário.', 'error')
                return redirect(url_for('main.alimentacao_restaurantes'))
            
            # Criar registro para cada funcionário selecionado
            registros_criados = 0
            for funcionario_id in funcionarios_ids:
                registro = RegistroAlimentacao(
                    funcionario_id=funcionario_id,
                    obra_id=form.obra_id.data if form.obra_id.data != 0 else None,
                    restaurante_id=form.restaurante_id.data if form.restaurante_id.data != 0 else None,
                    data=form.data.data,
                    tipo=form.tipo.data,
                    valor=form.valor.data,
                    observacoes=form.observacoes.data
                )
                db.session.add(registro)
                registros_criados += 1
            
            db.session.commit()
            
            flash(f'Lançamento criado com sucesso para {registros_criados} funcionário(s)!', 'success')
            
        except Exception as e:
            db.session.rollback()
            flash('Erro ao criar lançamento. Tente novamente.', 'error')
    else:
        flash('Erro na validação dos dados. Verifique as informações.', 'error')
    
    return redirect(url_for('main.alimentacao_restaurantes'))


@main_bp.route('/alimentacao/restaurante/<int:restaurante_id>')
@login_required
def detalhes_restaurante(restaurante_id):
    """Página de detalhes de um restaurante específico"""
    from models import Restaurante, RegistroAlimentacao, Funcionario
    from datetime import date, timedelta
    from sqlalchemy import func
    
    restaurante = Restaurante.query.get_or_404(restaurante_id)
    
    # Período padrão (mês atual)
    data_inicio = date.today().replace(day=1)
    data_fim = date.today()
    
    # KPIs do restaurante
    custo_total = db.session.query(func.sum(RegistroAlimentacao.valor)).filter(
        RegistroAlimentacao.restaurante_id == restaurante.id,
        RegistroAlimentacao.data >= data_inicio,
        RegistroAlimentacao.data <= data_fim
    ).scalar() or 0
    
    total_registros = RegistroAlimentacao.query.filter(
        RegistroAlimentacao.restaurante_id == restaurante.id,
        RegistroAlimentacao.data >= data_inicio,
        RegistroAlimentacao.data <= data_fim
    ).count()
    
    funcionarios_unicos = db.session.query(func.count(func.distinct(RegistroAlimentacao.funcionario_id))).filter(
        RegistroAlimentacao.restaurante_id == restaurante.id,
        RegistroAlimentacao.data >= data_inicio,
        RegistroAlimentacao.data <= data_fim
    ).scalar() or 0
    
    media_por_funcionario = custo_total / funcionarios_unicos if funcionarios_unicos > 0 else 0
    
    kpis = {
        'custo_total': custo_total,
        'total_registros': total_registros,
        'funcionarios_unicos': funcionarios_unicos,
        'media_por_funcionario': media_por_funcionario
    }
    
    # Histórico de lançamentos
    registros = RegistroAlimentacao.query.filter(
        RegistroAlimentacao.restaurante_id == restaurante.id
    ).join(Funcionario).order_by(RegistroAlimentacao.data.desc()).limit(50).all()
    
    # Dados para o formulário de lançamento
    funcionarios = Funcionario.query.filter_by(ativo=True).order_by(Funcionario.nome).all()
    from models import Obra
    obras = Obra.query.filter_by(status='Em andamento').order_by(Obra.nome).all()
    data_hoje = date.today().strftime('%Y-%m-%d')
    
    return render_template('alimentacao/detalhes_restaurante.html',
                         restaurante=restaurante,
                         kpis=kpis,
                         registros=registros,
                         funcionarios=funcionarios,
                         obras=obras,
                         data_hoje=data_hoje)

# ========== NOVAS ROTAS DASHBOARD v3.0 ==========

@main_bp.route('/funcionario/<int:id>/dashboard-v3')
@login_required
def dashboard_funcionario_v3(id):
    """
    Dashboard v3.0 com layout 4-4-2 e KPIs corretos
    Implementa especificação técnica completa
    """
    funcionario = Funcionario.query.get_or_404(id)
    
    # Obter parâmetros de filtro
    data_inicio_str = request.args.get('data_inicio')
    data_fim_str = request.args.get('data_fim')
    
    # Definir período padrão (mês atual)
    if data_inicio_str and data_fim_str:
        try:
            data_inicio = datetime.strptime(data_inicio_str, '%Y-%m-%d').date()
            data_fim = datetime.strptime(data_fim_str, '%Y-%m-%d').date()
        except ValueError:
            data_inicio = date.today().replace(day=1)
            data_fim = date.today()
    else:
        data_inicio = date.today().replace(day=1)
        data_fim = date.today()
    
    # Calcular KPIs usando o engine v3.0
    try:
        kpis = kpis_engine.calcular_kpis_funcionario(id, data_inicio, data_fim)
        if not kpis:
            flash('Erro ao calcular KPIs do funcionário', 'error')
            return redirect(url_for('main.funcionario_perfil', id=id))
    except Exception as e:
        flash(f'Erro ao processar dados: {str(e)}', 'error')
        # Criar KPIs vazios para evitar erro na template
        kpis = {
            'horas_trabalhadas': 0, 'horas_extras': 0, 'faltas': 0, 'atrasos_horas': 0,
            'custo_mensal': 0, 'absenteismo': 0, 'media_diaria': 0, 'horas_perdidas': 0,
            'produtividade': 0, 'custo_alimentacao': 0,
            'periodo': {'inicio': data_inicio, 'fim': data_fim, 'dias_uteis': 0}
        }
    
    return render_template('dashboard_funcionario_v3.html',
                         funcionario=funcionario,
                         kpis=kpis,
                         periodo=kpis.get('periodo', {'inicio': data_inicio, 'fim': data_fim, 'dias_uteis': 0}))

@main_bp.route('/api/funcionario/<int:id>/kpis')
@login_required
def api_kpis_funcionario(id):
    """
    API endpoint para atualização automática dos KPIs
    Retorna dados em JSON para atualização via JavaScript
    """
    funcionario = Funcionario.query.get_or_404(id)
    
    # Obter parâmetros de filtro
    data_inicio_str = request.args.get('data_inicio')
    data_fim_str = request.args.get('data_fim')
    
    if data_inicio_str and data_fim_str:
        try:
            data_inicio = datetime.strptime(data_inicio_str, '%Y-%m-%d').date()
            data_fim = datetime.strptime(data_fim_str, '%Y-%m-%d').date()
        except ValueError:
            return jsonify({'error': 'Formato de data inválido'}), 400
    else:
        data_inicio = date.today().replace(day=1)
        data_fim = date.today()
    
    # Calcular KPIs
    try:
        kpis = kpis_engine.calcular_kpis_funcionario(id, data_inicio, data_fim)
        if not kpis:
            return jsonify({'error': 'Funcionário não encontrado'}), 404
        
        return jsonify(kpis)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@main_bp.route('/funcionario/<int:id>/relatorio-export')
@login_required
def exportar_relatorio_funcionario(id):
    """
    Exporta relatório completo do funcionário em formato CSV
    """
    funcionario = Funcionario.query.get_or_404(id)
    
    # Obter parâmetros de filtro
    data_inicio_str = request.args.get('data_inicio')
    data_fim_str = request.args.get('data_fim')
    
    if data_inicio_str and data_fim_str:
        try:
            data_inicio = datetime.strptime(data_inicio_str, '%Y-%m-%d').date()
            data_fim = datetime.strptime(data_fim_str, '%Y-%m-%d').date()
        except ValueError:
            flash('Formato de data inválido', 'error')
            return redirect(url_for('main.dashboard_funcionario_v3', id=id))
    else:
        data_inicio = date.today().replace(day=1)
        data_fim = date.today()
    
    # Calcular KPIs
    try:
        kpis = kpis_engine.calcular_kpis_funcionario(id, data_inicio, data_fim)
        if not kpis:
            flash('Erro ao gerar relatório', 'error')
            return redirect(url_for('main.dashboard_funcionario_v3', id=id))
        
        # Gerar conteúdo CSV
        import csv
        from io import StringIO
        from flask import Response
        
        output = StringIO()
        writer = csv.writer(output)
        
        # Cabeçalho
        writer.writerow(['RELATÓRIO DE KPIs - ' + funcionario.nome])
        writer.writerow(['Código:', funcionario.codigo])
        writer.writerow(['Período:', f"{data_inicio.strftime('%d/%m/%Y')} a {data_fim.strftime('%d/%m/%Y')}"])
        writer.writerow([])
        
        # Dados dos KPIs
        writer.writerow(['INDICADOR', 'VALOR', 'UNIDADE'])
        writer.writerow(['Horas Trabalhadas', kpis['horas_trabalhadas'], 'horas'])
        writer.writerow(['Horas Extras', kpis['horas_extras'], 'horas'])
        writer.writerow(['Faltas', kpis['faltas'], 'dias'])
        writer.writerow(['Atrasos', kpis['atrasos_horas'], 'horas'])
        writer.writerow(['Custo Mensal', f"R$ {kpis['custo_mensal']:.2f}", 'reais'])
        writer.writerow(['Absenteísmo', f"{kpis['absenteismo']:.1f}%", 'percentual'])
        writer.writerow(['Média Diária', kpis['media_diaria'], 'horas/dia'])
        writer.writerow(['Horas Perdidas', kpis['horas_perdidas'], 'horas'])
        writer.writerow(['Produtividade', f"{kpis['produtividade']:.1f}%", 'percentual'])
        writer.writerow(['Custo Alimentação', f"R$ {kpis['custo_alimentacao']:.2f}", 'reais'])
        
        output.seek(0)
        
        # Retornar como download
        response = Response(
            output.getvalue(),
            mimetype='text/csv',
            headers={
                'Content-Disposition': f'attachment; filename=relatorio_kpis_{funcionario.codigo}_{data_inicio}_{data_fim}.csv'
            }
        )
        
        return response
        
    except Exception as e:
        flash(f'Erro ao gerar relatório: {str(e)}', 'error')
        return redirect(url_for('main.dashboard_funcionario_v3', id=id))
