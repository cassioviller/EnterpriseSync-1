from models import db
"""
Views para o Módulo 6 - Sistema de Folha de Pagamento Automática
Versão limpa e funcional
"""

from flask import Blueprint, render_template, request, jsonify, redirect, url_for, flash, make_response, send_file
from models import (FolhaPagamento, ParametrosLegais, Funcionario, BeneficioFuncionario, 
                    Adiantamento, CalculoHorasMensal)
from flask_login import login_required, current_user
from datetime import datetime, date, timedelta
from io import BytesIO
import calendar

folha_bp = Blueprint('folha', __name__)

def admin_required(f):
    """Decorator simples para admin - implementação temporária"""
    from functools import wraps
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated:
            return redirect(url_for('main.login'))
        # Verificação simplificada para Admin
        if hasattr(current_user, 'tipo_usuario') and current_user.tipo_usuario and current_user.tipo_usuario.name in ['ADMIN', 'SUPER_ADMIN']:
            return f(*args, **kwargs)
        flash('Acesso negado. Apenas administradores podem acessar esta página.', 'error')
        return redirect(url_for('main.dashboard'))
    return decorated_function

# ================================
# DASHBOARD PRINCIPAL
# ================================

@folha_bp.route('/dashboard')
@admin_required
def dashboard():
    """Dashboard principal da folha de pagamento"""
    
    # Importar modelos e db dentro da função para evitar importação circular
    
    # Mês atual
    hoje = date.today()
    mes_atual = hoje.replace(day=1)
    
    # Obter folhas do mês atual
    folhas_mes = FolhaPagamento.query.filter_by(
        admin_id=current_user.id,
        mes_referencia=mes_atual
    ).all()
    
    # Calcular métricas
    total_folha = sum(f.total_proventos or 0 for f in folhas_mes)
    total_liquido = sum(f.salario_liquido or 0 for f in folhas_mes)
    total_encargos = sum((f.inss or 0) + (f.fgts or 0) for f in folhas_mes)
    
    # Mês anterior para comparação
    mes_anterior = (mes_atual - timedelta(days=1)).replace(day=1)
    folhas_anterior = FolhaPagamento.query.filter_by(
        admin_id=current_user.id,
        mes_referencia=mes_anterior
    ).all()
    
    total_anterior = sum(f.total_proventos or 0 for f in folhas_anterior)
    variacao = ((total_folha - total_anterior) / total_anterior * 100) if total_anterior > 0 else 0
    
    # Status do processamento
    from models import Funcionario
    funcionarios_ativos = Funcionario.query.filter_by(
        admin_id=current_user.id,
        ativo=True
    ).count()
    
    folhas_processadas = len(folhas_mes)
    percentual_processado = (folhas_processadas / funcionarios_ativos * 100) if funcionarios_ativos > 0 else 0
    
    # Top 5 maiores salários
    top_salarios = sorted(folhas_mes, key=lambda x: x.total_proventos or 0, reverse=True)[:5]
    
    # Funcionários com mais horas extras
    top_extras = sorted(folhas_mes, key=lambda x: x.horas_extras or 0, reverse=True)[:5]
    
    # Verificar se parâmetros legais estão configurados
    parametros = ParametrosLegais.query.filter_by(
        admin_id=current_user.id,
        ano_vigencia=hoje.year,
        ativo=True
    ).first()
    
    parametros_configurados = parametros is not None
    if not parametros_configurados:
        flash('Parâmetros legais não configurados para este ano. Configure antes de processar a folha.', 'warning')
    
    return render_template('folha_pagamento/dashboard.html',
                         mes_referencia=mes_atual,
                         total_folha=total_folha,
                         total_liquido=total_liquido,
                         total_encargos=total_encargos,
                         variacao=variacao,
                         funcionarios_ativos=funcionarios_ativos,
                         folhas_processadas=folhas_processadas,
                         percentual_processado=percentual_processado,
                         top_salarios=top_salarios,
                         top_extras=top_extras,
                         parametros_configurados=parametros_configurados)

# ================================
# PROCESSAMENTO DA FOLHA
# ================================

@folha_bp.route('/processar/<int:ano>/<int:mes>', methods=['POST'])
@admin_required
def processar_folha_mes(ano, mes):
    """Processar folha de pagamento do mês - VERSÃO COMPLETA"""
    
    try:
        from services.folha_service import processar_folha_funcionario
        from event_manager import EventManager
        
        mes_referencia = date(ano, mes, 1)
        
        # Verificar se já foi processada
        folhas_existentes = FolhaPagamento.query.filter_by(
            admin_id=current_user.id,
            mes_referencia=mes_referencia
        ).count()
        
        reprocessar = request.form.get('reprocessar') == 'true'
        
        if folhas_existentes > 0 and not reprocessar:
            flash('Folha já processada para este mês. Use a opção "Reprocessar" se necessário.', 'warning')
            return redirect(url_for('folha.dashboard'))
        
        # Se reprocessar, deletar folhas existentes
        if reprocessar:
            FolhaPagamento.query.filter_by(
                admin_id=current_user.id,
                mes_referencia=mes_referencia
            ).delete()
            db.session.commit()
        
        # Buscar funcionários ativos
        funcionarios = Funcionario.query.filter_by(
            admin_id=current_user.id,
            ativo=True
        ).all()
        
        if not funcionarios:
            flash('Nenhum funcionário ativo encontrado para processar.', 'warning')
            return redirect(url_for('folha.dashboard'))
        
        # Processar cada funcionário
        folhas_criadas = 0
        erros = 0
        
        for funcionario in funcionarios:
            # Calcular folha do funcionário
            dados_folha = processar_folha_funcionario(funcionario, ano, mes)
            
            if dados_folha:
                # Criar registro de folha de pagamento
                folha = FolhaPagamento(
                    funcionario_id=funcionario.id,
                    mes_referencia=mes_referencia,
                    salario_base=dados_folha['salario_base'],
                    horas_extras=dados_folha['horas_extras'],
                    total_proventos=dados_folha['total_proventos'],
                    inss=dados_folha['inss'],
                    irrf=dados_folha['irrf'],
                    outros_descontos=dados_folha['outros_descontos'],
                    total_descontos=dados_folha['total_descontos'],
                    salario_liquido=dados_folha['salario_liquido'],
                    fgts=dados_folha['fgts'],
                    admin_id=current_user.id
                )
                
                db.session.add(folha)
                db.session.flush()  # CRÍTICO: Flush para gerar o ID antes de emitir evento
                folhas_criadas += 1
                
                # Emitir evento para contabilidade (agora com folha.id válido)
                try:
                    EventManager.emit('folha_processada', {
                        'folha_id': folha.id,
                        'funcionario_id': funcionario.id,
                        'mes_referencia': mes_referencia.isoformat(),
                        'valor_total': dados_folha['total_proventos'],
                        'encargos': dados_folha['encargos_patronais'],
                        'admin_id': current_user.id
                    })
                except Exception as e:
                    print(f"Erro ao emitir evento folha_processada: {e}")
            else:
                erros += 1
        
        # Commit final
        db.session.commit()
        
        # Mensagens de resultado
        if folhas_criadas > 0:
            flash(f'Folha processada com sucesso! {folhas_criadas} funcionários processados.', 'success')
        if erros > 0:
            flash(f'{erros} funcionários com erro no processamento.', 'warning')
        
        return redirect(url_for('folha.dashboard'))
        
    except Exception as e:
        db.session.rollback()
        print(f"ERRO AO PROCESSAR FOLHA: {e}")
        flash(f'Erro ao processar folha: {str(e)}', 'danger')
        return redirect(url_for('folha.dashboard'))

# ================================
# ROTAS SIMPLIFICADAS
# ================================

# ================================
# PARÂMETROS LEGAIS
# ================================

@folha_bp.route('/parametros-legais')
@admin_required
def parametros_legais():
    """Listar parâmetros legais"""
    parametros = ParametrosLegais.query.filter_by(
        admin_id=current_user.id
    ).order_by(ParametrosLegais.ano_vigencia.desc()).all()
    
    return render_template('folha_pagamento/parametros_legais.html', 
                         parametros=parametros)

@folha_bp.route('/parametros-legais/criar', methods=['GET', 'POST'])
@admin_required
def criar_parametros():
    """Criar novos parâmetros legais"""
    if request.method == 'POST':
        try:
            ano_vigencia = int(request.form.get('ano_vigencia'))
            
            # Verificar se já existe parâmetro para este ano
            existe = ParametrosLegais.query.filter_by(
                admin_id=current_user.id,
                ano_vigencia=ano_vigencia
            ).first()
            
            if existe:
                flash(f'Já existem parâmetros cadastrados para o ano {ano_vigencia}.', 'warning')
                return redirect(url_for('folha.parametros_legais'))
            
            # Criar novo parâmetro
            parametro = ParametrosLegais(
                ano_vigencia=ano_vigencia,
                admin_id=current_user.id,
                
                # INSS
                inss_faixa1_limite=float(request.form.get('inss_faixa1_limite', 1412.00)),
                inss_faixa1_percentual=float(request.form.get('inss_faixa1_percentual', 7.5)),
                inss_faixa2_limite=float(request.form.get('inss_faixa2_limite', 2666.68)),
                inss_faixa2_percentual=float(request.form.get('inss_faixa2_percentual', 9.0)),
                inss_faixa3_limite=float(request.form.get('inss_faixa3_limite', 4000.03)),
                inss_faixa3_percentual=float(request.form.get('inss_faixa3_percentual', 12.0)),
                inss_faixa4_limite=float(request.form.get('inss_faixa4_limite', 7786.02)),
                inss_faixa4_percentual=float(request.form.get('inss_faixa4_percentual', 14.0)),
                inss_teto=float(request.form.get('inss_teto', 908.85)),
                
                # IRRF
                irrf_isencao=float(request.form.get('irrf_isencao', 2259.20)),
                irrf_faixa1_limite=float(request.form.get('irrf_faixa1_limite', 2826.65)),
                irrf_faixa1_percentual=float(request.form.get('irrf_faixa1_percentual', 7.5)),
                irrf_faixa1_deducao=float(request.form.get('irrf_faixa1_deducao', 169.44)),
                irrf_faixa2_limite=float(request.form.get('irrf_faixa2_limite', 3751.05)),
                irrf_faixa2_percentual=float(request.form.get('irrf_faixa2_percentual', 15.0)),
                irrf_faixa2_deducao=float(request.form.get('irrf_faixa2_deducao', 381.44)),
                irrf_faixa3_limite=float(request.form.get('irrf_faixa3_limite', 4664.68)),
                irrf_faixa3_percentual=float(request.form.get('irrf_faixa3_percentual', 22.5)),
                irrf_faixa3_deducao=float(request.form.get('irrf_faixa3_deducao', 662.77)),
                irrf_faixa4_percentual=float(request.form.get('irrf_faixa4_percentual', 27.5)),
                irrf_faixa4_deducao=float(request.form.get('irrf_faixa4_deducao', 896.00)),
                irrf_dependente_valor=float(request.form.get('irrf_dependente_valor', 189.59)),
                
                # Outros
                fgts_percentual=float(request.form.get('fgts_percentual', 8.0)),
                salario_minimo=float(request.form.get('salario_minimo', 1412.00)),
                vale_transporte_percentual=float(request.form.get('vale_transporte_percentual', 6.0)),
                adicional_noturno_percentual=float(request.form.get('adicional_noturno_percentual', 20.0)),
                hora_extra_50_percentual=float(request.form.get('hora_extra_50_percentual', 50.0)),
                hora_extra_100_percentual=float(request.form.get('hora_extra_100_percentual', 100.0)),
                
                ativo=True
            )
            
            db.session.add(parametro)
            db.session.commit()
            
            flash(f'Parâmetros legais para {ano_vigencia} criados com sucesso!', 'success')
            return redirect(url_for('folha.parametros_legais'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao criar parâmetros: {str(e)}', 'danger')
            return redirect(url_for('folha.criar_parametros'))
    
    # GET - valores padrão 2025
    hoje = date.today()
    return render_template('folha_pagamento/parametros_form.html',
                         parametro=None,
                         ano_padrao=hoje.year)

@folha_bp.route('/parametros-legais/editar/<int:id>', methods=['GET', 'POST'])
@admin_required
def editar_parametros(id):
    """Editar parâmetros legais existentes"""
    parametro = ParametrosLegais.query.filter_by(
        id=id,
        admin_id=current_user.id
    ).first_or_404()
    
    if request.method == 'POST':
        try:
            # Atualizar INSS
            parametro.inss_faixa1_limite = float(request.form.get('inss_faixa1_limite'))
            parametro.inss_faixa1_percentual = float(request.form.get('inss_faixa1_percentual'))
            parametro.inss_faixa2_limite = float(request.form.get('inss_faixa2_limite'))
            parametro.inss_faixa2_percentual = float(request.form.get('inss_faixa2_percentual'))
            parametro.inss_faixa3_limite = float(request.form.get('inss_faixa3_limite'))
            parametro.inss_faixa3_percentual = float(request.form.get('inss_faixa3_percentual'))
            parametro.inss_faixa4_limite = float(request.form.get('inss_faixa4_limite'))
            parametro.inss_faixa4_percentual = float(request.form.get('inss_faixa4_percentual'))
            parametro.inss_teto = float(request.form.get('inss_teto'))
            
            # Atualizar IRRF
            parametro.irrf_isencao = float(request.form.get('irrf_isencao'))
            parametro.irrf_faixa1_limite = float(request.form.get('irrf_faixa1_limite'))
            parametro.irrf_faixa1_percentual = float(request.form.get('irrf_faixa1_percentual'))
            parametro.irrf_faixa1_deducao = float(request.form.get('irrf_faixa1_deducao'))
            parametro.irrf_faixa2_limite = float(request.form.get('irrf_faixa2_limite'))
            parametro.irrf_faixa2_percentual = float(request.form.get('irrf_faixa2_percentual'))
            parametro.irrf_faixa2_deducao = float(request.form.get('irrf_faixa2_deducao'))
            parametro.irrf_faixa3_limite = float(request.form.get('irrf_faixa3_limite'))
            parametro.irrf_faixa3_percentual = float(request.form.get('irrf_faixa3_percentual'))
            parametro.irrf_faixa3_deducao = float(request.form.get('irrf_faixa3_deducao'))
            parametro.irrf_faixa4_percentual = float(request.form.get('irrf_faixa4_percentual'))
            parametro.irrf_faixa4_deducao = float(request.form.get('irrf_faixa4_deducao'))
            parametro.irrf_dependente_valor = float(request.form.get('irrf_dependente_valor'))
            
            # Atualizar Outros
            parametro.fgts_percentual = float(request.form.get('fgts_percentual'))
            parametro.salario_minimo = float(request.form.get('salario_minimo'))
            parametro.vale_transporte_percentual = float(request.form.get('vale_transporte_percentual'))
            parametro.adicional_noturno_percentual = float(request.form.get('adicional_noturno_percentual'))
            parametro.hora_extra_50_percentual = float(request.form.get('hora_extra_50_percentual'))
            parametro.hora_extra_100_percentual = float(request.form.get('hora_extra_100_percentual'))
            
            parametro.updated_at = datetime.utcnow()
            
            db.session.commit()
            
            flash(f'Parâmetros legais de {parametro.ano_vigencia} atualizados com sucesso!', 'success')
            return redirect(url_for('folha.parametros_legais'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao atualizar parâmetros: {str(e)}', 'danger')
    
    return render_template('folha_pagamento/parametros_form.html',
                         parametro=parametro,
                         ano_padrao=None)

@folha_bp.route('/parametros-legais/toggle/<int:id>', methods=['POST'])
@admin_required
def toggle_parametros(id):
    """Ativar/desativar parâmetros legais"""
    parametro = ParametrosLegais.query.filter_by(
        id=id,
        admin_id=current_user.id
    ).first_or_404()
    
    parametro.ativo = not parametro.ativo
    db.session.commit()
    
    status = 'ativado' if parametro.ativo else 'desativado'
    flash(f'Parâmetros de {parametro.ano_vigencia} {status} com sucesso!', 'success')
    return redirect(url_for('folha.parametros_legais'))

# ================================
# BENEFÍCIOS
# ================================

@folha_bp.route('/beneficios')
@admin_required
def beneficios():
    """Listar benefícios de funcionários com agrupamento por tipo"""
    
    # Buscar todos os benefícios do admin com join do funcionário
    beneficios = BeneficioFuncionario.query.filter_by(
        admin_id=current_user.id
    ).join(Funcionario).order_by(
        BeneficioFuncionario.tipo_beneficio,
        Funcionario.nome
    ).all()
    
    # Agrupar por tipo de benefício
    beneficios_por_tipo = {}
    totais_por_tipo = {}
    
    tipos_beneficio = ['VR', 'VT', 'PLANO_SAUDE', 'PLANO_ODONTO', 'SEGURO_VIDA', 'AUXILIO_CRECHE', 'OUTRO']
    
    for tipo in tipos_beneficio:
        beneficios_tipo = [b for b in beneficios if b.tipo_beneficio == tipo]
        beneficios_por_tipo[tipo] = beneficios_tipo
        
        # Calcular total apenas dos ativos
        total = sum(float(b.valor or 0) for b in beneficios_tipo if b.ativo)
        totais_por_tipo[tipo] = total
    
    # Total geral de benefícios ativos
    total_geral = sum(totais_por_tipo.values())
    
    # Contar funcionários com benefícios
    funcionarios_com_beneficios = len(set(b.funcionario_id for b in beneficios if b.ativo))
    
    return render_template('folha_pagamento/beneficios.html',
                         beneficios=beneficios,
                         beneficios_por_tipo=beneficios_por_tipo,
                         totais_por_tipo=totais_por_tipo,
                         total_geral=total_geral,
                         funcionarios_com_beneficios=funcionarios_com_beneficios,
                         tipos_beneficio=tipos_beneficio)

@folha_bp.route('/beneficios/criar', methods=['GET', 'POST'])
@admin_required
def criar_beneficio():
    """Criar novo benefício"""
    
    if request.method == 'POST':
        try:
            funcionario_id = request.form.get('funcionario_id')
            tipo_beneficio = request.form.get('tipo_beneficio')
            valor = request.form.get('valor')
            percentual_desconto = request.form.get('percentual_desconto', 0)
            dias_por_mes = request.form.get('dias_por_mes', 22)
            data_inicio = request.form.get('data_inicio')
            data_fim = request.form.get('data_fim')
            observacoes = request.form.get('observacoes')
            ativo = request.form.get('ativo') == 'on'
            
            # Validações
            if not funcionario_id or not tipo_beneficio or not valor or not data_inicio:
                flash('Funcionário, tipo de benefício, valor e data de início são obrigatórios.', 'warning')
                return redirect(url_for('folha.criar_beneficio'))
            
            # Verificar se o funcionário pertence ao admin
            funcionario = Funcionario.query.filter_by(
                id=funcionario_id,
                admin_id=current_user.id
            ).first()
            
            if not funcionario:
                flash('Funcionário não encontrado.', 'danger')
                return redirect(url_for('folha.criar_beneficio'))
            
            # Criar benefício
            beneficio = BeneficioFuncionario(
                funcionario_id=funcionario_id,
                tipo_beneficio=tipo_beneficio,
                valor=float(valor),
                percentual_desconto=float(percentual_desconto),
                dias_por_mes=int(dias_por_mes),
                data_inicio=datetime.strptime(data_inicio, '%Y-%m-%d').date(),
                data_fim=datetime.strptime(data_fim, '%Y-%m-%d').date() if data_fim else None,
                observacoes=observacoes,
                ativo=ativo,
                admin_id=current_user.id
            )
            
            db.session.add(beneficio)
            db.session.commit()
            
            flash(f'Benefício {tipo_beneficio} criado com sucesso para {funcionario.nome}!', 'success')
            return redirect(url_for('folha.beneficios'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao criar benefício: {str(e)}', 'danger')
            return redirect(url_for('folha.criar_beneficio'))
    
    # GET - buscar funcionários ativos
    funcionarios = Funcionario.query.filter_by(
        admin_id=current_user.id,
        ativo=True
    ).order_by(Funcionario.nome).all()
    
    tipos_beneficio = [
        ('VR', 'Vale Refeição'),
        ('VT', 'Vale Transporte'),
        ('PLANO_SAUDE', 'Plano de Saúde'),
        ('PLANO_ODONTO', 'Plano Odontológico'),
        ('SEGURO_VIDA', 'Seguro de Vida'),
        ('AUXILIO_CRECHE', 'Auxílio Creche'),
        ('OUTRO', 'Outro')
    ]
    
    return render_template('folha_pagamento/beneficio_form.html',
                         beneficio=None,
                         funcionarios=funcionarios,
                         tipos_beneficio=tipos_beneficio,
                         hoje=date.today())

@folha_bp.route('/beneficios/editar/<int:id>', methods=['GET', 'POST'])
@admin_required
def editar_beneficio(id):
    """Editar benefício existente"""
    
    beneficio = BeneficioFuncionario.query.filter_by(
        id=id,
        admin_id=current_user.id
    ).first_or_404()
    
    if request.method == 'POST':
        try:
            tipo_beneficio = request.form.get('tipo_beneficio')
            valor = request.form.get('valor')
            percentual_desconto = request.form.get('percentual_desconto', 0)
            dias_por_mes = request.form.get('dias_por_mes', 22)
            data_inicio = request.form.get('data_inicio')
            data_fim = request.form.get('data_fim')
            observacoes = request.form.get('observacoes')
            ativo = request.form.get('ativo') == 'on'
            
            # Validações
            if not tipo_beneficio or not valor or not data_inicio:
                flash('Tipo de benefício, valor e data de início são obrigatórios.', 'warning')
                return redirect(url_for('folha.editar_beneficio', id=id))
            
            # Atualizar benefício
            beneficio.tipo_beneficio = tipo_beneficio
            beneficio.valor = float(valor)
            beneficio.percentual_desconto = float(percentual_desconto)
            beneficio.dias_por_mes = int(dias_por_mes)
            beneficio.data_inicio = datetime.strptime(data_inicio, '%Y-%m-%d').date()
            beneficio.data_fim = datetime.strptime(data_fim, '%Y-%m-%d').date() if data_fim else None
            beneficio.observacoes = observacoes
            beneficio.ativo = ativo
            beneficio.updated_at = datetime.utcnow()
            
            db.session.commit()
            
            flash(f'Benefício atualizado com sucesso!', 'success')
            return redirect(url_for('folha.beneficios'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao atualizar benefício: {str(e)}', 'danger')
            return redirect(url_for('folha.editar_beneficio', id=id))
    
    # GET - buscar funcionários ativos
    funcionarios = Funcionario.query.filter_by(
        admin_id=current_user.id,
        ativo=True
    ).order_by(Funcionario.nome).all()
    
    tipos_beneficio = [
        ('VR', 'Vale Refeição'),
        ('VT', 'Vale Transporte'),
        ('PLANO_SAUDE', 'Plano de Saúde'),
        ('PLANO_ODONTO', 'Plano Odontológico'),
        ('SEGURO_VIDA', 'Seguro de Vida'),
        ('AUXILIO_CRECHE', 'Auxílio Creche'),
        ('OUTRO', 'Outro')
    ]
    
    return render_template('folha_pagamento/beneficio_form.html',
                         beneficio=beneficio,
                         funcionarios=funcionarios,
                         tipos_beneficio=tipos_beneficio,
                         hoje=date.today())

@folha_bp.route('/beneficios/deletar/<int:id>', methods=['POST'])
@admin_required
def deletar_beneficio(id):
    """Deletar benefício"""
    
    beneficio = BeneficioFuncionario.query.filter_by(
        id=id,
        admin_id=current_user.id
    ).first_or_404()
    
    try:
        tipo = beneficio.tipo_beneficio
        funcionario_nome = beneficio.funcionario.nome
        
        db.session.delete(beneficio)
        db.session.commit()
        
        flash(f'Benefício {tipo} de {funcionario_nome} deletado com sucesso!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao deletar benefício: {str(e)}', 'danger')
    
    return redirect(url_for('folha.beneficios'))

# ================================
# ADIANTAMENTOS
# ================================

@folha_bp.route('/adiantamentos')
@admin_required
def adiantamentos():
    """Listar adiantamentos salariais com filtros por status"""
    
    # Filtro por status via query param
    status_filtro = request.args.get('status', 'TODOS')
    
    # Query base com join do funcionário
    query = Adiantamento.query.filter_by(
        admin_id=current_user.id
    ).join(Funcionario)
    
    # Aplicar filtro de status se não for TODOS
    if status_filtro != 'TODOS':
        query = query.filter(Adiantamento.status == status_filtro)
    
    # Ordenar por data de solicitação (mais recentes primeiro)
    adiantamentos_list = query.order_by(Adiantamento.data_solicitacao.desc()).all()
    
    # Calcular totais por status
    todos_adiantamentos = Adiantamento.query.filter_by(admin_id=current_user.id).all()
    
    totais = {
        'SOLICITADO': {
            'count': len([a for a in todos_adiantamentos if a.status == 'SOLICITADO']),
            'valor': sum(float(a.valor_total or 0) for a in todos_adiantamentos if a.status == 'SOLICITADO')
        },
        'APROVADO': {
            'count': len([a for a in todos_adiantamentos if a.status == 'APROVADO']),
            'valor': sum(float(a.valor_total or 0) for a in todos_adiantamentos if a.status == 'APROVADO')
        },
        'QUITADO': {
            'count': len([a for a in todos_adiantamentos if a.status == 'QUITADO']),
            'valor': sum(float(a.valor_total or 0) for a in todos_adiantamentos if a.status == 'QUITADO')
        },
        'CANCELADO': {
            'count': len([a for a in todos_adiantamentos if a.status == 'CANCELADO']),
            'valor': sum(float(a.valor_total or 0) for a in todos_adiantamentos if a.status == 'CANCELADO')
        }
    }
    
    return render_template('folha_pagamento/adiantamentos.html',
                         adiantamentos=adiantamentos_list,
                         status_filtro=status_filtro,
                         totais=totais)

@folha_bp.route('/adiantamentos/criar', methods=['GET', 'POST'])
@admin_required
def criar_adiantamento():
    """Criar novo adiantamento salarial"""
    
    if request.method == 'POST':
        try:
            funcionario_id = request.form.get('funcionario_id')
            valor_total = request.form.get('valor_total')
            parcelas = request.form.get('parcelas', 1)
            motivo = request.form.get('motivo')
            observacoes = request.form.get('observacoes')
            
            # Validações
            if not funcionario_id or not valor_total or not parcelas or not motivo:
                flash('Funcionário, valor total, parcelas e motivo são obrigatórios.', 'warning')
                return redirect(url_for('folha.criar_adiantamento'))
            
            # Verificar se o funcionário pertence ao admin
            funcionario = Funcionario.query.filter_by(
                id=funcionario_id,
                admin_id=current_user.id,
                ativo=True
            ).first()
            
            if not funcionario:
                flash('Funcionário não encontrado ou inativo.', 'danger')
                return redirect(url_for('folha.criar_adiantamento'))
            
            # Calcular valor da parcela
            valor_total_float = float(valor_total)
            parcelas_int = int(parcelas)
            valor_parcela = valor_total_float / parcelas_int
            
            # Criar adiantamento
            adiantamento = Adiantamento(
                funcionario_id=funcionario_id,
                valor_total=valor_total_float,
                data_solicitacao=date.today(),
                parcelas=parcelas_int,
                valor_parcela=valor_parcela,
                parcelas_pagas=0,
                status='SOLICITADO',
                motivo=motivo,
                observacoes=observacoes,
                admin_id=current_user.id
            )
            
            db.session.add(adiantamento)
            db.session.commit()
            
            flash(f'Adiantamento de R$ {valor_total_float:,.2f} solicitado para {funcionario.nome}!', 'warning')
            return redirect(url_for('folha.adiantamentos'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao criar adiantamento: {str(e)}', 'danger')
            return redirect(url_for('folha.criar_adiantamento'))
    
    # GET - buscar funcionários ativos
    funcionarios = Funcionario.query.filter_by(
        admin_id=current_user.id,
        ativo=True
    ).order_by(Funcionario.nome).all()
    
    return render_template('folha_pagamento/adiantamento_form.html',
                         adiantamento=None,
                         funcionarios=funcionarios,
                         hoje=date.today())

@folha_bp.route('/adiantamentos/aprovar/<int:id>', methods=['POST'])
@admin_required
def aprovar_adiantamento(id):
    """Aprovar adiantamento salarial"""
    
    # Validar propriedade (admin_id)
    adiantamento = Adiantamento.query.filter_by(
        id=id,
        admin_id=current_user.id
    ).first_or_404()
    
    try:
        # Verificar se já foi aprovado ou cancelado
        if adiantamento.status != 'SOLICITADO':
            flash('Este adiantamento não pode ser aprovado.', 'warning')
            return redirect(url_for('folha.adiantamentos'))
        
        # Atualizar status
        adiantamento.status = 'APROVADO'
        adiantamento.data_aprovacao = date.today()
        adiantamento.aprovado_por = current_user.id
        adiantamento.updated_at = datetime.utcnow()
        
        db.session.commit()
        
        flash(f'Adiantamento de R$ {float(adiantamento.valor_total):,.2f} para {adiantamento.funcionario.nome} aprovado com sucesso!', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao aprovar adiantamento: {str(e)}', 'danger')
    
    return redirect(url_for('folha.adiantamentos'))

@folha_bp.route('/adiantamentos/rejeitar/<int:id>', methods=['POST'])
@admin_required
def rejeitar_adiantamento(id):
    """Rejeitar adiantamento salarial"""
    
    # Validar propriedade (admin_id)
    adiantamento = Adiantamento.query.filter_by(
        id=id,
        admin_id=current_user.id
    ).first_or_404()
    
    try:
        # Verificar se pode ser rejeitado
        if adiantamento.status != 'SOLICITADO':
            flash('Este adiantamento não pode ser rejeitado.', 'warning')
            return redirect(url_for('folha.adiantamentos'))
        
        # Obter motivo da rejeição
        motivo_rejeicao = request.form.get('motivo_rejeicao', 'Sem motivo especificado')
        
        # Atualizar status e adicionar motivo nas observações
        adiantamento.status = 'CANCELADO'
        observacoes_atual = adiantamento.observacoes or ''
        adiantamento.observacoes = f"{observacoes_atual}\n\nREJEITADO EM {date.today().strftime('%d/%m/%Y')}: {motivo_rejeicao}"
        adiantamento.updated_at = datetime.utcnow()
        
        db.session.commit()
        
        flash(f'Adiantamento de R$ {float(adiantamento.valor_total):,.2f} para {adiantamento.funcionario.nome} rejeitado.', 'danger')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao rejeitar adiantamento: {str(e)}', 'danger')
    
    return redirect(url_for('folha.adiantamentos'))

# ================================
# RELATÓRIOS
# ================================

@folha_bp.route('/relatorios')
@admin_required
def relatorios():
    """Dashboard de relatórios - últimos 6 meses com folhas processadas"""
    
    hoje = date.today()
    mes_atual = hoje.replace(day=1)
    
    # Calcular últimos 6 meses
    meses_disponiveis = []
    for i in range(6):
        # Calcular mês (voltando i meses)
        if mes_atual.month - i <= 0:
            ano = mes_atual.year - 1
            mes = 12 + (mes_atual.month - i)
        else:
            ano = mes_atual.year
            mes = mes_atual.month - i
        
        mes_referencia = date(ano, mes, 1)
        
        # Buscar folhas deste mês
        folhas_mes = FolhaPagamento.query.filter_by(
            admin_id=current_user.id,
            mes_referencia=mes_referencia
        ).all()
        
        if folhas_mes:
            # Formatar nome do mês em português
            meses_pt = ['Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
                       'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']
            mes_nome = f"{meses_pt[mes - 1]}/{ano}"
            
            total_folhas = len(folhas_mes)
            total_bruto = sum(f.total_proventos or 0 for f in folhas_mes)
            total_liquido = sum(f.salario_liquido or 0 for f in folhas_mes)
            
            meses_disponiveis.append({
                'mes_referencia': mes_referencia,
                'mes_nome': mes_nome,
                'total_folhas': total_folhas,
                'total_bruto': total_bruto,
                'total_liquido': total_liquido,
                'ano': ano,
                'mes': mes
            })
    
    return render_template('folha_pagamento/relatorios.html',
                         meses_disponiveis=meses_disponiveis)

@folha_bp.route('/relatorios/holerite/<int:folha_id>')
@admin_required
def holerite_pdf(folha_id):
    """Gerar holerite em PDF para uma folha específica"""
    
    # Buscar folha e validar admin_id
    folha = FolhaPagamento.query.filter_by(
        id=folha_id,
        admin_id=current_user.id
    ).first_or_404()
    
    funcionario = folha.funcionario
    
    # Dados da empresa (pode vir de configurações)
    empresa = {
        'nome': 'EMPRESA EXEMPLO LTDA',
        'cnpj': '00.000.000/0001-00',
        'endereco': 'Rua Exemplo, 123 - Centro',
        'cidade': 'São Paulo - SP'
    }
    
    try:
        # Tentar usar reportlab (disponível)
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_RIGHT
        
        # Criar buffer para PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=2*cm, bottomMargin=2*cm)
        
        # Estilos
        styles = getSampleStyleSheet()
        style_title = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#2c3e50'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        style_normal = styles['Normal']
        style_right = ParagraphStyle('Right', parent=styles['Normal'], alignment=TA_RIGHT)
        
        # Elementos do PDF
        story = []
        
        # Cabeçalho da empresa
        story.append(Paragraph(f"<b>{empresa['nome']}</b>", style_title))
        story.append(Paragraph(f"CNPJ: {empresa['cnpj']}", styles['Normal']))
        story.append(Paragraph(f"{empresa['endereco']}", styles['Normal']))
        story.append(Paragraph(f"{empresa['cidade']}", styles['Normal']))
        story.append(Spacer(1, 0.5*cm))
        
        # Título do holerite
        meses_pt = ['Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
                   'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']
        mes_nome = meses_pt[folha.mes_referencia.month - 1]
        story.append(Paragraph(f"<b>RECIBO DE PAGAMENTO DE SALÁRIO</b>", style_title))
        story.append(Paragraph(f"<b>Referência: {mes_nome}/{folha.mes_referencia.year}</b>", styles['Normal']))
        story.append(Spacer(1, 0.5*cm))
        
        # Dados do funcionário
        data_funcionario = [
            ['Funcionário:', funcionario.nome, 'CPF:', funcionario.cpf],
            ['Cargo:', funcionario.funcao_ref.nome if funcionario.funcao_ref else 'N/A', 
             'Código:', funcionario.codigo]
        ]
        
        t_func = Table(data_funcionario, colWidths=[3*cm, 7*cm, 2.5*cm, 4.5*cm])
        t_func.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#ecf0f1')),
            ('BACKGROUND', (2, 0), (2, -1), colors.HexColor('#ecf0f1')),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        story.append(t_func)
        story.append(Spacer(1, 0.5*cm))
        
        # Tabela de proventos e descontos
        data_valores = [
            ['PROVENTOS', 'VALOR', 'DESCONTOS', 'VALOR']
        ]
        
        # Proventos
        proventos = [
            ('Salário Base', folha.salario_base or 0),
            ('Horas Extras', folha.horas_extras or 0),
        ]
        
        # Descontos
        descontos = [
            ('INSS', folha.inss or 0),
            ('IRRF', folha.irrf or 0),
            ('Outros Descontos', folha.outros_descontos or 0),
        ]
        
        max_linhas = max(len(proventos), len(descontos))
        for i in range(max_linhas):
            prov_desc = proventos[i][0] if i < len(proventos) else ''
            prov_val = f"R$ {proventos[i][1]:,.2f}" if i < len(proventos) else ''
            desc_desc = descontos[i][0] if i < len(descontos) else ''
            desc_val = f"R$ {descontos[i][1]:,.2f}" if i < len(descontos) else ''
            data_valores.append([prov_desc, prov_val, desc_desc, desc_val])
        
        t_valores = Table(data_valores, colWidths=[6*cm, 3*cm, 6*cm, 3*cm])
        t_valores.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#34495e')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('ALIGN', (3, 0), (3, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        story.append(t_valores)
        story.append(Spacer(1, 0.3*cm))
        
        # Totais
        data_totais = [
            ['TOTAL PROVENTOS', f"R$ {folha.total_proventos or 0:,.2f}",
             'TOTAL DESCONTOS', f"R$ {folha.total_descontos or 0:,.2f}"],
            ['', '', 'SALÁRIO LÍQUIDO', f"R$ {folha.salario_liquido or 0:,.2f}"]
        ]
        
        t_totais = Table(data_totais, colWidths=[6*cm, 3*cm, 6*cm, 3*cm])
        t_totais.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (1, 0), colors.HexColor('#27ae60')),
            ('BACKGROUND', (2, 0), (3, 0), colors.HexColor('#e74c3c')),
            ('BACKGROUND', (2, 1), (3, 1), colors.HexColor('#2c3e50')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.whitesmoke),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('ALIGN', (3, 0), (3, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        story.append(t_totais)
        story.append(Spacer(1, 0.5*cm))
        
        # FGTS
        story.append(Paragraph(f"<b>FGTS do Mês:</b> R$ {folha.fgts or 0:,.2f}", styles['Normal']))
        story.append(Spacer(1, 1*cm))
        
        # Assinatura
        story.append(Paragraph("_" * 50, style_right))
        story.append(Paragraph("Assinatura do Funcionário", style_right))
        
        # Construir PDF
        doc.build(story)
        
        # Retornar PDF
        buffer.seek(0)
        response = make_response(buffer.getvalue())
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename=holerite_{funcionario.codigo}_{folha.mes_referencia.strftime("%Y%m")}.pdf'
        
        return response
        
    except Exception as e:
        flash(f'Erro ao gerar PDF: {str(e)}', 'danger')
        return redirect(url_for('folha.relatorios'))

@folha_bp.route('/api/funcionarios/folha/<int:ano>/<int:mes>')
@admin_required
def api_funcionarios_folha(ano, mes):
    """API para listar funcionários com folha processada em um mês específico"""
    
    try:
        mes_referencia = date(ano, mes, 1)
        
        # Buscar folhas do mês
        folhas = FolhaPagamento.query.filter_by(
            admin_id=current_user.id,
            mes_referencia=mes_referencia
        ).all()
        
        funcionarios = []
        for folha in folhas:
            func = folha.funcionario
            funcionarios.append({
                'folha_id': folha.id,
                'nome': func.nome,
                'cpf': func.cpf,
                'cargo': func.funcao_ref.nome if func.funcao_ref else 'N/A',
                'salario_liquido': float(folha.salario_liquido or 0)
            })
        
        return jsonify({
            'success': True,
            'funcionarios': funcionarios
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@folha_bp.route('/relatorios/analitico/<int:ano>/<int:mes>')
@admin_required
def relatorio_excel(ano, mes):
    """Gerar relatório analítico em Excel para um mês específico"""
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        mes_referencia = date(ano, mes, 1)
        
        # Buscar folhas do mês e validar admin_id
        folhas = FolhaPagamento.query.filter_by(
            admin_id=current_user.id,
            mes_referencia=mes_referencia
        ).all()
        
        if not folhas:
            flash(f'Nenhuma folha encontrada para {mes:02d}/{ano}.', 'warning')
            return redirect(url_for('folha.relatorios'))
        
        # Criar workbook
        wb = Workbook()
        ws = wb.active
        
        # Nome da planilha
        meses_pt = ['Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
                   'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']
        ws.title = f"Folha {meses_pt[mes-1]} {ano}"
        
        # Estilos
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2c3e50', end_color='2c3e50', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        cell_alignment = Alignment(horizontal='left', vertical='center')
        number_alignment = Alignment(horizontal='right', vertical='center')
        
        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        # Cabeçalhos
        headers = [
            'Funcionário', 'CPF', 'Cargo', 'Salário Base', 'Horas Extras',
            'Total Proventos', 'INSS', 'IRRF', 'Outros Desc.',
            'Total Desc.', 'Líquido', 'FGTS'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Dados
        row_num = 2
        for folha in folhas:
            funcionario = folha.funcionario
            
            ws.cell(row=row_num, column=1, value=funcionario.nome).alignment = cell_alignment
            ws.cell(row=row_num, column=2, value=funcionario.cpf).alignment = cell_alignment
            ws.cell(row=row_num, column=3, value=funcionario.funcao_ref.nome if funcionario.funcao_ref else 'N/A').alignment = cell_alignment
            ws.cell(row=row_num, column=4, value=folha.salario_base or 0).alignment = number_alignment
            ws.cell(row=row_num, column=4).number_format = 'R$ #,##0.00'
            ws.cell(row=row_num, column=5, value=folha.horas_extras or 0).alignment = number_alignment
            ws.cell(row=row_num, column=5).number_format = 'R$ #,##0.00'
            ws.cell(row=row_num, column=6, value=folha.total_proventos or 0).alignment = number_alignment
            ws.cell(row=row_num, column=6).number_format = 'R$ #,##0.00'
            ws.cell(row=row_num, column=7, value=folha.inss or 0).alignment = number_alignment
            ws.cell(row=row_num, column=7).number_format = 'R$ #,##0.00'
            ws.cell(row=row_num, column=8, value=folha.irrf or 0).alignment = number_alignment
            ws.cell(row=row_num, column=8).number_format = 'R$ #,##0.00'
            ws.cell(row=row_num, column=9, value=folha.outros_descontos or 0).alignment = number_alignment
            ws.cell(row=row_num, column=9).number_format = 'R$ #,##0.00'
            ws.cell(row=row_num, column=10, value=folha.total_descontos or 0).alignment = number_alignment
            ws.cell(row=row_num, column=10).number_format = 'R$ #,##0.00'
            ws.cell(row=row_num, column=11, value=folha.salario_liquido or 0).alignment = number_alignment
            ws.cell(row=row_num, column=11).number_format = 'R$ #,##0.00'
            ws.cell(row=row_num, column=12, value=folha.fgts or 0).alignment = number_alignment
            ws.cell(row=row_num, column=12).number_format = 'R$ #,##0.00'
            
            # Aplicar bordas
            for col in range(1, 13):
                ws.cell(row=row_num, column=col).border = border
            
            row_num += 1
        
        # Ajustar largura das colunas
        column_widths = {
            'A': 30, 'B': 15, 'C': 25, 'D': 14, 'E': 14,
            'F': 16, 'G': 12, 'H': 12, 'I': 14,
            'J': 14, 'K': 14, 'L': 12
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Salvar em buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Nome do arquivo
        filename = f"folha_analitica_{ano}_{mes:02d}.xlsx"
        
        return send_file(
            buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        flash(f'Erro ao gerar relatório Excel: {str(e)}', 'danger')
        return redirect(url_for('folha.relatorios'))