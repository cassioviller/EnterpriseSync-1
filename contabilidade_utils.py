"""
Utilitários para o Módulo 7 - Sistema Contábil Completo
"""
import calendar
from datetime import date, timedelta
from decimal import Decimal
from sqlalchemy import func, case
from models import (
    db, PlanoContas, CentroCustoContabil, LancamentoContabil, PartidaContabil, 
    BalanceteMensal, DREMensal, BalancoPatrimonial, FluxoCaixaContabil, 
    ProvisaoMensal, SpedContabil, AuditoriaContabil, Proposta, 
    NotaFiscal, MovimentacaoEstoque, FolhaPagamento
)

# ===============================================================
# == MÓDULO 7: FUNÇÕES UTILITÁRIAS CONTÁBEIS
# ===============================================================

def criar_plano_contas_padrao(admin_id):
    """Cria o plano de contas brasileiro completo para um novo admin."""
    if PlanoContas.query.filter_by(admin_id=admin_id).first():
        return # Já existe

    contas = [
        # 1. ATIVO
        ('1', 'ATIVO', 'ATIVO', 'DEVEDORA', 1, None, False),
        ('1.1', 'ATIVO CIRCULANTE', 'ATIVO', 'DEVEDORA', 2, '1', False),
        ('1.1.01', 'DISPONÍVEL', 'ATIVO', 'DEVEDORA', 3, '1.1', False),
        ('1.1.01.001', 'Caixa', 'ATIVO', 'DEVEDORA', 4, '1.1.01', True),
        ('1.1.01.002', 'Bancos Conta Movimento', 'ATIVO', 'DEVEDORA', 4, '1.1.01', True),
        ('1.1.01.003', 'Aplicações Financeiras', 'ATIVO', 'DEVEDORA', 4, '1.1.01', True),
        
        ('1.1.02', 'CONTAS A RECEBER', 'ATIVO', 'DEVEDORA', 3, '1.1', False),
        ('1.1.02.001', 'Clientes', 'ATIVO', 'DEVEDORA', 4, '1.1.02', True),
        ('1.1.02.002', 'Duplicatas a Receber', 'ATIVO', 'DEVEDORA', 4, '1.1.02', True),
        ('1.1.02.003', 'Provisão para Devedores Duvidosos', 'ATIVO', 'CREDORA', 4, '1.1.02', True),
        
        ('1.1.03', 'ESTOQUES', 'ATIVO', 'DEVEDORA', 3, '1.1', False),
        ('1.1.03.001', 'Estoque de Materiais', 'ATIVO', 'DEVEDORA', 4, '1.1.03', True),
        ('1.1.03.002', 'Estoque de Produtos Acabados', 'ATIVO', 'DEVEDORA', 4, '1.1.03', True),
        ('1.1.03.003', 'Estoque em Trânsito', 'ATIVO', 'DEVEDORA', 4, '1.1.03', True),
        
        ('1.1.04', 'IMPOSTOS A RECUPERAR', 'ATIVO', 'DEVEDORA', 3, '1.1', False),
        ('1.1.04.001', 'ICMS a Recuperar', 'ATIVO', 'DEVEDORA', 4, '1.1.04', True),
        ('1.1.04.002', 'IPI a Recuperar', 'ATIVO', 'DEVEDORA', 4, '1.1.04', True),
        ('1.1.04.003', 'PIS a Recuperar', 'ATIVO', 'DEVEDORA', 4, '1.1.04', True),
        ('1.1.04.004', 'COFINS a Recuperar', 'ATIVO', 'DEVEDORA', 4, '1.1.04', True),
        
        # 2. PASSIVO
        ('2', 'PASSIVO', 'PASSIVO', 'CREDORA', 1, None, False),
        ('2.1', 'PASSIVO CIRCULANTE', 'PASSIVO', 'CREDORA', 2, '2', False),
        ('2.1.01', 'FORNECEDORES', 'PASSIVO', 'CREDORA', 3, '2.1', False),
        ('2.1.01.001', 'Fornecedores Nacionais', 'PASSIVO', 'CREDORA', 4, '2.1.01', True),
        ('2.1.01.002', 'Fornecedores Estrangeiros', 'PASSIVO', 'CREDORA', 4, '2.1.01', True),
        
        ('2.1.02', 'OBRIGAÇÕES TRABALHISTAS', 'PASSIVO', 'CREDORA', 3, '2.1', False),
        ('2.1.02.001', 'Salários a Pagar', 'PASSIVO', 'CREDORA', 4, '2.1.02', True),
        ('2.1.02.002', 'Provisão para Férias', 'PASSIVO', 'CREDORA', 4, '2.1.02', True),
        ('2.1.02.003', 'Provisão para 13º Salário', 'PASSIVO', 'CREDORA', 4, '2.1.02', True),
        ('2.1.02.004', 'FGTS a Recolher', 'PASSIVO', 'CREDORA', 4, '2.1.02', True),
        
        ('2.1.03', 'OBRIGAÇÕES FISCAIS', 'PASSIVO', 'CREDORA', 3, '2.1', False),
        ('2.1.03.007', 'INSS a Recolher', 'PASSIVO', 'CREDORA', 4, '2.1.03', True),
        ('2.1.03.008', 'IRRF a Recolher', 'PASSIVO', 'CREDORA', 4, '2.1.03', True),
        ('2.1.03.009', 'ISS a Recolher', 'PASSIVO', 'CREDORA', 4, '2.1.03', True),
        
        # 3. PATRIMÔNIO LÍQUIDO
        ('3', 'PATRIMÔNIO LÍQUIDO', 'PATRIMONIO', 'CREDORA', 1, None, False),
        ('3.1', 'CAPITAL SOCIAL', 'PATRIMONIO', 'CREDORA', 2, '3', False),
        ('3.1.01', 'Capital Social Subscrito', 'PATRIMONIO', 'CREDORA', 3, '3.1', True),
        
        # 4. RECEITAS
        ('4', 'RECEITAS', 'RECEITA', 'CREDORA', 1, None, False),
        ('4.1', 'RECEITA BRUTA', 'RECEITA', 'CREDORA', 2, '4', False),
        ('4.1.02', 'RECEITA DE SERVIÇOS', 'RECEITA', 'CREDORA', 3, '4.1', False),
        ('4.1.02.001', 'Prestação de Serviços', 'RECEITA', 'CREDORA', 4, '4.1.02', True),
        ('4.1.02.002', 'Serviços de Construção', 'RECEITA', 'CREDORA', 4, '4.1.02', True),
        
        # 5. CUSTOS
        ('5', 'CUSTOS', 'DESPESA', 'DEVEDORA', 1, None, False),
        ('5.1', 'CUSTO DOS SERVIÇOS PRESTADOS', 'DESPESA', 'DEVEDORA', 2, '5', False),
        ('5.1.01', 'Materiais Diretos', 'DESPESA', 'DEVEDORA', 3, '5.1', True),
        ('5.1.02', 'Mão de Obra Direta', 'DESPESA', 'DEVEDORA', 3, '5.1', True),
        ('5.2', 'CUSTOS INDIRETOS', 'DESPESA', 'DEVEDORA', 2, '5', False),
        ('5.2.01', 'Materiais Indiretos', 'DESPESA', 'DEVEDORA', 3, '5.2', True),
        
        # 6. DESPESAS
        ('6', 'DESPESAS', 'DESPESA', 'DEVEDORA', 1, None, False),
        ('6.1', 'DESPESAS OPERACIONAIS', 'DESPESA', 'DEVEDORA', 2, '6', False),
        ('6.1.01', 'DESPESAS COM PESSOAL', 'DESPESA', 'DEVEDORA', 3, '6.1', False),
        ('6.1.01.001', 'Salários', 'DESPESA', 'DEVEDORA', 4, '6.1.01', True),
        ('6.1.01.002', 'Encargos Sociais', 'DESPESA', 'DEVEDORA', 4, '6.1.01', True),
        
        ('6.1.02', 'DESPESAS ADMINISTRATIVAS', 'DESPESA', 'DEVEDORA', 3, '6.1', False),
        ('6.1.02.001', 'Material de Escritório', 'DESPESA', 'DEVEDORA', 4, '6.1.02', True),
        ('6.1.02.002', 'Telefone', 'DESPESA', 'DEVEDORA', 4, '6.1.02', True),
        ('6.1.02.003', 'Energia Elétrica', 'DESPESA', 'DEVEDORA', 4, '6.1.02', True),
    ]

    for codigo, nome, tipo_conta, natureza, nivel, conta_pai_codigo, aceita_lancamento in contas:
        conta = PlanoContas(
            codigo=codigo, nome=nome, tipo_conta=tipo_conta, natureza=natureza,
            nivel=nivel, conta_pai_codigo=conta_pai_codigo, 
            aceita_lancamento=aceita_lancamento, admin_id=admin_id
        )
        db.session.add(conta)
    db.session.commit()

def get_next_lancamento_numero(admin_id):
    """Obtém o próximo número de lançamento sequencial."""
    last_lancamento = LancamentoContabil.query.filter_by(admin_id=admin_id).order_by(LancamentoContabil.numero.desc()).first()
    return (last_lancamento.numero + 1) if last_lancamento else 1

def criar_lancamento_automatico(data, historico, valor, origem, origem_id, admin_id, partidas):
    """Função central para criar lançamentos contábeis automáticos."""
    # Validação de partidas
    total_debitos = sum(p['valor'] for p in partidas if p['tipo'] == 'DEBITO')
    total_creditos = sum(p['valor'] for p in partidas if p['tipo'] == 'CREDITO')
    if abs(total_debitos - total_creditos) > 0.01:
        raise ValueError("Lançamento desbalanceado: débitos não são iguais aos créditos.")

    lancamento = LancamentoContabil(
        numero=get_next_lancamento_numero(admin_id),
        data_lancamento=data,
        historico=historico,
        valor_total=valor,
        origem=origem,
        origem_id=origem_id,
        admin_id=admin_id
    )
    db.session.add(lancamento)
    db.session.flush() # Para obter o ID do lançamento

    for i, p in enumerate(partidas):
        partida = PartidaContabil(
            lancamento_id=lancamento.id,
            sequencia=i + 1,
            conta_codigo=p['conta'],
            centro_custo_id=p.get('centro_custo_id'),
            tipo_partida=p['tipo'],
            valor=p['valor'],
            historico_complementar=p.get('historico_comp'),
            admin_id=admin_id
        )
        db.session.add(partida)
    
    db.session.commit()
    return lancamento

# --- Funções de Integração Automática ---

def contabilizar_proposta_aprovada(proposta_id):
    """Gera lançamentos contábeis quando uma proposta é aprovada."""
    proposta = Proposta.query.get(proposta_id)
    if not proposta or proposta.status != 'APROVADA':
        return

    # Cria centro de custo para a obra se não existir
    centro_custo = CentroCustoContabil.query.filter_by(obra_id=proposta.obra_id, admin_id=proposta.admin_id).first()
    if not centro_custo:
        centro_custo = CentroCustoContabil(
            codigo=f"OBRA_{proposta.obra.id}", 
            nome=proposta.obra.nome, 
            tipo='OBRA', 
            obra_id=proposta.obra_id, 
            admin_id=proposta.admin_id
        )
        db.session.add(centro_custo)
        db.session.flush()

    partidas = [
        {'tipo': 'DEBITO', 'conta': '1.1.02.001', 'valor': proposta.valor_total, 'centro_custo_id': centro_custo.id},
        {'tipo': 'CREDITO', 'conta': '4.1.02.002', 'valor': proposta.valor_total, 'centro_custo_id': centro_custo.id}
    ]

    criar_lancamento_automatico(
        data=proposta.data_aprovacao,
        historico=f"Aprovação da Proposta #{proposta.id} - Cliente: {proposta.cliente_nome}",
        valor=proposta.valor_total,
        origem='MODULO_1',
        origem_id=proposta.id,
        admin_id=proposta.admin_id,
        partidas=partidas
    )

def contabilizar_entrada_material(nota_fiscal_id):
    """Gera lançamentos contábeis para entrada de material via NF."""
    nota = NotaFiscal.query.get(nota_fiscal_id)
    if not nota:
        return

    partidas = [
        {'tipo': 'DEBITO', 'conta': '1.1.03.001', 'valor': nota.valor_produtos},
        {'tipo': 'CREDITO', 'conta': '2.1.01.001', 'valor': nota.valor_total}
    ]
    if nota.valor_icms and nota.valor_icms > 0:
        partidas.append({'tipo': 'DEBITO', 'conta': '1.1.04.001', 'valor': nota.valor_icms})

    criar_lancamento_automatico(
        data=nota.data_emissao,
        historico=f"Entrada NF #{nota.numero} - Fornecedor: {nota.fornecedor_nome}",
        valor=nota.valor_total,
        origem='MODULO_4',
        origem_id=nota.id,
        admin_id=nota.admin_id,
        partidas=partidas
    )

def contabilizar_folha_pagamento(admin_id, mes_referencia):
    """Gera o lançamento contábil consolidado da folha de pagamento."""
    folhas = FolhaPagamento.query.filter_by(admin_id=admin_id, mes_referencia=mes_referencia).all()
    if not folhas:
        return

    # Totalizadores
    total_salario_bruto = sum(f.salario_bruto or 0 for f in folhas)
    total_inss_func = sum(f.inss or 0 for f in folhas)
    total_irrf = sum(f.irrf or 0 for f in folhas)
    total_fgts = sum(f.fgts or 0 for f in folhas)
    total_liquido = sum(f.salario_liquido or 0 for f in folhas)
    inss_empresa = total_salario_bruto * Decimal('0.20') # Simplificado

    partidas = [
        {'tipo': 'DEBITO', 'conta': '6.1.01.001', 'valor': total_salario_bruto}, # Despesa com Salários
        {'tipo': 'DEBITO', 'conta': '6.1.01.002', 'valor': inss_empresa + total_fgts}, # Despesa com Encargos
        {'tipo': 'CREDITO', 'conta': '2.1.02.001', 'valor': total_liquido}, # Salários a Pagar
        {'tipo': 'CREDITO', 'conta': '2.1.03.007', 'valor': total_inss_func + inss_empresa}, # INSS a Recolher
        {'tipo': 'CREDITO', 'conta': '2.1.02.004', 'valor': total_fgts}, # FGTS a Recolher
        {'tipo': 'CREDITO', 'conta': '2.1.03.008', 'valor': total_irrf}  # IRRF a Recolher
    ]

    criar_lancamento_automatico(
        data=mes_referencia.replace(day=calendar.monthrange(mes_referencia.year, mes_referencia.month)[1]),
        historico=f"Folha de Pagamento ref. {mes_referencia.strftime('%m/%Y')}",
        valor=total_salario_bruto + inss_empresa + total_fgts,
        origem='MODULO_6',
        origem_id=0, # Representa o mês todo
        admin_id=admin_id,
        partidas=[p for p in partidas if p['valor'] > 0]
    )

# --- Funções de Geração de Relatórios ---

def calcular_saldo_conta(conta_codigo, admin_id, data_inicio=None, data_fim=None):
    """Calcula o saldo de uma conta em um período."""
    query = db.session.query(func.sum(
        case((PartidaContabil.tipo_partida == 'DEBITO', PartidaContabil.valor), else_=-PartidaContabil.valor)
    )).join(LancamentoContabil).filter(
        PartidaContabil.conta_codigo == conta_codigo,
        PartidaContabil.admin_id == admin_id
    )
    
    if data_fim:
        query = query.filter(LancamentoContabil.data_lancamento <= data_fim)
    if data_inicio:
        query = query.filter(LancamentoContabil.data_lancamento >= data_inicio)
    
    saldo = query.scalar() or Decimal('0.0')
    return saldo

def gerar_razao_conta(admin_id, conta_codigo, data_inicio, data_fim):
    """
    Generate analytical ledger for a specific account.
    
    Args:
        admin_id: Admin/tenant ID
        conta_codigo: Account code (ex: '1.1.01.001')
        data_inicio: Start date
        data_fim: End date
        
    Returns:
        dict with account info, movements, and totals
    """
    from models import PlanoContas, PartidaContabil, LancamentoContabil
    from decimal import Decimal
    
    conta = PlanoContas.query.filter_by(
        codigo=conta_codigo,
        admin_id=admin_id
    ).first()
    
    if not conta:
        return None
    
    saldo_anterior = calcular_saldo_conta(conta_codigo, admin_id, None, data_inicio - timedelta(days=1))
    
    partidas = db.session.query(
        PartidaContabil, LancamentoContabil
    ).join(
        LancamentoContabil, 
        PartidaContabil.lancamento_id == LancamentoContabil.id
    ).filter(
        PartidaContabil.conta_codigo == conta_codigo,
        PartidaContabil.admin_id == admin_id,
        LancamentoContabil.data_lancamento >= data_inicio,
        LancamentoContabil.data_lancamento <= data_fim
    ).order_by(
        LancamentoContabil.data_lancamento.asc(),
        LancamentoContabil.numero.asc(),
        PartidaContabil.sequencia.asc()
    ).all()
    
    movimentos = []
    saldo_acumulado = saldo_anterior
    total_debitos = Decimal('0')
    total_creditos = Decimal('0')
    
    for partida, lancamento in partidas:
        valor_decimal = Decimal(str(partida.valor))
        
        if partida.tipo_partida == 'DEBITO':
            saldo_acumulado += valor_decimal
            total_debitos += valor_decimal
        else:
            saldo_acumulado -= valor_decimal
            total_creditos += valor_decimal
        
        movimentos.append({
            'data': lancamento.data_lancamento,
            'lancamento_numero': lancamento.numero,
            'lancamento_id': lancamento.id,
            'historico': lancamento.historico,
            'tipo': partida.tipo_partida,
            'valor': valor_decimal,
            'saldo_acumulado': saldo_acumulado
        })
    
    return {
        'conta': {
            'codigo': conta.codigo,
            'nome': conta.nome,
            'tipo_saldo': 'DEVEDOR' if conta.codigo.startswith(('1', '5', '6')) else 'CREDOR'
        },
        'periodo': {
            'data_inicio': data_inicio,
            'data_fim': data_fim
        },
        'saldo_anterior': saldo_anterior,
        'movimentos': movimentos,
        'totais': {
            'total_debitos': total_debitos,
            'total_creditos': total_creditos,
            'saldo_final': saldo_acumulado
        }
    }

def gerar_balancete_mensal(admin_id, mes_referencia):
    """Gera o balancete mensal para todas as contas."""
    # Obter todas as contas que aceitam lançamento
    contas = PlanoContas.query.filter_by(admin_id=admin_id, aceita_lancamento=True).all()
    
    primeiro_dia = mes_referencia.replace(day=1)
    ultimo_dia = date(mes_referencia.year, mes_referencia.month, 
                     calendar.monthrange(mes_referencia.year, mes_referencia.month)[1])
    
    for conta in contas:
        # Verificar se já existe balancete para esta conta/mês
        balancete_existente = BalanceteMensal.query.filter_by(
            conta_codigo=conta.codigo,
            mes_referencia=primeiro_dia,
            admin_id=admin_id
        ).first()
        
        if balancete_existente:
            continue
            
        # Calcular saldos
        saldo_anterior = calcular_saldo_conta(conta.codigo, admin_id, None, primeiro_dia - timedelta(days=1))
        saldo_atual = calcular_saldo_conta(conta.codigo, admin_id, None, ultimo_dia)
        
        # Calcular movimentação do mês
        debitos_mes = db.session.query(func.sum(PartidaContabil.valor)).join(LancamentoContabil).filter(
            PartidaContabil.conta_codigo == conta.codigo,
            PartidaContabil.admin_id == admin_id,
            PartidaContabil.tipo_partida == 'DEBITO',
            LancamentoContabil.data_lancamento >= primeiro_dia,
            LancamentoContabil.data_lancamento <= ultimo_dia
        ).scalar() or Decimal('0.0')
        
        creditos_mes = db.session.query(func.sum(PartidaContabil.valor)).join(LancamentoContabil).filter(
            PartidaContabil.conta_codigo == conta.codigo,
            PartidaContabil.admin_id == admin_id,
            PartidaContabil.tipo_partida == 'CREDITO',
            LancamentoContabil.data_lancamento >= primeiro_dia,
            LancamentoContabil.data_lancamento <= ultimo_dia
        ).scalar() or Decimal('0.0')
        
        balancete = BalanceteMensal(
            conta_codigo=conta.codigo,
            mes_referencia=primeiro_dia,
            saldo_anterior=saldo_anterior,
            debitos_mes=debitos_mes,
            creditos_mes=creditos_mes,
            saldo_atual=saldo_atual,
            admin_id=admin_id
        )
        db.session.add(balancete)
    
    db.session.commit()

def gerar_balanco_patrimonial(admin_id, data_referencia):
    """
    Gera o Balanço Patrimonial em uma data específica.
    
    Args:
        admin_id: ID do administrador
        data_referencia: Data de referência para o balanço (end of day)
        
    Returns:
        dict: Estrutura completa do balanço patrimonial com ATIVO, PASSIVO e PL
    """
    contas = PlanoContas.query.filter_by(admin_id=admin_id).all()
    
    ativo = {'circulante': {}, 'nao_circulante': {}, 'total': Decimal('0')}
    passivo = {'circulante': {}, 'nao_circulante': {}, 'total': Decimal('0')}
    patrimonio_liquido = {}
    
    for conta in contas:
        saldo = calcular_saldo_conta(conta.codigo, admin_id, data_inicio=None, data_fim=data_referencia)
        
        if saldo == 0:
            continue
        
        if conta.codigo.startswith('1.1'):
            ativo['circulante'][conta.codigo] = {
                'nome': conta.nome,
                'saldo': saldo,
                'nivel': conta.nivel
            }
            ativo['total'] += saldo
        elif conta.codigo.startswith('1.2'):
            ativo['nao_circulante'][conta.codigo] = {
                'nome': conta.nome,
                'saldo': saldo,
                'nivel': conta.nivel
            }
            ativo['total'] += saldo
        elif conta.codigo.startswith('2.1'):
            passivo['circulante'][conta.codigo] = {
                'nome': conta.nome,
                'saldo': abs(saldo),
                'nivel': conta.nivel
            }
            passivo['total'] += abs(saldo)
        elif conta.codigo.startswith('2.2'):
            passivo['nao_circulante'][conta.codigo] = {
                'nome': conta.nome,
                'saldo': abs(saldo),
                'nivel': conta.nivel
            }
            passivo['total'] += abs(saldo)
        elif conta.codigo.startswith('3'):
            patrimonio_liquido[conta.codigo] = {
                'nome': conta.nome,
                'saldo': abs(saldo),
                'nivel': conta.nivel
            }
    
    total_pl = sum(item['saldo'] for item in patrimonio_liquido.values())
    
    total_passivo_pl = passivo['total'] + total_pl
    balanceado = abs(ativo['total'] - total_passivo_pl) < Decimal('0.01')
    
    return {
        'data_referencia': data_referencia,
        'ativo': ativo,
        'passivo': passivo,
        'patrimonio_liquido': patrimonio_liquido,
        'total_pl': total_pl,
        'total_passivo_pl': total_passivo_pl,
        'balanceado': balanceado
    }

def executar_auditoria_automatica(admin_id):
    """Executa verificações automáticas de auditoria."""
    alertas = []
    
    # Verificar se há lançamentos desbalanceados
    lancamentos_desbalanceados = db.session.query(LancamentoContabil).filter_by(admin_id=admin_id).all()
    
    for lancamento in lancamentos_desbalanceados:
        total_debitos = sum(p.valor for p in lancamento.partidas if p.tipo_partida == 'DEBITO')
        total_creditos = sum(p.valor for p in lancamento.partidas if p.tipo_partida == 'CREDITO')
        
        if abs(total_debitos - total_creditos) > 0.01:
            auditoria = AuditoriaContabil(
                tipo_verificacao='LANÇAMENTO_DESBALANCEADO',
                resultado='NAO_CONFORME',
                observacoes=f'Lançamento #{lancamento.numero} desbalanceado: D={total_debitos} C={total_creditos}',
                valor_divergencia=abs(total_debitos - total_creditos),
                admin_id=admin_id
            )
            db.session.add(auditoria)
            alertas.append(auditoria)
    
    db.session.commit()
    return alertas

def calcular_dre_mensal(admin_id: int, ano: int, mes: int):
    """
    Calcula DRE completo baseado na estrutura contábil brasileira
    
    Estrutura DRE (11 seções):
    1. RECEITA BRUTA (4.1.x)
    2. (-) DEDUÇÕES (4.2.x)
    3. = RECEITA LÍQUIDA
    4. (-) CMV/CPV (5.1.03.x)
    5. = LUCRO BRUTO
    6. (-) DESPESAS OPERACIONAIS (5.1.01, 5.1.02, 5.1.04, 5.1.05, 5.1.06)
    7. = EBITDA
    8. (+/-) RESULTADO FINANCEIRO (4.3.01 - 5.2.01)
    9. = RESULTADO ANTES IR/CSLL
    10. (-) PROVISÃO IR E CSLL (5.3.01, 5.3.02)
    11. = LUCRO LÍQUIDO
    
    Args:
        admin_id: ID do admin
        ano: Ano de referência
        mes: Mês de referência (1-12)
        
    Returns:
        dict: Dicionário com todos os valores do DRE
    """
    from sqlalchemy import and_
    from decimal import Decimal
    
    try:
        # Definir período do mês
        data_inicio = date(ano, mes, 1)
        if mes == 12:
            data_fim = date(ano + 1, 1, 1) - timedelta(days=1)
        else:
            data_fim = date(ano, mes + 1, 1) - timedelta(days=1)
        
        # Função auxiliar para calcular saldo de contas específicas
        def calcular_valor_contas(prefixos: list, tipo_esperado: str = None):
            """
            Calcula o total de valores para lista de prefixos de contas
            
            Args:
                prefixos: Lista de prefixos de contas (ex: ['4.1.01', '4.1.02'])
                tipo_esperado: 'CREDITO' para receitas, 'DEBITO' para despesas/custos
            """
            total = Decimal('0')
            
            for prefixo in prefixos:
                partidas = PartidaContabil.query.join(
                    LancamentoContabil,
                    PartidaContabil.lancamento_id == LancamentoContabil.id
                ).filter(
                    and_(
                        PartidaContabil.admin_id == admin_id,
                        PartidaContabil.conta_codigo.like(f'{prefixo}%'),
                        LancamentoContabil.data_lancamento >= data_inicio,
                        LancamentoContabil.data_lancamento <= data_fim
                    )
                ).all()
                
                for partida in partidas:
                    valor = Decimal(str(partida.valor))
                    
                    # Se tipo esperado for especificado, só conta se for o tipo correto
                    if tipo_esperado:
                        if partida.tipo_partida == tipo_esperado:
                            total += valor
                    else:
                        # Senão, credita positivo e debita negativo
                        if partida.tipo_partida == 'CREDITO':
                            total += valor
                        else:
                            total -= valor
            
            return total
        
        # 1. RECEITA BRUTA (contas 4.1.x - CREDITO)
        receita_bruta = calcular_valor_contas(['4.1.01', '4.1.02'], 'CREDITO')
        
        # 2. DEDUÇÕES E ABATIMENTOS (contas 4.2.x - DEBITO)
        deducoes = calcular_valor_contas(['4.2.01', '4.2.02'], 'DEBITO')
        
        # 3. RECEITA LÍQUIDA
        receita_liquida = receita_bruta - deducoes
        
        # 4. CMV/CPV (contas 5.1.03.x - DEBITO)
        cmv = calcular_valor_contas(['5.1.03'], 'DEBITO')
        
        # 5. LUCRO BRUTO
        lucro_bruto = receita_liquida - cmv
        
        # 6. DESPESAS OPERACIONAIS (contas 5.1.x - DEBITO, exceto 5.1.03)
        despesas_pessoal = calcular_valor_contas(['5.1.01'], 'DEBITO')
        despesas_materiais = calcular_valor_contas(['5.1.02'], 'DEBITO')
        despesas_administrativas = calcular_valor_contas(['5.1.04'], 'DEBITO')
        despesas_comerciais = calcular_valor_contas(['5.1.05'], 'DEBITO')
        outras_despesas = calcular_valor_contas(['5.1.06'], 'DEBITO')
        
        total_despesas_operacionais = (
            despesas_pessoal + despesas_materiais + despesas_administrativas + 
            despesas_comerciais + outras_despesas
        )
        
        # 7. EBITDA
        ebitda = lucro_bruto - total_despesas_operacionais
        
        # 8. RESULTADO FINANCEIRO
        receitas_financeiras = calcular_valor_contas(['4.3.01'], 'CREDITO')
        despesas_financeiras = calcular_valor_contas(['5.2.01'], 'DEBITO')
        resultado_financeiro = receitas_financeiras - despesas_financeiras
        
        # 9. RESULTADO ANTES IR/CSLL
        resultado_antes_ir = ebitda + resultado_financeiro
        
        # 10. PROVISÃO IR E CSLL (contas 5.3.x - DEBITO)
        provisao_ir = calcular_valor_contas(['5.3.01'], 'DEBITO')
        provisao_csll = calcular_valor_contas(['5.3.02'], 'DEBITO')
        total_provisoes = provisao_ir + provisao_csll
        
        # 11. LUCRO LÍQUIDO
        lucro_liquido = resultado_antes_ir - total_provisoes
        
        # Calcular percentuais (sobre receita líquida)
        def calcular_percentual(valor, base):
            if base and base != 0:
                return float((valor / base) * 100)
            return 0.0
        
        margem_bruta = calcular_percentual(lucro_bruto, receita_liquida)
        margem_ebitda = calcular_percentual(ebitda, receita_liquida)
        margem_liquida = calcular_percentual(lucro_liquido, receita_liquida)
        
        # Montar estrutura completa do DRE
        dre_data = {
            'mes': mes,
            'ano': ano,
            'mes_referencia': data_inicio,
            
            # Seção 1-3: Receitas
            'receita_bruta': float(receita_bruta),
            'deducoes': float(deducoes),
            'receita_liquida': float(receita_liquida),
            
            # Seção 4-5: CMV e Lucro Bruto
            'cmv': float(cmv),
            'lucro_bruto': float(lucro_bruto),
            
            # Seção 6: Despesas Operacionais (detalhadas)
            'despesas_operacionais': {
                'pessoal': float(despesas_pessoal),
                'materiais': float(despesas_materiais),
                'administrativas': float(despesas_administrativas),
                'comerciais': float(despesas_comerciais),
                'outras': float(outras_despesas),
                'total': float(total_despesas_operacionais)
            },
            
            # Seção 7: EBITDA
            'ebitda': float(ebitda),
            
            # Seção 8: Resultado Financeiro
            'resultado_financeiro': {
                'receitas': float(receitas_financeiras),
                'despesas': float(despesas_financeiras),
                'total': float(resultado_financeiro)
            },
            
            # Seção 9-11: Resultado Final
            'resultado_antes_ir': float(resultado_antes_ir),
            'provisao_ir_csll': {
                'ir': float(provisao_ir),
                'csll': float(provisao_csll),
                'total': float(total_provisoes)
            },
            'lucro_liquido': float(lucro_liquido),
            
            # Percentuais
            'percentuais': {
                'margem_bruta': margem_bruta,
                'margem_ebitda': margem_ebitda,
                'margem_liquida': margem_liquida
            }
        }
        
        print(f"✅ DRE {mes}/{ano} calculada: Receita Líquida={receita_liquida}, Lucro Líquido={lucro_liquido}")
        return dre_data
        
    except Exception as e:
        print(f"❌ Erro ao calcular DRE: {e}")
        import traceback
        traceback.print_exc()
        return None

# ===============================================================
# == FUNÇÕES DE EXPORTAÇÃO PDF E EXCEL
# ===============================================================

def obter_dados_balancete(admin_id, mes, ano):
    """
    Obtém dados do balancete formatados para exportação
    
    Args:
        admin_id: ID do administrador
        mes: Mês (1-12)
        ano: Ano
        
    Returns:
        dict: Dicionário com contas e totais
    """
    from sqlalchemy import func
    
    # Definir período
    primeiro_dia = date(ano, mes, 1)
    ultimo_dia = date(ano, mes, calendar.monthrange(ano, mes)[1])
    fim_mes_anterior = primeiro_dia - timedelta(days=1)
    
    # Buscar TODAS as contas do plano de contas
    todas_contas = PlanoContas.query.filter_by(admin_id=admin_id).order_by(PlanoContas.codigo).all()
    
    contas_data = []
    total_debitos = Decimal('0')
    total_creditos = Decimal('0')
    total_saldo_devedor = Decimal('0')
    total_saldo_credor = Decimal('0')
    
    for conta in todas_contas:
        # Calcular saldo anterior
        partidas_anteriores = db.session.query(PartidaContabil).join(LancamentoContabil).filter(
            PartidaContabil.conta_codigo == conta.codigo,
            PartidaContabil.admin_id == admin_id,
            LancamentoContabil.data_lancamento <= fim_mes_anterior
        ).all()
        
        debito_anterior = sum(p.valor for p in partidas_anteriores if p.tipo_partida == 'DEBITO')
        credito_anterior = sum(p.valor for p in partidas_anteriores if p.tipo_partida == 'CREDITO')
        
        if conta.natureza == 'DEVEDORA':
            saldo_anterior = debito_anterior - credito_anterior
        else:
            saldo_anterior = credito_anterior - debito_anterior
        
        # Calcular movimentação do mês
        partidas_mes = db.session.query(PartidaContabil).join(LancamentoContabil).filter(
            PartidaContabil.conta_codigo == conta.codigo,
            PartidaContabil.admin_id == admin_id,
            LancamentoContabil.data_lancamento >= primeiro_dia,
            LancamentoContabil.data_lancamento <= ultimo_dia
        ).all()
        
        debitos_mes = sum(p.valor for p in partidas_mes if p.tipo_partida == 'DEBITO')
        creditos_mes = sum(p.valor for p in partidas_mes if p.tipo_partida == 'CREDITO')
        
        # Calcular saldo atual
        partidas_ate_hoje = db.session.query(PartidaContabil).join(LancamentoContabil).filter(
            PartidaContabil.conta_codigo == conta.codigo,
            PartidaContabil.admin_id == admin_id,
            LancamentoContabil.data_lancamento <= ultimo_dia
        ).all()
        
        debito_total = sum(p.valor for p in partidas_ate_hoje if p.tipo_partida == 'DEBITO')
        credito_total = sum(p.valor for p in partidas_ate_hoje if p.tipo_partida == 'CREDITO')
        
        if conta.natureza == 'DEVEDORA':
            saldo_atual = debito_total - credito_total
        else:
            saldo_atual = credito_total - debito_total
        
        # Só incluir contas com movimento
        tem_movimento = (abs(saldo_anterior) > Decimal('0.01') or 
                        abs(debitos_mes) > Decimal('0.01') or 
                        abs(creditos_mes) > Decimal('0.01'))
        
        if tem_movimento:
            contas_data.append({
                'codigo': conta.codigo,
                'nome': conta.nome,
                'nivel': conta.nivel,
                'natureza': conta.natureza,
                'debitos': debitos_mes,
                'creditos': creditos_mes,
                'saldo_devedor': saldo_atual if saldo_atual > 0 else Decimal('0'),
                'saldo_credor': abs(saldo_atual) if saldo_atual < 0 else Decimal('0')
            })
            
            total_debitos += debitos_mes
            total_creditos += creditos_mes
            
            if saldo_atual > 0:
                total_saldo_devedor += saldo_atual
            else:
                total_saldo_credor += abs(saldo_atual)
    
    return {
        'contas': contas_data,
        'totais': {
            'total_debitos': total_debitos,
            'total_creditos': total_creditos,
            'total_saldo_devedor': total_saldo_devedor,
            'total_saldo_credor': total_saldo_credor
        }
    }

def gerar_balancete_pdf(admin_id, mes, ano):
    """
    Gera PDF do Balancete de Verificação
    
    Args:
        admin_id: ID do administrador
        mes: Mês (1-12)
        ano: Ano
        
    Returns:
        BytesIO: Buffer com o PDF gerado
    """
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    from io import BytesIO
    from datetime import datetime
    
    try:
        # Obter dados
        dados = obter_dados_balancete(admin_id, mes, ano)
        contas = dados['contas']
        totais = dados['totais']
        
        # Criar buffer
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), 
                               topMargin=1.5*cm, bottomMargin=1.5*cm)
        elements = []
        
        # Estilos
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#1a1a1a'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        # Título
        title = Paragraph("BALANCETE DE VERIFICAÇÃO", title_style)
        elements.append(title)
        
        # Período
        period_style = ParagraphStyle('Period', parent=styles['Normal'], 
                                     fontSize=12, alignment=TA_CENTER)
        period = Paragraph(f"Período: {mes:02d}/{ano}", period_style)
        elements.append(period)
        
        # Data de geração
        gen_date = Paragraph(f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}", 
                            period_style)
        elements.append(gen_date)
        elements.append(Spacer(1, 0.8*cm))
        
        # Montar tabela
        table_data = [['Código', 'Conta', 'Débito', 'Crédito', 'Saldo Devedor', 'Saldo Credor']]
        
        for conta in contas:
            # Indentação baseada no nível
            indent = '  ' * (conta['nivel'] - 1)
            nome_indentado = indent + conta['nome']
            
            table_data.append([
                conta['codigo'],
                nome_indentado,
                f"R$ {float(conta['debitos']):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.') if conta['debitos'] > 0 else '-',
                f"R$ {float(conta['creditos']):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.') if conta['creditos'] > 0 else '-',
                f"R$ {float(conta['saldo_devedor']):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.') if conta['saldo_devedor'] > 0 else '-',
                f"R$ {float(conta['saldo_credor']):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.') if conta['saldo_credor'] > 0 else '-'
            ])
        
        # Linha de totais
        table_data.append([
            '',
            'TOTAIS',
            f"R$ {float(totais['total_debitos']):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'),
            f"R$ {float(totais['total_creditos']):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'),
            f"R$ {float(totais['total_saldo_devedor']):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'),
            f"R$ {float(totais['total_saldo_credor']):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
        ])
        
        # Criar tabela
        table = Table(table_data, colWidths=[3*cm, 8*cm, 4*cm, 4*cm, 4*cm, 4*cm])
        table.setStyle(TableStyle([
            # Cabeçalho
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            
            # Dados
            ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -2), 8),
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),
            ('ALIGN', (1, 1), (1, -1), 'LEFT'),
            ('ALIGN', (2, 1), (-1, -1), 'RIGHT'),
            
            # Linha de totais
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#ecf0f1')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -1), (-1, -1), 9),
            
            # Bordas
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BOX', (0, 0), (-1, -1), 1, colors.black),
        ]))
        
        elements.append(table)
        
        # Construir PDF
        doc.build(elements)
        
        buffer.seek(0)
        return buffer
        
    except Exception as e:
        print(f"❌ Erro ao gerar PDF do balancete: {e}")
        import traceback
        traceback.print_exc()
        raise

def gerar_balancete_excel(admin_id, mes, ano):
    """
    Gera arquivo Excel do Balancete de Verificação
    
    Args:
        admin_id: ID do administrador
        mes: Mês (1-12)
        ano: Ano
        
    Returns:
        BytesIO: Buffer com o arquivo Excel
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from io import BytesIO
    from datetime import datetime
    
    try:
        # Obter dados
        dados = obter_dados_balancete(admin_id, mes, ano)
        contas = dados['contas']
        totais = dados['totais']
        
        # Criar workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Balancete"
        
        # Cabeçalho
        ws['A1'] = "BALANCETE DE VERIFICAÇÃO"
        ws['A1'].font = Font(size=16, bold=True, color="1a1a1a")
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:F1')
        
        ws['A2'] = f"Período: {mes:02d}/{ano}"
        ws['A2'].font = Font(size=12)
        ws['A2'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A2:F2')
        
        ws['A3'] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws['A3'].font = Font(size=10, italic=True)
        ws['A3'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A3:F3')
        
        # Cabeçalhos das colunas (linha 5)
        headers = ['Código', 'Conta', 'Débito', 'Crédito', 'Saldo Devedor', 'Saldo Credor']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Dados
        row = 6
        for conta in contas:
            # Indentação no nome
            indent = '  ' * (conta['nivel'] - 1)
            
            ws.cell(row=row, column=1, value=conta['codigo'])
            ws.cell(row=row, column=2, value=indent + conta['nome'])
            ws.cell(row=row, column=3, value=float(conta['debitos']) if conta['debitos'] > 0 else None)
            ws.cell(row=row, column=4, value=float(conta['creditos']) if conta['creditos'] > 0 else None)
            ws.cell(row=row, column=5, value=float(conta['saldo_devedor']) if conta['saldo_devedor'] > 0 else None)
            ws.cell(row=row, column=6, value=float(conta['saldo_credor']) if conta['saldo_credor'] > 0 else None)
            
            # Bordas
            for col in range(1, 7):
                ws.cell(row=row, column=col).border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            
            row += 1
        
        # Linha de totais
        ws.cell(row=row, column=2, value='TOTAIS').font = Font(bold=True)
        ws.cell(row=row, column=2).fill = PatternFill(start_color="ecf0f1", end_color="ecf0f1", fill_type="solid")
        
        # Fórmulas de soma
        ws.cell(row=row, column=3, value=f'=SUM(C6:C{row-1})')
        ws.cell(row=row, column=4, value=f'=SUM(D6:D{row-1})')
        ws.cell(row=row, column=5, value=f'=SUM(E6:E{row-1})')
        ws.cell(row=row, column=6, value=f'=SUM(F6:F{row-1})')
        
        for col in range(2, 7):
            ws.cell(row=row, column=col).font = Font(bold=True)
            ws.cell(row=row, column=col).fill = PatternFill(start_color="ecf0f1", end_color="ecf0f1", fill_type="solid")
            ws.cell(row=row, column=col).border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Formatação de moeda
        for r in range(6, row+1):
            for c in [3, 4, 5, 6]:
                ws.cell(row=r, column=c).number_format = '"R$" #,##0.00'
                ws.cell(row=r, column=c).alignment = Alignment(horizontal='right')
        
        # Larguras das colunas
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 45
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 18
        
        # Congelar painéis
        ws.freeze_panes = 'A6'
        
        # Salvar em buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
        
    except Exception as e:
        print(f"❌ Erro ao gerar Excel do balancete: {e}")
        import traceback
        traceback.print_exc()
        raise

def gerar_dre_pdf(admin_id, mes, ano):
    """
    Gera PDF da DRE
    
    Args:
        admin_id: ID do administrador
        mes: Mês (1-12)
        ano: Ano
        
    Returns:
        BytesIO: Buffer com o PDF gerado
    """
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
    from io import BytesIO
    from datetime import datetime
    
    try:
        # Obter dados DRE
        dre_data = calcular_dre_mensal(admin_id, ano, mes)
        
        if not dre_data:
            raise ValueError("Não foi possível calcular DRE")
        
        # Criar buffer
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, 
                               topMargin=2*cm, bottomMargin=2*cm,
                               leftMargin=2*cm, rightMargin=2*cm)
        elements = []
        
        # Estilos
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#1a1a1a'),
            spaceAfter=20,
            alignment=TA_CENTER
        )
        
        # Título
        title = Paragraph("DEMONSTRAÇÃO DO RESULTADO DO EXERCÍCIO (DRE)", title_style)
        elements.append(title)
        
        # Período
        period_style = ParagraphStyle('Period', parent=styles['Normal'], 
                                     fontSize=12, alignment=TA_CENTER)
        period = Paragraph(f"Período: {mes:02d}/{ano}", period_style)
        elements.append(period)
        
        gen_date = Paragraph(f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}", 
                            period_style)
        elements.append(gen_date)
        elements.append(Spacer(1, 1*cm))
        
        # Função para formatar valor
        def fmt(valor):
            return f"R$ {valor:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
        
        # Montar tabela DRE
        table_data = []
        
        # Receita Bruta
        table_data.append(['RECEITA BRUTA', fmt(dre_data['receita_bruta'])])
        
        # Deduções
        if dre_data['deducoes'] > 0:
            table_data.append(['(-) Deduções e Abatimentos', f"({fmt(dre_data['deducoes'])})"])
        
        # Receita Líquida
        table_data.append(['= RECEITA LÍQUIDA', fmt(dre_data['receita_liquida'])])
        table_data.append(['', ''])
        
        # CMV
        if dre_data['cmv'] > 0:
            table_data.append(['(-) Custo dos Serviços Prestados (CMV)', f"({fmt(dre_data['cmv'])})"])
        
        # Lucro Bruto
        table_data.append(['= LUCRO BRUTO', fmt(dre_data['lucro_bruto'])])
        table_data.append(['', ''])
        
        # Despesas Operacionais
        desp_op = dre_data['despesas_operacionais']
        table_data.append(['(-) DESPESAS OPERACIONAIS', ''])
        if desp_op['pessoal'] > 0:
            table_data.append(['    Despesas com Pessoal', f"({fmt(desp_op['pessoal'])})"])
        if desp_op['materiais'] > 0:
            table_data.append(['    Despesas com Materiais', f"({fmt(desp_op['materiais'])})"])
        if desp_op['administrativas'] > 0:
            table_data.append(['    Despesas Administrativas', f"({fmt(desp_op['administrativas'])})"])
        if desp_op['comerciais'] > 0:
            table_data.append(['    Despesas Comerciais', f"({fmt(desp_op['comerciais'])})"])
        if desp_op['outras'] > 0:
            table_data.append(['    Outras Despesas', f"({fmt(desp_op['outras'])})"])
        
        # EBITDA
        table_data.append(['= EBITDA', fmt(dre_data['ebitda'])])
        table_data.append(['', ''])
        
        # Resultado Financeiro
        res_fin = dre_data['resultado_financeiro']
        if res_fin['receitas'] > 0:
            table_data.append(['(+) Receitas Financeiras', fmt(res_fin['receitas'])])
        if res_fin['despesas'] > 0:
            table_data.append(['(-) Despesas Financeiras', f"({fmt(res_fin['despesas'])})"])
        
        # Resultado Antes IR
        table_data.append(['= RESULTADO ANTES IR/CSLL', fmt(dre_data['resultado_antes_ir'])])
        table_data.append(['', ''])
        
        # Provisões
        prov = dre_data['provisao_ir_csll']
        if prov['total'] > 0:
            table_data.append(['(-) Provisão IR e CSLL', f"({fmt(prov['total'])})"])
        
        # Lucro Líquido
        table_data.append(['= LUCRO LÍQUIDO DO PERÍODO', fmt(dre_data['lucro_liquido'])])
        
        # Criar tabela
        table = Table(table_data, colWidths=[12*cm, 5*cm])
        table.setStyle(TableStyle([
            # Alinhamento
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            
            # Fontes
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            
            # Totalizadores em negrito
            ('FONTNAME', (0, 2), (-1, 2), 'Helvetica-Bold'),  # Receita Líquida
            ('FONTNAME', (0, 5), (-1, 5), 'Helvetica-Bold'),  # Lucro Bruto
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),  # Lucro Líquido
            
            # Linhas
            ('LINEABOVE', (0, 2), (-1, 2), 1, colors.black),
            ('LINEABOVE', (0, 5), (-1, 5), 1, colors.black),
            ('LINEABOVE', (0, -1), (-1, -1), 2, colors.black),
            ('LINEBELOW', (0, -1), (-1, -1), 2, colors.black),
            
            # Padding
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        
        elements.append(table)
        
        # Indicadores
        elements.append(Spacer(1, 1*cm))
        perc = dre_data['percentuais']
        indicadores_data = [
            ['INDICADORES', ''],
            ['Margem Bruta', f"{perc['margem_bruta']:.2f}%"],
            ['Margem EBITDA', f"{perc['margem_ebitda']:.2f}%"],
            ['Margem Líquida', f"{perc['margem_liquida']:.2f}%"]
        ]
        
        ind_table = Table(indicadores_data, colWidths=[12*cm, 5*cm])
        ind_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#ecf0f1')),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        
        elements.append(ind_table)
        
        # Construir PDF
        doc.build(elements)
        
        buffer.seek(0)
        return buffer
        
    except Exception as e:
        print(f"❌ Erro ao gerar PDF da DRE: {e}")
        import traceback
        traceback.print_exc()
        raise

def gerar_dre_excel(admin_id, mes, ano):
    """
    Gera arquivo Excel da DRE
    
    Args:
        admin_id: ID do administrador
        mes: Mês (1-12)
        ano: Ano
        
    Returns:
        BytesIO: Buffer com o arquivo Excel
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from io import BytesIO
    from datetime import datetime
    
    try:
        # Obter dados DRE
        dre_data = calcular_dre_mensal(admin_id, ano, mes)
        
        if not dre_data:
            raise ValueError("Não foi possível calcular DRE")
        
        # Criar workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "DRE"
        
        # Cabeçalho
        ws['A1'] = "DEMONSTRAÇÃO DO RESULTADO DO EXERCÍCIO (DRE)"
        ws['A1'].font = Font(size=16, bold=True, color="1a1a1a")
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:C1')
        
        ws['A2'] = f"Período: {mes:02d}/{ano}"
        ws['A2'].font = Font(size=12)
        ws['A2'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A2:C2')
        
        ws['A3'] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws['A3'].font = Font(size=10, italic=True)
        ws['A3'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A3:C3')
        
        # Cabeçalhos das colunas
        row = 5
        ws.cell(row=row, column=1, value='DESCRIÇÃO').font = Font(bold=True)
        ws.cell(row=row, column=2, value='VALOR (R$)').font = Font(bold=True)
        ws.cell(row=row, column=3, value='% RECEITA').font = Font(bold=True)
        
        for col in range(1, 4):
            ws.cell(row=row, column=col).fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
            ws.cell(row=row, column=col).font = Font(bold=True, color="FFFFFF")
            ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')
        
        row += 1
        
        # Helper para adicionar linha
        def add_row(descricao, valor, formula_perc=None, bold=False, indent=0, bg_color=None):
            nonlocal row
            
            # Descrição com indentação
            ws.cell(row=row, column=1, value='  ' * indent + descricao)
            
            # Valor
            if valor is not None:
                ws.cell(row=row, column=2, value=valor)
                ws.cell(row=row, column=2).number_format = '"R$" #,##0.00'
            
            # Percentual
            if formula_perc:
                ws.cell(row=row, column=3, value=formula_perc)
                ws.cell(row=row, column=3).number_format = '0.00"%"'
            
            # Formatação
            if bold:
                for col in range(1, 4):
                    ws.cell(row=row, column=col).font = Font(bold=True)
            
            if bg_color:
                for col in range(1, 3):
                    ws.cell(row=row, column=col).fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
            
            for col in range(1, 4):
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='left' if col == 1 else 'right')
            
            row += 1
        
        # Receita Bruta
        add_row('RECEITA BRUTA', dre_data['receita_bruta'], None, bold=True, bg_color="d5f4e6")
        
        # Deduções
        if dre_data['deducoes'] > 0:
            add_row('(-) Deduções e Abatimentos', -dre_data['deducoes'], None)
        
        # Receita Líquida
        receita_liq_row = row
        add_row('= RECEITA LÍQUIDA', dre_data['receita_liquida'], 100.0, bold=True, bg_color="ecf0f1")
        row += 1
        
        # CMV
        if dre_data['cmv'] > 0:
            add_row('(-) Custo dos Serviços Prestados', -dre_data['cmv'], 
                   f'=B{row}/B${receita_liq_row}*100')
        
        # Lucro Bruto
        add_row('= LUCRO BRUTO', dre_data['lucro_bruto'], 
               f'=B{row}/B${receita_liq_row}*100', bold=True, bg_color="ecf0f1")
        row += 1
        
        # Despesas Operacionais
        desp_op = dre_data['despesas_operacionais']
        add_row('(-) DESPESAS OPERACIONAIS', None, None, bold=True)
        
        if desp_op['pessoal'] > 0:
            add_row('Despesas com Pessoal', -desp_op['pessoal'], 
                   f'=B{row}/B${receita_liq_row}*100', indent=1)
        if desp_op['materiais'] > 0:
            add_row('Despesas com Materiais', -desp_op['materiais'], 
                   f'=B{row}/B${receita_liq_row}*100', indent=1)
        if desp_op['administrativas'] > 0:
            add_row('Despesas Administrativas', -desp_op['administrativas'], 
                   f'=B{row}/B${receita_liq_row}*100', indent=1)
        if desp_op['comerciais'] > 0:
            add_row('Despesas Comerciais', -desp_op['comerciais'], 
                   f'=B{row}/B${receita_liq_row}*100', indent=1)
        if desp_op['outras'] > 0:
            add_row('Outras Despesas', -desp_op['outras'], 
                   f'=B{row}/B${receita_liq_row}*100', indent=1)
        
        # EBITDA
        add_row('= EBITDA', dre_data['ebitda'], 
               f'=B{row}/B${receita_liq_row}*100', bold=True, bg_color="ecf0f1")
        row += 1
        
        # Resultado Financeiro
        res_fin = dre_data['resultado_financeiro']
        if res_fin['receitas'] > 0:
            add_row('(+) Receitas Financeiras', res_fin['receitas'], 
                   f'=B{row}/B${receita_liq_row}*100')
        if res_fin['despesas'] > 0:
            add_row('(-) Despesas Financeiras', -res_fin['despesas'], 
                   f'=B{row}/B${receita_liq_row}*100')
        
        # Resultado Antes IR
        add_row('= RESULTADO ANTES IR/CSLL', dre_data['resultado_antes_ir'], 
               f'=B{row}/B${receita_liq_row}*100', bold=True, bg_color="ecf0f1")
        row += 1
        
        # Provisões
        prov = dre_data['provisao_ir_csll']
        if prov['total'] > 0:
            add_row('(-) Provisão IR e CSLL', -prov['total'], 
                   f'=B{row}/B${receita_liq_row}*100')
        
        # Lucro Líquido
        add_row('= LUCRO LÍQUIDO DO PERÍODO', dre_data['lucro_liquido'], 
               f'=B{row}/B${receita_liq_row}*100', bold=True, bg_color="ffeaa7")
        
        # Adicionar indicadores
        row += 2
        ws.cell(row=row, column=1, value='INDICADORES').font = Font(bold=True, size=12)
        ws.merge_cells(f'A{row}:C{row}')
        row += 1
        
        perc = dre_data['percentuais']
        add_row('Margem Bruta', None, perc['margem_bruta'], bold=True)
        add_row('Margem EBITDA', None, perc['margem_ebitda'], bold=True)
        add_row('Margem Líquida', None, perc['margem_liquida'], bold=True)
        
        # Larguras das colunas
        ws.column_dimensions['A'].width = 45
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        
        # Congelar painéis
        ws.freeze_panes = 'A6'
        
        # Salvar em buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
        
    except Exception as e:
        print(f"❌ Erro ao gerar Excel da DRE: {e}")
        import traceback
        traceback.print_exc()
        raise