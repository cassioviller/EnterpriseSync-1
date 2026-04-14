"""
Serviço de Importação Excel para o SIGE v9.0
5 módulos: Funcionários, Diárias, Alimentação, Transporte, Custos

Fluxo por módulo:
  1. processar(ws, admin_id, defaults={}) → (validos, erros)
     - Lê planilha, valida campos, retorna preview sem salvar no DB
  2. importar(rows, admin_id) → {'criados': N, 'erros': [...]}
     - Recebe lista de dicts com dados já validados, salva no DB
"""
import logging
import re
from datetime import datetime, date

logger = logging.getLogger(__name__)

# ── Helpers compartilhados ─────────────────────────────────────────────────────

def _norm(v):
    if v is None:
        return ''
    return str(v).strip()

def _parse_data(v):
    if isinstance(v, (date, datetime)):
        return v.date() if isinstance(v, datetime) else v
    s = _norm(v)
    for fmt in ('%d/%m/%Y', '%d/%m/%y', '%Y-%m-%d', '%d-%m-%Y'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None

def _parse_float(v, default=0.0):
    if isinstance(v, (int, float)):
        return float(v)
    s = _norm(v).replace('R$', '').replace(' ', '')
    if not s:
        return default
    if ',' in s:
        s = s.replace('.', '').replace(',', '.')
    try:
        return float(s)
    except (ValueError, TypeError):
        return default

def _limpar_cpf(v):
    return re.sub(r'\D', '', _norm(v))[:11]

def _mapear_headers(row_vals):
    """Retorna {header_normalizado: col_index (0-based)}."""
    import unicodedata
    def n(s):
        s = _norm(s).lower()
        s = unicodedata.normalize('NFD', s)
        s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
        s = s.replace(' ', '_').replace('*', '').replace('-', '_').strip('_')
        return s
    return {n(str(cell)): idx for idx, cell in enumerate(row_vals)}

def _detectar_header_row(ws, must_contain):
    """Detecta qual linha tem o cabeçalho contendo pelo menos um dos termos."""
    import unicodedata
    def n(s):
        s = _norm(s).lower()
        s = unicodedata.normalize('NFD', s)
        s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
        s = s.replace(' ', '_').replace('*', '').strip('_')
        return s
    for r in range(1, min(6, ws.max_row + 1)):
        vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        normalized = [n(str(v)) for v in vals]
        if any(term in normalized for term in must_contain):
            return r, [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
    return None, None

def _cel(ws, row_num, hm, *keys):
    """Retorna o valor da primeira chave encontrada no header_map."""
    for k in keys:
        if k in hm:
            return ws.cell(row=row_num, column=hm[k] + 1).value
    return None


# ── MÓDULO 1: Funcionários ─────────────────────────────────────────────────────

class ImportacaoFuncionarios:
    """
    Suporta dois formatos:
    - SIGE: cabeçalho com headers minúsculos: nome, cpf, tipo_remuneracao, data_admissao, ...
    - REGISTRO_COLABORADORES: cabeçalho com headers maiúsculos: NOME, RG, CPF, DATA NASC, ...

    Detecção: raw header[0] == 'nome' (lowercase) → SIGE
              raw header[0] == 'NOME' (uppercase) → REGISTRO_COLABORADORES
    """

    def _detectar_formato(self, raw_headers):
        """
        Detecta formato pelo primeiro header não-vazio, após remover sufixos de marcação (*, !).

        Regra:
          • Primeiro char minúsculo → SIGE (ex: 'nome', 'nome *', 'nome*')
          • Primeiro char maiúsculo → REGISTRO_COLABORADORES (ex: 'NOME', 'NOME *')
          • Não inicia com 'nome'/variante → 'desconhecido'
        """
        import unicodedata

        def sem_acento(s):
            s = unicodedata.normalize('NFD', s)
            return ''.join(c for c in s if unicodedata.category(c) != 'Mn')

        if not raw_headers:
            return 'desconhecido'

        # Pega o primeiro header real (ignora None/vazio)
        primeiro_raw = None
        for h in raw_headers:
            v = str(h).strip() if h is not None else ''
            if v:
                primeiro_raw = v
                break
        if primeiro_raw is None:
            return 'desconhecido'

        # Remove sufixos de marcação (*, !, espaços) para comparar a raiz
        primeiro_clean = sem_acento(
            primeiro_raw.replace('*', '').replace('!', '').strip().lower()
        )
        if primeiro_clean != 'nome':
            return 'desconhecido'

        # Diferencia SIGE (minúsculo) de REGISTRO_COLABORADORES (maiúsculo)
        # usa o primeiro caractere alfabético do header original
        primeiro_alfa = next((c for c in primeiro_raw if c.isalpha()), '')
        if primeiro_alfa.islower():
            return 'sige'
        return 'colaboradores'

    def processar(self, ws, admin_id, defaults=None):
        """
        Retorna (validos, erros).
        validos = list[dict] prontos para importar
        erros   = list[dict] com 'linha' e 'motivo'
        """
        from models import Funcionario
        defaults = defaults or {}

        header_row, raw_headers = _detectar_header_row(ws, ['nome', 'cpf'])
        if not header_row:
            return [], [{'linha': '?', 'nome': '—', 'motivo': 'Cabeçalho não encontrado'}]

        hm = _mapear_headers(raw_headers)
        formato = self._detectar_formato(raw_headers)
        if formato == 'desconhecido':
            return [], [{'linha': 3, 'nome': '—',
                         'motivo': 'Formato não reconhecido. Use o template SIGE ou o modelo "Registro de Colaboradores".'}]

        validos, erros = [], []
        cpfs_vistos = set()

        for rn in range(header_row + 1, ws.max_row + 1):
            def c(*keys):
                return _cel(ws, rn, hm, *keys)

            nome = _norm(c('nome'))
            if not nome or nome.upper() in ('NOME', '-', '#N/A', 'NONE'):
                continue

            cpf_raw = c('cpf')
            if not cpf_raw:
                erros.append({'linha': rn, 'nome': nome, 'motivo': 'CPF não informado'})
                continue
            cpf = _limpar_cpf(cpf_raw)
            if len(cpf) < 11:
                erros.append({'linha': rn, 'nome': nome, 'motivo': f'CPF inválido: {cpf_raw}'})
                continue
            if cpf in cpfs_vistos:
                erros.append({'linha': rn, 'nome': nome, 'motivo': f'CPF duplicado na planilha: {cpf}'})
                continue
            cpfs_vistos.add(cpf)

            # CPF já cadastrado neste tenant → rejeitado (importação cria novos, não atualiza)
            existe = Funcionario.query.filter_by(cpf=cpf, admin_id=admin_id).first()
            if existe:
                erros.append({'linha': rn, 'nome': nome,
                              'motivo': f'CPF {cpf} já cadastrado (funcionário: {existe.nome}). '
                                        f'Use a edição individual para atualizar.'})
                continue

            if formato == 'colaboradores':
                # Tenta ler da planilha; se ausente, usa defaults do formulário
                rem_raw = _norm(c('remuneracao', 'tipo_remuneracao')).upper()
                if rem_raw:
                    tipo_rem = 'diaria' if ('DIARIA' in rem_raw or 'DIÁRIA' in rem_raw) else 'salario'
                else:
                    tipo_rem = _norm(defaults.get('tipo_remuneracao') or 'salario').lower()
                    tipo_rem = 'diaria' if 'diaria' in tipo_rem else 'salario'

                valor = _parse_float(c('valor', 'salario', 'valor_diaria'))
                if valor == 0:
                    valor = _parse_float(defaults.get('valor', '0'))
                # Para Registro de Colaboradores, o valor padrão é obrigatório
                if valor == 0:
                    erros.append({'linha': rn, 'nome': nome,
                                  'motivo': 'Valor não informado. Defina o valor padrão '
                                            'de salário/diária no formulário de importação.'})
                    continue

                data_admissao = _parse_data(c('data_admissao', 'admissao')) or (
                    _parse_data(defaults.get('data_admissao')) or date.today()
                )
            else:
                rem_raw = _norm(c('tipo_remuneracao')).lower()
                tipo_rem = 'diaria' if 'diaria' in rem_raw or 'diária' in rem_raw else 'salario'
                valor = _parse_float(c('valor', 'valor_diaria', 'salario'))
                data_admissao = _parse_data(c('data_admissao', 'admissao')) or date.today()

            status_raw = _norm(c('status')).upper()
            ativo = status_raw != 'INATIVO'

            # Resolução de funcao_id por nome
            funcao_raw = _norm(c('funcao', 'cargo', 'funcao_id'))
            funcao_id = None
            funcao_nome = None
            if funcao_raw:
                from models import Funcao
                funcao_obj = Funcao.query.filter(
                    Funcao.admin_id == admin_id,
                    Funcao.nome.ilike(funcao_raw)
                ).first()
                if funcao_obj:
                    funcao_id = funcao_obj.id
                    funcao_nome = funcao_obj.nome
                else:
                    funcao_nome = f'{funcao_raw} (não encontrada)'

            row = {
                'linha': rn,
                'nome': nome,
                'cpf': cpf,
                'email': _norm(c('email')) or None,
                'rg': _norm(c('rg')) or None,
                'telefone': _norm(c('telefone', 'tel')) or None,
                'endereco': _norm(c('endereco')) or None,
                'chave_pix': _norm(c('chave_pix', 'pix')) or None,
                'data_nascimento': _parse_data(c('data_nascimento', 'data_nasc')),
                'ativo': ativo,
                'tipo_remuneracao': tipo_rem,
                'valor': valor,
                'data_admissao': str(data_admissao),
                'valor_va': _parse_float(c('valor_va', 'va')),
                'valor_vt': _parse_float(c('valor_vt', 'vt')),
                'funcao_id': funcao_id,
                'funcao_nome': funcao_nome,
                'operacao': 'criar',  # Importação cria novos; CPFs existentes são rejeitados em preview
            }
            validos.append(row)

        return validos, erros

    def importar(self, rows, admin_id):
        """
        Apenas CRIA novos funcionários. CPFs já existentes devem ter sido rejeitados em processar().
        Se por alguma razão um CPF duplicado chegar aqui, é registrado como erro sem alteração.
        """
        from models import db, Funcionario
        criados, erros = 0, []

        # Busca o max código VV do tenant uma única vez, antes do loop.
        # O filtro admin_id está correto: cada tenant tem sua própria sequência independente.
        # Re-consultar dentro do loop não funcionaria porque os INSERTs estão em savepoint
        # (não commitados para a sessão principal) e a query retornaria sempre o mesmo valor.
        ultimo_tenant = (Funcionario.query
                         .filter(Funcionario.codigo.like('VV%'), Funcionario.admin_id == admin_id)
                         .order_by(Funcionario.codigo.desc()).first())
        try:
            proximo_num = int(ultimo_tenant.codigo[2:]) + 1 if ultimo_tenant else 1
        except Exception:
            proximo_num = 1

        for row in rows:
            sp = db.session.begin_nested()
            try:
                cpf = row['cpf']
                data_adm = _parse_data(row.get('data_admissao')) or date.today()

                # Dupla verificação: rejeita se CPF já existe (não atualiza)
                if Funcionario.query.filter_by(cpf=cpf, admin_id=admin_id).first():
                    erros.append({'linha': row.get('linha'), 'nome': row.get('nome'),
                                  'motivo': f'CPF {cpf} já cadastrado — ignorado na importação'})
                    continue

                codigo = f"VV{proximo_num:03d}"
                proximo_num += 1  # incrementa o contador local para o próximo do lote

                v = _parse_float(row.get('valor', 0))
                kwargs = dict(
                    nome=row['nome'], cpf=cpf, rg=row.get('rg'),
                    email=row.get('email'),
                    telefone=row.get('telefone'), endereco=row.get('endereco'),
                    chave_pix=row.get('chave_pix'),
                    data_nascimento=_parse_data(row.get('data_nascimento')),
                    ativo=row.get('ativo', True), admin_id=admin_id, codigo=codigo,
                    tipo_remuneracao=row['tipo_remuneracao'], data_admissao=data_adm,
                    valor_va=_parse_float(row.get('valor_va', 0)),
                    valor_vt=_parse_float(row.get('valor_vt', 0)),
                )
                if row.get('funcao_id'):
                    kwargs['funcao_id'] = row['funcao_id']
                if row['tipo_remuneracao'] == 'diaria':
                    kwargs['valor_diaria'] = v
                    kwargs['salario'] = 0
                else:
                    kwargs['salario'] = v
                    kwargs['valor_diaria'] = 0

                db.session.add(Funcionario(**kwargs))
                sp.commit()
                criados += 1

            except Exception as e:
                sp.rollback()
                erros.append({'linha': row.get('linha'), 'nome': row.get('nome'), 'motivo': str(e)})

        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            erros.append({'linha': 'COMMIT', 'motivo': str(e)})

        return {'criados': criados, 'atualizados': 0, 'erros': erros}


# ── MÓDULO 2: Diárias ──────────────────────────────────────────────────────────


# ── Tipos especiais de lançamento no campo "obra / centro de custo" ────────────
# A palavra-chave é detectada no campo obra da planilha.
# Qualquer valor não reconhecido como tipo especial é tratado como nome de obra.
_TIPOS_ESPECIAIS_DIARIA = {
    'FERIADO':    'sem_lancamento',   # registro apenas, sem custo
    'FALTOU':     'sem_lancamento',
    'DESCANSO':   'sem_lancamento',
    'FOLGA':      'sem_lancamento',
    'FALTA':      'sem_lancamento',    # falta → só registro, sem lançamento
    'ATESTADO':   'somente_diaria',   # atestado → lança só diária, sem VA e VT
}

_TIPO_LABELS = {
    'sem_lancamento': '🚫 Sem lançamento',
    'diaria_completa': '✅ Diária + VA + VT',
    'somente_diaria': '📋 Só diária (atestado)',
}

def _detectar_tipo_lancamento(obra_raw):
    """Retorna tipo especial ou None quando for obra normal."""
    if not obra_raw:
        return None
    return _TIPOS_ESPECIAIS_DIARIA.get(obra_raw.strip().upper())


class ImportacaoDiarias:
    """
    Suporta dois formatos de planilha:

    FORMATO SIGE (colunas: nome_funcionario, data, obra, status)
    - Valor da diária, VA e VT vêm do cadastro do funcionário

    FORMATO COLABORADORES (detectado pela palavra COLABORADOR na linha 1)
    - Nome do funcionário em: linha 1, col C (ou col A com rótulo COLABORADOR na mesma linha)
    - Cabeçalho: DATA | DIA DA SEMANA | CENTRO DE CUSTO | $ DIÁRIA | $ VT | $ VA
    - Valores lidos diretamente da planilha por linha
    - Múltiplos funcionários por arquivo (cada seção começa com COLABORADOR)

    Regras de lançamento por tipo de centro de custo:
    - Obra normal               → diária + VA + VT
    - FALTA                     → diária + VA + VT (sem desconto)
    - FERIADO / FALTOU / DESCANSO / FOLGA → registro apenas (sem lançamento financeiro)
    - ATESTADO                  → somente diária

    Se a obra não existir → é criada automaticamente com o nome (para editar depois).
    Se o funcionário não existir → é criado com o nome e os valores da planilha (CPF temporário).
    """

    # ── Detecção de formato ────────────────────────────────────────────────────
    def _is_formato_colaboradores(self, ws):
        """Retorna True se a planilha usa o formato multi-colaborador."""
        cell_a1 = _norm(ws.cell(1, 1).value).upper()
        return 'COLABORADOR' in cell_a1

    # ── Busca ou prepara funcionário (usado nos dois formatos) ─────────────────
    def _resolver_funcionario(self, nome, admin_id, valor_diaria=0, valor_va=0, valor_vt=0):
        """
        Busca funcionário por nome. Se não encontrado, prepara dict de criação.
        Retorna (func_id_ou_None, nome_normalizado, func_criar, aviso)
        """
        from models import Funcionario
        func = (Funcionario.query
                .filter(Funcionario.nome.ilike(nome), Funcionario.admin_id == admin_id)
                .first())
        if func:
            return func.id, func.nome, False, None
        # Não encontrado → será criado no importar()
        aviso = f'Funcionário "{nome}" não encontrado — será criado automaticamente'
        return None, nome.title(), True, aviso

    # ── Busca ou prepara obra (usado nos dois formatos) ────────────────────────
    def _resolver_obra(self, obra_raw, admin_id):
        """
        Busca obra por nome/código. Se não encontrada, prepara criação automática.
        Retorna (obra_id_ou_None, obra_nome, obra_criar)
        """
        from models import Obra
        if not obra_raw:
            return None, '(sem obra)', False
        obra = (Obra.query
                .filter(Obra.admin_id == admin_id)
                .filter((Obra.nome.ilike(obra_raw)) | (Obra.codigo == obra_raw))
                .first())
        if obra:
            return obra.id, obra.nome, False
        # Não encontrada → será criada automaticamente
        return None, obra_raw.title(), True

    # ── FORMATO COLABORADORES ──────────────────────────────────────────────────
    def _processar_colaboradores(self, ws, admin_id):
        """
        Formato: linha com COLABORADOR define o funcionário atual.
        A seguir, cabeçalho DATA | DIA DA SEMANA | CENTRO DE CUSTO | $ DIÁRIA | $ VT | $ VA
        Depois, linhas de dados.
        """
        validos, erros, avisos = [], [], []
        funcionario_nome = None
        em_dados = False
        col_data = col_cc = col_diaria = col_vt = col_va = None

        for rn in range(1, ws.max_row + 1):
            row = [ws.cell(rn, c).value for c in range(1, ws.max_column + 1)]

            # ── Detecta linha COLABORADOR ──────────────────────────────────────
            primeira = _norm(row[0]).upper()
            if 'COLABORADOR' in primeira:
                # Nome pode estar na coluna A, B ou C dependendo do layout
                for ci in range(len(row) - 1, -1, -1):
                    v = _norm(row[ci])
                    if v and v.upper() not in ('COLABORADOR',):
                        funcionario_nome = v.strip()
                        break
                em_dados = False
                col_data = col_cc = col_diaria = col_vt = col_va = None
                continue

            # ── Detecta linha de cabeçalho ──────────────────────────────────
            normalized_row = [_norm(v).upper() for v in row]
            if 'DATA' in normalized_row and ('DIÁRIA' in ' '.join(normalized_row) or 'DIARIA' in ' '.join(normalized_row)):
                col_data = next((i for i, v in enumerate(normalized_row) if v == 'DATA'), None)
                col_cc = next((i for i, v in enumerate(normalized_row)
                               if 'CENTRO' in v or 'CUSTO' in v or 'OBRA' in v), None)
                col_diaria = next((i for i, v in enumerate(normalized_row)
                                   if 'DIÁRIA' in v or 'DIARIA' in v), None)
                col_vt = next((i for i, v in enumerate(normalized_row) if 'VT' in v), None)
                col_va = next((i for i, v in enumerate(normalized_row) if 'VA' in v), None)
                em_dados = True
                continue

            # ── Processa linha de dados ────────────────────────────────────
            if not em_dados or funcionario_nome is None:
                continue

            data_ref = _parse_data(row[col_data] if col_data is not None else None)
            if not data_ref:
                continue  # linha sem data válida → pula silenciosamente (DOMINGO vazio etc.)

            obra_raw = _norm(row[col_cc] if col_cc is not None else None)
            valor_d = _parse_float(row[col_diaria] if col_diaria is not None else 0)
            valor_vt_ = _parse_float(row[col_vt] if col_vt is not None else 0)
            valor_va_ = _parse_float(row[col_va] if col_va is not None else 0)

            # Linhas sem obra e sem valores → dia de descanso, pula
            if not obra_raw and valor_d == 0 and valor_vt_ == 0 and valor_va_ == 0:
                continue

            tipo_lancamento = _detectar_tipo_lancamento(obra_raw)

            # Validação de valor: se não for tipo especial, precisa de valor
            if tipo_lancamento is None and valor_d <= 0:
                # Sem tipo especial e sem valor → pula silenciosamente
                continue

            func_id, func_nome, func_criar, aviso = self._resolver_funcionario(
                funcionario_nome, admin_id, valor_d, valor_va_, valor_vt_)
            # func_criar não é erro — será criado no importar(); apenas erros reais vão para avisos
            if aviso and not func_criar:
                avisos.append({'linha': rn, 'nome': func_nome, 'motivo': aviso})

            if tipo_lancamento:
                obra_id, obra_nome, obra_criar = None, obra_raw.title(), False
            else:
                obra_id, obra_nome, obra_criar = self._resolver_obra(obra_raw, admin_id)

            row_dict = {
                'linha': rn,
                'nome': func_nome,
                'funcionario_id': func_id,
                'func_criar': func_criar,
                'func_valores': {'valor_diaria': valor_d, 'valor_va': valor_va_, 'valor_vt': valor_vt_}
                                if func_criar else None,
                'data': str(data_ref),
                'valor_diaria': valor_d,
                'valor_va': valor_va_,
                'valor_vt': valor_vt_,
                'obra_id': obra_id,
                'obra_nome': obra_nome,
                'obra_criar': obra_criar,
                'obra_raw': obra_raw,
                'tipo_lancamento': tipo_lancamento or 'diaria_completa',
                'tipo_label': _TIPO_LABELS.get(tipo_lancamento or 'diaria_completa', ''),
                'status': 'PENDENTE',
            }
            validos.append(row_dict)

        # Avisos integrados como "erros" não-bloqueantes para exibição no preview
        return validos, avisos

    # ── FORMATO SIGE ──────────────────────────────────────────────────────────
    def _processar_sige(self, ws, admin_id):
        header_row, raw_headers = _detectar_header_row(ws, ['nome_funcionario', 'nome', 'funcionario'])
        if not header_row:
            return [], [{'linha': '?', 'nome': '—', 'motivo': 'Cabeçalho não encontrado'}]

        hm = _mapear_headers(raw_headers)
        validos, erros = [], []

        for rn in range(header_row + 1, ws.max_row + 1):
            def c(*keys):
                return _cel(ws, rn, hm, *keys)

            nome = _norm(c('nome_funcionario', 'nome', 'funcionario'))
            if not nome:
                continue

            data_ref = _parse_data(c('data', 'data_ponto'))
            if not data_ref:
                erros.append({'linha': rn, 'nome': nome, 'motivo': 'Data inválida'})
                continue

            obra_raw = _norm(c('obra', 'obra_id', 'codigo_obra', 'centro_custo', 'centro de custo'))
            tipo_lancamento = _detectar_tipo_lancamento(obra_raw)

            func_id, func_nome, func_criar, aviso = self._resolver_funcionario(nome, admin_id)
            if aviso and not func_criar:
                erros.append({'linha': rn, 'nome': nome, 'motivo': aviso})
                continue

            # Valor: da planilha se informado, senão do cadastro
            valor_d = _parse_float(c('valor_diaria', 'valor', '$ diária', 'diaria'))
            valor_va_ = _parse_float(c('valor_va', 'va', '$ va'))
            valor_vt_ = _parse_float(c('valor_vt', 'vt', '$ vt'))

            if not func_criar:
                from models import Funcionario
                func = Funcionario.query.get(func_id)
                if func:
                    if valor_d <= 0: valor_d = float(func.valor_diaria or 0)
                    if valor_va_ <= 0: valor_va_ = float(func.valor_va or 0)
                    if valor_vt_ <= 0: valor_vt_ = float(func.valor_vt or 0)

            if tipo_lancamento is None and valor_d <= 0:
                erros.append({'linha': rn, 'nome': func_nome,
                              'motivo': 'Funcionário não tem valor de diária cadastrado no perfil'})
                continue

            if tipo_lancamento:
                obra_id, obra_nome, obra_criar = None, (obra_raw.title() if obra_raw else ''), False
            else:
                obra_id, obra_nome, obra_criar = self._resolver_obra(obra_raw, admin_id)

            validos.append({
                'linha': rn,
                'nome': func_nome,
                'funcionario_id': func_id,
                'func_criar': func_criar,
                'func_valores': {'valor_diaria': valor_d, 'valor_va': valor_va_, 'valor_vt': valor_vt_}
                                if func_criar else None,
                'data': str(data_ref),
                'valor_diaria': valor_d,
                'valor_va': valor_va_,
                'valor_vt': valor_vt_,
                'obra_id': obra_id,
                'obra_nome': obra_nome,
                'obra_criar': obra_criar,
                'obra_raw': obra_raw,
                'tipo_lancamento': tipo_lancamento or 'diaria_completa',
                'tipo_label': _TIPO_LABELS.get(tipo_lancamento or 'diaria_completa', ''),
                'status': _norm(c('status')) or 'PENDENTE',
            })

        return validos, erros

    # ── Entry point ───────────────────────────────────────────────────────────
    def processar(self, ws, admin_id):
        if self._is_formato_colaboradores(ws):
            return self._processar_colaboradores(ws, admin_id)
        return self._processar_sige(ws, admin_id)

    # ── importar ──────────────────────────────────────────────────────────────
    def importar(self, rows, admin_id):
        from models import db, Funcionario, Obra
        from utils.financeiro_integration import registrar_custo_automatico
        import random, string
        criados, erros = 0, []

        # Cache de obras e funcionários criados nesta importação para reusar IDs
        obras_criadas = {}      # obra_raw_upper → obra_id
        funcs_criados = {}      # nome_upper → func_id

        for row in rows:
            try:
                data_ref = _parse_data(row['data'])
                linha = row['linha']
                nome = row['nome']
                tipo = row.get('tipo_lancamento', 'diaria_completa')

                # ── Tipo "sem_lancamento": só registro, sem custo ──────────
                if tipo == 'sem_lancamento':
                    criados += 1
                    continue

                # ── Resolver / criar funcionário ──────────────────────────
                func_id = row.get('funcionario_id')
                if row.get('func_criar') and not func_id:
                    nome_up = nome.upper()
                    if nome_up in funcs_criados:
                        func_id = funcs_criados[nome_up]
                    else:
                        # Gera CPF temporário único (até 14 chars)
                        cpf_tmp = f"TMP{random.randint(10000000000, 99999999999)}"
                        # Garante unicidade consultando o DB
                        while Funcionario.query.filter_by(cpf=cpf_tmp).first():
                            cpf_tmp = f"TMP{random.randint(10000000000, 99999999999)}"

                        # Busca max codigo VV global (constraint global)
                        ultimo_vv = (Funcionario.query
                                     .filter(Funcionario.codigo.like('VV%'))
                                     .order_by(Funcionario.codigo.desc()).first())
                        try:
                            proximo_vv = int(ultimo_vv.codigo[2:]) + 1 if ultimo_vv else 1
                        except Exception:
                            proximo_vv = 1
                        # Garante unicidade do código
                        while Funcionario.query.filter(
                            Funcionario.codigo == f"VV{proximo_vv:03d}",
                            Funcionario.admin_id == admin_id
                        ).first():
                            proximo_vv += 1

                        vals = row.get('func_valores') or {}
                        novo_func = Funcionario(
                            codigo=f"VV{proximo_vv:03d}",
                            nome=nome,
                            cpf=cpf_tmp,
                            tipo_remuneracao='diaria',
                            valor_diaria=float(vals.get('valor_diaria', row.get('valor_diaria', 0))),
                            valor_va=float(vals.get('valor_va', row.get('valor_va', 0))),
                            valor_vt=float(vals.get('valor_vt', row.get('valor_vt', 0))),
                            data_admissao=data_ref,
                            ativo=True,
                            admin_id=admin_id,
                        )
                        db.session.add(novo_func)
                        db.session.flush()  # obtém ID sem commit
                        func_id = novo_func.id
                        funcs_criados[nome_up] = func_id

                # ── Resolver / criar obra ─────────────────────────────────
                obra_id = row.get('obra_id')
                if row.get('obra_criar') and not obra_id:
                    obra_raw = row.get('obra_raw', row.get('obra_nome', ''))
                    obra_up = obra_raw.strip().upper()
                    if obra_up in obras_criadas:
                        obra_id = obras_criadas[obra_up]
                    else:
                        nova_obra = Obra(
                            nome=obra_raw.title(),
                            codigo=None,   # sem código — editável depois
                            data_inicio=data_ref,
                            admin_id=admin_id,
                            ativo=True,
                        )
                        db.session.add(nova_obra)
                        db.session.flush()
                        obra_id = nova_obra.id
                        obras_criadas[obra_up] = obra_id

                # ── Lançamentos financeiros ───────────────────────────────
                data_fmt = data_ref.strftime('%d/%m/%Y')

                # Diária — sempre (exceto sem_lancamento já tratado acima)
                filho = registrar_custo_automatico(
                    admin_id=admin_id,
                    tipo_categoria='SALARIO',
                    entidade_nome=nome,
                    entidade_id=func_id,
                    data=data_ref,
                    descricao=f'Diária - {nome} - {data_fmt}',
                    valor=row.get('valor_diaria', 0),
                    obra_id=obra_id,
                    origem_tabela='importacao_diaria',
                    origem_id=func_id,        # Integer — ID do funcionário
                )
                if not filho:
                    erros.append({'linha': linha, 'nome': nome,
                                  'motivo': 'Erro ao registrar custo da diária'})
                    db.session.rollback()
                    continue

                # VA e VT — somente se tipo for diaria_completa
                if tipo == 'diaria_completa':
                    if row.get('valor_va', 0) > 0:
                        registrar_custo_automatico(
                            admin_id=admin_id,
                            tipo_categoria='ALIMENTACAO',
                            entidade_nome=nome,
                            entidade_id=func_id,
                            data=data_ref,
                            descricao=f'VA - {nome} - {data_fmt}',
                            valor=row['valor_va'],
                            obra_id=obra_id,
                            origem_tabela='importacao_diaria_va',
                            origem_id=func_id,
                        )
                    if row.get('valor_vt', 0) > 0:
                        registrar_custo_automatico(
                            admin_id=admin_id,
                            tipo_categoria='TRANSPORTE',
                            entidade_nome=nome,
                            entidade_id=func_id,
                            data=data_ref,
                            descricao=f'VT - {nome} - {data_fmt}',
                            valor=row['valor_vt'],
                            obra_id=obra_id,
                            origem_tabela='importacao_diaria_vt',
                            origem_id=func_id,
                        )
                # tipo 'somente_diaria' (ATESTADO) → só diária, sem VA/VT

                db.session.commit()
                criados += 1

            except Exception as e:
                db.session.rollback()
                erros.append({'linha': row.get('linha'), 'nome': row.get('nome'), 'motivo': str(e)})

        return {'criados': criados, 'erros': erros}


# ── MÓDULO 3: Alimentação ──────────────────────────────────────────────────────

class ImportacaoAlimentacao:
    """
    Colunas: data, obra, valor_total, funcionarios (sep por ;), descricao, restaurante
    Cria AlimentacaoLancamento + associa funcionários por nome
    """

    def processar(self, ws, admin_id):
        from models import Funcionario, Obra

        header_row, raw_headers = _detectar_header_row(ws, ['data', 'valor_total', 'funcionarios'])
        if not header_row:
            return [], [{'linha': '?', 'nome': '—', 'motivo': 'Cabeçalho não encontrado'}]

        hm = _mapear_headers(raw_headers)
        validos, erros = [], []

        for rn in range(header_row + 1, ws.max_row + 1):
            def c(*keys):
                return _cel(ws, rn, hm, *keys)

            data_ref = _parse_data(c('data'))
            if not data_ref:
                continue

            valor = _parse_float(c('valor_total', 'valor'))
            if valor <= 0:
                erros.append({'linha': rn, 'nome': str(data_ref), 'motivo': 'Valor deve ser maior que 0'})
                continue

            obra_raw = _norm(c('obra', 'obra_id'))
            if not obra_raw:
                erros.append({'linha': rn, 'nome': str(data_ref),
                              'motivo': 'Obra obrigatória para lançamento de alimentação'})
                continue
            obra = (Obra.query.filter(Obra.admin_id == admin_id)
                    .filter((Obra.nome.ilike(obra_raw)) | (Obra.codigo == obra_raw))
                    .first())
            if not obra:
                erros.append({'linha': rn, 'nome': str(data_ref), 'motivo': f'Obra "{obra_raw}" não encontrada'})
                continue

            # Resolver lista de funcionários
            funcs_raw = _norm(c('funcionarios', 'funcionario'))
            func_ids = []
            func_nomes_ok = []
            func_nomes_erro = []
            if funcs_raw:
                for fn in [x.strip() for x in funcs_raw.split(';') if x.strip()]:
                    f = (Funcionario.query
                         .filter(Funcionario.nome.ilike(fn), Funcionario.admin_id == admin_id)
                         .first())
                    if f:
                        func_ids.append(f.id)
                        func_nomes_ok.append(f.nome)
                    else:
                        func_nomes_erro.append(fn)

            if func_nomes_erro:
                erros.append({'linha': rn, 'nome': str(data_ref),
                               'motivo': f'Funcionários não encontrados: {", ".join(func_nomes_erro)}'})
                continue

            validos.append({
                'linha': rn,
                'data': str(data_ref),
                'valor_total': valor,
                'obra_id': obra.id if obra else None,
                'obra_nome': obra.nome if obra else '(sem obra)',
                'descricao': _norm(c('descricao')) or 'Alimentação importada',
                'restaurante': _norm(c('restaurante', 'fornecedor')) or None,
                'funcionario_ids': func_ids,
                'funcionarios_nomes': '; '.join(func_nomes_ok),
            })

        return validos, erros

    def importar(self, rows, admin_id):
        from models import db, AlimentacaoLancamento, Funcionario
        from sqlalchemy import text
        criados, erros = 0, []

        for row in rows:
            sp = db.session.begin_nested()
            try:
                data_ref = _parse_data(row['data'])
                lanc = AlimentacaoLancamento(
                    data=data_ref,
                    valor_total=_parse_float(row['valor_total']),
                    descricao=row['descricao'],
                    obra_id=row.get('obra_id'),
                    admin_id=admin_id,
                )
                db.session.add(lanc)
                db.session.flush()

                # Associar funcionários (tabela M2M: alimentacao_funcionarios_assoc)
                for fid in row.get('funcionario_ids', []):
                    db.session.execute(
                        text('INSERT INTO alimentacao_funcionarios_assoc (lancamento_id, funcionario_id, admin_id) VALUES (:lid, :fid, :aid) ON CONFLICT DO NOTHING'),
                        {'lid': lanc.id, 'fid': fid, 'aid': admin_id}
                    )
                sp.commit()
                criados += 1
            except Exception as e:
                sp.rollback()
                erros.append({'linha': row['linha'], 'motivo': str(e)})

        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            erros.append({'linha': 'COMMIT', 'motivo': str(e)})

        return {'criados': criados, 'erros': erros}


# ── MÓDULO 4: Transporte ───────────────────────────────────────────────────────

class ImportacaoTransporte:
    """
    Colunas: data, nome_funcionario, categoria, valor, obra, descricao
    Cria GestaoCustoPai/Filho via registrar_custo_automatico(tipo='TRANSPORTE')
    """

    def processar(self, ws, admin_id):
        from models import Funcionario, Obra, CategoriaTransporte

        header_row, raw_headers = _detectar_header_row(ws, ['nome_funcionario', 'funcionario', 'nome'])
        if not header_row:
            return [], [{'linha': '?', 'nome': '—', 'motivo': 'Cabeçalho não encontrado'}]

        hm = _mapear_headers(raw_headers)
        validos, erros = [], []

        # Cache de categorias para evitar N queries
        categorias_cache = {
            c.nome.lower(): c
            for c in CategoriaTransporte.query.filter_by(admin_id=admin_id).all()
        }

        for rn in range(header_row + 1, ws.max_row + 1):
            def c(*keys):
                return _cel(ws, rn, hm, *keys)

            nome = _norm(c('nome_funcionario', 'nome', 'funcionario'))
            if not nome:
                continue

            data_ref = _parse_data(c('data'))
            if not data_ref:
                erros.append({'linha': rn, 'nome': nome, 'motivo': 'Data inválida'})
                continue

            valor = _parse_float(c('valor'))
            if valor <= 0:
                erros.append({'linha': rn, 'nome': nome, 'motivo': 'Valor deve ser maior que 0'})
                continue

            func = (Funcionario.query
                    .filter(Funcionario.nome.ilike(nome), Funcionario.admin_id == admin_id)
                    .first())
            if not func:
                erros.append({'linha': rn, 'nome': nome, 'motivo': f'Funcionário "{nome}" não encontrado'})
                continue

            obra_raw = _norm(c('obra'))
            obra = None
            if obra_raw:
                obra = (Obra.query.filter(Obra.admin_id == admin_id)
                        .filter((Obra.nome.ilike(obra_raw)) | (Obra.codigo == obra_raw))
                        .first())
                if not obra:
                    erros.append({'linha': rn, 'nome': nome,
                                  'motivo': f'Obra "{obra_raw}" não encontrada'})
                    continue

            # Busca CategoriaTransporte por nome (case-insensitive)
            categoria_raw = _norm(c('categoria')) or ''
            cat_obj = categorias_cache.get(categoria_raw.lower())
            if not cat_obj and categoria_raw:
                # Busca direta caso não esteja no cache (criado após o cache)
                cat_obj = CategoriaTransporte.query.filter(
                    CategoriaTransporte.nome.ilike(categoria_raw),
                    CategoriaTransporte.admin_id == admin_id,
                ).first()
            if not cat_obj:
                erros.append({
                    'linha': rn, 'nome': nome,
                    'motivo': (
                        f'Categoria de transporte "{categoria_raw}" não encontrada. '
                        f'Disponíveis: {", ".join(categorias_cache.keys()) or "nenhuma cadastrada"}'
                    ),
                })
                continue

            validos.append({
                'linha': rn,
                'nome': func.nome,
                'funcionario_id': func.id,
                'data': str(data_ref),
                'valor': valor,
                'categoria': cat_obj.nome,
                'categoria_id': cat_obj.id,
                'obra_id': obra.id if obra else None,
                'obra_nome': obra.nome if obra else '(sem obra)',
                'descricao': _norm(c('descricao')) or f'{cat_obj.nome} - {func.nome} - {data_ref.strftime("%d/%m/%Y")}',
            })

        return validos, erros

    def importar(self, rows, admin_id):
        from models import db
        from utils.financeiro_integration import registrar_custo_automatico
        criados, erros = 0, []

        for row in rows:
            try:
                data_ref = _parse_data(row['data'])
                filho = registrar_custo_automatico(
                    admin_id=admin_id,
                    tipo_categoria='TRANSPORTE',
                    entidade_nome=row['nome'],
                    entidade_id=row['funcionario_id'],
                    data=data_ref,
                    descricao=row['descricao'],
                    valor=_parse_float(row['valor']),
                    obra_id=row.get('obra_id'),
                    origem_tabela='importacao_transporte',
                    origem_id=row['linha'],
                )
                if filho:
                    db.session.commit()
                    criados += 1
                else:
                    erros.append({'linha': row['linha'], 'nome': row['nome'], 'motivo': 'Erro ao registrar custo'})
            except Exception as e:
                db.session.rollback()
                erros.append({'linha': row['linha'], 'nome': row['nome'], 'motivo': str(e)})

        return {'criados': criados, 'erros': erros}


# ── MÓDULO 5: Custos ───────────────────────────────────────────────────────────

class ImportacaoCustos:
    """
    Colunas: data, fornecedor, descricao, valor, obra, categoria, status,
             forma_pagamento, banco_conta, nf_numero, parcela, data_vencimento, observacoes
    Cria GestaoCustoPai diretamente
    """
    CATEGORIAS_VALIDAS = {
        'SALARIO', 'ALIMENTACAO', 'TRANSPORTE', 'MATERIAL',
        'COMPRA', 'DESPESA_GERAL', 'REEMBOLSO',
    }

    def processar(self, ws, admin_id):
        from models import Obra

        header_row, raw_headers = _detectar_header_row(ws, ['fornecedor', 'descricao', 'valor'])
        if not header_row:
            return [], [{'linha': '?', 'nome': '—', 'motivo': 'Cabeçalho não encontrado'}]

        hm = _mapear_headers(raw_headers)
        validos, erros = [], []

        for rn in range(header_row + 1, ws.max_row + 1):
            def c(*keys):
                return _cel(ws, rn, hm, *keys)

            fornecedor = _norm(c('fornecedor', 'nome', 'empresa'))
            descricao = _norm(c('descricao', 'historico', 'item'))
            if not descricao and not fornecedor:
                continue

            valor = _parse_float(c('valor', 'valor_total'))
            if valor <= 0:
                erros.append({'linha': rn, 'nome': fornecedor or descricao, 'motivo': 'Valor deve ser maior que 0'})
                continue

            data_ref = _parse_data(c('data', 'data_pagamento')) or date.today()
            data_venc = _parse_data(c('data_vencimento'))

            categoria = _norm(c('categoria', 'tipo', 'tipo_categoria')).upper()
            if categoria not in self.CATEGORIAS_VALIDAS:
                categoria = 'DESPESA_GERAL'

            obra_raw = _norm(c('obra', 'centro_custo', 'obra_id'))
            obra = None
            if obra_raw:
                obra = (Obra.query.filter(Obra.admin_id == admin_id)
                        .filter((Obra.nome.ilike(obra_raw)) | (Obra.codigo == obra_raw))
                        .first())
                if not obra:
                    erros.append({'linha': rn, 'nome': fornecedor or descricao,
                                  'motivo': f'Obra/centro de custo "{obra_raw}" não encontrado'})
                    continue

            status_raw = _norm(c('status')).upper()
            status = status_raw if status_raw in ('PAGO', 'PENDENTE', 'AUTORIZADO') else 'PENDENTE'

            validos.append({
                'linha': rn,
                'fornecedor': fornecedor or 'Importação',
                'descricao': descricao or fornecedor,
                'valor': valor,
                'data': str(data_ref),
                'data_vencimento': str(data_venc) if data_venc else None,
                'categoria': categoria,
                'obra_id': obra.id if obra else None,
                'obra_nome': obra.nome if obra else '(sem obra)',
                'status': status,
                'forma_pagamento': _norm(c('forma_pagamento')) or None,
                'banco_conta': _norm(c('banco_conta')) or None,
                'nf_numero': _norm(c('nf_numero', 'nota_fiscal')) or None,
                'parcela': _norm(c('parcela')) or None,
                'observacoes': _norm(c('observacoes', 'obs')) or None,
            })

        return validos, erros

    def importar(self, rows, admin_id):
        from models import db, GestaoCustoPai, GestaoCustoFilho
        criados, erros = 0, []

        for row in rows:
            sp = db.session.begin_nested()
            try:
                data_ref = _parse_data(row['data'])
                data_venc = _parse_data(row.get('data_vencimento'))
                valor = _parse_float(row['valor'])

                pai = GestaoCustoPai(
                    admin_id=admin_id,
                    tipo_categoria=row['categoria'],
                    entidade_nome=row['fornecedor'],
                    entidade_id=None,
                    valor_total=valor,
                    status=row.get('status', 'PENDENTE'),
                    forma_pagamento=row.get('forma_pagamento') or None,
                    data_vencimento=data_venc,
                    numero_documento=row.get('nf_numero') or None,
                    observacoes=row.get('observacoes') or None,
                    data_pagamento=data_ref if row.get('status') == 'PAGO' else None,
                )
                db.session.add(pai)
                db.session.flush()

                filho = GestaoCustoFilho(
                    pai_id=pai.id,
                    data_referencia=data_ref,
                    descricao=row.get('descricao') or row['fornecedor'],
                    valor=valor,
                    obra_id=row.get('obra_id'),
                    admin_id=admin_id,
                    origem_tabela='importacao_custos',
                    origem_id=row.get('linha'),
                )
                db.session.add(filho)
                sp.commit()
                criados += 1
            except Exception as e:
                sp.rollback()
                erros.append({'linha': row.get('linha', '?'), 'motivo': str(e)})

        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            erros.append({'linha': 'COMMIT', 'motivo': str(e)})

        return {'criados': criados, 'erros': erros}


# ── Importação Fluxo de Caixa ─────────────────────────────────────────────────

import unicodedata as _ud

def _normalizar(texto):
    """Lowercase sem acentos."""
    s = str(texto or '').lower()
    s = _ud.normalize('NFD', s)
    s = ''.join(c for c in s if _ud.category(c) != 'Mn')
    return s.strip()


# ---------------------------------------------------------------------------
# Mapeamento Plano de Contas → tipo_categoria (Nível 1)
# ---------------------------------------------------------------------------
_PLANO_PARA_CATEGORIA = {
    'salarios': 'SALARIO',
    'salario': 'SALARIO',
    'adiantamento salarial': 'SALARIO',
    'das simples nacional': 'TRIBUTOS',
    'simples nacional': 'TRIBUTOS',
    'inss': 'TRIBUTOS',
    'fgts': 'TRIBUTOS',
    'das': 'TRIBUTOS',
    'projeto': 'MAO_OBRA_DIRETA',
    'honorarios': 'MAO_OBRA_DIRETA',
    'honorario': 'MAO_OBRA_DIRETA',
    'beneficios': 'ALIMENTACAO',
    'beneficio': 'ALIMENTACAO',
    'consumo': 'ALUGUEL_UTILITIES',
    'despesa bancaria': 'OUTROS',
    'despesa financeira': 'OUTROS',
    'manutencao': 'OUTROS',
    'marketing': 'OUTROS',
    'retirada dos socios': 'OUTROS',
    'retirada de socios': 'OUTROS',
}

def _categoria_por_plano(plano):
    p = _normalizar(plano)
    for chave, cat in _PLANO_PARA_CATEGORIA.items():
        if chave in p:
            return cat
    return None


# ---------------------------------------------------------------------------
# Keywords por categoria (Nível 2b)
# ---------------------------------------------------------------------------
_KEYWORDS_CATEGORIA = [
    ('MATERIAL', [
        'cimento', 'ferro', 'aco', 'tijolo', 'areia', 'brita', 'tubo', 'fios',
        'cabos', 'tinta', 'parafuso', 'prego', 'ferragens', 'telha', 'gesso',
        'drywall', 'porcelanato', 'argamassa', 'madeira', 'vidro', 'esquadria',
        'aluminio', 'pvc', 'torneira', 'luminaria', 'lampada', 'conduite',
        'mangueira', 'epi', 'bota', 'luva', 'capacete', 'leroy', 'concrelagos',
        'loja do mecanico', 'lojas mm', 'materiais', 'ferramenta', 'parafusadeira',
    ]),
    ('MAO_OBRA_DIRETA', [
        'diaria', 'servico', 'mao de obra', 'pedreiro', 'ajudante', 'eletricista',
        'encanador', 'pintor', 'gesseiro', 'serralheria', 'vidraca', 'empreitada',
        'armacao', 'carpinteiro', 'medicao', 'pintura', 'hidraulica', 'eletrica',
        'medicao', 'medir', 'contratado', 'terceirizado',
    ]),
    ('ALIMENTACAO', [
        'vale alimentacao', 'almoco', 'refeicao', 'marmita', 'pao', 'lanche',
        'cesta basica', 'restaurante', 'cafe', 'marmitex', 'jantar', 'cambuca',
        'ifood', 'mercado livre', 'mercado', 'supermercado', 'alimentos',
    ]),
    ('TRANSPORTE', [
        'vale transporte', 'combustivel', 'gasolina', 'diesel', 'passagem', 'uber',
        'pedagio', 'estacionamento', 'frete', ' km ', '+km', 'km ', 'kilometro',
        'quilometro', 'rodagem',
    ]),
    ('ALUGUEL_UTILITIES', [
        'aluguel', ' luz', 'agua', 'internet', 'energia', 'enel', 'sabesp',
        'vivo', 'claro', 'tim', 'iptu', 'condominio', 'google', 'workspace',
        'oi ', 'net ', 'sky ', 'streaming',
    ]),
    ('TRIBUTOS', [
        'imposto', ' das', 'iss', 'taxa', 'prefeitura', 'simples nacional',
        'gps', 'fgts', 'darf', 'tributo', 'inss', 'contribuicao',
    ]),
    ('OUTROS', [
        'tarifa', 'juros', 'multa', 'seguro', 'mensalidade', 'papelaria',
        'limpeza', 'plotagem', 'bonus', 'gratificacao', 'mouse', 'notebook',
        'computador', 'celular', 'escritorio',
    ]),
]

# Keywords de contexto dentro de reembolsos (Nível 2a)
_KEYWORDS_REEMBOLSO_CONTEXTO = [
    ('TRANSPORTE', ['km', 'kilometro', 'quilometro', 'rodagem', 'combustivel',
                    'gasolina', 'uber', 'passagem', 'frete', 'pedagio']),
    ('ALIMENTACAO', ['almoco', 'refeicao', 'marmita', 'cafe', 'jantar', 'lanche',
                     'ifood', 'alimentos', 'refeicoes', 'lanchonete']),
    ('MATERIAL', ['material', 'cimento', 'ferragem', 'leroy', 'parafuso', 'epi',
                  'bota', 'loja do mecanico', 'parafusadeira', 'ferramenta',
                  'tinta', 'madeira', 'vidro']),
    ('MAO_OBRA_DIRETA', ['diaria', 'medicao', 'servico', 'empreitada', 'contratado']),
    ('SALARIO', ['salario', 'salário']),
    ('ALUGUEL_UTILITIES', ['luz', 'agua', 'internet', 'energia', 'enel', 'sabesp',
                            'google', 'workspace', 'aluguel', 'gas']),
    ('OUTROS', ['cartao', 'compras', 'diverso', 'reembolso', 'adiantamento']),
]


def _classificar_keywords(texto):
    """Aplica keywords gerais ao texto. Retorna tipo_categoria ou None."""
    t = _normalizar(texto)
    for cat, kws in _KEYWORDS_CATEGORIA:
        for kw in kws:
            if kw in t:
                return cat
    return None


def _classificar_reembolso_contexto(desc_completa):
    """
    Extrai contexto após '|' ou '-' e aplica keywords de reembolso.
    Padrão fixo 'km' → TRANSPORTE.
    Retorna (tipo_categoria, contexto_extraido) ou (None, '').
    """
    desc_norm = _normalizar(desc_completa)

    # Padrão fixo "reembolso + km"
    if 'km' in desc_norm:
        return 'TRANSPORTE', desc_norm

    # Extrair contexto: texto após '|' ou '-'
    contexto = desc_completa
    if '|' in desc_completa:
        contexto = desc_completa.split('|', 1)[-1]
    elif ' - ' in desc_completa:
        partes = desc_completa.split(' - ', 1)
        # Usar a segunda parte só se a primeira tiver "reembolso/adiantamento"
        primeira = _normalizar(partes[0])
        if any(x in primeira for x in ['reembolso', 'adiantamento']):
            contexto = partes[1]

    ctx_norm = _normalizar(contexto)
    for cat, kws in _KEYWORDS_REEMBOLSO_CONTEXTO:
        for kw in kws:
            if kw in ctx_norm:
                if cat != 'OUTROS':  # OUTROS é fallback, verificar outros primeiro
                    return cat, ctx_norm
    # Segunda passagem para OUTROS
    for kw in _KEYWORDS_REEMBOLSO_CONTEXTO[-1][1]:
        if kw in ctx_norm:
            return None, ctx_norm  # sem contexto suficiente → revisão manual

    return None, ctx_norm


def _eh_reembolso(desc):
    d = _normalizar(desc)
    return any(x in d for x in ['reembolso', 'adiantamento', 'reembolsos'])


def _eh_transferencia_interna(desc, valor):
    if valor is None:
        return True
    d = _normalizar(str(desc or ''))
    return ('transferencia de valores' in d or
            ('nubank' in d and 'itau' in d))


# ---------------------------------------------------------------------------
# Mapeamento Centro de Custo → obra_id (admin_id=63)
# ---------------------------------------------------------------------------
_CC_OBRA_MAP = {
    'gespi refeitorio': 243,
    'gespi - refeitorio': 243,
    'gespi suprimentos': 245,
    'gespi - suprimentos': 245,
    'gespi sala de controle': 245,
    'gespi - sala de controle': 245,
    'angela cid': 249,
    'clinica dgm': 255,
    'dgm': 255,
    'gattai': 254,
    'braganca': 254,
    'anderson': 251,
    'anderson - urbanova': 251,
    'steel home': 246,
    'steelhome': 246,
    'vereda': 247,
    'rafael': 256,
    'rafael - urbanova': 256,
    'urbanova': 256,
}
_CC_ADMIN = ['escritorio', 'guilherme e ariane', 'guilherme', 'ariane',
             'administrativo', 'geral', 'head', 'holding']


def _match_cc_obra(cc, obras_dict):
    """Retorna obra_id ou None (None = usar obra ADMINISTRATIVO)."""
    if not cc:
        return None
    cc_norm = _normalizar(cc)
    # Checa se é administrativo
    if any(x in cc_norm for x in _CC_ADMIN):
        return None
    # Lookup direto
    valid_obra_ids = set(obras_dict.values()) if obras_dict else set()
    for chave, oid in _CC_OBRA_MAP.items():
        if chave in cc_norm or cc_norm in chave:
            # Valida que o ID pertence às obras do admin atual
            if oid in valid_obra_ids:
                return oid
            # ID não pertence a este tenant: segue para fuzzy
            break
    # Fuzzy fallback com obras do banco
    try:
        from thefuzz import process as fuzz_process
        if obras_dict:
            match = fuzz_process.extractOne(cc_norm, obras_dict.keys(), score_cutoff=70)
            if match:
                return obras_dict[match[0]]
    except ImportError:
        pass
    return None


def _fuzzy_match_entidade(nome_excel, funcionarios, fornecedores):
    """
    Retorna (tipo, id, nome_banco, score) ou (None, None, None, 0).
    tipo = 'funcionario' | 'fornecedor'
    Prioridade: Funcionario > Fornecedor (regra de negócio para reembolsos).
    Fornecedores consideram nome, razao_social e nome_fantasia.
    Carregue funcionarios e fornecedores EM MEMÓRIA antes de chamar.

    funcionarios: list of (id, nome)
    fornecedores: list of (id, nome, razao_social, nome_fantasia)  — aliases opcionais
    """
    try:
        from thefuzz import fuzz
    except ImportError:
        return None, None, None, 0

    THRESHOLD = 85
    nome_norm = _normalizar(nome_excel)

    # ── 1. Tenta funcionário primeiro (prioridade) ──────────────────────────
    melhor_f_score = 0
    melhor_f_id = None
    melhor_f_nome = None
    for item in funcionarios:
        fid, fnome = item[0], item[1]
        score = fuzz.token_set_ratio(nome_norm, _normalizar(fnome))
        if score > melhor_f_score:
            melhor_f_score = score
            melhor_f_id = fid
            melhor_f_nome = fnome

    if melhor_f_score > THRESHOLD:
        return 'funcionario', melhor_f_id, melhor_f_nome, melhor_f_score

    # ── 2. Tenta fornecedor (considera todos os aliases) ────────────────────
    melhor_s_score = 0
    melhor_s_id = None
    melhor_s_nome = None
    for item in fornecedores:
        fid = item[0]
        aliases = [a for a in item[1:] if a]  # nome, razao_social, nome_fantasia
        for alias in aliases:
            score = fuzz.token_set_ratio(nome_norm, _normalizar(alias))
            if score > melhor_s_score:
                melhor_s_score = score
                melhor_s_id = fid
                melhor_s_nome = alias

    if melhor_s_score > THRESHOLD:
        return 'fornecedor', melhor_s_id, melhor_s_nome, melhor_s_score

    # Nenhum match suficiente: retorna o melhor score geral (sem vínculo)
    if melhor_f_score >= melhor_s_score:
        return None, None, None, melhor_f_score
    return None, None, None, melhor_s_score


class ImportacaoFluxoCaixa:
    """
    Parser + classificador para o arquivo Fluxo de Caixa Veks Engenharia.
    Processa abas 'Entrada' e 'Saída', retorna 4 listas.
    """

    def processar(self, arquivo_path_ou_file, admin_id,
                  data_inicio=None, data_fim=None):
        """
        Lê o Excel e retorna dict com 4 listas + metadados:
          entradas, saidas_auto, saidas_manual, ignorados,
          primeiro_dia (date | None), periodo_str (str),
          datas_disponiveis (sorted list of date strings in file)

        data_inicio / data_fim: date objects para filtrar o período.
        Se ambos None, processa todas as datas do arquivo.
        """
        import openpyxl
        from datetime import datetime as dt

        wb = openpyxl.load_workbook(arquivo_path_ou_file, data_only=True)

        # Carregar entidades em memória (um hit no BD)
        from models import Funcionario, Fornecedor
        funcionarios = [(f.id, f.nome) for f in
                        Funcionario.query.filter_by(admin_id=admin_id, ativo=True).all()]
        # Fornecedores: tupla (id, nome, razao_social, nome_fantasia) para fuzzy com 3 aliases
        fornecedores = []
        for f in Fornecedor.query.filter_by(admin_id=admin_id, ativo=True).all():
            fornecedores.append((f.id, f.nome or '', f.razao_social or '', f.nome_fantasia or ''))

        # Carregar obras em memória para match fuzzy de CC
        from models import Obra
        obras_qs = Obra.query.filter_by(admin_id=admin_id).all()
        obras_dict = {_normalizar(o.nome): o.id for o in obras_qs}

        # Pre-computar fornecedores MATERIAL com compras no período (para sugestao_apenas_pagamento)
        # Nota: a query de PedidoCompra é feita DEPOIS de coletar as datas do arquivo
        # para que, quando o usuário não passar data_inicio/data_fim, seja possível
        # usar o intervalo real das datas encontradas no arquivo.
        # Portanto, _forn_com_compras é populado após o primeiro parse das abas.
        _forn_material_ids = set()
        _forn_com_compras = set()  # será preenchido após o primeiro loop

        entradas = []
        saidas_auto = []
        saidas_manual = []
        ignorados = []
        primeiro_dia = None   # primeira data com lançamento no arquivo
        todas_datas = set()   # todas as datas encontradas no arquivo

        def _dentro_periodo(d):
            if data_inicio and d < data_inicio:
                return False
            if data_fim and d > data_fim:
                return False
            return True

        # ── Aba Entrada ──────────────────────────────────────────────────────
        if 'Entrada' in wb.sheetnames:
            ws_e = wb['Entrada']
            for row in ws_e.iter_rows(min_row=6, values_only=True):
                data_val = row[0]
                if not data_val or not isinstance(data_val, (dt, date)):
                    continue
                data_obj = data_val.date() if isinstance(data_val, dt) else data_val
                todas_datas.add(data_obj)
                if primeiro_dia is None or data_obj < primeiro_dia:
                    primeiro_dia = data_obj
                if not _dentro_periodo(data_obj):
                    continue

                plano = _norm(row[1]) if row[1] else ''
                cliente = _norm(row[2]) if row[2] else ''
                desc = _norm(row[3]) if row[3] else ''
                cc = _norm(row[4]) if row[4] else ''
                status_raw = _norm(row[9]) if len(row) > 9 else ''
                valor = None
                for i in range(10, min(len(row), 13)):
                    if row[i] and isinstance(row[i], (int, float)):
                        valor = float(row[i])
                        break
                    elif row[i]:
                        v = _parse_float(row[i])
                        if v > 0:
                            valor = v
                            break

                if not valor:
                    continue

                status = 'PAGO' if 'pago' in _normalizar(status_raw) else 'PENDENTE'
                obra_id = _match_cc_obra(cc, obras_dict)

                # Fuzzy match do cliente
                ent_tipo, ent_id, ent_nome_banco, ent_score = _fuzzy_match_entidade(
                    cliente, funcionarios, fornecedores)

                entradas.append({
                    'tipo': 'entrada',
                    'data': str(data_obj),
                    'plano_contas': plano,
                    'cliente': cliente,
                    'descricao': desc,
                    'cc': cc,
                    'obra_id': obra_id,
                    'valor': valor,
                    'status': status,
                    'entidade_tipo': ent_tipo,
                    'entidade_id': ent_id,
                    'entidade_nome_banco': ent_nome_banco,
                    'fuzzy_score': ent_score,
                })

        # ── Aba Saída ────────────────────────────────────────────────────────
        if 'Saída' in wb.sheetnames:
            ws_s = wb['Saída']
            for row in ws_s.iter_rows(min_row=6, values_only=True):
                data_val = row[1]
                if not data_val or not isinstance(data_val, (dt, date)):
                    continue
                data_obj = data_val.date() if isinstance(data_val, dt) else data_val
                todas_datas.add(data_obj)
                if primeiro_dia is None or data_obj < primeiro_dia:
                    primeiro_dia = data_obj
                if not _dentro_periodo(data_obj):
                    continue

                plano = _norm(row[2]) if len(row) > 2 and row[2] else ''
                fornecedor_nome = _norm(row[3]) if len(row) > 3 and row[3] else ''
                desc = _norm(row[4]) if len(row) > 4 and row[4] else ''
                cc = _norm(row[5]) if len(row) > 5 and row[5] else ''
                valor = None
                for i in range(10, min(len(row), 14)):
                    if row[i] and isinstance(row[i], (int, float)):
                        valor = float(row[i])
                        break
                    elif row[i]:
                        v = _parse_float(row[i])
                        if v > 0:
                            valor = v
                            break

                # Determinar status — procurar em colunas restantes
                status_raw = ''
                for i in range(6, min(len(row), 12)):
                    cell = _norm(row[i]) if row[i] else ''
                    if 'pago' in cell.lower() or 'aberto' in cell.lower():
                        status_raw = cell
                        break

                # Transferência interna → ignorar
                if _eh_transferencia_interna(desc + ' ' + plano, valor):
                    ignorados.append({
                        'data': str(data_obj),
                        'fornecedor': fornecedor_nome,
                        'descricao': desc,
                        'motivo': 'Transferência interna',
                    })
                    continue

                status = 'PAGO' if 'pago' in _normalizar(status_raw) else 'PENDENTE'
                obra_id = _match_cc_obra(cc, obras_dict)

                # Fuzzy match do fornecedor
                ent_tipo, ent_id, ent_nome_banco, ent_score = _fuzzy_match_entidade(
                    fornecedor_nome, funcionarios, fornecedores)
                obs_fuzzy = None
                if ent_score <= 85 and fornecedor_nome:
                    obs_fuzzy = f'[EXCEL] {fornecedor_nome} — vincular manualmente'

                texto_busca = (desc + ' ' + fornecedor_nome).strip()
                eh_reembolso = _eh_reembolso(desc)

                # Nível 1: Plano de Contas
                cat = _categoria_por_plano(plano)

                # Nível 2a: Reembolso com contexto
                precisa_revisao = False
                if cat is None and eh_reembolso:
                    cat, _ = _classificar_reembolso_contexto(desc)
                    if cat is None:
                        precisa_revisao = True

                # Nível 2b: keywords gerais
                if cat is None and not precisa_revisao:
                    cat = _classificar_keywords(texto_busca)
                    if cat is None:
                        precisa_revisao = True

                registro = {
                    'tipo': 'saida',
                    'data': str(data_obj),
                    'plano_contas': plano,
                    'fornecedor': fornecedor_nome,
                    'descricao': desc,
                    'cc': cc,
                    'obra_id': obra_id,
                    'valor': valor,
                    'status': status,
                    'tipo_categoria': cat,
                    'eh_reembolso': eh_reembolso,
                    'entidade_tipo': ent_tipo,
                    'entidade_id': ent_id,
                    'entidade_nome_banco': ent_nome_banco,
                    'fuzzy_score': ent_score,
                    'observacoes': obs_fuzzy,
                    'sugestao_apenas_pagamento': False,  # será ajustado após todos os loops
                }

                if precisa_revisao:
                    saidas_manual.append(registro)
                else:
                    saidas_auto.append(registro)

        # ── Sugestão "Apenas Pagamento" ──────────────────────────────────────────
        # Feita APÓS todos os loops para que todas_datas contenha o intervalo real do arquivo.
        # Quando o usuário não passa data_inicio/data_fim, usamos o min/max das datas encontradas.
        _ap_inicio = data_inicio or (min(todas_datas) if todas_datas else None)
        _ap_fim = data_fim or (max(todas_datas) if todas_datas else None)
        try:
            from models import PedidoCompra as PC
            _forn_material_ids = set(
                f.id for f in Fornecedor.query.filter(
                    Fornecedor.admin_id == admin_id,
                    Fornecedor.tipo_fornecedor == 'MATERIAL',
                ).all()
            )
            if _forn_material_ids:
                q_pc = PC.query.filter(
                    PC.admin_id == admin_id,
                    PC.fornecedor_id.in_(_forn_material_ids),
                )
                if _ap_inicio:
                    q_pc = q_pc.filter(PC.data_compra >= _ap_inicio)
                if _ap_fim:
                    q_pc = q_pc.filter(PC.data_compra <= _ap_fim)
                _forn_com_compras = set(
                    pc.fornecedor_id for pc in q_pc.all() if pc.fornecedor_id
                )
        except Exception as exc_ap:
            import logging as _log
            _log.getLogger(__name__).warning(
                f'[FLUXO processar] Falha ao calcular sugestao_apenas_pagamento: {exc_ap}'
            )

        # Atualizar sugestao_apenas_pagamento em todos os registros de saída
        for r in saidas_auto + saidas_manual:
            r['sugestao_apenas_pagamento'] = (
                r.get('entidade_tipo') == 'funcionario'
                or (
                    r.get('entidade_tipo') == 'fornecedor'
                    and r.get('entidade_id')
                    and r.get('entidade_id') in _forn_com_compras
                )
            )

        # Período descritivo
        datas_sorted = sorted(todas_datas)
        if data_inicio and data_fim:
            periodo_str = f"{data_inicio.strftime('%d/%m/%Y')} a {data_fim.strftime('%d/%m/%Y')}"
        elif datas_sorted:
            periodo_str = (f"{datas_sorted[0].strftime('%d/%m/%Y')} a "
                           f"{datas_sorted[-1].strftime('%d/%m/%Y')}")
        else:
            periodo_str = '—'

        return {
            'entradas': entradas,
            'saidas_auto': saidas_auto,
            'saidas_manual': saidas_manual,
            'ignorados': ignorados,
            'primeiro_dia': str(primeiro_dia) if primeiro_dia else None,
            'periodo_str': periodo_str,
            'datas_disponiveis': [str(d) for d in datas_sorted],
        }

    def importar(self, dados, admin_id):
        """
        Persiste os registros confirmados no BD.
        dados = {
          'entradas': [...],
          'saidas': [...],  # combinação de auto + manual já com categorias definidas
          'batch_id': str,
        }
        Retorna dict com totais por categoria + contagens.
        """
        import uuid
        from datetime import datetime as dt
        from decimal import Decimal
        from models import (db, GestaoCustoPai, GestaoCustoFilho, ContaPagar,
                            ContaReceber, FluxoCaixa, Obra)

        batch_id = dados.get('batch_id') or \
            f"import_{dt.now().strftime('%Y%m%d_%H%M')}_{uuid.uuid4().hex[:6]}"

        totais = {}
        n_entradas = 0
        n_saidas = 0
        n_fluxo = 0
        n_conta_pagar = 0
        n_apenas_pagamento = 0
        n_fornecedores_criados = 0
        erros = []
        duplicados = 0

        # Garantir obra administrativa
        from datetime import date as _date
        obra_adm = Obra.query.filter(
            Obra.admin_id == admin_id,
            Obra.nome.ilike('%ADMINISTRATIVO%')
        ).first()
        if not obra_adm:
            obra_adm = Obra(
                nome='000 - ADMINISTRATIVO / GERAL',
                admin_id=admin_id,
                data_inicio=_date.today(),
            )
            db.session.add(obra_adm)
            db.session.flush()
        obra_adm_id = obra_adm.id

        def _obra_efetiva(obra_id):
            return obra_id if obra_id else obra_adm_id

        def _ja_existe_saida(data_str, valor, fornecedor, aid):
            """Checa chave de não-duplicidade para saídas."""
            existing = GestaoCustoPai.query.filter(
                GestaoCustoPai.admin_id == aid,
                GestaoCustoPai.entidade_nome == fornecedor,
                GestaoCustoPai.valor_total == Decimal(str(valor)),
            ).join(
                GestaoCustoFilho,
                GestaoCustoFilho.pai_id == GestaoCustoPai.id
            ).filter(
                GestaoCustoFilho.data_referencia == _parse_data(data_str)
            ).first()
            return existing is not None

        def _ja_existe_entrada(data_str, valor, cliente, aid):
            existing = ContaReceber.query.filter(
                ContaReceber.admin_id == aid,
                ContaReceber.cliente_nome == cliente,
                ContaReceber.valor_original == Decimal(str(valor)),
                ContaReceber.data_emissao == _parse_data(data_str)
            ).first()
            return existing is not None

        try:
            # ── Auto-criar Fornecedores não reconhecidos ─────────────────────
            from models import Fornecedor as FornecedorModel

            def _inferir_tipo_fornecedor(cat, nome_lower, plano_lower=''):
                """Infere tipo_fornecedor a partir da categoria, plano de contas e nome."""
                if cat == 'MATERIAL':
                    return 'MATERIAL'
                if cat in ('MAO_OBRA_DIRETA', 'SERVICO'):
                    return 'PRESTADOR_SERVICO'
                # Keywords no plano de contas (mais específico que o nome)
                _plano_mat_kw = ('material', 'compra', 'estoque', 'suprimento')
                _plano_prest_kw = ('servico', 'servico', 'honorario', 'empreitada',
                                   'terceiro', 'prestac')
                for kw in _plano_mat_kw:
                    if kw in plano_lower:
                        return 'MATERIAL'
                for kw in _plano_prest_kw:
                    if kw in plano_lower:
                        return 'PRESTADOR_SERVICO'
                # Fallback: keywords no nome do fornecedor
                _mat_kw = ('material', 'ferrag', 'leroy', 'tintas', 'pvc',
                           'ferro', 'aco', 'cimento', 'madeira', 'vidro',
                           'ferramentas', 'parafus', 'tubos', 'telha', 'compra')
                _prest_kw = ('empreit', 'servic', 'construtora', 'reform',
                             'instalac', 'manutenc', 'eletric', 'hidraul',
                             'honorar', 'pedreiro', 'pintor', 'gesseiro',
                             'serralheria', 'terceiro')
                for kw in _mat_kw:
                    if kw in nome_lower:
                        return 'MATERIAL'
                for kw in _prest_kw:
                    if kw in nome_lower:
                        return 'PRESTADOR_SERVICO'
                return 'OUTRO'

            n_fornecedores_criados = 0
            _fornecedor_id_map = {}   # nome_lower → fornecedor_id (novos ou já existentes)

            # Coletar nomes únicos + melhor categoria + plano de contas por nome
            nomes_nao_matched = {}  # nome_lower → (nome_original, best_cat, best_plano)
            _PREF_CATS = ('MATERIAL', 'MAO_OBRA_DIRETA', 'SERVICO')
            for row in dados.get('saidas', []):
                ent_tipo = row.get('entidade_tipo')
                ent_id_row = row.get('entidade_id')
                nome = (row.get('fornecedor') or '').strip()
                nome_lower = nome.lower()
                if nome and not ent_tipo and not ent_id_row and nome_lower not in ('desconhecido', ''):
                    cat_row = row.get('tipo_categoria') or 'OUTROS'
                    plano_row = (row.get('plano_contas') or '').lower()
                    prev = nomes_nao_matched.get(nome_lower, (nome, 'OUTROS', ''))
                    best_cat = cat_row if cat_row in _PREF_CATS else prev[1]
                    best_plano = plano_row if plano_row else prev[2]
                    nomes_nao_matched[nome_lower] = (nome, best_cat, best_plano)

            for nome_lower, (nome_original, best_cat, best_plano) in nomes_nao_matched.items():
                existente = FornecedorModel.query.filter(
                    FornecedorModel.admin_id == admin_id,
                    db.func.lower(FornecedorModel.nome) == nome_lower
                ).first()
                if existente:
                    _fornecedor_id_map[nome_lower] = existente.id
                    continue
                tipo = _inferir_tipo_fornecedor(best_cat, nome_lower, best_plano)
                cnpj_placeholder = f'IMP-{uuid.uuid4().hex[:14]}'
                novo_forn = FornecedorModel(
                    nome=nome_original[:100],
                    cnpj=cnpj_placeholder,
                    tipo_fornecedor=tipo,
                    admin_id=admin_id,
                    ativo=True,
                )
                db.session.add(novo_forn)
                db.session.flush()
                _fornecedor_id_map[nome_lower] = novo_forn.id
                n_fornecedores_criados += 1

            # ── Saídas ──────────────────────────────────────────────────────
            # Erros em qualquer linha propagam para o bloco externo que faz rollback total
            for row in dados.get('saidas', []):
                cat = row.get('tipo_categoria') or 'OUTROS'
                valor = float(row.get('valor') or 0)
                fornecedor = row.get('fornecedor') or 'Desconhecido'
                data_str = row.get('data', '')
                status = row.get('status', 'PENDENTE')
                obra_id = _obra_efetiva(row.get('obra_id'))
                ent_id = row.get('entidade_id')
                ent_tipo_row = row.get('entidade_tipo')
                obs = row.get('observacoes') or ''
                apenas_pagamento = bool(row.get('apenas_pagamento', False))

                # Usar fornecedor auto-criado se entidade não estava vinculada
                if not ent_id and fornecedor.lower() in _fornecedor_id_map:
                    ent_id = _fornecedor_id_map[fornecedor.lower()]

                if _ja_existe_saida(data_str, valor, fornecedor, admin_id):
                    duplicados += 1
                    continue

                data_obj = _parse_data(data_str)

                banco_id_row = row.get('banco_id') or None

                if apenas_pagamento:
                    # ── Modo "Apenas Pagamento": cria apenas FluxoCaixa, sem GCP/GCF/ContaPagar
                    fc = FluxoCaixa(
                        admin_id=admin_id,
                        data_movimento=data_obj,
                        tipo_movimento='SAIDA',
                        categoria=cat,
                        valor=valor,
                        descricao=(row.get('descricao') or fornecedor)[:200],
                        obra_id=obra_id,
                        observacoes=obs or None,
                        import_batch_id=batch_id,
                        banco_id=banco_id_row,
                    )
                    db.session.add(fc)
                    n_fluxo += 1
                    n_apenas_pagamento += 1

                else:
                    # ── Modo normal: GCP + GCF + FluxoCaixa + ContaPagar (reembolso) ─
                    status_gcp = 'PAGO' if status == 'PAGO' else 'PENDENTE'

                    # GestaoCustoPai
                    gcp = GestaoCustoPai(
                        tipo_categoria=cat,
                        entidade_nome=fornecedor,
                        entidade_id=ent_id,
                        valor_total=Decimal(str(valor)),
                        status=status_gcp,
                        data_pagamento=data_obj if status == 'PAGO' else None,
                        observacoes=obs or None,
                        admin_id=admin_id,
                        import_batch_id=batch_id,
                    )
                    db.session.add(gcp)
                    db.session.flush()

                    # GestaoCustoFilho
                    gcf = GestaoCustoFilho(
                        pai_id=gcp.id,
                        descricao=(row.get('descricao') or fornecedor)[:300],
                        valor=Decimal(str(valor)),
                        data_referencia=data_obj,
                        obra_id=obra_id,
                        admin_id=admin_id,
                    )
                    db.session.add(gcf)

                    # FluxoCaixa para PAGO
                    if status == 'PAGO':
                        fc = FluxoCaixa(
                            admin_id=admin_id,
                            data_movimento=data_obj,
                            tipo_movimento='SAIDA',
                            categoria=cat,
                            valor=valor,
                            descricao=(row.get('descricao') or fornecedor)[:200],
                            obra_id=obra_id,
                            referencia_id=gcp.id,
                            referencia_tabela='gestao_custo_pai',
                            observacoes=obs or None,
                            import_batch_id=batch_id,
                        )
                        db.session.add(fc)
                        n_fluxo += 1

                    # ContaPagar para reembolsos
                    if row.get('eh_reembolso') and data_obj:
                        eh_forn = ent_tipo_row == 'fornecedor'
                        eh_func = ent_tipo_row == 'funcionario'

                        obs_parts = []
                        if cat:
                            obs_parts.append(f'Categoria: {cat}')
                        if eh_func and ent_id:
                            obs_parts.append(f'FUNCIONARIO_ID: {ent_id}')
                            obs_parts.append(f'Funcionario: {row.get("entidade_nome_banco") or fornecedor}')
                        if obs:
                            obs_parts.append(obs)
                        obs_final = '. '.join(obs_parts) or None

                        cp = ContaPagar(
                            descricao=f"[REEMBOLSO] {row.get('descricao') or fornecedor}",
                            valor_original=Decimal(str(valor)),
                            valor_pago=Decimal(str(valor)) if status == 'PAGO' else Decimal('0'),
                            saldo=Decimal('0') if status == 'PAGO' else Decimal(str(valor)),
                            data_emissao=data_obj,
                            data_vencimento=data_obj,
                            data_pagamento=data_obj if status == 'PAGO' else None,
                            status='PAGO' if status == 'PAGO' else 'PENDENTE',
                            obra_id=obra_id,
                            admin_id=admin_id,
                            fornecedor_id=ent_id if (ent_id and eh_forn) else None,
                            observacoes=obs_final,
                            origem_tipo='gestao_custo_pai',
                            origem_id=gcp.id,
                            import_batch_id=batch_id,
                        )
                        db.session.add(cp)
                        n_conta_pagar += 1

                n_saidas += 1
                totais[cat] = totais.get(cat, {'count': 0, 'valor': 0.0})
                totais[cat]['count'] += 1
                totais[cat]['valor'] += valor

            # ── Entradas ─────────────────────────────────────────────────────
            for row in dados.get('entradas', []):
                valor = float(row.get('valor') or 0)
                cliente = row.get('cliente') or 'Desconhecido'
                data_str = row.get('data', '')
                status = row.get('status', 'PENDENTE')
                obra_id = _obra_efetiva(row.get('obra_id'))

                if _ja_existe_entrada(data_str, valor, cliente, admin_id):
                    duplicados += 1
                    continue

                data_obj = _parse_data(data_str)

                # Observação estruturada para entradas sem vínculo automático
                cr_obs = None
                if not row.get('entidade_id'):
                    cr_obs = f'Vincular manualmente: cliente "{cliente}" não identificado automaticamente.'

                cr = ContaReceber(
                    cliente_nome=cliente,
                    descricao=row.get('descricao') or f'Entrada {data_str}',
                    valor_original=Decimal(str(valor)),
                    valor_recebido=Decimal(str(valor)) if status == 'PAGO' else Decimal('0'),
                    saldo=Decimal('0') if status == 'PAGO' else Decimal(str(valor)),
                    data_emissao=data_obj,
                    data_vencimento=data_obj,
                    data_recebimento=data_obj if status == 'PAGO' else None,
                    status='RECEBIDO' if status == 'PAGO' else 'PENDENTE',
                    obra_id=obra_id,
                    admin_id=admin_id,
                    observacoes=cr_obs,
                    import_batch_id=batch_id,
                )
                db.session.add(cr)
                db.session.flush()

                if status == 'PAGO':
                    fc = FluxoCaixa(
                        admin_id=admin_id,
                        data_movimento=data_obj,
                        tipo_movimento='ENTRADA',
                        categoria='receita',
                        valor=valor,
                        descricao=(row.get('descricao') or cliente)[:200],
                        obra_id=obra_id,
                        referencia_id=cr.id,
                        referencia_tabela='conta_receber',
                        import_batch_id=batch_id,
                    )
                    db.session.add(fc)
                    n_fluxo += 1

                n_entradas += 1

            db.session.commit()

        except Exception as e:
            db.session.rollback()
            logger.error(f'[FLUXO] Erro na importação — rollback total do lote {batch_id}: {e}', exc_info=True)
            erros.append({'linha': 'GERAL', 'motivo': str(e)})

        return {
            'batch_id': batch_id,
            'n_saidas': n_saidas,
            'n_entradas': n_entradas,
            'n_fluxo': n_fluxo,
            'n_conta_pagar': n_conta_pagar,
            'n_apenas_pagamento': n_apenas_pagamento,
            'n_fornecedores_criados': n_fornecedores_criados,
            'duplicados': duplicados,
            'totais': totais,
            'erros': erros,
        }


# ── Fábrica ────────────────────────────────────────────────────────────────────

MODULO_MAP = {
    'funcionarios': ImportacaoFuncionarios,
    'diarias': ImportacaoDiarias,
    'alimentacao': ImportacaoAlimentacao,
    'transporte': ImportacaoTransporte,
    'custos': ImportacaoCustos,
}

def get_importador(modulo):
    cls = MODULO_MAP.get(modulo)
    if not cls:
        raise ValueError(f'Módulo desconhecido: {modulo}')
    return cls()
