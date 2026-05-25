"""Exportação e importação de Fornecedores via planilha Excel.

- ``gerar_modelo_fornecedores_xlsx(admin_id)`` — exporta fornecedores ativos do
  tenant com cabeçalho, aba Instruções e dropdown tipo_fornecedor.
- ``importar_fornecedores_xlsx(stream, admin_id)`` — lê linhas, valida, faz
  upsert por CNPJ dentro do mesmo tenant, retorna dict com resultados e erros.
"""
from __future__ import annotations

import io
import logging
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from models import Fornecedor, db

logger = logging.getLogger(__name__)

MAX_UPLOAD_BYTES = 5 * 1024 * 1024
MAX_DATA_ROWS = 5000

TIPOS_FORNECEDOR_VALIDOS = {'MATERIAL', 'PRESTADOR_SERVICO', 'OUTRO'}
_TIPOS_FORNECEDOR_DROPDOWN = '"MATERIAL,PRESTADOR_SERVICO,OUTRO"'

FORNECEDOR_HEADERS = [
    'razao_social', 'nome_fantasia', 'cnpj', 'inscricao_estadual',
    'tipo_fornecedor', 'email', 'telefone', 'contato_responsavel',
    'endereco', 'cidade', 'estado', 'cep', 'chave_pix',
]


def _style_header(ws, ncols: int) -> None:
    fill = PatternFill('solid', fgColor='2563EB')
    font = Font(bold=True, color='FFFFFF')
    align = Alignment(horizontal='center', vertical='center')
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = align
    ws.row_dimensions[1].height = 22


def _autosize(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _str_or_none(value) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return s or None


def _clean_cnpj(value) -> str | None:
    """Remove formatação do CNPJ e retorna só dígitos, ou None se vazio."""
    s = _str_or_none(value)
    if not s:
        return None
    digits = ''.join(c for c in s if c.isdigit())
    return digits or None


def gerar_modelo_fornecedores_xlsx(admin_id: int) -> bytes:
    """Gera o modelo .xlsx de Fornecedores em memória com dados reais do tenant.

    Exporta todos os fornecedores ativos de ``admin_id``. Se não houver nenhum,
    exporta apenas o cabeçalho.
    """
    wb = Workbook()

    ws = wb.active
    ws.title = 'Fornecedores'
    ws.append(FORNECEDOR_HEADERS)
    _style_header(ws, len(FORNECEDOR_HEADERS))

    fornecedores = (
        Fornecedor.query
        .filter_by(admin_id=admin_id, ativo=True)
        .order_by(Fornecedor.razao_social)
        .all()
    )

    for f in fornecedores:
        ws.append([
            f.razao_social or '',
            f.nome_fantasia or '',
            f.cnpj or '',
            f.inscricao_estadual or '',
            f.tipo_fornecedor or 'OUTRO',
            f.email or '',
            f.telefone or '',
            f.contato_responsavel or '',
            f.endereco or '',
            f.cidade or '',
            f.estado or '',
            f.cep or '',
            f.chave_pix or '',
        ])

    _autosize(ws, [30, 25, 18, 16, 20, 28, 16, 22, 35, 18, 8, 12, 30])

    col_tipo = get_column_letter(FORNECEDOR_HEADERS.index('tipo_fornecedor') + 1)
    dv = DataValidation(
        type='list',
        formula1=_TIPOS_FORNECEDOR_DROPDOWN,
        allow_blank=True,
        showDropDown=False,
    )
    dv.sqref = f'{col_tipo}2:{col_tipo}{MAX_DATA_ROWS + 1}'
    ws.add_data_validation(dv)

    inst = wb.create_sheet('Instruções')
    inst.column_dimensions['A'].width = 25
    inst.column_dimensions['B'].width = 90
    inst['A1'] = 'Coluna'
    inst['B1'] = 'Descrição'
    _style_header(inst, 2)

    linhas_inst = [
        ('razao_social',
         'Obrigatório. Razão Social do fornecedor. Também é usada como nome de exibição.'),
        ('nome_fantasia',
         'Opcional. Nome fantasia ou apelido comercial do fornecedor.'),
        ('cnpj',
         'Obrigatório. CNPJ do fornecedor (somente números ou formatado como '
         '00.000.000/0001-00). Usado para identificar duplicatas: '
         'se já existir um fornecedor com o mesmo CNPJ no seu cadastro, '
         'ele será atualizado em vez de duplicado.'),
        ('inscricao_estadual',
         'Opcional. Inscrição Estadual do fornecedor.'),
        ('tipo_fornecedor',
         'Opcional. Tipo do fornecedor. Valores aceitos (use o dropdown): '
         'MATERIAL, PRESTADOR_SERVICO ou OUTRO. Padrão: OUTRO.'),
        ('email',
         'Opcional. Endereço de e-mail para contato.'),
        ('telefone',
         'Opcional. Número de telefone ou celular.'),
        ('contato_responsavel',
         'Opcional. Nome da pessoa responsável pelo contato neste fornecedor.'),
        ('endereco',
         'Opcional. Logradouro, número e complemento.'),
        ('cidade',
         'Opcional. Cidade do fornecedor.'),
        ('estado',
         'Opcional. Sigla do estado (UF), ex.: SP, RJ, MG.'),
        ('cep',
         'Opcional. CEP (somente números ou formatado como 00000-000).'),
        ('chave_pix',
         'Opcional. Chave PIX para pagamentos (CPF, CNPJ, e-mail, celular ou chave aleatória).'),
    ]
    for col, txt in linhas_inst:
        inst.append([col, txt])

    inst.append([])
    inst.append(['Importação',
                 'Fornecedores com CNPJ já existente no seu cadastro serão atualizados '
                 '(upsert). Fornecedores sem CNPJ cadastrado serão criados como novos.'])
    inst.append(['',
                 'Linhas com erro (CNPJ ausente, tipo inválido) são reportadas no '
                 'resumo de preview; as demais linhas continuam sendo processadas.'])
    inst.append(['',
                 'Categorias de fornecedor não são importadas pela planilha '
                 '(devem ser configuradas manualmente após a importação).'])

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def importar_fornecedores_xlsx(stream, admin_id: int) -> dict[str, Any]:
    """Lê um .xlsx e retorna um preview com status por linha.

    Não persiste nada no banco — apenas analisa e retorna:
    ``{'linhas': [...], 'erros': [...], 'novos': int, 'atualizados': int}``

    Cada item em ``linhas`` contém:
    - ``status``: 'novo' | 'atualizar' | 'erro'
    - ``razao_social``, ``cnpj``, ``tipo_fornecedor``, ``cidade``
    - ``dados``: dict com todos os campos prontos para upsert (ou None se erro)
    - ``motivo``: texto descritivo do erro (somente quando status == 'erro')

    Em caso de erro grave (arquivo ilegível, aba ausente), levanta ``ValueError``.
    """
    raw = stream.read(MAX_UPLOAD_BYTES + 1)
    if len(raw) > MAX_UPLOAD_BYTES:
        raise ValueError(
            f'Arquivo Excel maior que o limite permitido '
            f'({MAX_UPLOAD_BYTES // (1024 * 1024)} MB). Reduza o conteúdo e tente novamente.'
        )
    bio = io.BytesIO(raw)
    try:
        wb = load_workbook(bio, data_only=True, read_only=True)
    except Exception as e:
        raise ValueError(f'Não foi possível ler o arquivo Excel: {e}') from e

    ws = wb['Fornecedores'] if 'Fornecedores' in wb.sheetnames else wb.worksheets[0]
    n_rows = max(0, (ws.max_row or 1) - 1)
    if n_rows > MAX_DATA_ROWS:
        raise ValueError(
            f'A planilha excede o limite de {MAX_DATA_ROWS} linhas. '
            'Divida o arquivo em lotes menores antes de importar.'
        )

    header = [(_str_or_none(c.value) or '').lower() for c in ws[1]]
    if not header or header[0] not in ('razao_social', 'razão_social', 'razao social'):
        raise ValueError(
            'Cabeçalho inválido. A primeira coluna deve ser "razao_social". '
            'Baixe novamente o modelo Excel para verificar o formato.'
        )

    def _col(row, name):
        try:
            idx = header.index(name)
        except ValueError:
            return None
        return row[idx] if idx < len(row) else None

    existentes = {
        (f.cnpj or '').strip(): f
        for f in Fornecedor.query.filter_by(admin_id=admin_id).all()
        if f.cnpj
    }

    linhas = []
    erros = []
    novos = 0
    atualizados = 0
    # Track CNPJs seen within THIS batch to handle duplicates inside the same spreadsheet.
    # First occurrence of a new CNPJ → 'novo'; any later occurrence → 'atualizar' (the
    # first row will have created it, so the second row should update it).
    cnpjs_no_lote: set[str] = set()

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(c not in (None, '') for c in row):
            continue

        razao_social = _str_or_none(_col(row, 'razao_social'))
        cnpj = _clean_cnpj(_col(row, 'cnpj'))

        if not razao_social:
            item = {
                'status': 'erro',
                'linha': row_idx,
                'razao_social': '(vazio)',
                'cnpj': cnpj or '',
                'tipo_fornecedor': '',
                'cidade': '',
                'motivo': 'Razão Social é obrigatória.',
                'dados': None,
            }
            linhas.append(item)
            erros.append(item)
            continue

        if not cnpj:
            item = {
                'status': 'erro',
                'linha': row_idx,
                'razao_social': razao_social,
                'cnpj': '',
                'tipo_fornecedor': '',
                'cidade': '',
                'motivo': 'CNPJ é obrigatório.',
                'dados': None,
            }
            linhas.append(item)
            erros.append(item)
            continue

        tipo = (_str_or_none(_col(row, 'tipo_fornecedor')) or 'OUTRO').upper()
        if tipo not in TIPOS_FORNECEDOR_VALIDOS:
            item = {
                'status': 'erro',
                'linha': row_idx,
                'razao_social': razao_social,
                'cnpj': cnpj,
                'tipo_fornecedor': tipo,
                'cidade': '',
                'motivo': (f'Tipo "{tipo}" inválido. '
                           'Use MATERIAL, PRESTADOR_SERVICO ou OUTRO.'),
                'dados': None,
            }
            linhas.append(item)
            erros.append(item)
            continue

        cidade = _str_or_none(_col(row, 'cidade'))
        cep_raw = _str_or_none(_col(row, 'cep'))
        cep = ''.join(c for c in (cep_raw or '') if c.isdigit()) or None

        dados = {
            'razao_social': razao_social,
            'nome_fantasia': _str_or_none(_col(row, 'nome_fantasia')),
            'cnpj': cnpj,
            'inscricao_estadual': _str_or_none(_col(row, 'inscricao_estadual')),
            'tipo_fornecedor': tipo,
            'email': _str_or_none(_col(row, 'email')),
            'telefone': _str_or_none(_col(row, 'telefone')),
            'contato_responsavel': _str_or_none(_col(row, 'contato_responsavel')),
            'endereco': _str_or_none(_col(row, 'endereco')),
            'cidade': cidade,
            'estado': _str_or_none(_col(row, 'estado')),
            'cep': cep,
            'chave_pix': _str_or_none(_col(row, 'chave_pix')),
        }

        existe = existentes.get(cnpj)
        cnpj_ja_no_lote = cnpj in cnpjs_no_lote
        if existe or cnpj_ja_no_lote:
            status = 'atualizar'
            atualizados += 1
        else:
            status = 'novo'
            novos += 1
            cnpjs_no_lote.add(cnpj)

        linhas.append({
            'status': status,
            'linha': row_idx,
            'razao_social': razao_social,
            'cnpj': cnpj,
            'tipo_fornecedor': tipo,
            'cidade': cidade or '',
            'dados': dados,
            'motivo': None,
        })

    return {
        'linhas': linhas,
        'erros': erros,
        'novos': novos,
        'atualizados': atualizados,
    }


def confirmar_importacao_fornecedores(linhas_validas: list[dict], admin_id: int) -> dict[str, int]:
    """Persiste os fornecedores válidos no banco (upsert por CNPJ).

    Recebe apenas as linhas com status 'novo' ou 'atualizar' (sem erros).
    Retorna ``{'criados': int, 'atualizados': int}``.
    """
    existentes = {
        (f.cnpj or '').strip(): f
        for f in Fornecedor.query.filter_by(admin_id=admin_id).all()
        if f.cnpj
    }

    criados = 0
    atualizados = 0

    for item in linhas_validas:
        dados = item.get('dados')
        if not dados:
            continue

        cnpj = dados['cnpj']
        razao_social = dados['razao_social']

        f = existentes.get(cnpj)
        if f is None:
            f = Fornecedor(
                nome=razao_social,
                razao_social=razao_social,
                cnpj=cnpj,
                admin_id=admin_id,
            )
            db.session.add(f)
            # Register immediately so later rows with the same CNPJ in this batch
            # update this record instead of trying to create another duplicate.
            existentes[cnpj] = f
            criados += 1
        else:
            atualizados += 1

        f.nome = razao_social
        f.razao_social = razao_social
        f.nome_fantasia = dados.get('nome_fantasia')
        f.inscricao_estadual = dados.get('inscricao_estadual')
        f.tipo_fornecedor = dados.get('tipo_fornecedor', 'OUTRO')
        f.email = dados.get('email')
        f.telefone = dados.get('telefone')
        f.contato_responsavel = dados.get('contato_responsavel')
        f.endereco = dados.get('endereco')
        f.cidade = dados.get('cidade')
        f.estado = dados.get('estado')
        f.cep = dados.get('cep')
        f.chave_pix = dados.get('chave_pix')
        if not f.ativo:
            f.ativo = True

    db.session.commit()
    return {'criados': criados, 'atualizados': atualizados}
