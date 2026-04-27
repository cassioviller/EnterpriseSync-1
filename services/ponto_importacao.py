"""
Serviço de Importação de Pontos via Excel
Sistema SIGE v9.0
"""

from io import BytesIO
from datetime import datetime, date, time
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from typing import List, Dict, Tuple
import logging
import calendar

logger = logging.getLogger(__name__)


class PontoExcelService:
    """Serviço para gerar planilhas modelo e importar pontos"""
    
    # Definição de tipos de registro disponíveis
    TIPOS_REGISTRO = {
        'TRAB': 'Trabalho Normal',
        'FALTA': 'Falta',
        'FALTA_J': 'Falta Justificada',
        'SAB_TRAB': 'Sábado Trabalhado',
        'SAB_FOLGA': 'Sábado Folga',
        'DOM_TRAB': 'Domingo Trabalhado',
        'DOM_FOLGA': 'Domingo Folga',
        'FER_FOLGA': 'Feriado Folga',
        'FER_TRAB': 'Feriado Trabalhado',
        'ATESTADO': 'Atestado Médico',
        'FERIAS': 'Férias'
    }
    
    @staticmethod
    def gerar_planilha_modelo(funcionarios: List, obras: List = None, mes_referencia: date = None) -> BytesIO:
        """
        Gera planilha Excel modelo com uma aba por funcionário ativo
        
        Args:
            funcionarios: Lista de objetos Funcionario ativos
            mes_referencia: Mês de referência para gerar dias (padrão: mês atual)
        
        Returns:
            BytesIO com o arquivo Excel
        """
        if not mes_referencia:
            mes_referencia = date.today()
        
        # Criar workbook
        wb = Workbook()
        
        # Se não houver funcionários, criar uma planilha de aviso
        if not funcionarios:
            ws = wb.active
            ws.title = "Aviso"
            ws['A1'] = "⚠️ ATENÇÃO"
            ws['A1'].font = Font(bold=True, size=16, color="FF0000")
            ws['A3'] = "Não há funcionários ativos cadastrados no sistema."
            ws['A4'] = "Por favor, cadastre funcionários antes de gerar a planilha de importação."
            ws.column_dimensions['A'].width = 60
            
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return buffer
        
        # Criar aba de Legenda primeiro
        ws_legenda = wb.active
        ws_legenda.title = "📖 LEGENDA"
        PontoExcelService._criar_aba_legenda(ws_legenda, obras or [])
        
        # Estilos
        header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        kpi_header_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Criar uma aba para cada funcionário
        ano = mes_referencia.year
        mes = mes_referencia.month
        _, ultimo_dia = calendar.monthrange(ano, mes)
        
        for func in funcionarios:
            # Nome da aba: "Código - Nome" (limitado a 31 caracteres do Excel)
            sheet_name = f"{func.codigo} - {func.nome}"[:31]
            ws = wb.create_sheet(title=sheet_name)
            
            # Cabeçalho informativo
            ws['A1'] = f"REGISTRO DE PONTO - {func.nome.upper()}"
            ws['A1'].font = Font(bold=True, size=14)
            ws.merge_cells('A1:I1')
            
            ws['A2'] = f"Código: {func.codigo}"
            ws['B2'] = f"CPF: {func.cpf}"
            ws['D2'] = f"Cargo: {func.funcao_ref.nome if func.funcao_ref else 'N/A'}"
            
            # Horário padrão do funcionário (para cálculos)
            horario_entrada_padrao = "08:00"
            horario_saida_padrao = "17:00"
            if func.horario_trabalho:
                horario_entrada_padrao = func.horario_trabalho.entrada.strftime('%H:%M')
                horario_saida_padrao = func.horario_trabalho.saida.strftime('%H:%M')
            
            ws['A3'] = f"⏰ Horário Padrão:"
            ws['A3'].font = Font(bold=True, size=11, color="1565C0")
            ws['B3'] = f"Entrada: {horario_entrada_padrao}"
            ws['D3'] = f"Saída: {horario_saida_padrao}"
            ws['F3'] = f"(Usado para calcular atrasos e horas extras)"
            ws['F3'].font = Font(italic=True, size=9, color="666666")
            
            # Armazenar horários em células ocultas para fórmulas (como valores de tempo, não strings)
            # Converter strings HH:MM para valores de tempo do Excel
            from datetime import datetime as dt
            entrada_time = dt.strptime(horario_entrada_padrao, '%H:%M').time()
            saida_time = dt.strptime(horario_saida_padrao, '%H:%M').time()
            
            ws['Z1'] = entrada_time
            ws['Z1'].number_format = 'HH:MM'
            ws['Z2'] = saida_time
            ws['Z2'].number_format = 'HH:MM'
            
            # Linha em branco
            current_row = 5
            
            # Cabeçalho da tabela (ADICIONADAS colunas "Tipo", "Obra ID", "Min. Atraso" e "Min. HE")
            headers = ['Data', 'Tipo', 'Obra ID', 'Entrada', 'Saída', 'Início Almoço', 'Fim Almoço', 'Min. Atraso', 'Min. HE']
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
                
                # Destaque especial para colunas calculadas
                if header in ['Min. Atraso', 'Min. HE']:
                    cell.fill = PatternFill(start_color="FF6F00", end_color="FF6F00", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF", size=10)
            
            # Criar lista de tipos para dropdown
            tipos_lista = ','.join(PontoExcelService.TIPOS_REGISTRO.keys())
            dv_tipos = DataValidation(type="list", formula1=f'"{tipos_lista}"', allow_blank=True)
            dv_tipos.error = 'Selecione um tipo válido da lista'
            dv_tipos.errorTitle = 'Tipo Inválido'
            ws.add_data_validation(dv_tipos)
            
            # Criar lista de obras para dropdown
            dv_obras = None
            if obras:
                obras_ids = ','.join(str(obra.id) for obra in obras)
                dv_obras = DataValidation(type="list", formula1=f'"{obras_ids}"', allow_blank=True)
                dv_obras.error = 'Selecione um ID de obra válido da aba LEGENDA'
                dv_obras.errorTitle = 'Obra Inválida'
                ws.add_data_validation(dv_obras)
            
            # Linha inicial dos dados
            data_start_row = current_row + 1
            
            # Gerar dias do mês
            current_row += 1
            for dia in range(1, ultimo_dia + 1):
                data_dia = date(ano, mes, dia)
                
                # Data (coluna A)
                cell = ws.cell(row=current_row, column=1, value=data_dia.strftime('%d/%m/%Y'))
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
                
                # Tipo (coluna B com dropdown)
                cell = ws.cell(row=current_row, column=2, value='TRAB')
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
                dv_tipos.add(cell)  # Adicionar validação de dados
                
                # Obra ID (coluna C com dropdown)
                cell = ws.cell(row=current_row, column=3, value='')
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
                if dv_obras:
                    dv_obras.add(cell)  # Adicionar validação de dados
                
                # Horários vazios (colunas D, E, F, G)
                for col in range(4, 8):
                    cell = ws.cell(row=current_row, column=col, value='')
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center')
                
                # Coluna H: Minutos de Atraso (calculado automaticamente)
                # Fórmula: SE entrada > entrada_padrão, calcular diferença em minutos, senão 0
                formula_atraso = f'=IF(AND(D{current_row}<>"",D{current_row}>$Z$1),(D{current_row}-$Z$1)*1440,0)'
                cell = ws.cell(row=current_row, column=8, value=formula_atraso)
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
                cell.font = Font(color="FF6F00", bold=True)
                
                # Coluna I: Minutos de Horas Extras (calculado automaticamente)
                # Fórmula: SE saída > saída_padrão, calcular diferença em minutos, senão 0
                formula_he = f'=IF(AND(E{current_row}<>"",E{current_row}>$Z$2),(E{current_row}-$Z$2)*1440,0)'
                cell = ws.cell(row=current_row, column=9, value=formula_he)
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
                cell.font = Font(color="1565C0", bold=True)
                
                current_row += 1
            
            data_end_row = current_row - 1
            
            # SEÇÃO DE KPIs (após os dados)
            current_row += 2
            PontoExcelService._criar_secao_kpis(
                ws, current_row, data_start_row, data_end_row, 
                kpi_header_fill, header_font, border
            )
            
            # Instruções no final
            current_row += 15  # Pular as linhas dos KPIs
            ws[f'A{current_row}'] = "INSTRUÇÕES:"
            ws[f'A{current_row}'].font = Font(bold=True, color="FF0000")
            
            current_row += 1
            instrucoes = [
                "• Preencha os horários no formato HH:MM (exemplo: 08:00)",
                "• Selecione o TIPO do dia na coluna B (use o dropdown)",
                "• Preencha o OBRA ID na coluna C (veja IDs na aba 📖 LEGENDA)",
                "• Deixe horários em branco para dias de falta/folga",
                "• Os KPIs serão calculados automaticamente",
                "• Não altere as datas ou fórmulas",
                "• Não exclua ou adicione linhas na tabela de dados"
            ]
            for instrucao in instrucoes:
                ws[f'A{current_row}'] = instrucao
                current_row += 1
            
            # Ajustar largura das colunas
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 15
            ws.column_dimensions['G'].width = 15
            ws.column_dimensions['H'].width = 15
            ws.column_dimensions['I'].width = 15
            
            # Ocultar coluna Z (horários padrão)
            ws.column_dimensions['Z'].hidden = True
        
        # Salvar em BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer
    
    @staticmethod
    def _criar_aba_legenda(ws, obras):
        """Cria aba de legenda com dicionário de códigos e lista de obras"""
        # Título
        ws['A1'] = "📖 DICIONÁRIO - TIPOS DE PONTO E OBRAS"
        ws['A1'].font = Font(bold=True, size=16, color="1565C0")
        ws.merge_cells('A1:C1')
        
        ws['A3'] = "Use esta legenda para preencher as colunas 'Tipo' e 'Obra ID' nas abas de cada funcionário"
        ws['A3'].font = Font(italic=True, size=11)
        ws.merge_cells('A3:C3')
        
        # Cabeçalho da tabela
        row = 5
        headers = ['CÓDIGO', 'DESCRIÇÃO', 'QUANDO USAR']
        header_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Dados da legenda
        row += 1
        legenda_dados = [
            ('TRAB', 'Trabalho Normal', 'Dia útil trabalhado normalmente'),
            ('FALTA', 'Falta', 'Falta não justificada (gera desconto)'),
            ('FALTA_J', 'Falta Justificada', 'Falta com justificativa (atestado, etc)'),
            ('SAB_TRAB', 'Sábado Trabalhado', 'Trabalho em sábado (hora extra 50%)'),
            ('SAB_FOLGA', 'Sábado Folga', 'Sábado de folga/descanso'),
            ('DOM_TRAB', 'Domingo Trabalhado', 'Trabalho em domingo (hora extra 100%)'),
            ('DOM_FOLGA', 'Domingo Folga', 'Domingo de folga/descanso'),
            ('FER_FOLGA', 'Feriado Folga', 'Feriado não trabalhado'),
            ('FER_TRAB', 'Feriado Trabalhado', 'Trabalho em feriado (hora extra 100%)'),
            ('ATESTADO', 'Atestado Médico', 'Afastamento médico'),
            ('FERIAS', 'Férias', 'Período de férias')
        ]
        
        for codigo, desc, quando in legenda_dados:
            ws.cell(row=row, column=1, value=codigo).border = border
            ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=1).font = Font(bold=True)
            
            ws.cell(row=row, column=2, value=desc).border = border
            ws.cell(row=row, column=3, value=quando).border = border
            row += 1
        
        # Instruções de uso
        row += 2
        ws[f'A{row}'] = "💡 DICA: Como usar a legenda"
        ws[f'A{row}'].font = Font(bold=True, size=12, color="FF6F00")
        ws.merge_cells(f'A{row}:C{row}')
        
        row += 1
        dicas = [
            "1. Clique na célula da coluna 'Tipo' de qualquer funcionário",
            "2. Aparecerá uma seta dropdown - clique nela",
            "3. Selecione o código desejado da lista",
            "4. OU copie o código desta legenda e cole diretamente",
            "",
            "⚠️ IMPORTANTE:",
            "• Deixe horários em BRANCO para faltas/folgas/férias",
            "• Para dias trabalhados, preencha Entrada e Saída",
            "• Para horas extras, preencha normalmente e use o tipo correto (SAB_TRAB, DOM_TRAB, FER_TRAB)"
        ]
        
        for dica in dicas:
            ws[f'A{row}'] = dica
            if dica.startswith('⚠️'):
                ws[f'A{row}'].font = Font(bold=True, color="FF0000")
            ws.merge_cells(f'A{row}:C{row}')
            row += 1
        
        # SEÇÃO DE OBRAS
        row += 3
        ws[f'A{row}'] = "🏗️ OBRAS DISPONÍVEIS"
        ws[f'A{row}'].font = Font(bold=True, size=14, color="2E7D32")
        ws.merge_cells(f'A{row}:C{row}')
        
        row += 1
        ws[f'A{row}'] = "Use os IDs abaixo para preencher a coluna 'Obra ID' nas abas de funcionários"
        ws[f'A{row}'].font = Font(italic=True, size=11)
        ws.merge_cells(f'A{row}:C{row}')
        
        row += 2
        # Cabeçalho da tabela de obras
        headers_obras = ['ID', 'NOME DA OBRA', 'CLIENTE']
        for col, header in enumerate(headers_obras, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        row += 1
        # Dados das obras
        if obras:
            for obra in obras:
                ws.cell(row=row, column=1, value=obra.id).border = border
                ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
                ws.cell(row=row, column=1).font = Font(bold=True, size=12, color="2E7D32")
                
                ws.cell(row=row, column=2, value=obra.nome).border = border
                ws.cell(row=row, column=3, value=obra.cliente_nome_efetivo or 'N/A').border = border
                row += 1
        else:
            ws.cell(row=row, column=1, value="Nenhuma obra cadastrada").border = border
            ws.merge_cells(f'A{row}:C{row}')
            ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=1).font = Font(italic=True, color="999999")
            row += 1
        
        # Instruções sobre Obra ID
        row += 2
        ws[f'A{row}'] = "💡 DICA: Copie o ID da obra e cole na coluna 'Obra ID' do funcionário"
        ws[f'A{row}'].font = Font(bold=True, size=11, color="FF6F00")
        ws.merge_cells(f'A{row}:C{row}')
        
        # Ajustar larguras
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 35
    
    @staticmethod
    def _criar_secao_kpis(ws, start_row, data_start_row, data_end_row, kpi_fill, header_font, border):
        """Cria seção de KPIs com fórmulas Excel automáticas"""
        row = start_row
        
        # Título da seção
        ws[f'A{row}'] = "📊 INDICADORES DO MÊS (KPIs - Calculados Automaticamente)"
        ws[f'A{row}'].font = Font(bold=True, size=13, color="1565C0")
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 2
        
        # Cabeçalho dos KPIs
        ws.cell(row=row, column=1, value="INDICADOR").fill = kpi_fill
        ws.cell(row=row, column=1).font = header_font
        ws.cell(row=row, column=1).border = border
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
        
        ws.cell(row=row, column=2, value="VALOR").fill = kpi_fill
        ws.cell(row=row, column=2).font = header_font
        ws.cell(row=row, column=2).border = border
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
        
        row += 1
        
        # KPI 1: Total de Dias Trabalhados
        ws.cell(row=row, column=1, value="Dias Trabalhados").border = border
        ws.cell(row=row, column=2, value=f'=COUNTIF(B{data_start_row}:B{data_end_row},"TRAB")+COUNTIF(B{data_start_row}:B{data_end_row},"SAB_TRAB")+COUNTIF(B{data_start_row}:B{data_end_row},"DOM_TRAB")+COUNTIF(B{data_start_row}:B{data_end_row},"FER_TRAB")').border = border
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
        row += 1
        
        # KPI 2: Total de Faltas
        ws.cell(row=row, column=1, value="Total de Faltas").border = border
        ws.cell(row=row, column=2, value=f'=COUNTIF(B{data_start_row}:B{data_end_row},"FALTA")').border = border
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=2).font = Font(color="FF0000", bold=True)
        row += 1
        
        # KPI 3: Faltas Justificadas
        ws.cell(row=row, column=1, value="Faltas Justificadas").border = border
        ws.cell(row=row, column=2, value=f'=COUNTIF(B{data_start_row}:B{data_end_row},"FALTA_J")+COUNTIF(B{data_start_row}:B{data_end_row},"ATESTADO")').border = border
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
        row += 1
        
        # KPI 4: Sábados Trabalhados
        ws.cell(row=row, column=1, value="Sábados Trabalhados (HE 50%)").border = border
        ws.cell(row=row, column=2, value=f'=COUNTIF(B{data_start_row}:B{data_end_row},"SAB_TRAB")').border = border
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=2).font = Font(color="0066CC", bold=True)
        row += 1
        
        # KPI 5: Domingos/Feriados Trabalhados
        ws.cell(row=row, column=1, value="Domingos/Feriados Trab (HE 100%)").border = border
        ws.cell(row=row, column=2, value=f'=COUNTIF(B{data_start_row}:B{data_end_row},"DOM_TRAB")+COUNTIF(B{data_start_row}:B{data_end_row},"FER_TRAB")').border = border
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=2).font = Font(color="0066CC", bold=True)
        row += 1
        
        # KPI 6: Dias de Férias
        ws.cell(row=row, column=1, value="Dias de Férias").border = border
        ws.cell(row=row, column=2, value=f'=COUNTIF(B{data_start_row}:B{data_end_row},"FERIAS")').border = border
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
        row += 1
        
        # KPI 7: Total Minutos de Atraso
        ws.cell(row=row, column=1, value="Total Minutos de Atraso").border = border
        ws.cell(row=row, column=2, value=f'=SUM(H{data_start_row}:H{data_end_row})').border = border
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=2).font = Font(color="FF0000", bold=True)
        row += 1
        
        # KPI 8: Total Minutos de Horas Extras
        ws.cell(row=row, column=1, value="Total Minutos de Horas Extras").border = border
        ws.cell(row=row, column=2, value=f'=SUM(I{data_start_row}:I{data_end_row})').border = border
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=2).font = Font(color="FF6F00", bold=True)
        row += 1
        
        # KPI 9: Total Horas Extras (em formato HH:MM)
        ws.cell(row=row, column=1, value="Horas Extras (HH:MM)").border = border
        ws.cell(row=row, column=2, value=f'=TEXT(SUM(I{data_start_row}:I{data_end_row})/1440,"[HH]:MM")').border = border
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=2).font = Font(color="0066CC", bold=True)
        row += 1
        
        # KPI 10: Total Atrasos (em formato HH:MM)
        ws.cell(row=row, column=1, value="Atrasos (HH:MM)").border = border
        ws.cell(row=row, column=2, value=f'=TEXT(SUM(H{data_start_row}:H{data_end_row})/1440,"[HH]:MM")').border = border
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=2).font = Font(color="CC0000", bold=True)
        row += 1
        
        # Nota explicativa
        row += 2
        ws[f'A{row}'] = "💡 Os valores acima são calculados automaticamente conforme você preenche os dados"
        ws[f'A{row}'].font = Font(italic=True, size=10, color="666666")
        ws.merge_cells(f'A{row}:D{row}')
        
        # Ajustar larguras
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 15
    
    @staticmethod
    def validar_e_importar(
        excel_file: BytesIO,
        funcionarios_map: Dict[str, int],
        admin_id: int
    ) -> Tuple[List[Dict], List[str]]:
        """
        Valida e prepara dados do Excel para importação
        
        Args:
            excel_file: Arquivo Excel em BytesIO
            funcionarios_map: Dict {codigo_funcionario: funcionario_id}
            admin_id: ID do admin (multi-tenant)
        
        Returns:
            Tupla (registros_validos, erros)
        """
        registros_validos = []
        erros = []
        
        try:
            wb = load_workbook(excel_file)
        except Exception as e:
            erros.append(f"Erro ao ler arquivo Excel: {str(e)}")
            return [], erros
        
        # Processar cada aba (uma por funcionário)
        for sheet_name in wb.sheetnames:
            try:
                # Pular abas especiais (Legenda, Aviso, etc)
                if sheet_name in ["📖 LEGENDA", "Aviso"] or ' - ' not in sheet_name:
                    continue  # Não é aba de funcionário, pular silenciosamente
                
                ws = wb[sheet_name]
                
                codigo_func = sheet_name.split(' - ')[0].strip()
                
                # Verificar se funcionário existe
                if codigo_func not in funcionarios_map:
                    erros.append(f"Aba '{sheet_name}': funcionário com código '{codigo_func}' não encontrado ou inativo")
                    continue
                
                funcionario_id = funcionarios_map[codigo_func]
                
                # Encontrar linha do cabeçalho (procurar por "Data")
                header_row = None
                for row in range(1, 10):  # Procurar nas primeiras 10 linhas
                    if ws.cell(row=row, column=1).value == 'Data':
                        header_row = row
                        break
                
                if not header_row:
                    erros.append(f"Aba '{sheet_name}': cabeçalho não encontrado")
                    continue
                
                # Processar dados a partir da linha após o cabeçalho
                for row in range(header_row + 1, ws.max_row + 1):
                    data_cell = ws.cell(row=row, column=1).value
                    
                    # Pular linhas vazias
                    if not data_cell:
                        continue
                    
                    # Converter data
                    try:
                        if isinstance(data_cell, datetime):
                            data_ponto = data_cell.date()
                        elif isinstance(data_cell, date):
                            data_ponto = data_cell
                        elif isinstance(data_cell, str):
                            data_ponto = datetime.strptime(data_cell, '%d/%m/%Y').date()
                        else:
                            erros.append(f"Aba '{sheet_name}', linha {row}: formato de data inválido '{data_cell}'")
                            continue
                    except Exception as e:
                        erros.append(f"Aba '{sheet_name}', linha {row}: erro ao converter data '{data_cell}': {str(e)}")
                        continue
                    
                    # Ler tipo de registro (coluna B)
                    tipo_registro = ws.cell(row=row, column=2).value
                    if tipo_registro:
                        tipo_registro = str(tipo_registro).strip().upper()
                    else:
                        tipo_registro = 'TRAB'  # Default
                    
                    # Ler obra_id (coluna C)
                    obra_id_cell = ws.cell(row=row, column=3).value
                    obra_id = None
                    if obra_id_cell:
                        try:
                            obra_id = int(obra_id_cell)
                        except (ValueError, TypeError):
                            logger.warning(f"Aba '{sheet_name}', linha {row}: ID de obra inválido '{obra_id_cell}', ignorando")
                    
                    # Ler horários (AJUSTADO para colunas D, E, F, G)
                    entrada = PontoExcelService._parse_time(ws.cell(row=row, column=4).value)
                    saida = PontoExcelService._parse_time(ws.cell(row=row, column=5).value)
                    almoco_inicio = PontoExcelService._parse_time(ws.cell(row=row, column=6).value)
                    almoco_fim = PontoExcelService._parse_time(ws.cell(row=row, column=7).value)
                    
                    # Para tipos de folga/falta, não requer horários
                    tipos_sem_horario = ['FALTA', 'FALTA_J', 'SAB_FOLGA', 'DOM_FOLGA', 
                                         'FER_FOLGA', 'ATESTADO', 'FERIAS']
                    
                    # Pular se tipo sem horário e realmente não tem horários
                    if tipo_registro in tipos_sem_horario and not any([entrada, saida, almoco_inicio, almoco_fim]):
                        # Registrar como dia sem trabalho (será tratado na importação)
                        registros_validos.append({
                            'funcionario_id': funcionario_id,
                            'data': data_ponto,
                            'tipo_registro': tipo_registro,
                            'obra_id': obra_id,
                            'admin_id': admin_id
                        })
                        continue
                    
                    # Pular se todos os horários estiverem vazios para tipo de trabalho
                    if not any([entrada, saida, almoco_inicio, almoco_fim]):
                        continue
                    
                    # Validar lógica de horários
                    if entrada and saida and entrada >= saida:
                        erros.append(
                            f"Aba '{sheet_name}', linha {row}, data {data_ponto.strftime('%d/%m/%Y')}: "
                            f"Horário de entrada ({entrada}) deve ser menor que saída ({saida})"
                        )
                        continue
                    
                    if almoco_inicio and almoco_fim and almoco_inicio >= almoco_fim:
                        erros.append(
                            f"Aba '{sheet_name}', linha {row}, data {data_ponto.strftime('%d/%m/%Y')}: "
                            f"Início do almoço ({almoco_inicio}) deve ser menor que fim ({almoco_fim})"
                        )
                        continue
                    
                    # Registro válido
                    registros_validos.append({
                        'funcionario_id': funcionario_id,
                        'data': data_ponto,
                        'hora_entrada': entrada,
                        'hora_saida': saida,
                        'hora_almoco_saida': almoco_inicio,
                        'hora_almoco_retorno': almoco_fim,
                        'obra_id': obra_id,
                        'admin_id': admin_id,
                        'tipo_registro': tipo_registro
                    })
                    
            except Exception as e:
                erros.append(f"Erro ao processar aba '{sheet_name}': {str(e)}")
                logger.error(f"Erro ao processar aba '{sheet_name}': {e}", exc_info=True)
        
        return registros_validos, erros
    
    @staticmethod
    def _parse_time(value) -> time | None:
        """Converte valor de célula Excel para objeto time"""
        if not value:
            return None
        
        try:
            if isinstance(value, time):
                return value
            elif isinstance(value, datetime):
                return value.time()
            elif isinstance(value, str):
                # Tentar formatos HH:MM ou HH:MM:SS
                value = value.strip()
                if ':' in value:
                    parts = value.split(':')
                    if len(parts) >= 2:
                        hora = int(parts[0])
                        minuto = int(parts[1])
                        segundo = int(parts[2]) if len(parts) > 2 else 0
                        return time(hora, minuto, segundo)
            elif isinstance(value, (int, float)):
                # Excel pode armazenar tempo como fração de dia
                total_seconds = int(value * 86400)
                hours = total_seconds // 3600
                minutes = (total_seconds % 3600) // 60
                seconds = total_seconds % 60
                return time(hours, minutes, seconds)
        except Exception as e:
            logger.warning(f"Erro ao converter horário '{value}': {e}")
        
        return None
