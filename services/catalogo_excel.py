"""Task #23 — Modelos Excel para download e importação dos catálogos.

Dois catálogos são contemplados:

- **Insumos** (`Insumo` + `PrecoBaseInsumo`): gera um modelo `.xlsx` com
  cabeçalho, exemplos e aba de instruções; importa um arquivo preenchido
  fazendo upsert por ``(admin_id, nome)``. Quando a coluna ``preco_base``
  vem preenchida, cria/atualiza a vigência atual em ``PrecoBaseInsumo``
  seguindo o mesmo padrão da rota ``catalogo.insumo_novo_preco``.

- **Templates de Cronograma** (`CronogramaTemplate` +
  `CronogramaTemplateItem`): gera um modelo `.xlsx` com aba de cabeçalho do
  template, aba de itens (incluindo ``parent_ordem`` para hierarquia) e
  aba de instruções. A importação faz upsert do template por
  ``(admin_id, nome)``, **substituindo** os itens (apaga os antigos e recria
  na ordem da planilha resolvendo ``parent_item_id`` via mapa
  ``ordem -> id``).

Todas as funções respeitam ``admin_id`` (multi-tenant). Erros de validação
não corrompem o banco — em caso de falha grave, faz rollback. Validações
linha-a-linha não derrubam o import; a linha é rejeitada com motivo.
"""
from __future__ import annotations

import io
import logging
from datetime import date, timedelta
from decimal import Decimal, InvalidOperation
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from models import (
    CronogramaTemplate,
    CronogramaTemplateItem,
    Insumo,
    PrecoBaseInsumo,
    Servico,
    db,
)

logger = logging.getLogger(__name__)


TIPOS_INSUMO_VALIDOS = {'MATERIAL', 'MAO_OBRA', 'EQUIPAMENTO'}
RESPONSAVEIS_VALIDOS = {'empresa', 'terceiros'}

# Limites de upload para mitigar DoS / zip-bomb / OOM em cargas grandes.
MAX_UPLOAD_BYTES = 5 * 1024 * 1024  # 5 MB
MAX_DATA_ROWS = 5000  # Linhas de dados (excluindo cabeçalho)


def _ler_workbook_seguro(arquivo) -> Workbook:
    """Lê um arquivo `.xlsx` com proteções contra abuso.

    - Limita o tamanho do upload a ``MAX_UPLOAD_BYTES``.
    - Usa ``read_only=True`` para reduzir o consumo de memória.

    Levanta ``ValueError`` se o arquivo for grande demais ou ilegível.
    """
    raw = arquivo.read(MAX_UPLOAD_BYTES + 1)
    if len(raw) > MAX_UPLOAD_BYTES:
        raise ValueError(
            f'Arquivo Excel maior que o limite permitido '
            f'({MAX_UPLOAD_BYTES // (1024 * 1024)} MB). Reduza o conteúdo '
            'e tente novamente.'
        )
    bio = io.BytesIO(raw)
    try:
        return load_workbook(bio, data_only=True, read_only=True)
    except Exception as e:
        raise ValueError(f'Não foi possível ler o arquivo Excel: {e}') from e


def _checar_max_linhas(quantidade: int, contexto: str = '') -> None:
    """Aborta se exceder ``MAX_DATA_ROWS``."""
    if quantidade > MAX_DATA_ROWS:
        sufixo = f' (aba {contexto})' if contexto else ''
        raise ValueError(
            f'A planilha excede o limite de {MAX_DATA_ROWS} linhas{sufixo}. '
            'Divida o arquivo em lotes menores antes de importar.'
        )


# ──────────────────────────────────────────────────────────────────────
# Helpers comuns
# ──────────────────────────────────────────────────────────────────────
def _decimal_or_none(value) -> Decimal | None:
    """Converte string/num pt-BR/EN para Decimal. Retorna None se vazio
    ou inválido."""
    if value is None or value == '':
        return None
    s = str(value).strip()
    if not s:
        return None
    s = s.replace('R$', '').replace('r$', '').replace(' ', '').replace('\xa0', '')
    if ',' in s and '.' in s:
        s = s.replace('.', '').replace(',', '.')
    elif ',' in s:
        s = s.replace(',', '.')
    try:
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return None


def _str_or_none(value) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return s or None


def _int_or_none(value) -> int | None:
    if value is None or value == '':
        return None
    try:
        return int(float(str(value).replace(',', '.')))
    except (ValueError, TypeError):
        return None


def _style_header(ws, ncols: int):
    fill = PatternFill('solid', fgColor='2563EB')
    font = Font(bold=True, color='FFFFFF')
    align = Alignment(horizontal='center', vertical='center')
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = align
    ws.row_dimensions[1].height = 22


def _autosize(ws, widths: list[int]):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ──────────────────────────────────────────────────────────────────────
# INSUMOS — modelo
# ──────────────────────────────────────────────────────────────────────
INSUMO_HEADERS = [
    'nome', 'tipo', 'unidade', 'descricao', 'coeficiente_padrao', 'preco_base',
]


def gerar_modelo_insumos_xlsx() -> bytes:
    """Gera o modelo `.xlsx` de Insumos em memória.

    Retorna bytes prontos para serem enviados como `send_file`.
    """
    wb = Workbook()

    # Aba 1 — Insumos
    ws = wb.active
    ws.title = 'Insumos'
    ws.append(INSUMO_HEADERS)
    _style_header(ws, len(INSUMO_HEADERS))

    exemplos = [
        ['Pedreiro', 'MAO_OBRA', 'h', 'Hora de pedreiro contratado', '1', '35,00'],
        ['Cimento CP-II 50kg', 'MATERIAL', 'sc', 'Saco de cimento Portland', '1', '38,50'],
        ['Betoneira 400L', 'EQUIPAMENTO', 'h', 'Hora de uso de betoneira', '1', '12,00'],
    ]
    for linha in exemplos:
        ws.append(linha)
    _autosize(ws, [28, 14, 10, 36, 18, 14])

    # Aba 2 — Instruções
    inst = wb.create_sheet('Instruções')
    inst.column_dimensions['A'].width = 28
    inst.column_dimensions['B'].width = 90
    inst['A1'] = 'Coluna'
    inst['B1'] = 'Descrição'
    _style_header(inst, 2)

    linhas_inst = [
        ('nome', 'Obrigatório. Nome único do insumo dentro do seu cadastro. '
                 'Se já existir um insumo com o mesmo nome, ele será atualizado.'),
        ('tipo', 'Obrigatório. Valores aceitos: MATERIAL, MAO_OBRA ou '
                 'EQUIPAMENTO (sem acento, em maiúsculas).'),
        ('unidade', 'Obrigatório. Unidade de medida do insumo. Ex.: h '
                    '(hora), kg, m, m², m³, sc (saco), un, l, etc.'),
        ('descricao', 'Opcional. Texto livre para descrever o insumo.'),
        ('coeficiente_padrao', 'Opcional. Coeficiente padrão de consumo '
                               'sugerido ao adicionar este insumo na composição '
                               'de um serviço. Padrão: 1. Aceita "1,5" ou "1.5".'),
        ('preco_base', 'Opcional. Preço unitário em reais. Se preenchido, '
                       'será criada uma nova vigência de preço com início '
                       'na data de hoje. Aceita "35,00", "35.00" ou "1.234,56". '
                       'Se vazio, o preço atual do insumo (caso exista) é '
                       'mantido.'),
    ]
    for col, txt in linhas_inst:
        inst.append([col, txt])

    inst.append([])
    inst.append(['Importação', 'Insumos com nome igual a outro já existente '
                 'do seu cadastro (ignorando maiúsculas/minúsculas) serão '
                 'atualizados em vez de duplicados.'])
    inst.append(['', 'Linhas com erro são reportadas no resumo após o import; '
                 'as demais linhas continuam sendo processadas.'])
    inst.append(['', 'Não é necessário deixar a aba "Insumos" exatamente '
                 'igual ao modelo — a ordem das colunas é o que importa.'])

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


# ──────────────────────────────────────────────────────────────────────
# INSUMOS — importação
# ──────────────────────────────────────────────────────────────────────
def importar_insumos_xlsx(arquivo, admin_id: int) -> dict[str, Any]:
    """Lê um `.xlsx` enviado e faz upsert dos insumos por (admin_id, nome).

    Retorna ``{'created': int, 'updated': int, 'rejected': [{linha, motivo}],
    'preco_atualizado': int}``.

    Em caso de erro grave (arquivo corrompido, aba ausente, arquivo grande
    demais), levanta ``ValueError`` sem tocar no banco.
    """
    wb = _ler_workbook_seguro(arquivo)

    # Aceita primeira aba ou aba "Insumos"
    ws = wb['Insumos'] if 'Insumos' in wb.sheetnames else wb.worksheets[0]
    _checar_max_linhas(max(0, (ws.max_row or 1) - 1), 'Insumos')

    # Validação do cabeçalho
    header = [(_str_or_none(c.value) or '').lower() for c in ws[1]]
    if not header or header[0] != 'nome':
        raise ValueError(
            'Cabeçalho inválido. A primeira coluna deve ser "nome". '
            'Baixe novamente o modelo Excel para verificar o formato.'
        )

    def _col(row, name):
        try:
            idx = header.index(name)
        except ValueError:
            return None
        return row[idx] if idx < len(row) else None

    created = 0
    updated = 0
    preco_atualizado = 0
    rejected: list[dict[str, Any]] = []
    servicos_a_recalcular: set[int] = set()
    hoje = date.today()

    # Cache de insumos existentes do tenant para upsert por nome (case-insensitive)
    existentes = {
        (i.nome or '').strip().lower(): i
        for i in Insumo.query.filter_by(admin_id=admin_id).all()
    }

    try:
        for row_idx, row in enumerate(
            ws.iter_rows(min_row=2, values_only=True), start=2
        ):
            # Pula linhas totalmente vazias
            if not any(c not in (None, '') for c in row):
                continue

            nome = _str_or_none(_col(row, 'nome'))
            tipo = (_str_or_none(_col(row, 'tipo')) or '').upper()
            unidade = _str_or_none(_col(row, 'unidade'))

            if not nome:
                rejected.append({'linha': row_idx, 'motivo': 'Nome é obrigatório.'})
                continue
            if tipo not in TIPOS_INSUMO_VALIDOS:
                rejected.append({
                    'linha': row_idx,
                    'motivo': (f'Tipo "{tipo or "(vazio)"}" inválido. '
                               'Use MATERIAL, MAO_OBRA ou EQUIPAMENTO.'),
                })
                continue
            if not unidade:
                rejected.append({
                    'linha': row_idx,
                    'motivo': 'Unidade é obrigatória.',
                })
                continue

            descricao = _str_or_none(_col(row, 'descricao'))
            coef_dec = _decimal_or_none(_col(row, 'coeficiente_padrao'))
            if coef_dec is None:
                coef_dec = Decimal('1')
            if coef_dec < 0:
                rejected.append({
                    'linha': row_idx,
                    'motivo': 'coeficiente_padrao não pode ser negativo.',
                })
                continue

            preco_dec = _decimal_or_none(_col(row, 'preco_base'))
            if preco_dec is not None and preco_dec < 0:
                rejected.append({
                    'linha': row_idx,
                    'motivo': 'preco_base não pode ser negativo.',
                })
                continue

            chave = nome.strip().lower()
            ins = existentes.get(chave)
            if ins is None:
                ins = Insumo(
                    admin_id=admin_id,
                    nome=nome.strip(),
                    tipo=tipo,
                    unidade=unidade,
                    descricao=descricao,
                    coeficiente_padrao=coef_dec,
                )
                db.session.add(ins)
                db.session.flush()
                existentes[chave] = ins
                created += 1
            else:
                ins.nome = nome.strip()
                ins.tipo = tipo
                ins.unidade = unidade
                # Só sobrescreve descrição se vier preenchida (não apaga o
                # texto existente quando a coluna está vazia na planilha).
                if descricao is not None:
                    ins.descricao = descricao
                ins.coeficiente_padrao = coef_dec
                # Reativa caso esteja desativado (via "excluir" suave)
                if not ins.ativo:
                    ins.ativo = True
                updated += 1

            # Atualizar vigência de preço, se vier
            if preco_dec is not None and preco_dec > 0:
                _aplicar_nova_vigencia_preco(ins, admin_id, preco_dec, hoje)
                preco_atualizado += 1
                # Recalcular serviços que usam esse insumo
                for c in ins.composicoes:
                    servicos_a_recalcular.add(c.servico_id)

        db.session.flush()

        # Recalcular serviços afetados pelos novos preços
        if servicos_a_recalcular:
            from services.orcamento_service import recalcular_servico_preco
            for sid in servicos_a_recalcular:
                s = Servico.query.get(sid)
                if s:
                    recalcular_servico_preco(s)

        db.session.commit()
    except Exception:
        db.session.rollback()
        raise

    return {
        'created': created,
        'updated': updated,
        'rejected': rejected,
        'preco_atualizado': preco_atualizado,
        'servicos_recalculados': len(servicos_a_recalcular),
    }


def _aplicar_nova_vigencia_preco(
    ins: Insumo, admin_id: int, valor: Decimal, vigencia: date
) -> None:
    """Encerra a vigência atual de preço e abre uma nova.

    Espelha o comportamento de ``catalogo.insumo_novo_preco``: se já existe
    um preço aberto, é fechado em D-1 (ou no próprio dia, em caso de import
    retroativo). Cria nova linha em ``PrecoBaseInsumo``.
    """
    atual = (
        PrecoBaseInsumo.query
        .filter_by(insumo_id=ins.id, vigencia_fim=None)
        .order_by(PrecoBaseInsumo.vigencia_inicio.desc())
        .first()
    )
    if atual:
        # Se o preço existente é igual ao novo e cobre hoje, não cria duplicata.
        if Decimal(str(atual.valor or 0)) == valor and atual.vigencia_inicio <= vigencia:
            return
        if atual.vigencia_inicio < vigencia:
            atual.vigencia_fim = vigencia - timedelta(days=1)
        else:
            atual.vigencia_fim = atual.vigencia_inicio
    db.session.add(PrecoBaseInsumo(
        admin_id=admin_id,
        insumo_id=ins.id,
        valor=valor,
        vigencia_inicio=vigencia,
        observacao='Importação Excel',
    ))


# ──────────────────────────────────────────────────────────────────────
# CRONOGRAMA — modelo
# ──────────────────────────────────────────────────────────────────────
TEMPLATE_HEADERS = ['nome', 'descricao', 'categoria']
ITEM_HEADERS = [
    'ordem', 'parent_ordem', 'nome_tarefa', 'duracao_dias',
    'quantidade_prevista', 'responsavel',
]


def gerar_modelo_cronograma_xlsx() -> bytes:
    """Gera o modelo `.xlsx` de Templates de Cronograma em memória."""
    wb = Workbook()

    # Aba 1 — Template (cabeçalho)
    ws_t = wb.active
    ws_t.title = 'Template'
    ws_t.append(TEMPLATE_HEADERS)
    _style_header(ws_t, len(TEMPLATE_HEADERS))
    ws_t.append([
        'Obra Residencial Padrão',
        'Cronograma padrão para casa térrea de até 200 m²',
        'Residencial',
    ])
    _autosize(ws_t, [38, 60, 22])

    # Aba 2 — Itens
    ws_i = wb.create_sheet('Itens')
    ws_i.append(ITEM_HEADERS)
    _style_header(ws_i, len(ITEM_HEADERS))
    exemplos = [
        [1, '', 'Fundação', 10, '', 'empresa'],
        [2, 1, 'Locação da obra', 2, '', 'empresa'],
        [3, 1, 'Sapatas e baldrames', 8, 12, 'empresa'],
        [4, '', 'Estrutura', 20, '', 'empresa'],
        [5, 4, 'Pilares', 7, 8, 'empresa'],
        [6, 4, 'Vigas e laje', 13, '', 'terceiros'],
    ]
    for linha in exemplos:
        ws_i.append(linha)
    _autosize(ws_i, [10, 14, 38, 14, 22, 14])

    # Aba 3 — Instruções
    inst = wb.create_sheet('Instruções')
    inst.column_dimensions['A'].width = 28
    inst.column_dimensions['B'].width = 90
    inst['A1'] = 'Coluna'
    inst['B1'] = 'Descrição'
    _style_header(inst, 2)

    linhas = [
        ('Aba "Template"', 'Apenas a primeira linha de dados é considerada. '
                           'Define nome, descrição e categoria do template.'),
        ('nome (Template)', 'Obrigatório. Se já existir um template com o '
                            'mesmo nome no seu cadastro, ele será atualizado '
                            '(os itens existentes são substituídos pelos da '
                            'planilha).'),
        ('descricao', 'Opcional. Texto livre.'),
        ('categoria', 'Opcional. Ex.: Residencial, Comercial, Reforma.'),
        ('Aba "Itens"', 'Cada linha vira um CronogramaTemplateItem. A ordem '
                       'na planilha define a ordem do item no template.'),
        ('ordem', 'Obrigatório. Número inteiro único dentro da planilha. '
                  'Use para identificar o item ao definir parent_ordem '
                  '(hierarquia).'),
        ('parent_ordem', 'Opcional. Para criar subitens, informe a "ordem" '
                         'do item-pai. Deixe em branco para itens de '
                         'primeiro nível.'),
        ('nome_tarefa', 'Obrigatório. Nome da tarefa.'),
        ('duracao_dias', 'Obrigatório. Inteiro maior ou igual a 1.'),
        ('quantidade_prevista', 'Opcional. Número.'),
        ('responsavel', 'Opcional. Valores aceitos: empresa ou terceiros. '
                        'Padrão: empresa.'),
    ]
    for col, txt in linhas:
        inst.append([col, txt])

    inst.append([])
    inst.append(['Importação', 'Reimportar uma planilha com o mesmo nome '
                 'de template substitui completamente os itens antigos pelos '
                 'da nova planilha.'])
    inst.append(['', 'Linhas de itens com erro são reportadas no resumo após '
                 'o import; o template é criado/atualizado apenas se ao menos '
                 'uma linha de item for válida.'])
    inst.append(['', 'parent_ordem só pode apontar para uma "ordem" que '
                 'aparece em outra linha da mesma planilha.'])

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


# ──────────────────────────────────────────────────────────────────────
# CRONOGRAMA — importação
# ──────────────────────────────────────────────────────────────────────
def importar_cronograma_xlsx(arquivo, admin_id: int) -> dict[str, Any]:
    """Importa um template de cronograma a partir de `.xlsx`.

    Retorna ``{'template_nome', 'criado_ou_atualizado', 'itens_count',
    'rejected'}``. ``criado_ou_atualizado`` é "created" ou "updated".
    """
    wb = _ler_workbook_seguro(arquivo)

    if 'Template' not in wb.sheetnames or 'Itens' not in wb.sheetnames:
        raise ValueError(
            'O arquivo precisa ter as abas "Template" e "Itens". '
            'Baixe novamente o modelo Excel para verificar o formato.'
        )

    ws_t = wb['Template']
    ws_i = wb['Itens']
    _checar_max_linhas(max(0, (ws_i.max_row or 1) - 1), 'Itens')

    # Cabeçalho do template
    header_t = [(_str_or_none(c.value) or '').lower() for c in ws_t[1]]
    if not header_t or header_t[0] != 'nome':
        raise ValueError('Aba "Template": a primeira coluna deve ser "nome".')

    # Primeira linha de dados
    primeira_linha_dados = None
    for r in ws_t.iter_rows(min_row=2, values_only=True):
        if any(c not in (None, '') for c in r):
            primeira_linha_dados = r
            break
    if not primeira_linha_dados:
        raise ValueError('Aba "Template" não contém nenhuma linha de dados.')

    def _col_t(row, name):
        try:
            idx = header_t.index(name)
        except ValueError:
            return None
        return row[idx] if idx < len(row) else None

    nome_tmpl = _str_or_none(_col_t(primeira_linha_dados, 'nome'))
    if not nome_tmpl:
        raise ValueError('O campo "nome" do template é obrigatório.')

    descricao_tmpl = _str_or_none(_col_t(primeira_linha_dados, 'descricao'))
    categoria_tmpl = _str_or_none(_col_t(primeira_linha_dados, 'categoria'))

    # Cabeçalho dos itens
    header_i = [(_str_or_none(c.value) or '').lower() for c in ws_i[1]]
    obrigatorios = {'ordem', 'nome_tarefa', 'duracao_dias'}
    faltando = obrigatorios - set(header_i)
    if faltando:
        raise ValueError(
            'Aba "Itens": colunas obrigatórias ausentes: '
            f'{", ".join(sorted(faltando))}.'
        )

    def _col_i(row, name):
        try:
            idx = header_i.index(name)
        except ValueError:
            return None
        return row[idx] if idx < len(row) else None

    # Validação linha-a-linha + acumulação
    rejected: list[dict[str, Any]] = []
    itens_validos: list[dict[str, Any]] = []
    ordens_vistas: set[int] = set()

    for row_idx, row in enumerate(
        ws_i.iter_rows(min_row=2, values_only=True), start=2
    ):
        if not any(c not in (None, '') for c in row):
            continue

        ordem = _int_or_none(_col_i(row, 'ordem'))
        nome_tarefa = _str_or_none(_col_i(row, 'nome_tarefa'))
        dur = _int_or_none(_col_i(row, 'duracao_dias'))

        if ordem is None:
            rejected.append({'linha': row_idx, 'motivo': 'ordem é obrigatória e deve ser numérica.'})
            continue
        if ordem in ordens_vistas:
            rejected.append({'linha': row_idx, 'motivo': f'ordem {ordem} duplicada na planilha.'})
            continue
        if not nome_tarefa:
            rejected.append({'linha': row_idx, 'motivo': 'nome_tarefa é obrigatório.'})
            continue
        if dur is None or dur < 1:
            rejected.append({
                'linha': row_idx,
                'motivo': 'duracao_dias deve ser inteiro >= 1.',
            })
            continue

        parent_ordem = _int_or_none(_col_i(row, 'parent_ordem'))
        qtd_dec = _decimal_or_none(_col_i(row, 'quantidade_prevista'))
        responsavel = (_str_or_none(_col_i(row, 'responsavel')) or 'empresa').lower()
        if responsavel not in RESPONSAVEIS_VALIDOS:
            rejected.append({
                'linha': row_idx,
                'motivo': f'responsavel "{responsavel}" inválido. Use empresa ou terceiros.',
            })
            continue

        ordens_vistas.add(ordem)
        itens_validos.append({
            'linha': row_idx,
            'ordem': ordem,
            'parent_ordem': parent_ordem,
            'nome_tarefa': nome_tarefa,
            'duracao_dias': dur,
            'quantidade_prevista': float(qtd_dec) if qtd_dec is not None else None,
            'responsavel': responsavel,
        })

    # Validar parent_ordem -> ordens existentes
    itens_finais: list[dict[str, Any]] = []
    for it in itens_validos:
        if it['parent_ordem'] is not None and it['parent_ordem'] not in ordens_vistas:
            rejected.append({
                'linha': it['linha'],
                'motivo': f'parent_ordem {it["parent_ordem"]} não corresponde a nenhuma "ordem" da planilha.',
            })
            continue
        if it['parent_ordem'] is not None and it['parent_ordem'] == it['ordem']:
            rejected.append({
                'linha': it['linha'],
                'motivo': 'parent_ordem não pode apontar para a própria linha.',
            })
            continue
        itens_finais.append(it)

    if not itens_finais:
        raise ValueError(
            'Nenhum item válido para importar. Verifique a aba "Itens" e '
            'corrija os erros antes de tentar novamente.'
        )

    # Topo-sort DRY-RUN: detecta ciclos ANTES de tocar no banco. Garante que
    # um template existente nunca seja "apagado e deixado vazio" por causa
    # de hierarquia inválida.
    _ordens_resolvidas: set[int] = set()
    _pendentes_dry = list(itens_finais)
    while _pendentes_dry:
        _avancou = False
        _restantes = []
        for _it in _pendentes_dry:
            if _it['parent_ordem'] is None or _it['parent_ordem'] in _ordens_resolvidas:
                _ordens_resolvidas.add(_it['ordem'])
                _avancou = True
            else:
                _restantes.append(_it)
        if not _avancou:
            linhas_bloqueadas = ', '.join(str(it['linha']) for it in _restantes)
            raise ValueError(
                'Hierarquia inválida na aba "Itens": ciclo de parent_ordem '
                f'detectado nas linhas {linhas_bloqueadas}. Nenhuma alteração '
                'foi feita no template.'
            )
        _pendentes_dry = _restantes

    try:
        # Upsert do template por (admin_id, nome) — case-insensitive
        existente = (
            CronogramaTemplate.query
            .filter(
                CronogramaTemplate.admin_id == admin_id,
                db.func.lower(CronogramaTemplate.nome) == nome_tmpl.lower(),
            )
            .first()
        )

        if existente:
            criado_ou_atualizado = 'updated'
            tmpl = existente
            tmpl.nome = nome_tmpl
            tmpl.descricao = descricao_tmpl
            tmpl.categoria = categoria_tmpl
            if not tmpl.ativo:
                tmpl.ativo = True
            # Apagar itens antigos
            for it in list(tmpl.itens):
                db.session.delete(it)
            db.session.flush()
        else:
            criado_ou_atualizado = 'created'
            tmpl = CronogramaTemplate(
                nome=nome_tmpl,
                descricao=descricao_tmpl,
                categoria=categoria_tmpl,
                ativo=True,
                admin_id=admin_id,
            )
            db.session.add(tmpl)
            db.session.flush()

        # Topo-sort definitivo: cria pais antes dos filhos para resolver
        # `parent_item_id` a partir dos ids recém-criados. Ciclos já foram
        # rejeitados no DRY-RUN acima, então este laço sempre converge.
        ordem_para_id: dict[int, int] = {}
        pendentes = list(itens_finais)
        ordem_seq = 0  # ordem no template (monotônica, por inserção)

        while pendentes:
            ainda_pendentes: list[dict[str, Any]] = []
            for it in pendentes:
                parent_id = None
                if it['parent_ordem'] is not None:
                    parent_id = ordem_para_id.get(it['parent_ordem'])
                    if parent_id is None:
                        ainda_pendentes.append(it)
                        continue
                novo = CronogramaTemplateItem(
                    template_id=tmpl.id,
                    parent_item_id=parent_id,
                    nome_tarefa=it['nome_tarefa'],
                    ordem=ordem_seq,
                    duracao_dias=it['duracao_dias'],
                    quantidade_prevista=it['quantidade_prevista'],
                    responsavel=it['responsavel'],
                    admin_id=admin_id,
                )
                db.session.add(novo)
                db.session.flush()
                ordem_para_id[it['ordem']] = novo.id
                ordem_seq += 1
            pendentes = ainda_pendentes

        db.session.commit()
    except Exception:
        db.session.rollback()
        raise

    return {
        'template_nome': nome_tmpl,
        'criado_ou_atualizado': criado_ou_atualizado,
        'itens_count': ordem_seq,
        'rejected': rejected,
    }
