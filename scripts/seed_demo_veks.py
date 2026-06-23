"""
Seed DEMO — entidades do fluxo de caixa Veks num tenant, para testar a importação
(import → confirmar → salvar) com as entidades já reconhecidas.

Lê o Excel '1. FLUXO DE CAIXA_Veks Engenharia.xlsx' e cria, idempotente:
  • Obras           — distintos do CENTRO DE CUSTO
  • Funcionários    — quem aparece com 'diária' + os nomes do escritório
                      (abel/paulo/ana/cassio/gabriel(a), exceto empresas)
  • Fornecedores    — os demais nomes distintos da coluna FORNECEDOR

Uso:  python scripts/seed_demo_veks.py <admin_id|username>
"""
import os
import re
import sys
from datetime import date

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from main import app
from models import db, Usuario, Funcionario, Fornecedor, Obra, BancoEmpresa
from services.importacao_excel import _obter_ou_criar_cliente_placeholder, _normalizar

XLSX = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                    "1. FLUXO DE CAIXA_Veks Engenharia.xlsx")

# Marcadores de "empresa" (não vira funcionário mesmo casando um nome).
_EMP = re.compile(r"(ltda|eireli|s/?a\b|epp|me\b|comerci|industri|materia|distribu|"
                  r"servic|engenharia|construtora|tintas|locadora|locac|equipament|"
                  r"transport|autopec|auto pec|peças|pecas|supermerc|restaurante|"
                  r"pizzaria|advogad|tecnologia|telecom|banco|caixa|prefeitura|"
                  r"detran|concession|copias|cópias|vidros|esquadrias|flores|"
                  r"baterias|cartuch|mercado|conveniencia|forros|divisor|"
                  r"casa d|deposito|depósito|ferragens|&|administracao|administração)",
                  re.I)
_NOMES = re.compile(r"\b(abel|paulo|ana|cassio|cássio|gabriel|gabriela)\b", re.I)


def _coletar():
    import openpyxl
    from datetime import datetime as dt
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    forn = {}        # nome -> tem_diaria
    obras = set()
    bancos = {}      # nome_normalizado -> nome representativo
    for row in wb["Saída"].iter_rows(min_row=6, values_only=True):
        dv = row[1]
        if not dv or not isinstance(dv, (dt, date)):
            continue
        fo = str(row[3] or "").strip()
        de = str(row[4] or "").lower()
        cc = str(row[5] or "").strip() if len(row) > 5 else ""
        bk = str(row[6] or "").strip() if len(row) > 6 else ""
        if fo:
            forn.setdefault(fo, False)
            if "diari" in de or "diári" in de:
                forn[fo] = True
        if cc and cc != "???" and len(cc) > 2:
            obras.add(cc[:100])
        if bk and bk != "???":
            bancos.setdefault(_normalizar(bk).replace("banco ", "").strip(), bk[:100])
    for row in wb["Entrada"].iter_rows(min_row=6, values_only=True):
        dv = row[0]
        if not dv or not isinstance(dv, (dt, date)):
            continue
        cc = str(row[4] or "").strip() if len(row) > 4 else ""
        bk = str(row[5] or "").strip() if len(row) > 5 else ""
        if cc and cc != "???" and len(cc) > 2:
            obras.add(cc[:100])
        if bk and bk != "???":
            bancos.setdefault(_normalizar(bk).replace("banco ", "").strip(), bk[:100])

    funcionarios, fornecedores = [], []
    for nome, tem_diaria in forn.items():
        eh_empresa = bool(_EMP.search(nome))
        eh_func = tem_diaria or (bool(_NOMES.search(nome)) and not eh_empresa)
        (funcionarios if eh_func else fornecedores).append(nome)
    return sorted(obras), sorted(funcionarios), sorted(fornecedores), bancos


def seed(admin_id):
    obras, funcionarios, fornecedores, bancos = _coletar()
    cli_id = _obter_ou_criar_cliente_placeholder(admin_id)

    # índices do que já existe (idempotência por nome, case-insensitive)
    forn_ex = {f.nome.strip().lower() for f in
               Fornecedor.query.filter_by(admin_id=admin_id).all()}
    func_ex = {f.nome.strip().lower() for f in
               Funcionario.query.filter_by(admin_id=admin_id).all()}
    obra_ex = {o.nome.strip().lower() for o in
               Obra.query.filter_by(admin_id=admin_id).all()}
    banco_ex = {_normalizar(b.nome_banco).replace("banco ", "").strip() for b in
                BancoEmpresa.query.filter_by(admin_id=admin_id).all()}
    seq = Funcionario.query.filter_by(admin_id=admin_id).count() + 1

    n_obra = n_func = n_forn = n_banco = 0
    for i, (chave, nome) in enumerate(bancos.items()):
        if chave in banco_ex:
            continue
        db.session.add(BancoEmpresa(admin_id=admin_id, nome_banco=nome,
                                    agencia="0001", conta=f"{i + 1:06d}", ativo=True))
        banco_ex.add(chave); n_banco += 1
    for nome in obras:
        if nome.strip().lower() in obra_ex:
            continue
        db.session.add(Obra(nome=nome, admin_id=admin_id, cliente_id=cli_id,
                            data_inicio=date(2024, 1, 1), percentual_administracao=0,
                            ativo=True))
        obra_ex.add(nome.strip().lower()); n_obra += 1

    for i, nome in enumerate(funcionarios):
        if nome.strip().lower() in func_ex:
            continue
        k = seq + i
        db.session.add(Funcionario(
            admin_id=admin_id, nome=nome[:100], codigo=f"VK{k:04d}"[:10],
            cpf=f"{90000000000 + k:011d}"[:14], data_admissao=date(2024, 1, 1),
            tipo_remuneracao="diario", ativo=True))
        func_ex.add(nome.strip().lower()); n_func += 1

    for i, nome in enumerate(fornecedores):
        if nome.strip().lower() in forn_ex:
            continue
        db.session.add(Fornecedor(
            admin_id=admin_id, nome=nome[:100],
            cnpj=f"{90000000000000 + i:014d}"[:18], ativo=True))
        forn_ex.add(nome.strip().lower()); n_forn += 1

    db.session.commit()
    return dict(obras=len(obras), funcionarios=len(funcionarios),
                fornecedores=len(fornecedores), bancos=len(bancos),
                criadas_obras=n_obra, criados_func=n_func, criados_forn=n_forn,
                criados_banco=n_banco)


def _resolver_admin(arg):
    if str(arg).isdigit():
        return int(arg)
    u = Usuario.query.filter_by(username=arg).first()
    return u.id if u else None


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("uso: python scripts/seed_demo_veks.py <admin_id|username>")
        sys.exit(1)
    with app.app_context():
        admin_id = _resolver_admin(sys.argv[1])
        if not admin_id:
            print("admin não encontrado:", sys.argv[1]); sys.exit(1)
        r = seed(admin_id)
        print(f"[seed_demo_veks] admin_id={admin_id}")
        print(f"  Bancos:       {r['criados_banco']} criados / {r['bancos']} no arquivo")
        print(f"  Obras:        {r['criadas_obras']} criadas / {r['obras']} no arquivo")
        print(f"  Funcionários: {r['criados_func']} criados / {r['funcionarios']} no arquivo")
        print(f"  Fornecedores: {r['criados_forn']} criados / {r['fornecedores']} no arquivo")
