"""Gera a planilha-mestra de importação da obra Baia REV10 (17 itens) e
calcula a validação item a item (custo e venda) sem tocar no banco.

Saída:
  - obra_kabod/IMPORTACAO_Baia_REV10_completa.xlsx  (abas Insumos, Composicoes,
    Validacao, LEIA-ME)
  - tabela de validação no stdout
"""
from decimal import Decimal, ROUND_HALF_UP
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

INSUMO_HEADERS = ['nome', 'tipo', 'unidade', 'descricao', 'coeficiente_padrao',
                  'preco_base', 'fator_comercial', 'unidade_comercial', 'tipo_medicao']
COMP_HEADERS = ['servico_nome', 'servico_unidade', 'categoria',
                'insumo_nome', 'coeficiente', 'unidade_insumo', 'observacao']
CATEGORIA = 'Obra Baia REV10'
# Venda total do orçamento ORIGINAL = soma da coluna O (TOTAL venda) da aba
# 'Proposta Comercial' da REV10. A venda é calibrada por markup UNIFORME para o
# total bater exatamente com este número (os custos/insumos/coef. são os nossos).
VENDA_ORIGINAL = Decimal('1720796.75')


def D(s):
    return Decimal(str(s).replace('.', '').replace(',', '.')) if isinstance(s, str) else Decimal(str(s))


# nome -> (tipo, unidade, preco_base)
INSUMOS = {
    'Aço LSF galvanizado Z275': ('MATERIAL', 'kg', '13,50'),
    'Fixadores/parafusos LSF': ('MATERIAL', 'vb', '15000,00'),
    'Stain': ('MATERIAL', 'vb', '800,00'),
    'Ferragens/trilho portão': ('MATERIAL', 'un', '125,00'),
    'Fixadores cercado': ('MATERIAL', 'un', '100,00'),
    'Concreto usinado': ('MATERIAL', 'm3', '550,00'),
    'Extra/ferramenta concreto': ('MATERIAL', 'm2', '3,00'),
    'Ponto de luz (material)': ('MATERIAL', 'un', '400,00'),
    'Adicional elétrico': ('MATERIAL', 'vb', '1500,00'),
    'Quadro/diversos elétrico': ('MATERIAL', 'vb', '610,00'),
    'Material ponto hidráulico': ('MATERIAL', 'un', '867,10'),
    # --- 1.17 decomposto (substitui a verba 'pacote complementar') ---
    'Concreto armado fundação': ('MATERIAL', 'm3', '480,00'),
    'Bomba de concreto': ('EQUIPAMENTO', 'un', '1800,00'),
    'Aço CA-50 fundação': ('MATERIAL', 'kg', '5,50'),
    'Madeira de forma': ('MATERIAL', 'vb', '9000,00'),
    'Miscelâneas fundação': ('MATERIAL', 'vb', '5500,00'),
    'Equipe fundação (dia)': ('MAO_OBRA', 'dia', '1170,00'),
    'Visita técnica (Guilherme)': ('MAO_OBRA', 'vb', '2914,00'),
    'Material infra elétrica': ('MATERIAL', 'vb', '11000,00'),
    'M.O. infra elétrica': ('MAO_OBRA', 'vb', '6400,00'),
    'Material infra hidráulica': ('MATERIAL', 'vb', '9400,00'),
    'M.O. infra hidráulica': ('MAO_OBRA', 'vb', '9600,00'),
    'Painel lã de rocha 32kg/m³ 50mm': ('MATERIAL', 'm2', '25,00'),
    'M.O. instalação isolamento': ('MAO_OBRA', 'm2', '10,00'),
    'Forro PVC preto (material)': ('MATERIAL', 'm2', '42,00'),
    'M.O. instalação forro PVC': ('MAO_OBRA', 'm2', '24,00'),
    'Encarregado': ('MAO_OBRA', 'h', '36,36'),
    'Montador líder': ('MAO_OBRA', 'h', '31,82'),
    'Montador': ('MAO_OBRA', 'h', '26,14'),
    'Ajudante': ('MAO_OBRA', 'h', '20,46'),
    'M.O. pintura aço': ('MAO_OBRA', 'm2', '35,00'),
    'M.O. pintura/Stain portão': ('MAO_OBRA', 'm2', '35,00'),
    'M.O. montagem portão': ('MAO_OBRA', 'un', '350,00'),
    'M.O. placa cimentícia + basecoat': ('MAO_OBRA', 'm2', '45,00'),
    'M.O. régua de pinus': ('MAO_OBRA', 'm2', '45,00'),
    'M.O. pintura interna (cimento queimado)': ('MAO_OBRA', 'm2', '45,00'),
    'M.O. Stain externa': ('MAO_OBRA', 'm2', '30,00'),
    'M.O. verticalização pilar': ('MAO_OBRA', 'un', '568,18'),
    'M.O. concretagem': ('MAO_OBRA', 'm2', '25,00'),
    'Taxa fixa concretagem': ('MAO_OBRA', 'm2', '5,00'),
    'M.O. revestimento pedra moledo': ('MAO_OBRA', 'm2', '230,00'),
    'M.O. instalação elétrica': ('MAO_OBRA', 'vb', '6500,00'),
    'M.O. instalação telha shingle': ('MAO_OBRA', 'm2', '85,00'),
    'M.O. cercado': ('MAO_OBRA', 'un', '450,00'),
    'M.O. pintura cercado': ('MAO_OBRA', 'un', '400,00'),
    'M.O. ponto hidráulico': ('MAO_OBRA', 'un', '100,00'),
    'M.O. pacote complementar': ('MAO_OBRA', 'vb', '40242,00'),
}

# (codigo, nome, unidade, qty, [(insumo, coef, obs)], bdi_override)
SERVICOS = [
    ('1.1', 'Estrutura Aço LSF Z275', 'kg', '21900', [
        ('Aço LSF galvanizado Z275', '1,07', '1kg + 7% perda'),
        ('Fixadores/parafusos LSF', '0,000046', 'verba 15000 / 21900 kg'),
        ('Encarregado', '0,022', '1 pessoa'),
        ('Montador líder', '0,022', '1 pessoa'),
        ('Montador', '0,044', '2 pessoas'),
        ('Ajudante', '0,044', '2 pessoas'),
    ], {'Lmat': '0.25'}),
    ('1.2', 'Pintura do aço estrutural', 'm2', '1173', [
        ('M.O. pintura aço', '1', 'taxa 35/m2 — converter p/ h depois')], {}),
    ('1.3', 'Pintura/Stain portão Pinus', 'm2', '161', [
        ('Stain', '0,006211', 'verba 800 global / 161 m2'),
        ('M.O. pintura/Stain portão', '1', 'taxa 35/m2')], {}),
    ('1.4', 'Portão em Pinus', 'un', '48', [
        ('Ferragens/trilho portão', '1', 'madeira do cliente'),
        ('M.O. montagem portão', '1', 'taxa 350/un')], {}),
    ('1.5', 'Fechamento interno placa cimentícia', 'm2', '900', [
        ('M.O. placa cimentícia + basecoat', '1', 'material do cliente; taxa 45/m2')], {}),
    ('1.6', 'Fechamento externo régua pinus', 'm2', '660', [
        ('M.O. régua de pinus', '1', 'madeira do cliente; taxa 45/m2')], {}),
    ('1.7', 'Pintura fechamentos internos', 'm2', '900', [
        ('M.O. pintura interna (cimento queimado)', '1', 'unidade corrigida vb->m2')], {}),
    ('1.8', 'Pintura Stain paredes externas', 'm2', '660', [
        ('Stain', '0,001515', 'verba 800 global / 660 m2'),
        ('M.O. Stain externa', '1', 'taxa 30/m2')], {}),
    ('1.9', 'Verticalização de pilares roliços', 'un', '32', [
        ('M.O. verticalização pilar', '1', 'base 44 pilares, qtd 32 — validar')], {}),
    ('1.10', 'Corredores em concreto', 'm2', '500,4', [
        ('Concreto usinado', '0,20', 'esp. 0,20 m'),
        ('Extra/ferramenta concreto', '1', 'verba 1500 -> 3/m2'),
        ('M.O. concretagem', '1', 'taxa 25/m2'),
        ('Taxa fixa concretagem', '1', 'verba 2500 -> 5/m2')], {}),
    ('1.11', 'Revestimento pedra moledo', 'm2', '40', [
        ('M.O. revestimento pedra moledo', '1', 'pedra do cliente; taxa 230/m2')], {}),
    ('1.12', 'Instalação de pontos de luz', 'vb', '1', [
        ('Ponto de luz (material)', '12', '12 pontos x 400 — validar qtd'),
        ('Adicional elétrico', '1', 'verba'),
        ('Quadro/diversos elétrico', '1', 'verba'),
        ('M.O. instalação elétrica', '1', 'verba 6500')], {}),
    ('1.13', 'Telha Shingle', 'm2', '1173', [
        ('M.O. instalação telha shingle', '1', 'shingle não incluso (ver 1.17)')], {}),
    ('1.14', 'Cercado das baias', 'un', '24', [
        ('Fixadores cercado', '1', 'réguas do cliente'),
        ('M.O. cercado', '1', 'taxa 450/un')], {}),
    ('1.15', 'Pintura Stain do cercado', 'un', '24', [
        ('Stain', '0,041667', 'verba 800 global / 24 un'),
        ('M.O. pintura cercado', '1', 'taxa 400/un')], {}),
    ('1.16', 'Ponto hidráulico por baia', 'un', '24', [
        ('Material ponto hidráulico', '1', 'corrigido: material x24 (planilha conta 1x)'),
        ('M.O. ponto hidráulico', '1', 'taxa 100/un')], {}),
    # 1.17 DECOMPOSTO (era verba R$92.564). BDI do pacote: Lmat 0.22 / Lmo 0.15.
    ('1.17a', 'Fundação - sapatas, radier e aço', 'vb', '1', [
        ('Concreto armado fundação', '58,23', '58,23 m3 x R$480 (aba Fundação)'),
        ('Bomba de concreto', '3', '3 x R$1800'),
        ('Aço CA-50 fundação', '2743', '2.743 kg x R$5,50'),
        ('Madeira de forma', '1', 'verba R$9000'),
        ('Miscelâneas fundação', '1,5', '1,5 x R$5500'),
        ('Equipe fundação (dia)', '20', '20 dias x R$1170/dia (enxugado de 40)'),
        ('Visita técnica (Guilherme)', '1', 'verba R$2914')],
        {'Lmat': '0.22', 'Lmo': '0.15'}),
    ('1.17b', 'Infra Elétrica', 'vb', '1', [
        ('Material infra elétrica', '1', '44 pts ilum+tomadas+quadro; SINAPI 104473/104480 (projetos)'),
        ('M.O. infra elétrica', '1', 'instalação; SINAPI-SP')],
        {'Lmat': '0.22', 'Lmo': '0.15'}),
    ('1.17c', 'Infra Hidráulica', 'vb', '1', [
        ('Material infra hidráulica', '1', '24 esgoto baias+banheiro+rede AF; SINAPI 104678'),
        ('M.O. infra hidráulica', '1', 'instalação + M.O. louças R$3436 (Memorial)')],
        {'Lmat': '0.22', 'Lmo': '0.15'}),
    ('1.17d', 'Isolamento lã de rocha', 'm2', '196,142', [
        ('Painel lã de rocha 32kg/m³ 50mm', '1', 'mercado ~R$25/m2 (pct 4,32m2)'),
        ('M.O. instalação isolamento', '1', 'estimado R$10/m2')],
        {'Lmat': '0.22', 'Lmo': '0.15'}),
    ('1.17e', 'Forro PVC preto', 'm2', '196,142', [
        ('Forro PVC preto (material)', '1', 'SINAPI 96111 R$66/m2 all-in'),
        ('M.O. instalação forro PVC', '1', 'parcela M.O.')],
        {'Lmat': '0.22', 'Lmo': '0.15'}),
]


def calc():
    """Custos da decomposição + venda por MARKUP UNIFORME (venda total = original).

    rows = [(cod, nome, un, qty, custo_unit, custo_total, venda_total)].
    """
    tmp = []
    tot_custo = Decimal('0')
    for cod, nome, un, qty_s, comps, _bdi in SERVICOS:
        qty = D(qty_s)
        cu = Decimal('0')  # custo por unidade de serviço
        for ins, coef_s, _ in comps:
            _tipo, _u, preco = INSUMOS[ins]
            cu += D(coef_s) * D(preco)
        custo_total = cu * qty
        tot_custo += custo_total
        tmp.append((cod, nome, un, qty, cu, custo_total))
    # markup uniforme p/ a venda total igualar a do orçamento original
    fator = VENDA_ORIGINAL / tot_custo
    rows = [(cod, nome, un, qty, cu, ct, ct * fator) for cod, nome, un, qty, cu, ct in tmp]
    return rows, tot_custo, VENDA_ORIGINAL


def money(x):
    return x.quantize(Decimal('0.01'), ROUND_HALF_UP)


def build_xlsx(rows, tot_custo, tot_venda):
    wb = Workbook()
    fill = PatternFill('solid', fgColor='2F5496')

    def header(ws, hs, widths):
        ws.append(hs)
        for c in ws[1]:
            c.font = Font(bold=True, color='FFFFFF'); c.fill = fill
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + i)].width = w

    ws1 = wb.active; ws1.title = 'Insumos'
    header(ws1, INSUMO_HEADERS, [40, 12, 9, 44, 16, 14, 13, 14, 12])
    for nome, (tipo, un, preco) in INSUMOS.items():
        ws1.append([nome, tipo, un, '', '1', preco, '1', '', 'UNITARIO'])

    ws2 = wb.create_sheet('Composicoes')
    header(ws2, COMP_HEADERS, [36, 14, 16, 40, 14, 13, 40])
    for cod, nome, un, qty_s, comps, bdi in SERVICOS:
        for ins, coef_s, obs in comps:
            ws2.append([nome, un, CATEGORIA, ins, coef_s,
                        INSUMOS[ins][1], f'[{cod}] {obs}'])

    ws3 = wb.create_sheet('Validacao')
    header(ws3, ['item', 'serviço', 'un', 'quantidade', 'custo unit (R$)',
                 'custo total (R$)', 'venda total (R$)', 'margem (R$)'],
           [8, 38, 8, 14, 16, 18, 18, 16])
    for cod, nome, un, qty, cu, ct, vt in rows:
        ws3.append([cod, nome, un, float(qty), float(money(cu)),
                    float(money(ct)), float(money(vt)), float(money(vt - ct))])
    ws3.append([])
    ws3.append(['', 'TOTAL', '', '', '', float(money(tot_custo)),
                float(money(tot_venda)), float(money(tot_venda - tot_custo))])
    for c in ws3[ws3.max_row]:
        c.font = Font(bold=True)

    ws4 = wb.create_sheet('LEIA-ME')
    for line in [
        ['Importação da obra Baia REV10 — 21 serviços (1.17 decomposto em 1.17a-e). '
         'Composição = CUSTO; BDI/lucro por item.'],
        [''],
        ['Como importar:'],
        ['  1) Catálogo de Insumos  -> Importar Excel  (aba "Insumos")'],
        ['  2) Catálogo de Serviços -> Importar Composições  (aba "Composicoes")'],
        ['  3) Quantidades (aba "Validacao") entram ao montar o Orçamento (Passo B4).'],
        [''],
        ['Correções aplicadas (vs planilha REV10):'],
        ['  1.16 material x24 (planilha contava 1x) · 1.7 unidade vb->m2'],
        ['  1.3/1.8/1.15 Stain modelado como verba global (custo)'],
        [''],
        ['1.17 DECOMPOSTO (era verba R$92.564) em 5 serviços = R$148.211 (custo):'],
        ['  1.17a Fundação R$92.001 (enxugada: equipe 20 dias, sem escavação dupla)'],
        ['  1.17b Infra Elétrica R$17.400 · 1.17c Infra Hidráulica R$19.000 (projetos+SINAPI)'],
        ['  1.17d Isolamento lã de rocha R$6.865 · 1.17e Forro PVC R$12.945 (mercado/SINAPI)'],
        ['  NÃO recadastrados (já em outros itens): aço->1.1, fechamento->1.5/1.6, stain->1.8/1.15'],
        ['  NÃO incluso (cliente): material das louças/metais R$40.940'],
        [''],
        ['Pendentes de calibração sua:'],
        ['  1.9 base 44 vs qtd 32 · 1.12 12 pontos vs descrição (baias+pilares)'],
        ['  1.3 erro da planilha: R$128k fantasma (Stain global em H vs x161 em J)'],
        [''],
        ['ATENÇÃO ao importar Insumos: nomes genéricos (Montador, Ajudante, '
         'Concreto usinado...) podem já existir no seu catálogo e o preço será '
         'atualizado (cria nova vigência). Revise antes se necessário.'],
    ]:
        ws4.append(line)
    ws4.column_dimensions['A'].width = 100

    out = 'obra_kabod/IMPORTACAO_Baia_REV10_completa.xlsx'
    wb.save(out)
    return out


if __name__ == '__main__':
    rows, tc, tv = calc()
    out = build_xlsx(rows, tc, tv)
    print(f'gerado: {out}\n')
    print(f'{"item":5} {"serviço":36} {"un":3} {"qtd":>9} '
          f'{"custo/un":>10} {"custo tot":>13} {"venda tot":>13}')
    print('-' * 100)
    for cod, nome, un, qty, cu, ct, vt in rows:
        print(f'{cod:5} {nome[:36]:36} {un:3} {float(qty):>9.1f} '
              f'{float(cu):>10.2f} {float(ct):>13.2f} {float(vt):>13.2f}')
    print('-' * 100)
    print(f'{"":5} {"TOTAL":36} {"":3} {"":>9} {"":>10} '
          f'{float(tc):>13.2f} {float(tv):>13.2f}')
    print(f'\nInsumos distintos: {len(INSUMOS)} | Serviços: {len(SERVICOS)} | '
          f'Linhas de composição: {sum(len(s[4]) for s in SERVICOS)}')
