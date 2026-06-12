"""Gera uma APRESENTAÇÃO (não importação) do orçamento Baia REV10: cada serviço
seguido, logo abaixo, dos insumos vinculados a ele — para visualizar/apresentar.

Mantém os custos/insumos/coeficientes da nossa decomposição. A VENDA é calibrada
por um markup UNIFORME para que a VENDA TOTAL seja igual à do orçamento original
(R$1.720.796,75 = soma da coluna O da proposta REV10). O % de markup é mostrado
em cada insumo.

Reaproveita os dados do gerador de importação (INSUMOS, SERVICOS) sem alterá-lo.
Não toca nas planilhas de importação.

Saída:
  - APRESENTACAO_Baia_REV10.xlsx
  - APRESENTACAO_Baia_REV10.pdf
"""
from decimal import Decimal, ROUND_HALF_UP
import importlib.util
import os

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle, Paragraph,
                                Spacer, KeepTogether)

# --- carrega os dados do gerador de importação (sem rodá-lo) ---
_HERE = os.path.dirname(os.path.abspath(__file__))
_spec = importlib.util.spec_from_file_location(
    'gerar_importacao_baia_rev10', os.path.join(_HERE, 'gerar_importacao_baia_rev10.py'))
_gi = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(_gi)
INSUMOS, SERVICOS, D = _gi.INSUMOS, _gi.SERVICOS, _gi.D

# venda total do orçamento ORIGINAL = soma da coluna O (TOTAL venda) da aba
# 'Proposta Comercial' da planilha REV10. A venda da apresentação é calibrada
# para bater exatamente com este número.
VENDA_ORIGINAL = Decimal('1720796.75')

AZUL = '2F5496'
AZUL_CLARO = 'D9E1F2'
CINZA = 'F2F2F2'


def money(x):
    q = Decimal(x).quantize(Decimal('0.01'), ROUND_HALF_UP)
    return 'R$ ' + f'{q:,.2f}'.replace(',', '@').replace('.', ',').replace('@', '.')


def num(x):
    s = f'{Decimal(x):,.4f}'.rstrip('0').rstrip('.')
    if s in ('', '-0'):
        s = '0'
    return s.replace(',', '@').replace('.', ',').replace('@', '.')


def pct(x):
    q = Decimal(x).quantize(Decimal('0.01'), ROUND_HALF_UP)
    s = f'{q:,.2f}'.replace(',', '@').replace('.', ',').replace('@', '.')
    return ('+' if q >= 0 else '') + s + '%'


def computar():
    """Custos da decomposição + venda por markup uniforme (venda total = original)."""
    blocos = []
    tot_custo = Decimal('0')
    for cod, nome, un, qty_s, comps, _bdi in SERVICOS:
        qty = D(qty_s)
        linhas = []
        custo_total = Decimal('0')
        for ins, coef_s, obs in comps:
            tipo, un_ins, preco = INSUMOS[ins]
            coef = D(coef_s); p = D(preco)
            ct = coef * p * qty
            custo_total += ct
            linhas.append({'insumo': ins, 'tipo': tipo, 'un': un_ins,
                           'coef': coef, 'preco': p, 'custo_total': ct, 'obs': obs})
        blocos.append({'cod': cod, 'nome': nome, 'un': un, 'qty': qty,
                       'custo_total': custo_total, 'linhas': linhas})
        tot_custo += custo_total

    # markup UNIFORME para a venda total igualar a original
    fator = VENDA_ORIGINAL / tot_custo
    markup = (fator - 1) * 100
    for b in blocos:
        b['venda'] = b['custo_total'] * fator
        b['markup'] = markup
        for ln in b['linhas']:
            ln['venda'] = ln['custo_total'] * fator
            ln['markup'] = markup
    return blocos, tot_custo, VENDA_ORIGINAL, markup


COLS = ['', 'Descrição', 'Tipo', 'Un', 'Coef', 'Preço unit.', 'Custo total',
        '% markup', 'Venda total', 'Observação']
NC = len(COLS)


# ============================== EXCEL ==============================
def build_xlsx(blocos, tot_custo, tot_venda, markup, out):
    wb = Workbook(); ws = wb.active; ws.title = 'Orçamento Baia REV10'
    thin = Side(style='thin', color='BFBFBF')
    borda = Border(left=thin, right=thin, top=thin, bottom=thin)
    widths = [6, 40, 11, 5, 11, 14, 15, 10, 15, 44]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    r = 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NC)
    c = ws.cell(r, 1, 'ORÇAMENTO — Baias de Bovinos (Kabod Cabana) · REV10')
    c.font = Font(bold=True, size=15, color='FFFFFF'); c.fill = PatternFill('solid', fgColor=AZUL)
    c.alignment = Alignment(horizontal='center', vertical='center'); ws.row_dimensions[r].height = 26
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NC)
    c = ws.cell(r, 1, f'Custo total: {money(tot_custo)}     ·     '
                      f'Venda total: {money(tot_venda)} (= orçamento original)     ·     '
                      f'markup uniforme {pct(markup)}')
    c.font = Font(bold=True, size=11, color=AZUL); c.alignment = Alignment(horizontal='center')
    ws.row_dimensions[r].height = 20
    r += 2

    for b in blocos:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NC)
        c = ws.cell(r, 1, f'  {b["cod"]}   {b["nome"]}        ({b["un"]} · qtd {num(b["qty"])})'
                          f'        CUSTO {money(b["custo_total"])}   ·   VENDA {money(b["venda"])}')
        c.font = Font(bold=True, size=11, color='FFFFFF'); c.fill = PatternFill('solid', fgColor=AZUL)
        c.alignment = Alignment(horizontal='left', vertical='center'); ws.row_dimensions[r].height = 20
        r += 1
        for i, h in enumerate(COLS, 1):
            cc = ws.cell(r, i, h)
            cc.font = Font(bold=True, size=9, color=AZUL); cc.fill = PatternFill('solid', fgColor=AZUL_CLARO)
            cc.alignment = Alignment(horizontal='center' if i not in (2, 10) else 'left'); cc.border = borda
        r += 1
        for j, ln in enumerate(b['linhas']):
            vals = ['•', ln['insumo'], ln['tipo'], ln['un'], num(ln['coef']),
                    money(ln['preco']), money(ln['custo_total']), pct(ln['markup']),
                    money(ln['venda']), ln['obs']]
            fill = CINZA if j % 2 else 'FFFFFF'
            for i, v in enumerate(vals, 1):
                cc = ws.cell(r, i, v)
                cc.fill = PatternFill('solid', fgColor=fill); cc.border = borda
                cc.font = Font(size=9, color='7030A0' if ln['tipo'] == 'MAO_OBRA' else '000000')
                if i in (5, 6, 7, 8, 9):
                    cc.alignment = Alignment(horizontal='right')
                elif i in (1, 3, 4):
                    cc.alignment = Alignment(horizontal='center')
                else:
                    cc.alignment = Alignment(horizontal='left', wrap_text=(i == 10))
            r += 1
        r += 1

    ws.freeze_panes = 'A4'
    wb.save(out)
    return out


# ============================== PDF ==============================
def build_pdf(blocos, tot_custo, tot_venda, markup, out):
    doc = SimpleDocTemplate(out, pagesize=landscape(A4),
                            leftMargin=10 * mm, rightMargin=10 * mm,
                            topMargin=12 * mm, bottomMargin=12 * mm)
    ss = getSampleStyleSheet()
    h_title = ParagraphStyle('t', parent=ss['Title'], fontSize=16, textColor=colors.HexColor('#' + AZUL))
    h_sub = ParagraphStyle('s', parent=ss['Normal'], fontSize=10, textColor=colors.HexColor('#' + AZUL),
                           alignment=1, spaceAfter=8)
    svc_style = ParagraphStyle('svc', parent=ss['Normal'], fontSize=10.5, textColor=colors.white, leading=14)
    obs_style = ParagraphStyle('obs', parent=ss['Normal'], fontSize=7, textColor=colors.HexColor('#555555'), leading=8.5)
    ins_style = ParagraphStyle('ins', parent=ss['Normal'], fontSize=8, leading=9.5)

    story = [Paragraph('Orçamento — Baias de Bovinos (Kabod Cabana) · REV10', h_title),
             Paragraph(f'Custo total: {money(tot_custo)} &nbsp;·&nbsp; Venda total: '
                       f'<b>{money(tot_venda)}</b> (= orçamento original) &nbsp;·&nbsp; '
                       f'markup uniforme <b>{pct(markup)}</b>', h_sub),
             Spacer(1, 4)]

    col_w = [7 * mm, 58 * mm, 23 * mm, 9 * mm, 16 * mm, 22 * mm, 24 * mm, 16 * mm, 24 * mm, 38 * mm]
    for b in blocos:
        faixa = Table([[Paragraph(
            f'<b>{b["cod"]}</b> &nbsp; <b>{b["nome"]}</b> &nbsp;&nbsp;({b["un"]} · qtd {num(b["qty"])}) '
            f'&nbsp;&nbsp; CUSTO <b>{money(b["custo_total"])}</b> &nbsp;·&nbsp; VENDA '
            f'<b>{money(b["venda"])}</b>', svc_style)]], colWidths=[sum(col_w)])
        faixa.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#' + AZUL)),
            ('LEFTPADDING', (0, 0), (-1, -1), 6), ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4)]))

        data = [COLS[:1] + ['Insumo'] + COLS[2:]]
        for ln in b['linhas']:
            data.append(['•', Paragraph(ln['insumo'], ins_style), ln['tipo'], ln['un'],
                         num(ln['coef']), money(ln['preco']), money(ln['custo_total']),
                         pct(ln['markup']), money(ln['venda']), Paragraph(ln['obs'], obs_style)])
        t = Table(data, colWidths=col_w, repeatRows=1)
        sty = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#' + AZUL_CLARO)),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#' + AZUL)),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'), ('FONTSIZE', (0, 0), (-1, 0), 7),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#BFBFBF')),
            ('ALIGN', (4, 1), (8, -1), 'RIGHT'), ('ALIGN', (0, 0), (0, -1), 'CENTER'),
            ('ALIGN', (2, 0), (3, -1), 'CENTER'), ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 2), ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ]
        for i, ln in enumerate(b['linhas'], 1):
            if ln['tipo'] == 'MAO_OBRA':
                sty.append(('TEXTCOLOR', (1, i), (1, i), colors.HexColor('#7030A0')))
            if i % 2 == 0:
                sty.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor('#F2F2F2')))
        t.setStyle(TableStyle(sty))
        story.append(KeepTogether([faixa, t, Spacer(1, 7)]))

    doc.build(story)
    return out


if __name__ == '__main__':
    blocos, tc, tv, mk = computar()
    x = build_xlsx(blocos, tc, tv, mk, 'APRESENTACAO_Baia_REV10.xlsx')
    p = build_pdf(blocos, tc, tv, mk, 'APRESENTACAO_Baia_REV10.pdf')
    print(f'gerado: {x}')
    print(f'gerado: {p}')
    print(f'\n{len(blocos)} serviços · custo {money(tc)} · venda {money(tv)} '
          f'(= original) · markup uniforme {pct(mk)}')
