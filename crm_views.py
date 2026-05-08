"""
Task #42 — Blueprint do CRM de Leads.

Multi-tenant: cada admin vê só os leads da sua empresa.
Funcionário comum (TipoUsuario.FUNCIONARIO) vê apenas leads onde ele é o
responsável (mapeamento via CrmResponsavel.nome == Usuario.nome — fallback
porque CrmResponsavel é cadastro livre, sem FK obrigatória pra Usuario).

Rotas:
  /crm/                             — Kanban (default)
  /crm/lista                        — Visão tabela
  /crm/novo                         — Criar lead (form)
  /crm/<id>/editar                  — Editar lead (form com timeline)
  /crm/<id>/excluir                 — Excluir lead (POST)
  /crm/<id>/mudar_status            — POST drag-and-drop kanban
  /crm/<id>/gerar_proposta          — Redireciona p/ nova proposta pré-preenchida
  /crm/<id>/criar_obra              — Redireciona p/ nova obra pré-preenchida
  /crm/cadastros                    — CRUD listas mestras (tabs)
  /crm/cadastros/<lista>/criar      — POST adicionar item
  /crm/cadastros/<lista>/<id>/editar         — POST renomear
  /crm/cadastros/<lista>/<id>/toggle_ativo   — POST ativar/desativar
  /crm/cadastros/<lista>/<id>/excluir        — POST excluir (bloqueado se em uso)
"""
from __future__ import annotations

import io
import logging
import re
from datetime import datetime, date
from decimal import Decimal, InvalidOperation

from flask import (
    Blueprint, render_template, request, redirect, url_for,
    flash, jsonify, abort, send_file,
)
from flask_login import login_required, current_user
from sqlalchemy import or_

from app import db
from models import (
    Lead, LeadHistorico, LeadStatus,
    CrmResponsavel, CrmOrigem, CrmCadencia, CrmSituacao,
    CrmTipoMaterial, CrmTipoObra, CrmMotivoPerda,
    Cliente, Proposta, Obra, Usuario, TipoUsuario,
)

logger = logging.getLogger(__name__)

crm_bp = Blueprint('crm', __name__, url_prefix='/crm')


# ---------------------------------------------------------------------------
# Mapa de listas mestras (lookup por slug usado na URL)
# ---------------------------------------------------------------------------
LISTAS_MESTRAS = {
    'responsavel':    {'model': CrmResponsavel,  'label': 'Responsáveis',     'campo_lead': 'responsavel_id'},
    'origem':         {'model': CrmOrigem,       'label': 'Origens',          'campo_lead': 'origem_id'},
    'cadencia':       {'model': CrmCadencia,     'label': 'Cadências',        'campo_lead': 'cadencia_id'},
    'situacao':       {'model': CrmSituacao,     'label': 'Situações',        'campo_lead': 'situacao_id'},
    'tipo_material':  {'model': CrmTipoMaterial, 'label': 'Tipos de Material', 'campo_lead': 'tipo_material_id'},
    'tipo_obra':      {'model': CrmTipoObra,     'label': 'Tipos de Obra',    'campo_lead': 'tipo_obra_id'},
    'motivo_perda':   {'model': CrmMotivoPerda,  'label': 'Motivos de Perda', 'campo_lead': 'motivo_perda_id'},
}


# ---------------------------------------------------------------------------
# Helpers de tenant / permissão
# ---------------------------------------------------------------------------
def get_admin_id():
    """Retorna admin_id do usuário atual (padrão consolidado do sistema)."""
    if not current_user.is_authenticated:
        return None
    if hasattr(current_user, 'admin_id') and current_user.admin_id:
        return current_user.admin_id
    return current_user.id


def is_admin_user():
    """True se o usuário atual é ADMIN (vê todos os leads do tenant)."""
    if not current_user.is_authenticated:
        return False
    return current_user.tipo_usuario in (
        TipoUsuario.ADMIN, TipoUsuario.SUPER_ADMIN, TipoUsuario.GESTOR_EQUIPES,
    )


def _query_leads_visiveis(admin_id):
    """Query base de leads visíveis ao usuário atual:
    - admin/gestor: todos do tenant
    - funcionário comum: apenas onde CrmResponsavel.nome == current_user.nome
      (heurística porque CrmResponsavel é cadastro livre)."""
    q = Lead.query.filter_by(admin_id=admin_id)
    if is_admin_user():
        return q
    nome_user = (getattr(current_user, 'nome', '') or '').strip()
    if not nome_user:
        return q.filter(False)  # nada visível
    sub = (
        db.session.query(CrmResponsavel.id)
        .filter(CrmResponsavel.admin_id == admin_id)
        .filter(CrmResponsavel.nome == nome_user)
        .subquery()
    )
    return q.filter(Lead.responsavel_id.in_(sub))


def _pode_acessar_lead(lead):
    """Guarda centralizada de autorização ao nível do objeto Lead.
    Usar em TODAS as rotas que mexem com um lead específico (editar,
    excluir, mudar_status, gerar_proposta, criar_obra)."""
    if not lead:
        return False
    admin_id = get_admin_id()
    if not admin_id or lead.admin_id != admin_id:
        return False
    if is_admin_user():
        return True
    nome_user = (getattr(current_user, 'nome', '') or '').strip()
    if not nome_user:
        return False
    return bool(lead.responsavel and lead.responsavel.nome == nome_user)


def _validar_fk_tenant(slug, fk_id, admin_id):
    """Garante que `fk_id` (vindo do form) referencia uma linha da lista
    `slug` que pertence a `admin_id`. Retorna o id válido OU None.

    Bloqueia injeção de IDs de outro tenant em campos como
    `responsavel_id`, `origem_id` etc.
    """
    if fk_id in (None, ''):
        return None
    try:
        fk_id = int(fk_id)
    except (ValueError, TypeError):
        return None
    info = LISTAS_MESTRAS.get(slug)
    if not info:
        return None
    existe = (
        db.session.query(info['model'].id)
        .filter_by(id=fk_id, admin_id=admin_id)
        .first()
    )
    return fk_id if existe else None


# ---------------------------------------------------------------------------
# Helpers de parsing
# ---------------------------------------------------------------------------
def _normalizar_telefone(valor):
    if not valor:
        return None
    digits = re.sub(r'\D', '', valor)
    return digits or None


def _to_int(valor):
    try:
        v = int(valor) if valor not in (None, '', 'None') else None
        return v
    except (ValueError, TypeError):
        return None


def _to_decimal(valor):
    if valor in (None, '', 'None'):
        return None
    if isinstance(valor, (int, float, Decimal)):
        return Decimal(str(valor))
    s = str(valor).strip().replace('R$', '').replace(' ', '')
    # Aceita "1.234,56" e "1234.56"
    if s.count(',') == 1 and (s.count('.') == 0 or s.rfind('.') < s.rfind(',')):
        s = s.replace('.', '').replace(',', '.')
    try:
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return None


def _to_date(valor):
    if not valor:
        return None
    if isinstance(valor, date):
        return valor
    for fmt in ('%Y-%m-%d', '%d/%m/%Y'):
        try:
            return datetime.strptime(str(valor).strip(), fmt).date()
        except ValueError:
            continue
    return None


# ---------------------------------------------------------------------------
# Helper: registrar histórico
# ---------------------------------------------------------------------------
def _registrar_historico(lead, campo, valor_antes, valor_depois, descricao=None):
    """Adiciona uma linha em lead_historico (sem commitar — quem chama
    decide o commit)."""
    if str(valor_antes or '') == str(valor_depois or '') and not descricao:
        return None
    entrada = LeadHistorico(
        lead_id=lead.id,
        admin_id=lead.admin_id,
        campo=campo,
        valor_antes=str(valor_antes) if valor_antes is not None else None,
        valor_depois=str(valor_depois) if valor_depois is not None else None,
        descricao=descricao,
        usuario_id=(current_user.id if current_user.is_authenticated else None),
    )
    db.session.add(entrada)
    return entrada


# ---------------------------------------------------------------------------
# Helper: vincular ou criar Cliente
# ---------------------------------------------------------------------------
def vincular_ou_criar_cliente(lead, admin_id):
    """Procura cliente do mesmo admin com mesmo telefone normalizado OU
    e-mail (case-insensitive); se achar, vincula e completa campos vazios.
    Se não, cria novo. Retorna (cliente, "vinculado"|"criado"|"insuficiente"|"existente").

    Não comita — quem chama controla a transação. A rotina pressupõe que
    `lead.id` não precisa estar materializado (atribui vínculo direto).
    """
    if lead.cliente_id:
        return (lead.cliente, 'existente')
    nome = (lead.nome or '').strip()
    tel_norm = _normalizar_telefone(lead.contato)
    email_norm = (lead.email or '').strip().lower() or None
    if not nome or (not tel_norm and not email_norm):
        return (None, 'insuficiente')

    cliente = None
    q = Cliente.query.filter_by(admin_id=admin_id)
    if tel_norm:
        # Busca por dígitos: comparamos o telefone do cliente também
        # normalizado (regex em SQL é caro; faremos em Python sobre o
        # subset mais provável).
        candidatos_tel = q.filter(Cliente.telefone.isnot(None)).all()
        for c in candidatos_tel:
            if _normalizar_telefone(c.telefone) == tel_norm:
                cliente = c
                break
    if not cliente and email_norm:
        cliente = (
            Cliente.query.filter_by(admin_id=admin_id)
            .filter(db.func.lower(Cliente.email) == email_norm)
            .first()
        )

    if cliente:
        # Completa campos vazios sem sobrescrever
        if not (cliente.email or '').strip() and lead.email:
            cliente.email = lead.email
        if not (cliente.telefone or '').strip() and lead.contato:
            cliente.telefone = lead.contato
        if not (cliente.endereco or '').strip():
            partes = [p for p in (lead.localizacao, lead.detalhes_localizacao) if p]
            if partes:
                cliente.endereco = ' — '.join(partes)
        lead.cliente_id = cliente.id
        return (cliente, 'vinculado')

    # Cria novo
    partes = [p for p in (lead.localizacao, lead.detalhes_localizacao) if p]
    cliente = Cliente(
        nome=nome,
        email=lead.email or None,
        telefone=lead.contato or None,
        endereco=(' — '.join(partes) if partes else None),
        admin_id=admin_id,
    )
    db.session.add(cliente)
    db.session.flush()  # garantir id
    lead.cliente_id = cliente.id
    return (cliente, 'criado')


# ---------------------------------------------------------------------------
# Helper: aplicar automações na transição de status
# ---------------------------------------------------------------------------
STATUS_VALIDOS = {s.value for s in LeadStatus}


def _aplicar_automacoes_status(lead, status_anterior):
    """Chamado APÓS atualizar lead.status. Faz:
    - Status muda para "Enviado" → preenche data_envio (se vazia).
    - Status muda → atualiza status_changed_at.
    Retorna lista de mensagens (flash) para o usuário.
    """
    avisos = []
    novo = lead.status
    if novo != status_anterior:
        lead.status_changed_at = datetime.utcnow()
        _registrar_historico(lead, 'status', status_anterior, novo)

    if novo == LeadStatus.ENVIADO.value and not lead.data_envio:
        lead.data_envio = date.today()
        avisos.append('Data de envio preenchida automaticamente.')
    return avisos


# ---------------------------------------------------------------------------
# Listas auxiliares para popular dropdowns (apenas itens ATIVOS)
# ---------------------------------------------------------------------------
def _listas_para_form(admin_id):
    """Retorna dict com items ativos de cada lista, para popular selects.
    Itens inativos são EXCLUÍDOS para que não apareçam em novos cadastros,
    mas continuam disponíveis para exibição em leads antigos.
    """
    out = {}
    for slug, info in LISTAS_MESTRAS.items():
        out[slug] = (
            info['model'].query
            .filter_by(admin_id=admin_id, ativo=True)
            .order_by(info['model'].nome.asc())
            .all()
        )
    return out


# ===========================================================================
# ROTAS — KANBAN E LISTA
# ===========================================================================
@crm_bp.route('/')
@login_required
def kanban():
    admin_id = get_admin_id()
    if not admin_id:
        flash('Erro de autenticação', 'danger')
        return redirect(url_for('main.index'))

    filtros = _coletar_filtros()
    q = _query_leads_visiveis(admin_id)
    q = _aplicar_filtros(q, filtros)

    leads = q.order_by(Lead.status_changed_at.desc()).all()

    # Agrupa por status (na ordem do enum)
    colunas = []
    for status_enum in LeadStatus:
        # "Congelado" e "Perdido" aparecem por último na visão padrão,
        # mas mostramos todos no kanban.
        leads_col = [l for l in leads if l.status == status_enum.value]
        valor_total = sum(
            (l.valor_proposta or Decimal('0')) for l in leads_col
        )
        colunas.append({
            'enum': status_enum,
            'nome': status_enum.value,
            'leads': leads_col,
            'count': len(leads_col),
            'valor_total': valor_total,
        })

    listas = _listas_para_form(admin_id)
    return render_template(
        'crm/kanban.html',
        colunas=colunas,
        listas=listas,
        filtros=filtros,
        is_admin=is_admin_user(),
    )


@crm_bp.route('/lista')
@login_required
def lista():
    admin_id = get_admin_id()
    if not admin_id:
        flash('Erro de autenticação', 'danger')
        return redirect(url_for('main.index'))

    filtros = _coletar_filtros()
    q = _query_leads_visiveis(admin_id)
    q = _aplicar_filtros(q, filtros)

    leads = q.order_by(Lead.data_chegada.desc(), Lead.id.desc()).all()
    listas = _listas_para_form(admin_id)
    return render_template(
        'crm/lista.html',
        leads=leads,
        listas=listas,
        filtros=filtros,
        status_enum=LeadStatus,
        is_admin=is_admin_user(),
    )


def _coletar_filtros():
    return {
        'busca': (request.args.get('busca') or '').strip(),
        'responsavel_id': _to_int(request.args.get('responsavel_id')),
        'origem_id': _to_int(request.args.get('origem_id')),
        'tipo_material_id': _to_int(request.args.get('tipo_material_id')),
        'tipo_obra_id': _to_int(request.args.get('tipo_obra_id')),
        'mes_chegada': (request.args.get('mes_chegada') or '').strip(),
    }


def _aplicar_filtros(q, filtros):
    if filtros['busca']:
        like = f"%{filtros['busca']}%"
        q = q.filter(or_(
            Lead.nome.ilike(like),
            Lead.contato.ilike(like),
            Lead.email.ilike(like),
            Lead.demanda.ilike(like),
            Lead.localizacao.ilike(like),
        ))
    if filtros['responsavel_id']:
        q = q.filter(Lead.responsavel_id == filtros['responsavel_id'])
    if filtros['origem_id']:
        q = q.filter(Lead.origem_id == filtros['origem_id'])
    if filtros['tipo_material_id']:
        q = q.filter(Lead.tipo_material_id == filtros['tipo_material_id'])
    if filtros['tipo_obra_id']:
        q = q.filter(Lead.tipo_obra_id == filtros['tipo_obra_id'])
    if filtros['mes_chegada']:
        # formato esperado YYYY-MM
        try:
            ano, mes = filtros['mes_chegada'].split('-')
            ano = int(ano); mes = int(mes)
            q = q.filter(
                db.extract('year', Lead.data_chegada) == ano,
                db.extract('month', Lead.data_chegada) == mes,
            )
        except (ValueError, TypeError):
            pass
    return q


# ===========================================================================
# ROTAS — CRUD DO LEAD
# ===========================================================================
@crm_bp.route('/novo', methods=['GET', 'POST'])
@login_required
def novo():
    admin_id = get_admin_id()
    if not admin_id:
        flash('Erro de autenticação', 'danger')
        return redirect(url_for('main.index'))

    listas = _listas_para_form(admin_id)

    if request.method == 'POST':
        return _salvar_lead(None, listas, admin_id)

    # GET: pré-popular a partir de query string (vinda de outra tela)
    pre = {
        'nome': request.args.get('nome', ''),
        'contato': request.args.get('contato', ''),
        'email': request.args.get('email', ''),
        'demanda': request.args.get('demanda', ''),
    }
    return render_template(
        'crm/lead_form.html',
        lead=None,
        listas=listas,
        valores=pre,
        status_enum=LeadStatus,
    )


@crm_bp.route('/<int:lead_id>/editar', methods=['GET', 'POST'])
@login_required
def editar(lead_id):
    admin_id = get_admin_id()
    if not admin_id:
        flash('Erro de autenticação', 'danger')
        return redirect(url_for('main.index'))

    lead = Lead.query.filter_by(id=lead_id, admin_id=admin_id).first_or_404()
    if not _pode_acessar_lead(lead):
        abort(403)

    listas = _listas_para_form(admin_id)
    if request.method == 'POST':
        return _salvar_lead(lead, listas, admin_id)
    return render_template(
        'crm/lead_form.html',
        lead=lead,
        listas=listas,
        valores=lead,
        status_enum=LeadStatus,
    )


def _salvar_lead(lead, listas, admin_id):
    """Lógica unificada de criar/editar lead a partir do form POST."""
    novo_status = (request.form.get('status') or LeadStatus.EM_FILA.value).strip()
    if novo_status not in STATUS_VALIDOS:
        novo_status = LeadStatus.EM_FILA.value

    # Validar TODAS as FKs contra o tenant (anti-cross-tenant injection)
    fk_responsavel    = _validar_fk_tenant('responsavel',    request.form.get('responsavel_id'),    admin_id)
    fk_origem         = _validar_fk_tenant('origem',         request.form.get('origem_id'),         admin_id)
    fk_cadencia       = _validar_fk_tenant('cadencia',       request.form.get('cadencia_id'),       admin_id)
    fk_situacao       = _validar_fk_tenant('situacao',       request.form.get('situacao_id'),       admin_id)
    fk_tipo_material  = _validar_fk_tenant('tipo_material',  request.form.get('tipo_material_id'),  admin_id)
    fk_tipo_obra      = _validar_fk_tenant('tipo_obra',      request.form.get('tipo_obra_id'),      admin_id)
    motivo_perda_id   = _validar_fk_tenant('motivo_perda',   request.form.get('motivo_perda_id'),   admin_id)

    # Validações
    nome = (request.form.get('nome') or '').strip()
    if not nome:
        flash('Nome é obrigatório.', 'danger')
        return render_template(
            'crm/lead_form.html', lead=lead, listas=listas,
            valores=request.form, status_enum=LeadStatus,
        )
    if novo_status == LeadStatus.PERDIDO.value and not motivo_perda_id:
        flash('Para marcar como Perdido, selecione o Motivo da Perda.', 'danger')
        return render_template(
            'crm/lead_form.html', lead=lead, listas=listas,
            valores=request.form, status_enum=LeadStatus,
        )

    is_new = lead is None
    status_anterior = None if is_new else lead.status

    if is_new:
        lead = Lead(
            admin_id=admin_id,
            criado_por_id=current_user.id,
            data_chegada=_to_date(request.form.get('data_chegada')) or date.today(),
        )
        db.session.add(lead)
    else:
        nova_data_chegada = _to_date(request.form.get('data_chegada'))
        if nova_data_chegada:
            lead.data_chegada = nova_data_chegada

    lead.nome = nome
    lead.contato = (request.form.get('contato') or '').strip() or None
    lead.email = (request.form.get('email') or '').strip() or None
    lead.responsavel_id = fk_responsavel
    lead.origem_id = fk_origem
    lead.cadencia_id = fk_cadencia
    lead.situacao_id = fk_situacao
    lead.tipo_material_id = fk_tipo_material
    lead.tipo_obra_id = fk_tipo_obra
    lead.motivo_perda_id = motivo_perda_id
    lead.localizacao = (request.form.get('localizacao') or '').strip() or None
    lead.detalhes_localizacao = (request.form.get('detalhes_localizacao') or '').strip() or None
    lead.demanda = (request.form.get('demanda') or '').strip() or None
    lead.pasta = (request.form.get('pasta') or '').strip() or None
    lead.valor_proposta = _to_decimal(request.form.get('valor_proposta'))
    lead.observacao = (request.form.get('observacao') or '').strip() or None
    lead.data_envio = _to_date(request.form.get('data_envio'))
    lead.data_retomada = _to_date(request.form.get('data_retomada'))
    lead.status = novo_status

    # Garantir id antes de mexer em FK / histórico
    db.session.flush()

    # Vínculo automático de Cliente
    cliente_msg = None
    if not lead.cliente_id:
        cliente, status_vinc = vincular_ou_criar_cliente(lead, admin_id)
        if status_vinc == 'criado':
            cliente_msg = f'Novo cliente criado: {cliente.nome}'
            _registrar_historico(lead, 'sistema', None, cliente.nome,
                                 descricao=f'Cliente criado automaticamente: {cliente.nome}')
        elif status_vinc == 'vinculado':
            cliente_msg = f'Cliente vinculado: {cliente.nome}'
            _registrar_historico(lead, 'sistema', None, cliente.nome,
                                 descricao=f'Cliente existente vinculado: {cliente.nome}')

    # Automações de status
    if is_new:
        _registrar_historico(lead, 'status', None, novo_status,
                             descricao='Lead criado.')
        lead.status_changed_at = datetime.utcnow()
        if novo_status == LeadStatus.ENVIADO.value and not lead.data_envio:
            lead.data_envio = date.today()
    else:
        avisos = _aplicar_automacoes_status(lead, status_anterior)
        for a in avisos:
            flash(a, 'info')

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        logger.exception('Erro ao salvar lead')
        flash(f'Erro ao salvar lead: {e}', 'danger')
        return render_template(
            'crm/lead_form.html', lead=lead, listas=listas,
            valores=request.form, status_enum=LeadStatus,
        )

    if cliente_msg:
        flash(cliente_msg, 'success')
    flash(f'Lead "{lead.nome}" salvo com sucesso.', 'success')
    return redirect(url_for('crm.editar', lead_id=lead.id))


@crm_bp.route('/<int:lead_id>/excluir', methods=['POST'])
@login_required
def excluir(lead_id):
    admin_id = get_admin_id()
    if not admin_id or not is_admin_user():
        flash('Apenas administradores podem excluir leads.', 'danger')
        return redirect(url_for('crm.kanban'))

    lead = Lead.query.filter_by(id=lead_id, admin_id=admin_id).first_or_404()
    nome = lead.nome
    try:
        db.session.delete(lead)
        db.session.commit()
        flash(f'Lead "{nome}" excluído.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao excluir: {e}', 'danger')
    return redirect(url_for('crm.kanban'))


# ===========================================================================
# DRAG-AND-DROP DO KANBAN
# ===========================================================================
@crm_bp.route('/<int:lead_id>/mudar_status', methods=['POST'])
@login_required
def mudar_status(lead_id):
    """Endpoint chamado pelo Sortable.js no kanban.
    Aceita JSON ou form-data com {status: <novo_status>, motivo_perda_id?: <id>}.
    """
    admin_id = get_admin_id()
    if not admin_id:
        return jsonify({'ok': False, 'msg': 'Não autenticado'}), 401

    lead = Lead.query.filter_by(id=lead_id, admin_id=admin_id).first()
    if not lead:
        return jsonify({'ok': False, 'msg': 'Lead não encontrado'}), 404

    if not _pode_acessar_lead(lead):
        return jsonify({'ok': False, 'msg': 'Sem permissão'}), 403

    payload = request.get_json(silent=True) or request.form
    novo_status = (payload.get('status') or '').strip()
    if novo_status not in STATUS_VALIDOS:
        return jsonify({'ok': False, 'msg': f'Status inválido: {novo_status}'}), 400

    if novo_status == LeadStatus.PERDIDO.value:
        # Aceitar motivo embutido no payload OU lead já com motivo —
        # validando contra o tenant em ambos os casos
        novo_motivo = _validar_fk_tenant('motivo_perda', payload.get('motivo_perda_id'), admin_id)
        motivo_id = novo_motivo or lead.motivo_perda_id
        if not motivo_id:
            return jsonify({
                'ok': False,
                'requer_motivo': True,
                'msg': 'Para marcar como Perdido, informe o Motivo da Perda.',
            }), 422
        lead.motivo_perda_id = motivo_id

    status_anterior = lead.status
    lead.status = novo_status
    avisos = _aplicar_automacoes_status(lead, status_anterior)

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        logger.exception('Erro ao mudar status')
        return jsonify({'ok': False, 'msg': f'Erro ao salvar: {e}'}), 500

    return jsonify({
        'ok': True,
        'lead_id': lead.id,
        'status': lead.status,
        'avisos': avisos,
    })


# ===========================================================================
# AÇÕES RÁPIDAS — pré-preencher Proposta / Obra
# ===========================================================================
@crm_bp.route('/<int:lead_id>/gerar_proposta')
@login_required
def gerar_proposta(lead_id):
    """Redireciona para a tela de nova proposta com os campos pré-preenchidos
    via query string. NÃO cria proposta silenciosamente — só pré-preenche."""
    admin_id = get_admin_id()
    if not admin_id:
        flash('Erro de autenticação', 'danger')
        return redirect(url_for('main.index'))
    lead = Lead.query.filter_by(id=lead_id, admin_id=admin_id).first_or_404()
    if not _pode_acessar_lead(lead):
        abort(403)

    # Garante que existe Cliente vinculado
    if not lead.cliente_id:
        cliente, vinc = vincular_ou_criar_cliente(lead, admin_id)
        if cliente:
            try:
                db.session.commit()
                flash(
                    f'Cliente {("criado" if vinc=="criado" else "vinculado")}: {cliente.nome}',
                    'success',
                )
            except Exception:
                db.session.rollback()

    flash(
        f'Crie a proposta para o lead "{lead.nome}". '
        'Após salvar, vincule esta proposta ao lead pela edição do lead.',
        'info',
    )
    # Procura o blueprint propostas.nova_proposta
    try:
        url = url_for('propostas.nova_proposta',
                      cliente_id=lead.cliente_id,
                      lead_id=lead.id)
    except Exception:
        url = '/propostas/nova'
    return redirect(url)


@crm_bp.route('/<int:lead_id>/criar_obra')
@login_required
def criar_obra(lead_id):
    """Redireciona para a tela de nova obra com pré-preenchimento."""
    admin_id = get_admin_id()
    if not admin_id:
        flash('Erro de autenticação', 'danger')
        return redirect(url_for('main.index'))
    lead = Lead.query.filter_by(id=lead_id, admin_id=admin_id).first_or_404()
    if not _pode_acessar_lead(lead):
        abort(403)

    if not lead.cliente_id:
        cliente, vinc = vincular_ou_criar_cliente(lead, admin_id)
        if cliente:
            try:
                db.session.commit()
                flash(
                    f'Cliente {("criado" if vinc=="criado" else "vinculado")}: {cliente.nome}',
                    'success',
                )
            except Exception:
                db.session.rollback()

    flash(
        f'Crie a obra para o lead "{lead.nome}".',
        'info',
    )
    # Procura rota nova obra
    for endpoint in ('main.nova_obra', 'obras.nova', 'obras.criar'):
        try:
            url = url_for(endpoint, cliente_id=lead.cliente_id, lead_id=lead.id)
            return redirect(url)
        except Exception:
            continue
    return redirect(url_for('crm.editar', lead_id=lead.id))


# ===========================================================================
# CADASTROS DAS LISTAS MESTRAS
# ===========================================================================
@crm_bp.route('/cadastros')
@login_required
def cadastros():
    admin_id = get_admin_id()
    if not admin_id or not is_admin_user():
        flash('Apenas administradores podem editar os Cadastros do CRM.', 'danger')
        return redirect(url_for('crm.kanban'))

    dados = {}
    for slug, info in LISTAS_MESTRAS.items():
        dados[slug] = {
            'label': info['label'],
            'itens': (
                info['model'].query
                .filter_by(admin_id=admin_id)
                .order_by(info['model'].ativo.desc(), info['model'].nome.asc())
                .all()
            ),
        }
    return render_template('crm/cadastros.html', dados=dados, listas_mestras=LISTAS_MESTRAS)


@crm_bp.route('/cadastros/<slug>/criar', methods=['POST'])
@login_required
def cadastros_criar(slug):
    admin_id = get_admin_id()
    if not admin_id or not is_admin_user():
        return _redir_cadastros_erro('Sem permissão.')
    info = LISTAS_MESTRAS.get(slug)
    if not info:
        return _redir_cadastros_erro('Lista inválida.')
    nome = (request.form.get('nome') or '').strip()
    if not nome:
        flash('Nome é obrigatório.', 'danger')
        return redirect(url_for('crm.cadastros') + f'#{slug}')
    if len(nome) > 120:
        flash('Nome muito longo (máx 120).', 'danger')
        return redirect(url_for('crm.cadastros') + f'#{slug}')
    existe = info['model'].query.filter_by(admin_id=admin_id, nome=nome).first()
    if existe:
        flash(f'Já existe "{nome}" em {info["label"]}.', 'warning')
        return redirect(url_for('crm.cadastros') + f'#{slug}')
    item = info['model'](admin_id=admin_id, nome=nome, ativo=True)
    db.session.add(item)
    try:
        db.session.commit()
        flash(f'Adicionado "{nome}" em {info["label"]}.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao adicionar: {e}', 'danger')
    return redirect(url_for('crm.cadastros') + f'#{slug}')


@crm_bp.route('/cadastros/<slug>/<int:item_id>/editar', methods=['POST'])
@login_required
def cadastros_editar(slug, item_id):
    admin_id = get_admin_id()
    if not admin_id or not is_admin_user():
        return _redir_cadastros_erro('Sem permissão.')
    info = LISTAS_MESTRAS.get(slug)
    if not info:
        return _redir_cadastros_erro('Lista inválida.')
    item = info['model'].query.filter_by(id=item_id, admin_id=admin_id).first_or_404()
    novo_nome = (request.form.get('nome') or '').strip()
    if not novo_nome:
        flash('Nome é obrigatório.', 'danger')
        return redirect(url_for('crm.cadastros') + f'#{slug}')
    if novo_nome != item.nome:
        conflito = (
            info['model'].query
            .filter_by(admin_id=admin_id, nome=novo_nome)
            .filter(info['model'].id != item.id)
            .first()
        )
        if conflito:
            flash(f'Já existe outro item com o nome "{novo_nome}".', 'warning')
            return redirect(url_for('crm.cadastros') + f'#{slug}')
        item.nome = novo_nome
        try:
            db.session.commit()
            flash(f'Renomeado para "{novo_nome}".', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Erro: {e}', 'danger')
    return redirect(url_for('crm.cadastros') + f'#{slug}')


@crm_bp.route('/cadastros/<slug>/<int:item_id>/toggle_ativo', methods=['POST'])
@login_required
def cadastros_toggle(slug, item_id):
    admin_id = get_admin_id()
    if not admin_id or not is_admin_user():
        return _redir_cadastros_erro('Sem permissão.')
    info = LISTAS_MESTRAS.get(slug)
    if not info:
        return _redir_cadastros_erro('Lista inválida.')
    item = info['model'].query.filter_by(id=item_id, admin_id=admin_id).first_or_404()
    item.ativo = not item.ativo
    try:
        db.session.commit()
        flash(
            f'"{item.nome}" {"reativado" if item.ativo else "desativado"}.',
            'success',
        )
    except Exception as e:
        db.session.rollback()
        flash(f'Erro: {e}', 'danger')
    return redirect(url_for('crm.cadastros') + f'#{slug}')


@crm_bp.route('/cadastros/<slug>/<int:item_id>/excluir', methods=['POST'])
@login_required
def cadastros_excluir(slug, item_id):
    admin_id = get_admin_id()
    if not admin_id or not is_admin_user():
        return _redir_cadastros_erro('Sem permissão.')
    info = LISTAS_MESTRAS.get(slug)
    if not info:
        return _redir_cadastros_erro('Lista inválida.')
    item = info['model'].query.filter_by(id=item_id, admin_id=admin_id).first_or_404()
    campo_lead = info['campo_lead']
    em_uso = (
        Lead.query
        .filter_by(admin_id=admin_id)
        .filter(getattr(Lead, campo_lead) == item.id)
        .first()
    )
    if em_uso:
        flash(
            f'Não é possível excluir "{item.nome}" — está em uso por leads. '
            'Use "Desativar" para parar de aparecer em novos leads.',
            'warning',
        )
        return redirect(url_for('crm.cadastros') + f'#{slug}')
    nome = item.nome
    try:
        db.session.delete(item)
        db.session.commit()
        flash(f'"{nome}" excluído.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro: {e}', 'danger')
    return redirect(url_for('crm.cadastros') + f'#{slug}')


def _redir_cadastros_erro(msg):
    flash(msg, 'danger')
    return redirect(url_for('crm.cadastros'))


# ===========================================================================
# IMPORTAÇÃO / EXPORTAÇÃO EXCEL
# ===========================================================================

CABECALHOS_MODELO = [
    'Status', 'Data de Chegada', 'Data de Envio', 'Pasta', 'Nome',
    'Localizaçao', 'Detalhes Loc.', 'Tipo de obra', 'Contato Lead',
    'Responsável', 'Demanda', 'Origem', 'Cadência', 'Situação',
    'Tipo de Material', 'Valor da Proposta', 'Motivo da Perda', 'Observação',
]

EXEMPLO_LINHAS = [
    [
        'Em andamento', '15/03/2025', '', 'Pasta A', 'João Silva',
        'São Paulo', 'Zona Sul', 'Residencial', '(11) 99999-1234',
        'Maria', 'Reforma completa', 'Indicação', 'Semanal', 'Visita agendada',
        'Concreto', '85000', '', 'Cliente muito interessado',
    ],
    [
        'Enviado', '20/04/2025', '25/04/2025', '', 'Empresa ABC Ltda',
        'Campinas', 'Centro', 'Comercial', '(19) 3333-5678',
        'Carlos', 'Construção nova sede', 'Google', 'Quinzenal', 'Proposta enviada',
        'Alvenaria', '320000', '', '',
    ],
]

STATUS_MAP = {
    'perdido': 'Perdido',
    'ganho': 'Aprovado',
    'aprovado': 'Aprovado',
    'congelado': 'Congelado',
    'andamento': 'Em andamento',
    'em andamento': 'Em andamento',
    'enviado': 'Enviado',
    'validação': 'Validação',
    'validacao': 'Validação',
    'feedback': 'Feedback',
}

LOOKUP_COLS = {
    'Responsável':      ('responsavel',   CrmResponsavel),
    'Origem':           ('origem',        CrmOrigem),
    'Cadência':         ('cadencia',      CrmCadencia),
    'Situação':         ('situacao',      CrmSituacao),
    'Tipo de Material': ('tipo_material', CrmTipoMaterial),
    'Tipo de obra':     ('tipo_obra',     CrmTipoObra),
    'Motivo da Perda':  ('motivo_perda',  CrmMotivoPerda),
}


def _resolve_ou_criar_lookup(model, nome_val, admin_id, cache):
    """Busca ou cria item de lista mestra. cache = {tablename:nome_lower: id}."""
    if not nome_val:
        return None
    nome_val = str(nome_val).strip()
    if not nome_val:
        return None
    ck = f'{model.__tablename__}:{nome_val.lower()}'
    if ck in cache:
        return cache[ck]
    item = model.query.filter_by(admin_id=admin_id, nome=nome_val).first()
    if not item:
        item = model(admin_id=admin_id, nome=nome_val, ativo=True)
        db.session.add(item)
        db.session.flush()
    cache[ck] = item.id
    return item.id


def _normalizar_status(raw):
    if not raw:
        return LeadStatus.EM_FILA.value
    key = str(raw).strip().lower()
    mapped = STATUS_MAP.get(key)
    if mapped:
        for s in LeadStatus:
            if s.value == mapped:
                return s.value
    return LeadStatus.EM_FILA.value


@crm_bp.route('/exportar_modelo')
@login_required
def exportar_modelo():
    """Gera e devolve um .xlsx com aba 'Lead.2026' (modelo de importação)."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    admin_id = get_admin_id()
    if not admin_id or not is_admin_user():
        flash('Apenas administradores podem exportar o modelo.', 'danger')
        return redirect(url_for('crm.lista'))

    wb = openpyxl.Workbook()

    # ---- Aba Lead.2026 ----
    ws = wb.active
    ws.title = 'Lead.2026'

    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=11)

    for col_idx, cab in enumerate(CABECALHOS_MODELO, start=1):
        cell = ws.cell(row=1, column=col_idx, value=cab)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

    example_fill = PatternFill(start_color='EBF3FB', end_color='EBF3FB', fill_type='solid')
    for row_idx, linha in enumerate(EXEMPLO_LINHAS, start=2):
        for col_idx, val in enumerate(linha, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = example_fill

    ws.freeze_panes = 'A2'
    ws.row_dimensions[1].height = 30

    # ---- Aba Instruções ----
    wi = wb.create_sheet('Instruções')
    wi.column_dimensions['A'].width = 22
    wi.column_dimensions['B'].width = 60

    titulo_font = Font(bold=True, size=12)
    secao_font = Font(bold=True, size=11, color='1F4E79')

    wi['A1'] = 'INSTRUÇÕES DE IMPORTAÇÃO'
    wi['A1'].font = titulo_font

    wi['A3'] = 'Campo obrigatório:'
    wi['A3'].font = secao_font
    wi['B3'] = 'Nome  (linhas sem Nome são ignoradas)'

    wi['A5'] = 'Datas:'
    wi['A5'].font = secao_font
    wi['B5'] = 'Formato DD/MM/AAAA  (ex: 15/03/2025)'

    wi['A7'] = 'Valor da Proposta:'
    wi['A7'].font = secao_font
    wi['B7'] = 'Número sem símbolo (ex: 85000 ou 85000.50)'

    wi['A9'] = 'Valores válidos de Status:'
    wi['A9'].font = secao_font

    status_info = [
        ('Em fila', '(padrão quando em branco ou não reconhecido)'),
        ('Em andamento', ''),
        ('Enviado', ''),
        ('Validação', ''),
        ('Aprovado', '(também aceita "Ganho")'),
        ('Feedback', ''),
        ('Congelado', ''),
        ('Perdido', ''),
    ]
    for i, (val, obs) in enumerate(status_info, start=10):
        wi.cell(row=i, column=1, value=val)
        wi.cell(row=i, column=2, value=obs)

    linha_obs = 10 + len(status_info) + 1
    wi.cell(row=linha_obs, column=1, value='Campos lookup:').font = secao_font
    wi.cell(row=linha_obs, column=2,
            value='Responsável, Origem, Cadência, Situação, Tipo de Material, '
                  'Tipo de obra, Motivo da Perda — se o valor não existir nos cadastros, '
                  'será criado automaticamente.')

    # ---- Enviar como download ----
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        download_name='modelo_leads_crm.xlsx',
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@crm_bp.route('/importar', methods=['POST'])
@login_required
def importar():
    """Importa leads a partir de um .xlsx com aba 'Lead.2026'."""
    import openpyxl

    admin_id = get_admin_id()
    if not admin_id or not is_admin_user():
        flash('Apenas administradores podem importar leads.', 'danger')
        return redirect(url_for('crm.lista'))

    arquivo = request.files.get('arquivo')
    if not arquivo or not arquivo.filename:
        flash('Selecione um arquivo .xlsx para importar.', 'danger')
        return redirect(url_for('crm.lista'))

    try:
        wb = openpyxl.load_workbook(arquivo, read_only=True, data_only=True)
    except Exception as e:
        flash(f'Não foi possível abrir o arquivo: {e}', 'danger')
        return redirect(url_for('crm.lista'))

    if 'Lead.2026' not in wb.sheetnames:
        flash(
            'O arquivo não contém a aba "Lead.2026". '
            'Use o modelo correto (botão "Baixar Modelo").',
            'danger',
        )
        return redirect(url_for('crm.lista'))

    ws = wb['Lead.2026']
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        flash('A aba "Lead.2026" está vazia.', 'warning')
        return redirect(url_for('crm.lista'))

    # Mapear cabeçalhos pelo nome (case-insensitive)
    header_row = rows[0]
    col_map = {}
    for idx, h in enumerate(header_row):
        if h is not None:
            col_map[str(h).strip().lower()] = idx

    def _get_raw(row, col_name):
        """Retorna o valor nativo da célula (preserva date/datetime do Excel)."""
        idx = col_map.get(col_name.lower())
        if idx is None:
            return None
        return row[idx] if idx < len(row) else None

    def _get_str(row, col_name):
        """Retorna o valor como string limpa (None quando vazio)."""
        val = _get_raw(row, col_name)
        if val is None:
            return None
        s = str(val).strip()
        return s if s else None

    lookup_cache = {}
    importados = 0
    ignorados_sem_nome = 0
    duplicatas = 0
    erros = 0

    for row in rows[1:]:
        if all(c is None for c in row):
            continue

        nome = _get_str(row, 'nome')
        if not nome:
            ignorados_sem_nome += 1
            continue

        contato_raw = _get_str(row, 'contato lead')
        data_chegada = _to_date(_get_raw(row, 'data de chegada'))
        if not data_chegada:
            data_chegada = date.today()

        dup_q = Lead.query.filter_by(
            admin_id=admin_id,
            nome=nome,
            data_chegada=data_chegada,
        )
        if contato_raw:
            dup_q = dup_q.filter(Lead.contato == contato_raw)
        else:
            dup_q = dup_q.filter(
                or_(Lead.contato.is_(None), Lead.contato == '')
            )
        if dup_q.first():
            duplicatas += 1
            continue

        sp = db.session.begin_nested()
        try:
            lead = Lead(
                admin_id=admin_id,
                nome=nome,
                data_chegada=data_chegada,
                data_envio=_to_date(_get_raw(row, 'data de envio')),
                pasta=_get_str(row, 'pasta'),
                localizacao=_get_str(row, 'localizaçao'),
                detalhes_localizacao=_get_str(row, 'detalhes loc.'),
                contato=contato_raw,
                demanda=_get_str(row, 'demanda'),
                valor_proposta=_to_decimal(_get_str(row, 'valor da proposta')),
                observacao=_get_str(row, 'observação'),
                status=_normalizar_status(_get_str(row, 'status')),
                criado_por_id=current_user.id,
            )

            resp_nome = _get_str(row, 'responsável')
            if resp_nome:
                lead.responsavel_id = _resolve_ou_criar_lookup(
                    CrmResponsavel, resp_nome, admin_id, lookup_cache)

            orig_nome = _get_str(row, 'origem')
            if orig_nome:
                lead.origem_id = _resolve_ou_criar_lookup(
                    CrmOrigem, orig_nome, admin_id, lookup_cache)

            cad_nome = _get_str(row, 'cadência')
            if cad_nome:
                lead.cadencia_id = _resolve_ou_criar_lookup(
                    CrmCadencia, cad_nome, admin_id, lookup_cache)

            sit_nome = _get_str(row, 'situação')
            if sit_nome:
                lead.situacao_id = _resolve_ou_criar_lookup(
                    CrmSituacao, sit_nome, admin_id, lookup_cache)

            mat_nome = _get_str(row, 'tipo de material')
            if mat_nome:
                lead.tipo_material_id = _resolve_ou_criar_lookup(
                    CrmTipoMaterial, mat_nome, admin_id, lookup_cache)

            obra_nome = _get_str(row, 'tipo de obra')
            if obra_nome:
                lead.tipo_obra_id = _resolve_ou_criar_lookup(
                    CrmTipoObra, obra_nome, admin_id, lookup_cache)

            motivo_nome = _get_str(row, 'motivo da perda')
            if motivo_nome:
                lead.motivo_perda_id = _resolve_ou_criar_lookup(
                    CrmMotivoPerda, motivo_nome, admin_id, lookup_cache)

            if lead.status == LeadStatus.ENVIADO.value and not lead.data_envio:
                lead.data_envio = date.today()

            db.session.add(lead)
            db.session.flush()
            sp.commit()
            importados += 1

        except Exception as e:
            logger.warning(f'Erro ao importar linha (nome={nome}): {e}')
            sp.rollback()
            erros += 1

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        logger.exception('Erro ao commitar importação de leads')
        flash(f'Erro ao salvar leads: {e}', 'danger')
        return redirect(url_for('crm.lista'))

    flash(
        f'{importados} leads importados, {ignorados_sem_nome} ignorados (sem nome), '
        f'{duplicatas} já existiam' +
        (f', {erros} com erro (ver log)' if erros else '') + '.',
        'success' if importados > 0 else 'warning',
    )
    return redirect(url_for('crm.lista'))
