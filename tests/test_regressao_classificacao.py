"""
Fase C (TDD) — Teste de REGRESSÃO da classificação de Fluxo de Caixa.

Prova que o motor NOVO (classificador + REGRAS_SISTEMA semeado) classifica
IGUAL ao motor ANTIGO (`_classificar_categoria_nomeada`) sobre o Excel real.
Ambos recebem o MESMO tem_obra, então qualquer divergência é lógica de
classificação — que não pode mudar (ADR-0002).

Módulo puro (sem DB): lê o .xlsx e chama as duas funções puras.
"""
import os
import sys
from collections import Counter
from datetime import datetime as dt, date

import pytest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from services.importacao_excel import _classificar_categoria_nomeada, _normalizar
from services.classificador_cadastro import classificar, Lancamento, Contexto
from services.seed_palavras_chave import (
    regras_sistema, FALLBACK_ENTRADA, FALLBACK_SAIDA,
)

XLSX = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "1. FLUXO DE CAIXA_Veks Engenharia.xlsx",
)

_CC_ADMIN = ['escritorio', 'guilherme e ariane', 'guilherme', 'ariane',
             'administrativo', 'geral', 'head', 'holding']


def _tem_obra(cc):
    """Proxy determinístico de tem_obra a partir do centro de custo.
    Alimenta IGUALMENTE os dois motores, então não enviesa a comparação."""
    cc_norm = _normalizar(cc)
    if not cc_norm:
        return False
    return not any(x in cc_norm for x in _CC_ADMIN)


def _ler_lancamentos():
    import openpyxl
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    lancs = []  # (tipo, plano, descricao, entidade, tem_obra)

    if 'Entrada' in wb.sheetnames:
        for row in wb['Entrada'].iter_rows(min_row=6, values_only=True):
            dv = row[0]
            if not dv or not isinstance(dv, (dt, date)):
                continue
            plano = row[1] or ''
            cliente = row[2] or ''
            desc = row[3] or ''
            lancs.append(('ENTRADA', str(plano), str(desc), str(cliente), False))

    if 'Saída' in wb.sheetnames:
        for row in wb['Saída'].iter_rows(min_row=6, values_only=True):
            dv = row[1]
            if not dv or not isinstance(dv, (dt, date)):
                continue
            plano = row[2] if len(row) > 2 and row[2] else ''
            fornecedor = row[3] if len(row) > 3 and row[3] else ''
            desc = row[4] if len(row) > 4 and row[4] else ''
            cc = row[5] if len(row) > 5 and row[5] else ''
            lancs.append(('SAIDA', str(plano), str(desc), str(fornecedor), _tem_obra(cc)))
    return lancs


@pytest.fixture(scope="module")
def lancamentos():
    if not os.path.exists(XLSX):
        pytest.skip(f"Excel não encontrado: {XLSX}")
    return _ler_lancamentos()


def _classificar_novo(ctx, tipo, plano, descricao, entidade, tem_obra):
    lanc = Lancamento(descricao=descricao, fornecedor=entidade, plano=plano,
                      tem_obra=tem_obra, tipo=tipo)
    v = classificar(lanc, ctx)
    if not v.eh_pendente:
        return v.categoria_nome
    return FALLBACK_ENTRADA if tipo == 'ENTRADA' else FALLBACK_SAIDA


def test_motor_novo_reproduz_o_antigo_no_excel_real(lancamentos):
    ctx = Contexto(regras=regras_sistema(), memoria_exata={})

    # Invariante RESCUE-ONLY: o motor novo pode MELHORAR (resgatar um lançamento
    # que o legado jogou no fallback genérico para uma categoria específica), mas
    # NUNCA pode mudar uma classificação específica do legado. Regras de resgate são
    # de baixa precedência (ficam por último no seed), então só pegam o que cairia
    # em 'Outras Saídas'/'Outros Recebimentos'. Ver seed _SAIDA (resgates curados).
    FALLBACK = {FALLBACK_ENTRADA, FALLBACK_SAIDA}

    resgates = Counter()      # (fallback antigo, novo específico) → contagem — PERMITIDO
    regressoes = Counter()    # (antigo específico, novo) → contagem — PROIBIDO
    exemplos = {}
    total = 0

    for tipo, plano, descricao, entidade, tem_obra in lancamentos:
        total += 1
        antigo = _classificar_categoria_nomeada(tipo, plano, descricao, entidade, tem_obra)
        novo = _classificar_novo(ctx, tipo, plano, descricao, entidade, tem_obra)
        if antigo == novo:
            continue
        if antigo in FALLBACK:
            resgates[(antigo, novo)] += 1          # legado desistiu; novo classificou → OK
        else:
            regressoes[(antigo, novo)] += 1        # legado tinha categoria específica → NÃO pode mudar
        exemplos.setdefault((antigo, novo), (descricao[:60], entidade[:30], plano[:30]))

    if resgates:
        print(f"\n=== RESGATES (melhorias, permitidos): {sum(resgates.values())} ===")
        for (antigo, novo), n in resgates.most_common(20):
            print(f"  {n:4d}  [{antigo}] → [{novo}]  ex: {exemplos[(antigo, novo)]}")
    if regressoes:
        print(f"\n=== REGRESSÕES (proibidas): {sum(regressoes.values())} ===")
        for (antigo, novo), n in regressoes.most_common(30):
            print(f"  {n:4d}  [{antigo}] → [{novo}]  ex: {exemplos[(antigo, novo)]}")

    n_reg = sum(regressoes.values())
    assert n_reg == 0, (f"{n_reg}/{total} classificações ESPECÍFICAS do legado foram "
                        f"alteradas (regressão). Resgates de fallback são permitidos.")
