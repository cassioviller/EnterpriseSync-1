"""Onda 3 — o ponto que não vira hora, e o RDO que cobra quem saiu.

Quatro defeitos de UM tema só: **um lançamento que entra e não sai, ou uma
hora que entra e não vira número.** O arreio de tenant é
`tests/helpers_tenant.py` e o de dinheiro é `tests/helpers_dinheiro.py` — o
banco de desenvolvimento é compartilhado, então TODA contagem aqui é escopada
pelo `admin_id` do tenant do próprio teste. Contagem global mediria o mundo.
"""
import os
import sys
from datetime import date, time
from io import BytesIO

import pytest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import main  # noqa: F401 — registra blueprints e handlers de evento
from app import app, db
from helpers_dinheiro import (custo_diario, filhos_mao_de_obra,  # noqa: F401
                              linhas_mao_de_obra, mao_de_obra, rdo_da_obra,
                              soma)
from helpers_tenant import cliente_de, um_tenant  # noqa: F401

pytestmark = pytest.mark.integration

# Dia sem nenhum RegistroPonto semeado — ver o docstring de `um_tenant`.
DIA = date(2026, 6, 15)
OUTRO_DIA = date(2026, 6, 16)


@pytest.fixture(autouse=True)
def _config():
    app.config['TESTING'] = True
    app.config['WTF_CSRF_ENABLED'] = False
    if not app.secret_key:
        app.secret_key = 'test-onda3-ponto-rdo'
    yield


def _registro(tenant, data):
    """O RegistroPonto do tenant naquele dia — escopado, sempre relido."""
    from models import RegistroPonto
    db.session.expire_all()
    return RegistroPonto.query.filter_by(
        admin_id=tenant.admin_id, funcionario_id=tenant.funcionario_id,
        data=data).first()


# ---------------------------------------------------------------------------
# Defeito 1 — a importação Excel que nunca calculava hora (ponto_views.py:1487)
# ---------------------------------------------------------------------------

def _planilha_de_ponto(codigo, linhas):
    """Uma planilha no formato que `PontoExcelService.validar_e_importar` lê.

    Colunas: A=Data, B=tipo, C=obra_id, D=entrada, E=saída, F=almoço início,
    G=almoço fim (`services/ponto_importacao.py:576-595`). O nome da aba
    precisa de ' - ' e do código antes dele — é assim que o parser acha o
    funcionário. O sufixo é curto de propósito: openpyxl recusa título de aba
    com mais de 31 caracteres.
    """
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = f'{codigo} - F'
    ws.cell(row=1, column=1, value='Data')
    for i, (dia, tipo, obra_id, ent, sai, alm_i, alm_f) in enumerate(linhas):
        linha = 2 + i
        ws.cell(row=linha, column=1, value=dia.strftime('%d/%m/%Y'))
        ws.cell(row=linha, column=2, value=tipo)
        ws.cell(row=linha, column=3, value=obra_id)
        ws.cell(row=linha, column=4, value=ent)
        ws.cell(row=linha, column=5, value=sai)
        ws.cell(row=linha, column=6, value=alm_i)
        ws.cell(row=linha, column=7, value=alm_f)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def test_importacao_excel_calcula_horas_e_preserva_obra_no_update():
    """🔴 O mês importado marcava 0h — nos dois ramos da rota.

    `processar_importacao` (`ponto_views.py:1496-1516`) gravava os horários e
    commitava sem passar por `PontoService._calcular_horas`: a folha lia
    `horas_trabalhadas = 0` e cobrava todo dia como falta cheia, e nenhum
    custo de obra saía. O ramo de ATUALIZAÇÃO ainda descartava `obra_id` e
    `tipo_registro` — reimportar a planilha corrigida apagava a obra do
    registro que já existia.
    """
    from models import Funcionario, RegistroPonto

    with app.app_context():
        t = um_tenant('onda3_imp', data_ref=DIA, com_fatos=False)
        codigo = db.session.get(Funcionario, t.funcionario_id).codigo

        # O dia que já tem registro (sem obra e sem horário): o ramo de update.
        db.session.add(RegistroPonto(
            funcionario_id=t.funcionario_id, admin_id=t.admin_id,
            data=OUTRO_DIA))
        db.session.commit()

        planilha = _planilha_de_ponto(codigo, [
            (DIA, 'TRAB', t.obra_id, '08:00', '17:00', '12:00', '13:00'),
            (OUTRO_DIA, 'TRAB', t.obra_id, '08:00', '17:00', '12:00', '13:00'),
        ])

        resposta = cliente_de(t.admin_id).post(
            '/ponto/importar/processar',
            data={'arquivo': (planilha, 'ponto.xlsx')},
            content_type='multipart/form-data')
        assert resposta.status_code in (200, 302), (
            f'rota respondeu {resposta.status_code}')

        novo = _registro(t, DIA)
        assert novo is not None, 'o registro novo não foi importado'
        assert novo.hora_entrada == time(8, 0)
        assert novo.horas_trabalhadas == pytest.approx(8.0), (
            f'8h-17h com 1h de almoço são 8h; veio '
            f'{novo.horas_trabalhadas} — a folha cobraria falta cheia')

        atualizado = _registro(t, OUTRO_DIA)
        assert atualizado is not None
        assert atualizado.hora_entrada == time(8, 0)
        assert atualizado.horas_trabalhadas == pytest.approx(8.0), (
            f'o ramo de atualização também precisa calcular; veio '
            f'{atualizado.horas_trabalhadas}')
        assert atualizado.obra_id == t.obra_id, (
            'o ramo de atualização descartou a obra da planilha — sem obra '
            'não há custo de obra')
        assert atualizado.tipo_registro == 'TRAB', (
            f'o ramo de atualização descartou o tipo; veio '
            f'{atualizado.tipo_registro!r}')


# ---------------------------------------------------------------------------
# Defeito 2 — as duas rotas de ponto facial (ponto_views.py:2446)
# ---------------------------------------------------------------------------

def _jornada_aberta(tenant, data):
    """Registro do dia com entrada e almoço batidos, faltando só a saída.

    O ponto facial bate UM horário por chamada. Sem entrada gravada antes, a
    saída sozinha não tem o que calcular — e o teste mediria a ausência de
    dado em vez do cálculo.
    """
    from models import RegistroPonto
    registro = RegistroPonto(
        funcionario_id=tenant.funcionario_id, admin_id=tenant.admin_id,
        obra_id=tenant.obra_id, data=data,
        hora_entrada=time(8, 0), hora_almoco_saida=time(12, 0),
        hora_almoco_retorno=time(13, 0))
    db.session.add(registro)
    db.session.commit()
    return registro


def _relogio_fixo(monkeypatch, data, hora):
    """Pina o relógio das rotas de ponto — sem isso a hora da saída é a hora
    em que a suíte roda, e 8h de jornada viram um número diferente por
    execução."""
    import ponto_views
    monkeypatch.setattr(ponto_views, 'get_date_brasil', lambda: data)
    monkeypatch.setattr(ponto_views, 'get_time_brasil', lambda: hora)


def _foto_falsa(tenant):
    """Foto cadastrada no funcionário — o suficiente para as rotas o
    considerarem candidato. A comparação facial em si é dublada."""
    from models import Funcionario
    funcionario = db.session.get(Funcionario, tenant.funcionario_id)
    funcionario.foto_base64 = 'ZmFrZS1mb3RvLWRlLXRlc3Rl'
    db.session.commit()
    return funcionario


def test_registrar_ponto_facial_calcula_as_horas(monkeypatch):
    """🔴 `/api/registrar-facial` commitava sem `PontoService._calcular_horas`.

    O ponto batido pelo rosto ficava com `horas_trabalhadas = 0`: a mesma
    jornada valia 8h quando batida pelo `PontoService` (`ponto_service.py:138`)
    e 0h quando batida pela câmera.
    """
    import ponto_views

    with app.app_context():
        t = um_tenant('onda3_face', data_ref=DIA, com_fatos=False)
        _foto_falsa(t)
        _jornada_aberta(t, DIA)
        _relogio_fixo(monkeypatch, DIA, time(17, 0))
        monkeypatch.setattr(ponto_views, 'validar_qualidade_foto',
                            lambda *a, **k: (True, 'ok'))
        monkeypatch.setattr(ponto_views, 'comparar_faces_deepface',
                            lambda *a, **k: (True, 0.10, None))

        resposta = cliente_de(t.admin_id).post(
            '/ponto/api/registrar-facial',
            json={'funcionario_id': t.funcionario_id,
                  'foto_base64': 'ZmFrZQ==',
                  'obra_id': t.obra_id,
                  'tipo_ponto': 'saida'})
        assert resposta.status_code == 200, resposta.get_data(as_text=True)
        assert resposta.get_json()['success'] is True

        registro = _registro(t, DIA)
        assert registro.hora_saida == time(17, 0)
        assert registro.horas_trabalhadas == pytest.approx(8.0), (
            f'8h-17h com 1h de almoço são 8h; veio '
            f'{registro.horas_trabalhadas}')


def test_identificar_e_registrar_calcula_horas_e_emite_evento(monkeypatch):
    """🔴 `/api/identificar-e-registrar` não calculava hora NEM emitia evento.

    Além do zero de horas, esta era a única rota de ponto que nunca emitia
    `ponto_registrado` — o evento que lança o custo do diarista
    (`event_manager.py`). Ponto batido no totem da obra não virava custo
    nenhum.
    """
    import ponto_views
    from event_manager import EventManager

    emitidos = []

    with app.app_context():
        t = um_tenant('onda3_ident', data_ref=DIA, com_fatos=False)
        _foto_falsa(t)
        _jornada_aberta(t, DIA)
        _relogio_fixo(monkeypatch, DIA, time(17, 0))
        monkeypatch.setattr(ponto_views, 'validar_qualidade_foto_avancada',
                            lambda *a, **k: (True, 'ok', {}))
        monkeypatch.setattr(ponto_views, 'identificar_por_cache',
                            lambda *a, **k: (None, None, 'cache vazio'))
        monkeypatch.setattr(ponto_views, 'reconhecer_com_multiplas_fotos',
                            lambda *a, **k: (True, 0.10, 'foto principal'))
        monkeypatch.setattr(
            EventManager, 'emit',
            classmethod(lambda cls, nome, dados, admin_id, **k:
                        emitidos.append((nome, dados, admin_id))))

        resposta = cliente_de(t.admin_id).post(
            '/ponto/api/identificar-e-registrar',
            json={'foto_base64': 'ZmFrZQ==',
                  'obra_id': t.obra_id,
                  'tipo_ponto': 'saida'})
        assert resposta.status_code == 200, resposta.get_data(as_text=True)
        assert resposta.get_json()['success'] is True

        registro = _registro(t, DIA)
        assert registro.hora_saida == time(17, 0)
        assert registro.horas_trabalhadas == pytest.approx(8.0), (
            f'8h-17h com 1h de almoço são 8h; veio '
            f'{registro.horas_trabalhadas}')

        pontos = [e for e in emitidos if e[0] == 'ponto_registrado']
        assert pontos, ('a rota não emitiu ponto_registrado — o custo do '
                        'diarista nunca é lançado')
        assert pontos[0][1]['registro_id'] == registro.id
        assert pontos[0][1]['tipo_ponto'] == 'saida'
        assert pontos[0][2] == t.admin_id


# ---------------------------------------------------------------------------
# Defeito 3 — a edição unificada que não estorna (views/rdo.py:2867)
# ---------------------------------------------------------------------------

def _servico_do_tenant(tenant):
    """Um `Servico` ativo do tenant.

    `POST /rdo/salvar` sem campo de subatividade cai num fallback que pega o
    PRIMEIRO `Servico` do admin; sem nenhum, ela dá flash+redirect ANTES de
    parsear a mão de obra, e o teste mediria o abandono da rota.
    """
    from models import Servico
    servico = Servico(nome=f'Servico {tenant.marca}', categoria='geral',
                      unidade_medida='un', custo_unitario=100.0, ativo=True,
                      admin_id=tenant.admin_id)
    db.session.add(servico)
    db.session.commit()
    return servico


def _segundo_funcionario(tenant):
    """O colega que vai SAIR do RDO na edição."""
    import uuid
    from models import Funcionario
    funcionario = Funcionario(
        codigo=f'B{uuid.uuid4().hex[:8].upper()}',
        nome=f'Colega {tenant.marca}',
        cpf=f'{uuid.uuid4().int % 10**11:011d}',
        data_admissao=date(2026, 1, 2), admin_id=tenant.admin_id, ativo=True,
        salario=3000.0, tipo_remuneracao='salario', valor_diaria=0.0)
    db.session.add(funcionario)
    db.session.commit()
    return funcionario


def _custos_obra_do_rdo(tenant, rdo_id):
    """`CustoObra` do RDO — a TERCEIRA tabela que o Submeter escreve.

    `event_manager.py:932-961` grava uma linha por trabalhador, chaveada em
    (`rdo_id`, `funcionario_id`, `data`, `admin_id`). Ela não passa por
    `remover_custos_rdo` nem por `remover_custo_diario_rdo` — quem estorna só
    aquelas duas deixa esta de pé. Escopada pelo tenant, como todo coletor
    deste arquivo.
    """
    from models import CustoObra
    db.session.expire_all()
    return (CustoObra.query
            .filter_by(admin_id=tenant.admin_id, rdo_id=rdo_id)
            .order_by(CustoObra.id).all())


def _submeter(tenant, rdo_id):
    """`POST /rdo/<id>/finalizar` — o Submeter, que é quem lança o custo."""
    resposta = cliente_de(tenant.admin_id).post(f'/rdo/{rdo_id}/finalizar')
    assert resposta.status_code in (200, 302), (
        f'finalizar respondeu {resposta.status_code}')
    db.session.expire_all()


def test_edicao_unificada_estorna_o_custo_de_quem_saiu():
    """🔴 A edição apagava `RDOMaoObra` em bloco e deixava o custo de pé.

    `rdo_salvar_unificado` (`views/rdo.py:2864`) fazia
    `RDOMaoObra.query.filter_by(rdo_id=...).delete()` **sem**
    `remover_custos_rdo`/`remover_custo_diario_rdo` — e como os dois casam o
    lançamento pelo id da linha que acabou de ser apagada, o custo do
    trabalhador removido virava órfão insuprimível: ele seguia sendo cobrado
    na obra para sempre.
    """
    from models import RDOMaoObra

    with app.app_context():
        t = um_tenant('onda3_edit', data_ref=DIA, com_fatos=False,
                      tipo_remuneracao='salario', valor_diaria=0.0,
                      salario=3000.0)
        _servico_do_tenant(t)
        colega = _segundo_funcionario(t)

        rdo = rdo_da_obra(t, DIA)
        mao_de_obra(rdo, t, horas=8.0)
        db.session.add(RDOMaoObra(
            rdo_id=rdo.id, admin_id=t.admin_id, funcionario_id=colega.id,
            funcao_exercida='Servente', horas_trabalhadas=8.0))
        db.session.commit()
        rdo_id = rdo.id

        _submeter(t, rdo_id)
        antes = filhos_mao_de_obra(t)
        assert len(antes) == 2, (
            f'o cenário precisa dos DOIS custos lançados; veio {len(antes)}')
        valor_de_um = soma(antes) / 2
        assert len(_custos_obra_do_rdo(t, rdo_id)) == 2, (
            'o cenário precisa das DUAS linhas de CustoObra lançadas')

        # A edição que remove o colega: só o funcionário do tenant volta no
        # formulário (campos achatados do path A, `views/rdo.py:3307`).
        cli = cliente_de(t.admin_id)
        resposta = cli.post('/rdo/salvar', data={
            'rdo_id': str(rdo_id),
            'obra_id': str(t.obra_id),
            'data_relatorio': DIA.isoformat(),
            f'funcionario_{t.funcionario_id}_nome': f'Funcionario {t.marca}',
            f'funcionario_{t.funcionario_id}_horas': '8',
        })
        assert resposta.status_code in (200, 302), (
            f'salvar respondeu {resposta.status_code}')

        equipe = linhas_mao_de_obra(rdo_id)
        assert [l.funcionario_id for l in equipe] == [t.funcionario_id], (
            'o cenário depende de a edição ter mesmo removido o colega')

        depois = filhos_mao_de_obra(t)
        assert len(depois) == 1, (
            f'o custo do trabalhador removido continua no razão: esperava 1 '
            f'lançamento de mão de obra, encontrou {len(depois)}')
        assert soma(depois) == pytest.approx(valor_de_um, abs=0.01)
        assert len(custo_diario(rdo_id)) == 1, (
            'o RDOCustoDiario do removido também precisa sair — é dele que o '
            'lançamento nasce de novo')
        assert [c.funcionario_id for c in _custos_obra_do_rdo(t, rdo_id)] == [
            t.funcionario_id], (
            'o CustoObra do trabalhador removido continua cobrando a obra — '
            'ele não passa por remover_custos_rdo nem por '
            'remover_custo_diario_rdo')


# ---------------------------------------------------------------------------
# Defeito 4 — reabrir desfazia o cronograma e deixava o custo (views/rdo.py:1888)
# ---------------------------------------------------------------------------

def test_reabrir_rdo_estorna_o_custo_de_mao_de_obra():
    """🔴 O Reabrir desfazia o percentual e deixava o custo no razão.

    A simetria Submeter/Reabrir foi desenhada em
    `docs/superpowers/plans/2026-08-24-rascunho-nao-move-cronograma.md`
    (`95eb585f`) e implementada só na metade do AVANÇO: `reabrir_rdo` chamava
    `recalcular_percentuais_do_rdo` e nada mais. O custo de mão de obra que o
    Submeter lançou ficava de pé com o RDO de volta em `rascunho` — o estado
    que, por `services.rdo_ciclo_vida.publica_custos`, não pode ter custo
    nenhum. Reprocessar é estornar e refazer.
    """
    from models import RDO
    from services.rdo_ciclo_vida import PREENCHIDO, RASCUNHO, estado_de

    with app.app_context():
        t = um_tenant('onda3_reab', data_ref=DIA, com_fatos=False,
                      tipo_remuneracao='salario', valor_diaria=0.0,
                      salario=3000.0)
        rdo = rdo_da_obra(t, DIA)
        mao_de_obra(rdo, t, horas=8.0)
        rdo_id = rdo.id

        _submeter(t, rdo_id)
        assert estado_de(db.session.get(RDO, rdo_id)) == PREENCHIDO
        assert len(filhos_mao_de_obra(t)) == 1, (
            'o cenário precisa do custo lançado pelo Submeter')
        assert len(custo_diario(rdo_id)) == 1
        assert len(_custos_obra_do_rdo(t, rdo_id)) == 1

        resposta = cliente_de(t.admin_id).post(
            f'/rdo/{rdo_id}/reabrir', data={'motivo': 'horas erradas'})
        assert resposta.status_code in (200, 302), (
            f'reabrir respondeu {resposta.status_code}')

        db.session.expire_all()
        assert estado_de(db.session.get(RDO, rdo_id)) == RASCUNHO, (
            'o cenário depende de a reabertura ter acontecido')
        assert filhos_mao_de_obra(t) == [], (
            'o RDO voltou para rascunho e o custo de mão de obra continua no '
            'razão — rascunho não lança custo')
        assert custo_diario(rdo_id) == [], (
            'o RDOCustoDiario é a fonte do lançamento: sem estorná-lo, o '
            'próximo Submeter reaproveita o valor velho')
        assert _custos_obra_do_rdo(t, rdo_id) == [], (
            'o CustoObra do RDO reaberto continua cobrando a obra — rascunho '
            'não lança custo em tabela nenhuma')

        # A outra metade de "reprocessar é estornar e REFAZER": o estorno não
        # pode deixar o custo irrecuperável. Submetendo de novo, ele volta —
        # uma vez só, nas TRÊS tabelas.
        _submeter(t, rdo_id)
        assert len(filhos_mao_de_obra(t)) == 1, (
            'depois de estornar, o Submeter seguinte precisa relançar o '
            'custo — e uma vez só')
        assert len(custo_diario(rdo_id)) == 1, (
            'o RDOCustoDiario também precisa voltar')
        assert len(_custos_obra_do_rdo(t, rdo_id)) == 1, (
            'o CustoObra também precisa voltar')
