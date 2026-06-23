"""
Catálogos — Central de dados auxiliares gerenciados pelo administrador.
Blueprint: catalogos_bp  prefix: /catalogos
"""
import io
import logging
from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file, jsonify
from flask_login import login_required, current_user
from app import db
from models import (
    GrupoFinanceiro, CategoriaFluxoCaixa, CategoriaFornecedor, CategoriaReembolso,
    FluxoCaixa, TipoUsuario, PalavraChaveCategoria, PalavraChaveSugestao
)

logger = logging.getLogger(__name__)

catalogos_bp = Blueprint('catalogos', __name__, url_prefix='/catalogos')


def _get_admin_id():
    if not current_user.is_authenticated:
        return None
    if current_user.tipo_usuario not in (TipoUsuario.SUPER_ADMIN, TipoUsuario.ADMIN):
        return None
    if hasattr(current_user, 'admin_id') and current_user.admin_id:
        return current_user.admin_id
    return current_user.id


def _require_admin():
    admin_id = _get_admin_id()
    if not admin_id:
        flash('Acesso restrito a administradores.', 'danger')
        return None, redirect(url_for('main.index'))
    return admin_id, None


# ============================================================
# HUB
# ============================================================

@catalogos_bp.route('/')
@login_required
def hub():
    admin_id, err = _require_admin()
    if err:
        return err
    from services.dropdown_service import get_grupos_por_modulo, MODULOS_LABELS
    grupos_por_modulo = get_grupos_por_modulo(admin_id)
    return render_template('catalogos/hub.html',
                           grupos_por_modulo=grupos_por_modulo,
                           modulos_labels=MODULOS_LABELS)


# ============================================================
# CategoriaFluxoCaixa
# ============================================================

@catalogos_bp.route('/categorias-fluxo-caixa')
@login_required
def categorias_fluxo_caixa():
    admin_id, err = _require_admin()
    if err:
        return err
    cats = CategoriaFluxoCaixa.query.filter_by(admin_id=admin_id).order_by(CategoriaFluxoCaixa.nome).all()
    return render_template('catalogos/categorias_fluxo_caixa.html', categorias=cats)


@catalogos_bp.route('/categorias-fluxo-caixa/criar', methods=['GET', 'POST'])
@login_required
def categorias_fluxo_caixa_criar():
    admin_id, err = _require_admin()
    if err:
        return err
    if request.method == 'POST':
        nome = request.form.get('nome', '').strip()
        tipo = request.form.get('tipo', 'SAIDA').strip()
        grupo_id_raw = request.form.get('grupo_financeiro_id', '').strip()
        descricao = request.form.get('descricao', '').strip()
        if not nome:
            flash('Nome é obrigatório.', 'danger')
            return redirect(url_for('catalogos.categorias_fluxo_caixa_criar'))
        if tipo not in ('ENTRADA', 'SAIDA'):
            tipo = 'SAIDA'
        grupo_id = int(grupo_id_raw) if grupo_id_raw.isdigit() else None
        grupo_nome = None
        if grupo_id:
            gf = GrupoFinanceiro.query.filter_by(id=grupo_id, admin_id=admin_id).first()
            if gf:
                grupo_nome = gf.nome
            else:
                grupo_id = None
        cat = CategoriaFluxoCaixa(
            nome=nome, tipo=tipo,
            grupo_financeiro=grupo_nome,
            grupo_financeiro_id=grupo_id,
            descricao=descricao or None,
            ativo=True, admin_id=admin_id
        )
        db.session.add(cat)
        db.session.commit()
        flash(f'Categoria "{nome}" criada com sucesso!', 'success')
        return redirect(url_for('catalogos.categorias_fluxo_caixa'))
    grupos = GrupoFinanceiro.query.filter_by(admin_id=admin_id, ativo=True).order_by(GrupoFinanceiro.tipo, GrupoFinanceiro.nome).all()
    return render_template('catalogos/categorias_fluxo_caixa_form.html', categoria=None, grupos=grupos)


@catalogos_bp.route('/categorias-fluxo-caixa/editar/<int:id>', methods=['GET', 'POST'])
@login_required
def categorias_fluxo_caixa_editar(id):
    admin_id, err = _require_admin()
    if err:
        return err
    cat = CategoriaFluxoCaixa.query.filter_by(id=id, admin_id=admin_id).first_or_404()
    if request.method == 'POST':
        nome = request.form.get('nome', '').strip()
        tipo = request.form.get('tipo', 'SAIDA').strip()
        grupo_id_raw = request.form.get('grupo_financeiro_id', '').strip()
        descricao = request.form.get('descricao', '').strip()
        if not nome:
            flash('Nome é obrigatório.', 'danger')
            return redirect(url_for('catalogos.categorias_fluxo_caixa_editar', id=id))
        if tipo not in ('ENTRADA', 'SAIDA'):
            tipo = cat.tipo
        grupo_id = int(grupo_id_raw) if grupo_id_raw.isdigit() else None
        grupo_nome = None
        if grupo_id:
            gf = GrupoFinanceiro.query.filter_by(id=grupo_id, admin_id=admin_id).first()
            if gf:
                grupo_nome = gf.nome
            else:
                grupo_id = None
        cat.nome = nome
        cat.tipo = tipo
        cat.grupo_financeiro = grupo_nome
        cat.grupo_financeiro_id = grupo_id
        cat.descricao = descricao or None
        db.session.commit()
        flash(f'Categoria "{nome}" atualizada com sucesso!', 'success')
        return redirect(url_for('catalogos.categorias_fluxo_caixa'))
    grupos = GrupoFinanceiro.query.filter_by(admin_id=admin_id, ativo=True).order_by(GrupoFinanceiro.tipo, GrupoFinanceiro.nome).all()
    return render_template('catalogos/categorias_fluxo_caixa_form.html', categoria=cat, grupos=grupos)


@catalogos_bp.route('/categorias-fluxo-caixa/toggle/<int:id>', methods=['POST'])
@login_required
def categorias_fluxo_caixa_toggle(id):
    admin_id, err = _require_admin()
    if err:
        return err
    cat = CategoriaFluxoCaixa.query.filter_by(id=id, admin_id=admin_id).first_or_404()
    cat.ativo = not cat.ativo
    db.session.commit()
    estado = 'ativada' if cat.ativo else 'desativada'
    flash(f'Categoria "{cat.nome}" {estado}.', 'success')
    return redirect(url_for('catalogos.categorias_fluxo_caixa'))


@catalogos_bp.route('/categorias-fluxo-caixa/exportar-modelo', methods=['GET'])
@login_required
def categorias_fluxo_caixa_exportar_modelo():
    admin_id, err = _require_admin()
    if err:
        return err
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Categorias'
    headers = ['Nome', 'Tipo', 'Grupo Financeiro', 'Descrição']
    header_fill = PatternFill(start_color='1e3a5f', end_color='1e3a5f', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    note_cell = ws.cell(row=2, column=1, value='NOTA: O campo Tipo deve ser exatamente ENTRADA ou SAIDA (sem acento).')
    note_cell.font = Font(italic=True, color='888888')
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 28
    ws.column_dimensions['D'].width = 45
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name='modelo_categorias_fluxo_caixa.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@catalogos_bp.route('/categorias-fluxo-caixa/exportar-atuais', methods=['GET'])
@login_required
def categorias_fluxo_caixa_exportar_atuais():
    admin_id, err = _require_admin()
    if err:
        return err
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    categorias = CategoriaFluxoCaixa.query.filter_by(admin_id=admin_id, ativo=True).order_by(
        CategoriaFluxoCaixa.tipo, CategoriaFluxoCaixa.nome
    ).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Categorias'
    headers = ['Nome', 'Tipo', 'Grupo Financeiro', 'Descrição', 'Ativo']
    header_fill = PatternFill(start_color='1e3a5f', end_color='1e3a5f', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    for cat in categorias:
        ws.append([
            cat.nome,
            cat.tipo,
            cat.grupo_financeiro or '',
            cat.descricao or '',
            'TRUE' if cat.ativo else 'FALSE',
        ])
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 28
    ws.column_dimensions['D'].width = 45
    ws.column_dimensions['E'].width = 8
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name=f'categorias_fluxo_caixa_{admin_id}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@catalogos_bp.route('/categorias-fluxo-caixa/importar', methods=['POST'])
@login_required
def categorias_fluxo_caixa_importar():
    admin_id, err = _require_admin()
    if err:
        return err
    arquivo = request.files.get('arquivo')
    if not arquivo or arquivo.filename == '':
        flash('Nenhum arquivo selecionado.', 'warning')
        return redirect(url_for('catalogos.categorias_fluxo_caixa'))
    ext = arquivo.filename.rsplit('.', 1)[-1].lower() if '.' in arquivo.filename else ''
    if ext != 'xlsx':
        flash('Formato inválido. Use arquivo .xlsx.', 'danger')
        return redirect(url_for('catalogos.categorias_fluxo_caixa'))
    try:
        import openpyxl
        wb = openpyxl.load_workbook(arquivo, data_only=True)
        ws = wb.active
        header_row = None
        col_nome = col_tipo = col_grupo = col_desc = None
        for rn in range(1, min(5, ws.max_row + 1)):
            vals = [str(ws.cell(row=rn, column=c).value or '').strip().lower() for c in range(1, ws.max_column + 1)]
            if 'nome' in vals:
                header_row = rn
                col_nome = next((i + 1 for i, v in enumerate(vals) if v == 'nome'), None)
                col_tipo = next((i + 1 for i, v in enumerate(vals) if v == 'tipo'), None)
                col_grupo = next((i + 1 for i, v in enumerate(vals) if 'grupo' in v), None)
                col_desc = next((i + 1 for i, v in enumerate(vals) if 'descri' in v), None)
                break
        if not header_row or not col_nome or not col_tipo:
            flash('Cabeçalho não encontrado. Use o modelo disponível para download.', 'danger')
            return redirect(url_for('catalogos.categorias_fluxo_caixa'))
        criadas = 0
        atualizadas = 0
        ignoradas = 0
        for rn in range(header_row + 1, ws.max_row + 1):
            nome = str(ws.cell(row=rn, column=col_nome).value or '').strip()
            tipo = str(ws.cell(row=rn, column=col_tipo).value or '').strip().upper()
            if not nome or nome.lower().startswith('nota:'):
                continue
            if tipo not in ('ENTRADA', 'SAIDA'):
                ignoradas += 1
                continue
            grupo_nome = str(ws.cell(row=rn, column=col_grupo).value or '').strip() if col_grupo else ''
            desc = str(ws.cell(row=rn, column=col_desc).value or '').strip() if col_desc else ''
            # Resolver grupo financeiro (cria se não existir)
            grupo_id = None
            if grupo_nome:
                gf = GrupoFinanceiro.query.filter(
                    GrupoFinanceiro.admin_id == admin_id,
                    db.func.lower(GrupoFinanceiro.nome) == grupo_nome.lower(),
                    GrupoFinanceiro.tipo == tipo,
                ).first()
                if not gf:
                    gf = GrupoFinanceiro(nome=grupo_nome, tipo=tipo, ativo=True, admin_id=admin_id)
                    db.session.add(gf)
                    db.session.flush()
                grupo_id = gf.id
            existe = CategoriaFluxoCaixa.query.filter(
                CategoriaFluxoCaixa.admin_id == admin_id,
                db.func.lower(CategoriaFluxoCaixa.nome) == nome.lower(),
                CategoriaFluxoCaixa.tipo == tipo,
            ).first()
            if existe:
                # Upsert: atualiza campos editáveis sem alterar nome, tipo, created_at
                if grupo_nome:
                    existe.grupo_financeiro = grupo_nome
                    existe.grupo_financeiro_id = grupo_id
                if desc:
                    existe.descricao = desc
                atualizadas += 1
                continue
            cat = CategoriaFluxoCaixa(
                nome=nome, tipo=tipo,
                grupo_financeiro=grupo_nome or None,
                grupo_financeiro_id=grupo_id,
                descricao=desc or None,
                ativo=True, admin_id=admin_id,
            )
            db.session.add(cat)
            criadas += 1
        db.session.commit()
        partes = []
        if criadas:
            partes.append(f'{criadas} nova(s) inserida(s)')
        if atualizadas:
            partes.append(f'{atualizadas} atualizada(s)')
        if ignoradas:
            partes.append(f'{ignoradas} ignorada(s) (tipo inválido)')
        flash(', '.join(partes) + '.' if partes else 'Nenhuma linha processada.', 'success' if (criadas or atualizadas) else 'warning')
    except Exception as e:
        logger.error(f'[IMPORTAR_CATEGORIAS] {e}', exc_info=True)
        flash(f'Erro ao processar arquivo: {e}', 'danger')
    return redirect(url_for('catalogos.categorias_fluxo_caixa'))


@catalogos_bp.route('/categorias-fluxo-caixa/deletar/<int:id>', methods=['POST'])
@login_required
def categorias_fluxo_caixa_deletar(id):
    admin_id, err = _require_admin()
    if err:
        return err
    cat = CategoriaFluxoCaixa.query.filter_by(id=id, admin_id=admin_id).first_or_404()
    em_uso = FluxoCaixa.query.filter_by(admin_id=admin_id, categoria_fluxo_caixa_id=cat.id).first()
    if em_uso:
        flash(f'Não é possível excluir: a categoria "{cat.nome}" está em uso por lançamentos de Fluxo de Caixa.', 'danger')
        return redirect(url_for('catalogos.categorias_fluxo_caixa'))
    db.session.delete(cat)
    db.session.commit()
    flash(f'Categoria "{cat.nome}" excluída.', 'success')
    return redirect(url_for('catalogos.categorias_fluxo_caixa'))


# ============================================================
# CategoriaFornecedor
# ============================================================

@catalogos_bp.route('/categorias-fornecedor')
@login_required
def categorias_fornecedor():
    admin_id, err = _require_admin()
    if err:
        return err
    cats = CategoriaFornecedor.query.filter_by(admin_id=admin_id).order_by(CategoriaFornecedor.nome).all()
    return render_template('catalogos/categorias_fornecedor.html', categorias=cats)


@catalogos_bp.route('/categorias-fornecedor/criar', methods=['GET', 'POST'])
@login_required
def categorias_fornecedor_criar():
    admin_id, err = _require_admin()
    if err:
        return err
    if request.method == 'POST':
        nome = request.form.get('nome', '').strip()
        descricao = request.form.get('descricao', '').strip()
        if not nome:
            flash('Nome é obrigatório.', 'danger')
            return redirect(url_for('catalogos.categorias_fornecedor_criar'))
        cat = CategoriaFornecedor(nome=nome, descricao=descricao or None, ativo=True, admin_id=admin_id)
        db.session.add(cat)
        db.session.commit()
        flash(f'Categoria "{nome}" criada com sucesso!', 'success')
        return redirect(url_for('catalogos.categorias_fornecedor'))
    return render_template('catalogos/categorias_fornecedor_form.html', categoria=None)


@catalogos_bp.route('/categorias-fornecedor/editar/<int:id>', methods=['GET', 'POST'])
@login_required
def categorias_fornecedor_editar(id):
    admin_id, err = _require_admin()
    if err:
        return err
    cat = CategoriaFornecedor.query.filter_by(id=id, admin_id=admin_id).first_or_404()
    if request.method == 'POST':
        nome = request.form.get('nome', '').strip()
        descricao = request.form.get('descricao', '').strip()
        if not nome:
            flash('Nome é obrigatório.', 'danger')
            return redirect(url_for('catalogos.categorias_fornecedor_editar', id=id))
        cat.nome = nome
        cat.descricao = descricao or None
        db.session.commit()
        flash(f'Categoria "{nome}" atualizada com sucesso!', 'success')
        return redirect(url_for('catalogos.categorias_fornecedor'))
    return render_template('catalogos/categorias_fornecedor_form.html', categoria=cat)


@catalogos_bp.route('/categorias-fornecedor/toggle/<int:id>', methods=['POST'])
@login_required
def categorias_fornecedor_toggle(id):
    admin_id, err = _require_admin()
    if err:
        return err
    cat = CategoriaFornecedor.query.filter_by(id=id, admin_id=admin_id).first_or_404()
    cat.ativo = not cat.ativo
    db.session.commit()
    estado = 'ativada' if cat.ativo else 'desativada'
    flash(f'Categoria "{cat.nome}" {estado}.', 'success')
    return redirect(url_for('catalogos.categorias_fornecedor'))


@catalogos_bp.route('/categorias-fornecedor/deletar/<int:id>', methods=['POST'])
@login_required
def categorias_fornecedor_deletar(id):
    admin_id, err = _require_admin()
    if err:
        return err
    cat = CategoriaFornecedor.query.filter_by(id=id, admin_id=admin_id).first_or_404()
    from sqlalchemy import text as _text
    em_uso = db.session.execute(
        _text("SELECT 1 FROM fornecedor_categorias WHERE categoria_fornecedor_id = :cid LIMIT 1"),
        {'cid': cat.id}
    ).first()
    if em_uso:
        flash(f'Não é possível excluir: a categoria "{cat.nome}" está vinculada a fornecedores.', 'danger')
        return redirect(url_for('catalogos.categorias_fornecedor'))
    db.session.delete(cat)
    db.session.commit()
    flash(f'Categoria "{cat.nome}" excluída.', 'success')
    return redirect(url_for('catalogos.categorias_fornecedor'))


# ============================================================
# CategoriaReembolso
# ============================================================

@catalogos_bp.route('/categorias-reembolso')
@login_required
def categorias_reembolso():
    admin_id, err = _require_admin()
    if err:
        return err
    cats = CategoriaReembolso.query.filter_by(admin_id=admin_id).order_by(CategoriaReembolso.nome).all()
    return render_template('catalogos/categorias_reembolso.html', categorias=cats)


@catalogos_bp.route('/categorias-reembolso/criar', methods=['GET', 'POST'])
@login_required
def categorias_reembolso_criar():
    admin_id, err = _require_admin()
    if err:
        return err
    if request.method == 'POST':
        nome = request.form.get('nome', '').strip()
        descricao = request.form.get('descricao', '').strip()
        if not nome:
            flash('Nome é obrigatório.', 'danger')
            return redirect(url_for('catalogos.categorias_reembolso_criar'))
        cat = CategoriaReembolso(nome=nome, descricao=descricao or None, ativo=True, admin_id=admin_id)
        db.session.add(cat)
        db.session.commit()
        flash(f'Categoria "{nome}" criada com sucesso!', 'success')
        return redirect(url_for('catalogos.categorias_reembolso'))
    return render_template('catalogos/categorias_reembolso_form.html', categoria=None)


@catalogos_bp.route('/categorias-reembolso/editar/<int:id>', methods=['GET', 'POST'])
@login_required
def categorias_reembolso_editar(id):
    admin_id, err = _require_admin()
    if err:
        return err
    cat = CategoriaReembolso.query.filter_by(id=id, admin_id=admin_id).first_or_404()
    if request.method == 'POST':
        nome = request.form.get('nome', '').strip()
        descricao = request.form.get('descricao', '').strip()
        if not nome:
            flash('Nome é obrigatório.', 'danger')
            return redirect(url_for('catalogos.categorias_reembolso_editar', id=id))
        cat.nome = nome
        cat.descricao = descricao or None
        db.session.commit()
        flash(f'Categoria "{nome}" atualizada com sucesso!', 'success')
        return redirect(url_for('catalogos.categorias_reembolso'))
    return render_template('catalogos/categorias_reembolso_form.html', categoria=cat)


@catalogos_bp.route('/categorias-reembolso/toggle/<int:id>', methods=['POST'])
@login_required
def categorias_reembolso_toggle(id):
    admin_id, err = _require_admin()
    if err:
        return err
    cat = CategoriaReembolso.query.filter_by(id=id, admin_id=admin_id).first_or_404()
    cat.ativo = not cat.ativo
    db.session.commit()
    estado = 'ativada' if cat.ativo else 'desativada'
    flash(f'Categoria "{cat.nome}" {estado}.', 'success')
    return redirect(url_for('catalogos.categorias_reembolso'))


@catalogos_bp.route('/categorias-reembolso/deletar/<int:id>', methods=['POST'])
@login_required
def categorias_reembolso_deletar(id):
    admin_id, err = _require_admin()
    if err:
        return err
    cat = CategoriaReembolso.query.filter_by(id=id, admin_id=admin_id).first_or_404()
    from sqlalchemy import text as _text
    col_exists = db.session.execute(_text("""
        SELECT 1 FROM information_schema.columns
        WHERE table_name = 'reembolso_funcionario'
          AND column_name = 'categoria_reembolso_id'
        LIMIT 1
    """)).first()
    if col_exists:
        em_uso = db.session.execute(
            _text("SELECT 1 FROM reembolso_funcionario WHERE categoria_reembolso_id = :cid LIMIT 1"),
            {'cid': cat.id}
        ).first()
        if em_uso:
            flash(f'Não é possível excluir: a categoria "{cat.nome}" está vinculada a reembolsos existentes.', 'danger')
            return redirect(url_for('catalogos.categorias_reembolso'))
    db.session.delete(cat)
    db.session.commit()
    flash(f'Categoria "{cat.nome}" excluída.', 'success')
    return redirect(url_for('catalogos.categorias_reembolso'))


# ============================================================
# GrupoFinanceiro
# ============================================================

@catalogos_bp.route('/grupos-financeiros')
@login_required
def grupos_financeiros():
    admin_id, err = _require_admin()
    if err:
        return err
    grupos = GrupoFinanceiro.query.filter_by(admin_id=admin_id).order_by(GrupoFinanceiro.tipo, GrupoFinanceiro.nome).all()
    return render_template('catalogos/grupos_financeiros.html', grupos=grupos)


@catalogos_bp.route('/grupos-financeiros/api')
@login_required
def grupos_financeiros_api():
    admin_id, err = _require_admin()
    if err:
        return jsonify([])
    tipo = request.args.get('tipo', '').strip().upper()
    q = GrupoFinanceiro.query.filter_by(admin_id=admin_id, ativo=True)
    if tipo in ('ENTRADA', 'SAIDA'):
        q = q.filter_by(tipo=tipo)
    grupos = q.order_by(GrupoFinanceiro.nome).all()
    return jsonify([{'id': g.id, 'nome': g.nome, 'tipo': g.tipo} for g in grupos])


@catalogos_bp.route('/grupos-financeiros/criar', methods=['GET', 'POST'])
@login_required
def grupos_financeiros_criar():
    admin_id, err = _require_admin()
    if err:
        return err
    if request.method == 'POST':
        nome = request.form.get('nome', '').strip()
        tipo = request.form.get('tipo', 'SAIDA').strip()
        descricao = request.form.get('descricao', '').strip()
        if not nome:
            flash('Nome é obrigatório.', 'danger')
            return redirect(url_for('catalogos.grupos_financeiros_criar'))
        if tipo not in ('ENTRADA', 'SAIDA'):
            tipo = 'SAIDA'
        existente = GrupoFinanceiro.query.filter(
            GrupoFinanceiro.admin_id == admin_id,
            db.func.lower(GrupoFinanceiro.nome) == nome.lower(),
            GrupoFinanceiro.tipo == tipo,
        ).first()
        if existente:
            flash(f'Já existe um grupo "{nome}" do tipo {tipo}.', 'warning')
            return redirect(url_for('catalogos.grupos_financeiros_criar'))
        g = GrupoFinanceiro(nome=nome, tipo=tipo, descricao=descricao or None, ativo=True, admin_id=admin_id)
        db.session.add(g)
        db.session.commit()
        flash(f'Grupo financeiro "{nome}" criado com sucesso!', 'success')
        return redirect(url_for('catalogos.grupos_financeiros'))
    return render_template('catalogos/grupos_financeiros_form.html', grupo=None)


@catalogos_bp.route('/grupos-financeiros/editar/<int:id>', methods=['GET', 'POST'])
@login_required
def grupos_financeiros_editar(id):
    admin_id, err = _require_admin()
    if err:
        return err
    grupo = GrupoFinanceiro.query.filter_by(id=id, admin_id=admin_id).first_or_404()
    if request.method == 'POST':
        nome = request.form.get('nome', '').strip()
        tipo = request.form.get('tipo', 'SAIDA').strip()
        descricao = request.form.get('descricao', '').strip()
        if not nome:
            flash('Nome é obrigatório.', 'danger')
            return redirect(url_for('catalogos.grupos_financeiros_editar', id=id))
        if tipo not in ('ENTRADA', 'SAIDA'):
            tipo = grupo.tipo
        duplicado = GrupoFinanceiro.query.filter(
            GrupoFinanceiro.admin_id == admin_id,
            GrupoFinanceiro.id != id,
            db.func.lower(GrupoFinanceiro.nome) == nome.lower(),
            GrupoFinanceiro.tipo == tipo,
        ).first()
        if duplicado:
            flash(f'Já existe um grupo "{nome}" do tipo {tipo}.', 'warning')
            return redirect(url_for('catalogos.grupos_financeiros_editar', id=id))
        grupo.nome = nome
        grupo.tipo = tipo
        grupo.descricao = descricao or None
        db.session.commit()
        flash(f'Grupo financeiro "{nome}" atualizado com sucesso!', 'success')
        return redirect(url_for('catalogos.grupos_financeiros'))
    return render_template('catalogos/grupos_financeiros_form.html', grupo=grupo)


@catalogos_bp.route('/grupos-financeiros/toggle/<int:id>', methods=['POST'])
@login_required
def grupos_financeiros_toggle(id):
    admin_id, err = _require_admin()
    if err:
        return err
    grupo = GrupoFinanceiro.query.filter_by(id=id, admin_id=admin_id).first_or_404()
    grupo.ativo = not grupo.ativo
    db.session.commit()
    estado = 'ativado' if grupo.ativo else 'desativado'
    flash(f'Grupo financeiro "{grupo.nome}" {estado}.', 'success')
    return redirect(url_for('catalogos.grupos_financeiros'))


@catalogos_bp.route('/grupos-financeiros/deletar/<int:id>', methods=['POST'])
@login_required
def grupos_financeiros_deletar(id):
    admin_id, err = _require_admin()
    if err:
        return err
    grupo = GrupoFinanceiro.query.filter_by(id=id, admin_id=admin_id).first_or_404()
    em_uso = CategoriaFluxoCaixa.query.filter_by(admin_id=admin_id, grupo_financeiro_id=grupo.id).first()
    if em_uso:
        flash(f'Não é possível excluir: o grupo "{grupo.nome}" está em uso por categorias de fluxo de caixa.', 'danger')
        return redirect(url_for('catalogos.grupos_financeiros'))
    db.session.delete(grupo)
    db.session.commit()
    flash(f'Grupo financeiro "{grupo.nome}" excluído.', 'success')
    return redirect(url_for('catalogos.grupos_financeiros'))


# ============================================================
# PalavraChaveCategoria — Regras de Classificação (Fluxo de Caixa)
# ============================================================

_CAMPOS_ALVO = ('qualquer', 'descricao', 'fornecedor', 'plano')


def _norm_kw_csv(texto):
    """Normaliza uma lista de gatilhos separados por vírgula (lower + sem acento),
    preservando ordem e descartando vazios. Devolve string CSV ou None."""
    from services.classificador_cadastro import _norm
    itens = [_norm(p) for p in (texto or '').split(',')]
    itens = [p for p in itens if p]
    return ','.join(itens) if itens else None


def _conflito_regra(admin_id, palavras, prioridade, campo_alvo, tipo,
                    categoria_id, excluir_id=None):
    """Detecta conflito: regra ATIVA do tenant com mesma palavra + mesma
    prioridade + mesma especificidade (campo_alvo) + mesmo tipo, apontando para
    categoria DIFERENTE. Devolve a regra conflitante ou None (§8)."""
    q = PalavraChaveCategoria.query.filter_by(
        admin_id=admin_id, ativo=True, palavras=palavras,
        prioridade=prioridade, campo_alvo=campo_alvo, tipo=tipo)
    if excluir_id:
        q = q.filter(PalavraChaveCategoria.id != excluir_id)
    for r in q.all():
        if r.categoria_fluxo_caixa_id != categoria_id:
            return r
    return None


@catalogos_bp.route('/palavras-chave')
@login_required
def palavras_chave():
    admin_id, err = _require_admin()
    if err:
        return err
    # Filtros: busca (palavra), categoria, origem
    busca = (request.args.get('q') or '').strip().lower()
    f_cat = (request.args.get('categoria') or '').strip()
    f_origem = (request.args.get('origem') or '').strip()

    q = PalavraChaveCategoria.query.filter_by(admin_id=admin_id)
    if f_cat.isdigit():
        q = q.filter_by(categoria_fluxo_caixa_id=int(f_cat))
    if f_origem in ('sistema', 'usuario'):
        q = q.filter_by(origem=f_origem)
    regras = q.order_by(PalavraChaveCategoria.tipo,
                        PalavraChaveCategoria.prioridade).all()
    if busca:
        regras = [r for r in regras if busca in (r.palavras or '').lower()]

    cat_nome = {c.id: c.nome for c in
                CategoriaFluxoCaixa.query.filter_by(admin_id=admin_id).all()}
    categorias = CategoriaFluxoCaixa.query.filter_by(
        admin_id=admin_id, ativo=True).order_by(CategoriaFluxoCaixa.tipo,
                                                 CategoriaFluxoCaixa.nome).all()
    sugestoes = PalavraChaveSugestao.query.filter_by(
        admin_id=admin_id, dismissed=False).order_by(
        PalavraChaveSugestao.ocorrencias.desc()).all()
    return render_template('catalogos/palavras_chave.html',
                           regras=regras, cat_nome=cat_nome,
                           categorias=categorias, sugestoes=sugestoes,
                           filtros={'q': busca, 'categoria': f_cat, 'origem': f_origem})


@catalogos_bp.route('/palavras-chave/criar', methods=['POST'])
@login_required
def palavras_chave_criar():
    admin_id, err = _require_admin()
    if err:
        return err
    palavras = _norm_kw_csv(request.form.get('palavras'))
    if not palavras:
        flash('Informe ao menos um gatilho (palavra).', 'danger')
        return redirect(url_for('catalogos.palavras_chave'))
    try:
        categoria_id = int(request.form.get('categoria_id'))
        prioridade = int(request.form.get('prioridade', 50))
    except (TypeError, ValueError):
        flash('Categoria e prioridade são obrigatórias.', 'danger')
        return redirect(url_for('catalogos.palavras_chave'))

    campo_alvo = request.form.get('campo_alvo', 'qualquer')
    if campo_alvo not in _CAMPOS_ALVO:
        campo_alvo = 'qualquer'
    tipo = request.form.get('tipo', 'SAIDA').upper()
    if tipo not in ('ENTRADA', 'SAIDA'):
        tipo = 'SAIDA'

    cat = CategoriaFluxoCaixa.query.filter_by(id=categoria_id, admin_id=admin_id).first()
    if not cat:
        flash('Categoria não encontrada.', 'danger')
        return redirect(url_for('catalogos.palavras_chave'))

    conflito = _conflito_regra(admin_id, palavras, prioridade, campo_alvo, tipo, categoria_id)
    if conflito:
        flash(f'Conflito: já existe uma regra "{palavras}" (prioridade {prioridade}, '
              f'{campo_alvo}) apontando para outra categoria. Ajuste a prioridade ou '
              f'a palavra antes de salvar.', 'warning')
        return redirect(url_for('catalogos.palavras_chave'))

    excecoes = _norm_kw_csv(request.form.get('excecoes'))
    gatilho_extra = _norm_kw_csv(request.form.get('gatilho_extra'))
    campo_extra = request.form.get('campo_extra', 'qualquer')
    if campo_extra not in _CAMPOS_ALVO:
        campo_extra = 'qualquer'
    condicao_obra = request.form.get('condicao_obra', 'indiferente')
    if condicao_obra not in ('indiferente', 'com_obra', 'sem_obra'):
        condicao_obra = 'indiferente'

    db.session.add(PalavraChaveCategoria(
        admin_id=admin_id, categoria_fluxo_caixa_id=categoria_id, palavras=palavras,
        campo_alvo=campo_alvo, excecoes=excecoes, gatilho_extra=gatilho_extra,
        campo_extra=campo_extra, condicao_obra=condicao_obra, prioridade=prioridade,
        tipo=tipo, origem='usuario', ativo=True))
    db.session.commit()
    flash(f'Regra "{palavras}" → {cat.nome} criada.', 'success')
    return redirect(url_for('catalogos.palavras_chave'))


@catalogos_bp.route('/palavras-chave/<int:id>/editar', methods=['POST'])
@login_required
def palavras_chave_editar(id):
    admin_id, err = _require_admin()
    if err:
        return err
    regra = PalavraChaveCategoria.query.filter_by(id=id, admin_id=admin_id).first_or_404()
    palavras = _norm_kw_csv(request.form.get('palavras')) or regra.palavras
    try:
        categoria_id = int(request.form.get('categoria_id', regra.categoria_fluxo_caixa_id))
        prioridade = int(request.form.get('prioridade', regra.prioridade))
    except (TypeError, ValueError):
        flash('Categoria e prioridade inválidas.', 'danger')
        return redirect(url_for('catalogos.palavras_chave'))
    campo_alvo = request.form.get('campo_alvo', regra.campo_alvo)
    if campo_alvo not in _CAMPOS_ALVO:
        campo_alvo = regra.campo_alvo
    tipo = request.form.get('tipo', regra.tipo).upper()
    if tipo not in ('ENTRADA', 'SAIDA'):
        tipo = regra.tipo

    cat = CategoriaFluxoCaixa.query.filter_by(id=categoria_id, admin_id=admin_id).first()
    if not cat:
        flash('Categoria não encontrada.', 'danger')
        return redirect(url_for('catalogos.palavras_chave'))

    conflito = _conflito_regra(admin_id, palavras, prioridade, campo_alvo, tipo,
                               categoria_id, excluir_id=regra.id)
    if conflito:
        flash(f'Conflito: já existe uma regra "{palavras}" (prioridade {prioridade}) '
              f'para outra categoria. Ajuste antes de salvar.', 'warning')
        return redirect(url_for('catalogos.palavras_chave'))

    regra.palavras = palavras
    regra.categoria_fluxo_caixa_id = categoria_id
    regra.campo_alvo = campo_alvo
    regra.prioridade = prioridade
    regra.tipo = tipo
    regra.excecoes = _norm_kw_csv(request.form.get('excecoes'))
    db.session.commit()
    flash('Regra atualizada.', 'success')
    return redirect(url_for('catalogos.palavras_chave'))


@catalogos_bp.route('/palavras-chave/<int:id>/toggle', methods=['POST'])
@login_required
def palavras_chave_toggle(id):
    admin_id, err = _require_admin()
    if err:
        return err
    regra = PalavraChaveCategoria.query.filter_by(id=id, admin_id=admin_id).first_or_404()
    regra.ativo = not regra.ativo
    db.session.commit()
    flash(f'Regra {"ativada" if regra.ativo else "desativada"}.', 'success')
    return redirect(url_for('catalogos.palavras_chave'))


@catalogos_bp.route('/palavras-chave/<int:id>/excluir', methods=['POST'])
@login_required
def palavras_chave_excluir(id):
    admin_id, err = _require_admin()
    if err:
        return err
    regra = PalavraChaveCategoria.query.filter_by(id=id, admin_id=admin_id).first_or_404()
    db.session.delete(regra)
    db.session.commit()
    flash('Regra excluída.', 'success')
    return redirect(url_for('catalogos.palavras_chave'))


@catalogos_bp.route('/palavras-chave/sugestoes/cadastrar', methods=['POST'])
@login_required
def palavras_chave_sugestoes_cadastrar():
    """Promove Sugestões selecionadas a Regras (lote). Cada termo vira uma Regra
    do usuário (campo_alvo='fornecedor', prioridade 40) para a categoria escolhida."""
    admin_id, err = _require_admin()
    if err:
        return err
    # Form: para cada sugestão marcada, cat_<sugestao_id> = categoria_id escolhida.
    criadas = 0
    for sug in PalavraChaveSugestao.query.filter_by(admin_id=admin_id).all():
        cat_raw = (request.form.get(f'cat_{sug.id}') or '').strip()
        if not cat_raw.isdigit():
            continue
        categoria_id = int(cat_raw)
        cat = CategoriaFluxoCaixa.query.filter_by(id=categoria_id, admin_id=admin_id).first()
        if not cat:
            continue
        palavras = _norm_kw_csv(sug.termo)
        if not palavras:
            continue
        if _conflito_regra(admin_id, palavras, 40, 'fornecedor', sug.tipo, categoria_id):
            continue
        ja = PalavraChaveCategoria.query.filter_by(
            admin_id=admin_id, palavras=palavras, campo_alvo='fornecedor',
            tipo=sug.tipo, origem='usuario').first()
        if ja:
            continue
        db.session.add(PalavraChaveCategoria(
            admin_id=admin_id, categoria_fluxo_caixa_id=categoria_id, palavras=palavras,
            campo_alvo='fornecedor', prioridade=40, tipo=sug.tipo,
            origem='usuario', ativo=True))
        sug.dismissed = True
        criadas += 1
    db.session.commit()
    flash(f'{criadas} regra(s) criada(s) a partir das sugestões.', 'success')
    return redirect(url_for('catalogos.palavras_chave'))
